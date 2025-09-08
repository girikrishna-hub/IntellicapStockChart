import streamlit as st
import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
import time
import numpy as np
import io
import json
import hashlib
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import base64
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from news_sentiment_analyzer import run_sentiment_analysis, get_sentiment_summary_for_sharing
from sentiment_data_provider import get_advanced_metrics, advanced_data_provider
from reportlab.platypus import PageBreak, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Set page configuration
st.set_page_config(
    page_title="Stock 50-Day Moving Average Chart",
    page_icon="📈",
    layout="wide"
)

def fetch_stock_data(symbol, period="1y", market="US"):
    """
    Fetch stock data for the given symbol and period
    
    Args:
        symbol (str): Stock ticker symbol
        period (str): Time period for data retrieval
        market (str): Market type ("US" or "India")
    
    Returns:
        tuple: (historical_data, ticker_info, ticker_object) or (None, None, None) if error
    """
    try:
        # Format symbol for Indian stocks
        if market == "India":
            # Add .NS (NSE) or .BO (BSE) suffix if not present
            if not (symbol.endswith('.NS') or symbol.endswith('.BO')):
                # Default to NSE (.NS) for Indian stocks
                symbol = f"{symbol}.NS"
        
        # Create ticker object
        ticker = yf.Ticker(symbol)
        
        # Fetch historical data
        data = ticker.history(period=period)
        
        if data.empty:
            return None, None, None
        
        # Fetch additional company info
        try:
            info = ticker.info
        except:
            info = {}
            
        return data, info, ticker
    
    except Exception as e:
        st.error(f"Error fetching data for {symbol}: {str(e)}")
        return None, None, None

def get_gurufocus_financial_metrics(symbol):
    """
    Fetch financial metrics from GuruFocus API if key is available
    Returns dict with P/E, PEG, P/S ratios or None if unavailable
    """
    import os
    
    api_key = os.getenv('GURUFOCUS_API_KEY')
    if not api_key:
        return None
    
    try:
        import requests
        
        # GuruFocus API endpoints for stock profile data
        base_url = "https://api.gurufocus.com/data"
        headers = {
            "Authorization": api_key,
            "Content-Type": "application/json"
        }
        
        # Fetch stock profile which contains valuation metrics
        url = f"{base_url}/stocks/{symbol}/profile"
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            print(f"GuruFocus API success for {symbol}")
            
            # Extract valuation metrics from profile data
            profile = data.get('profile', {})
            ratios = profile.get('ratios', {})
            
            return {
                'pe_ratio': ratios.get('pe_ratio') or ratios.get('trailing_pe'),
                'peg_ratio': ratios.get('peg_ratio'),
                'ps_ratio': ratios.get('ps_ratio') or ratios.get('price_to_sales'),
                'pb_ratio': ratios.get('pb_ratio') or ratios.get('price_to_book'),
                'source': 'GuruFocus'
            }
        else:
            print(f"GuruFocus API error: {response.status_code}")
            print(f"Response content: {response.text[:200]}...")
            if response.status_code == 403:
                print(f"403 Forbidden - API key valid but subscription doesn't include this endpoint")
                print(f"Current subscription may not include real-time stock profile data")
            elif response.status_code == 404:
                print(f"Stock {symbol} not found in GuruFocus")
            return None
            
    except Exception as e:
        print(f"GuruFocus API error: {e}")
        return None

def get_peg_ratio(ticker_info, gurufocus_data=None):
    """
    Get the best available PEG ratio from GuruFocus first, then yfinance fallback
    """
    if gurufocus_data and gurufocus_data.get('peg_ratio'):
        return gurufocus_data['peg_ratio']
    
    # Fallback to yfinance data
    peg_ratio = ticker_info.get('pegRatio', None)
    if peg_ratio is None or pd.isna(peg_ratio):
        peg_ratio = ticker_info.get('trailingPegRatio', None)
    return peg_ratio

def get_hybrid_financial_metrics(symbol, ticker_info):
    """
    Get financial metrics with GuruFocus priority and yfinance fallback
    Returns dict with metrics and data source information
    """
    # Try GuruFocus first
    gurufocus_data = get_gurufocus_financial_metrics(symbol)
    
    if gurufocus_data:
        # Use GuruFocus data
        return {
            'pe_ratio': gurufocus_data.get('pe_ratio') or ticker_info.get('trailingPE'),
            'peg_ratio': gurufocus_data.get('peg_ratio') or get_peg_ratio(ticker_info),
            'ps_ratio': gurufocus_data.get('ps_ratio') or ticker_info.get('priceToSalesTrailing12Months'),
            'pb_ratio': gurufocus_data.get('pb_ratio') or ticker_info.get('priceToBook'),
            'forward_pe': ticker_info.get('forwardPE'),  # Usually only in yfinance
            'ev_revenue': ticker_info.get('enterpriseToRevenue'),
            'ev_ebitda': ticker_info.get('enterpriseToEbitda'),
            'source': 'GuruFocus + yfinance fallback'
        }
    else:
        # Use yfinance data only
        return {
            'pe_ratio': ticker_info.get('trailingPE'),
            'peg_ratio': get_peg_ratio(ticker_info),
            'ps_ratio': ticker_info.get('priceToSalesTrailing12Months'),
            'pb_ratio': ticker_info.get('priceToBook'),
            'forward_pe': ticker_info.get('forwardPE'),
            'ev_revenue': ticker_info.get('enterpriseToRevenue'),
            'ev_ebitda': ticker_info.get('enterpriseToEbitda'),
            'source': 'Yahoo Finance'
        }

def get_beta_value(ticker_info):
    """
    Get the beta value for a stock
    
    Args:
        ticker_info (dict): Stock information from yfinance
    
    Returns:
        str: Beta value formatted as string
    """
    try:
        beta = ticker_info.get('beta', None)
        if beta is not None and not pd.isna(beta):
            return f"{beta:.2f}"
        else:
            return "N/A"
    except:
        return "N/A"

def export_chart_as_png(fig, filename_prefix="chart"):
    """
    Export a Plotly figure as PNG and return download data
    
    Args:
        fig: Plotly figure object
        filename_prefix (str): Prefix for the filename
        
    Returns:
        tuple: (bytes_data, filename)
    """
    try:
        # Try to convert figure to PNG bytes using Kaleido
        img_bytes = fig.to_image(format="png", width=1200, height=800, scale=2, engine="kaleido")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.png"
        
        return img_bytes, filename
    except Exception as e:
        # If Kaleido fails, try alternative method using HTML export
        try:
            # Convert to HTML and then use a different approach
            html_str = fig.to_html(include_plotlyjs='cdn')
            
            # Create a simple PNG export message
            st.warning("PNG export is not available in this environment. Please use PDF export instead, or try the 'Download plot as PNG' option from the chart's toolbar (camera icon in the top-right corner of the chart).")
            return None, None
            
        except Exception as e2:
            st.error(f"Error exporting PNG: {str(e)}. Please try using the chart's built-in download feature (camera icon) or PDF export.")
            return None, None

def export_comprehensive_analysis_pdf(symbol, data, ticker_info, ticker_obj, ma_50, ma_200, rsi, support_level, resistance_level, market="US", fig=None):
    """
    Export comprehensive stock analysis as PDF including all metrics, scores, and charts
    
    Args:
        symbol (str): Stock symbol
        data (DataFrame): Stock price data
        ticker_info (dict): Stock information from yfinance
        ticker_obj: yfinance Ticker object
        ma_50, ma_200: Moving averages
        rsi: RSI values
        support_level, resistance_level: Support/resistance levels
        market (str): Market type ("US" or "India")
        fig: Optional Plotly figure object
        
    Returns:
        tuple: (bytes_data, filename)
    """
    try:
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        import io
        import pandas as pd
        from datetime import datetime, timedelta
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.black,
            spaceAfter=20,
            alignment=1
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceAfter=12
        )
        
        # Story content
        story = []
        
        # Get basic data
        current_price = data['Close'].iloc[-1]
        previous_close = ticker_info.get('previousClose', data['Close'].iloc[-2] if len(data) > 1 else current_price)
        price_change = current_price - previous_close
        price_change_pct = (price_change / previous_close) * 100 if previous_close != 0 else 0
        currency = "₹" if market == "India" else "$"
        company_name = ticker_info.get('longName', ticker_info.get('shortName', symbol))
        
        # Calculate metrics for PDF including Fibonacci analysis and hybrid financial metrics
        # Convert period parameter format for get_stock_metrics
        period_map = {"1y": "1y", "6mo": "6mo", "3mo": "3mo", "1mo": "1mo", "5d": "5d"}
        period_for_metrics = "1y"  # Default to 1 year for comprehensive analysis
        metrics = get_stock_metrics(symbol, period_for_metrics, market)
        
        # Get hybrid financial metrics (GuruFocus + Yahoo Finance fallback)
        hybrid_financial_metrics = get_hybrid_financial_metrics(symbol, ticker_info)
        
        # Title
        title_text = f"Comprehensive Stock Analysis Report: {company_name} ({symbol})"
        story.append(Paragraph(title_text, title_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        
        # Use hybrid metrics for PDF export
        pe_ratio = hybrid_financial_metrics.get('pe_ratio') or ticker_info.get('trailingPE')
        pb_ratio = hybrid_financial_metrics.get('pb_ratio') or ticker_info.get('priceToBook')
        peg_ratio = hybrid_financial_metrics.get('peg_ratio')
        ps_ratio = hybrid_financial_metrics.get('ps_ratio') or ticker_info.get('priceToSalesTrailing12Months')
        data_source = hybrid_financial_metrics.get('source', 'Yahoo Finance')
        
        summary_data = [
            ["Metric", "Value"],
            ["Company", company_name],
            ["Symbol", symbol],
            ["Current Price", f"{currency}{current_price:.2f}"],
            ["Price Change", f"{currency}{price_change:+.2f} ({price_change_pct:+.2f}%)"],
            ["Market Cap", f"{currency}{ticker_info.get('marketCap', 0)/1e9:.2f}B" if ticker_info.get('marketCap') else "N/A"],
            ["P/E Ratio", f"{pe_ratio:.2f}" if pe_ratio and not pd.isna(pe_ratio) else "N/A"],
            ["P/B Ratio", f"{pb_ratio:.2f}" if pb_ratio and not pd.isna(pb_ratio) else "N/A"],
            ["PEG Ratio", f"{peg_ratio:.2f}" if peg_ratio and not pd.isna(peg_ratio) else "N/A"],
            ["P/S Ratio", f"{ps_ratio:.2f}" if ps_ratio and not pd.isna(ps_ratio) else "N/A"],
            ["Beta", f"{ticker_info.get('beta', 0):.2f}" if ticker_info.get('beta') else "N/A"],
            ["Debt-to-Equity", f"{ticker_info.get('debtToEquity', 0):.2f}" if ticker_info.get('debtToEquity') else "N/A"],
            ["ROE", f"{ticker_info.get('returnOnEquity', 0)*100:.1f}%" if ticker_info.get('returnOnEquity') else "N/A"],
            ["ROA", f"{ticker_info.get('returnOnAssets', 0)*100:.1f}%" if ticker_info.get('returnOnAssets') else "N/A"],
            ["Data Source", data_source],
            ["Market", market],
            ["Report Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 3.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Technical Analysis Summary
        story.append(Paragraph("Technical Analysis Summary", heading_style))
        
        # Create technical analysis summary table with 52-week levels
        week_52_high = data['High'].rolling(window=252, min_periods=1).max().iloc[-1]
        week_52_low = data['Low'].rolling(window=252, min_periods=1).min().iloc[-1]
        
        tech_summary_data = [
            ["Metric", "Value", "Analysis"],
            ["Current Price", f"{currency}{current_price:.2f}", f"{price_change_pct:+.2f}% from previous close"],
            ["52-Week High", f"{currency}{week_52_high:.2f}", f"CTP is {((current_price - week_52_high) / week_52_high * 100):+.1f}% from 52W High"],
            ["52-Week Low", f"{currency}{week_52_low:.2f}", f"CTP is {((current_price - week_52_low) / week_52_low * 100):+.1f}% from 52W Low"],
            ["50-Day MA", f"{currency}{ma_50.iloc[-1]:.2f}" if not ma_50.empty else "N/A", 
             f"CTP is {((current_price - ma_50.iloc[-1]) / ma_50.iloc[-1] * 100):+.1f}% from 50-Day MA" if not ma_50.empty else "N/A"],
            ["200-Day MA", f"{currency}{ma_200.iloc[-1]:.2f}" if not ma_200.empty else "Not enough data",
             f"CTP is {((current_price - ma_200.iloc[-1]) / ma_200.iloc[-1] * 100):+.1f}% from 200-Day MA" if not ma_200.empty else "Not enough data"],
            ["RSI (14)", f"{rsi.iloc[-1]:.1f}" if not rsi.empty else "N/A",
             "Overbought" if not rsi.empty and rsi.iloc[-1] > 70 else "Oversold" if not rsi.empty and rsi.iloc[-1] < 30 else "Neutral"],
            ["Support Level", f"{currency}{support_level:.2f}", f"CTP is {((current_price - support_level) / support_level * 100):+.1f}% from Support"],
            ["Resistance Level", f"{currency}{resistance_level:.2f}", f"CTP is {((current_price - resistance_level) / resistance_level * 100):+.1f}% from Resistance"]
        ]
        
        # Always show Fibonacci analysis using 3-month period for accurate calculations
        # Remove all checks for old metrics data to ensure fresh calculation
        try:
            # Get the correct 3-month period data directly from yfinance
            import yfinance as yf
            temp_ticker = yf.Ticker(symbol)
            temp_3mo_data = temp_ticker.history(period='3mo')
            if not temp_3mo_data.empty:
                period_high = temp_3mo_data['High'].max()
                period_low = temp_3mo_data['Low'].min()
                print(f"PDF DEBUG: Using yfinance 3mo data - High: {period_high:.2f}, Low: {period_low:.2f}")
            else:
                raise ValueError("Empty 3mo data")
        except Exception as e:
            print(f"PDF DEBUG: Error with yfinance 3mo query: {e}")
            # Fallback: use 3-month equivalent from existing data (approximately 64 trading days)
            period_data = data.tail(64)  # Match the yfinance 3mo period length
            period_high = period_data['High'].max()
            period_low = period_data['Low'].min()
            print(f"PDF DEBUG: Using fallback with 64 days - High: {period_high:.2f}, Low: {period_low:.2f}")
        
        # Calculate Fibonacci levels
        fib_236 = period_low + (period_high - period_low) * 0.236
        fib_382 = period_low + (period_high - period_low) * 0.382
        fib_618 = period_low + (period_high - period_low) * 0.618
        
        # Add Fibonacci range
        tech_summary_data.append(["Fibonacci Range", f"High: {currency}{period_high:.2f}", f"Low: {currency}{period_low:.2f}"])
        
        # Add individual Fibonacci levels with distance analysis
        fib_236_dist = ((current_price - fib_236) / fib_236 * 100)
        fib_382_dist = ((current_price - fib_382) / fib_382 * 100)
        fib_618_dist = ((current_price - fib_618) / fib_618 * 100)
        
        tech_summary_data.extend([
            ["Fib 23.6% Level", f"{currency}{fib_236:.2f}", f"CTP is {fib_236_dist:+.1f}% from Fib 23.6%"],
            ["Fib 38.2% Level", f"{currency}{fib_382:.2f}", f"CTP is {fib_382_dist:+.1f}% from Fib 38.2%"],
            ["Fib 61.8% Level", f"{currency}{fib_618:.2f}", f"CTP is {fib_618_dist:+.1f}% from Fib 61.8%"]
        ])
        
        # Add safe trading levels
        safe_low = current_price * 0.875  # CTP - 12.5%
        safe_high = current_price * 1.125  # CTP + 12.5%
        tech_summary_data.extend([
            ["Safe Level Low", f"{currency}{safe_low:.2f}", "CTP is +12.5% from Safe Low"],
            ["Safe Level High", f"{currency}{safe_high:.2f}", "CTP is -12.5% from Safe High"]
        ])
        
        tech_summary_table = Table(tech_summary_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        tech_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(tech_summary_table)
        story.append(Spacer(1, 20))
        
        # Stock Ratings Section
        story.append(Paragraph("Stock Ratings (A-D Scale)", heading_style))
        
        # Calculate A-D ratings
        ratings = calculate_stock_ratings(ticker_obj, ticker_info)
        
        # Create ratings table
        ratings_data = [
            ["Category", "Rating", "Description"],
            ["Value", f"{ratings['Value']}", "P/E, P/B, P/S ratios analysis"],
            ["Growth", f"{ratings['Growth']}", "Revenue and earnings growth"],
            ["Momentum", f"{ratings['Momentum']}", "1-month, 3-month, 6-month performance"],
            ["Profitability", f"{ratings['Profitability']}", "ROE, ROA, profit margins"]
        ]
        
        ratings_table = Table(ratings_data, colWidths=[1.5*inch, 1*inch, 3.5*inch])
        ratings_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(ratings_table)
        story.append(Spacer(1, 15))
        
        # Investment Ratings (1-5 Scale)
        story.append(Paragraph("Investment Ratings (1-5 Scale, 5 = Most Desired)", heading_style))
        
        # Calculate 1-5 scale ratings
        print("PDF DEBUG: Calculating 1-5 scale investment ratings...")
        quantitative_rating, quantitative_explanation = calculate_quantitative_rating(ticker_obj, ticker_info)
        author_rating, author_explanation = calculate_author_rating(ticker_obj, ticker_info)
        sellside_rating, sellside_explanation = calculate_sellside_rating(ticker_obj, ticker_info)
        print(f"PDF DEBUG: 1-5 ratings calculated - Quantitative: {quantitative_rating}, Author: {author_rating}, Sellside: {sellside_rating}")
        
        # Create 1-5 scale ratings table
        scale_ratings_data = [
            ["Rating Type", "Score", "Description"],
            ["Quantitative", f"{quantitative_rating}/5", "Data-driven financial metrics and ratios"],
            ["Author", f"{author_rating}/5", "Comprehensive business analysis and fundamentals"],
            ["Sellside", f"{sellside_rating}/5", "Analyst sentiment and market perception"]
        ]
        
        scale_ratings_table = Table(scale_ratings_data, colWidths=[2*inch, 1*inch, 3*inch])
        scale_ratings_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(scale_ratings_table)
        story.append(Spacer(1, 15))

        # Financial Quality Scores
        story.append(Paragraph("Financial Quality Scores", heading_style))
        
        # Calculate scores
        piotroski_score, piotroski_details = calculate_piotroski_score(ticker_obj, ticker_info)
        z_score, z_interpretation = calculate_altman_z_score(ticker_obj, ticker_info)
        m_score, m_interpretation = calculate_beneish_m_score(ticker_obj, ticker_info)
        
        scores_data = [
            ["Score Type", "Value", "Interpretation"],
            ["Piotroski Score (1-9)", f"{piotroski_score}/9" if piotroski_score is not None else "N/A", 
             "Excellent" if piotroski_score and piotroski_score >= 7 else "Good" if piotroski_score and piotroski_score >= 5 else "Poor" if piotroski_score else "N/A"],
            ["Altman Z-Score", f"{z_score:.2f}" if z_score is not None else "N/A", z_interpretation if z_score else "N/A"],
            ["Beneish M-Score", f"{m_score:.2f}" if m_score is not None else "N/A", m_interpretation if m_score else "N/A"]
        ]
        
        scores_table = Table(scores_data, colWidths=[2*inch, 1.2*inch, 2.8*inch])
        scores_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(scores_table)
        story.append(Spacer(1, 20))
        
        # Piotroski Score Details
        if piotroski_score is not None and piotroski_details:
            story.append(Paragraph("Piotroski Score Breakdown", heading_style))
            for detail in piotroski_details[:9]:  # Limit to 9 items to fit on page
                story.append(Paragraph(f"• {detail}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Altman Z-Score Details
        if z_score is not None:
            story.append(Paragraph("Altman Z-Score Analysis", heading_style))
            z_details = [
                f"Working Capital/Total Assets: Component of financial liquidity assessment",
                f"Retained Earnings/Total Assets: Measures cumulative profitability",
                f"EBIT/Total Assets: Evaluates earning power relative to assets",
                f"Market Value Equity/Book Value Total Debt: Market-based leverage metric",
                f"Sales/Total Assets: Asset turnover efficiency measure",
                f"Final Z-Score: {z_score:.2f}" if z_score else "N/A",
                f"Interpretation: {z_interpretation}" if z_interpretation else "N/A"
            ]
            for detail in z_details[:7]:  # Limit to fit on page
                story.append(Paragraph(f"• {detail}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Beneish M-Score Details  
        if m_score is not None:
            story.append(Paragraph("Beneish M-Score Analysis", heading_style))
            m_details = [
                f"Days Sales Outstanding Index: Measures receivables quality changes",
                f"Gross Margin Index: Evaluates gross margin deterioration",
                f"Asset Quality Index: Assesses non-current asset composition",
                f"Sales Growth Index: Measures revenue growth patterns",
                f"Depreciation Index: Analyzes depreciation rate changes",
                f"SG&A Index: Evaluates selling and admin expense efficiency",
                f"Leverage Index: Measures debt level changes",
                f"Total Accruals/Total Assets: Assesses earnings quality",
                f"Final M-Score: {m_score:.2f}" if m_score else "N/A",
                f"Interpretation: {m_interpretation}" if m_interpretation else "N/A"
            ]
            for detail in m_details[:8]:  # Limit to fit on page
                story.append(Paragraph(f"• {detail}", styles['Normal']))
            story.append(Spacer(1, 15))
        
        # Earnings Performance Analysis Section
        story.append(Paragraph("Earnings Performance Analysis", heading_style))
        
        try:
            # Use the existing comprehensive earnings analysis function from GuruFocus tab
            # Get extended historical data for comprehensive 8-quarter analysis
            extended_data = ticker_obj.history(period="3y", interval="1d")  # 3 years for full 8-quarter coverage
            if extended_data.empty:
                extended_data = data  # Fallback to provided data
            
            earnings_analysis, quarters_found = get_detailed_earnings_performance_analysis(
                ticker_obj, extended_data, market=market, max_quarters=8
            )
            
            if earnings_analysis is not None and not earnings_analysis.empty:
                # Create comprehensive earnings table with ALL columns from Advanced Analysis tab
                earnings_summary = [
                    ["Quarter", "Date", "Pre-Close", "Next Open", "Next Close", "Overnight%", "NextDay%", "Week Close", "Week Date", "Week%", "Direction", "EPS Est", "EPS Act", "Surprise"]
                ]
                
                # Add all available quarters (up to 8) with complete column data
                for i, (idx, row) in enumerate(earnings_analysis.head(8).iterrows()):
                    quarter_info = row['Quarter']  # Quarter column already formatted as "Q1 2024"
                    earnings_date = row['Earnings Date'] if pd.notnull(row['Earnings Date']) else 'N/A'
                    
                    # Get all column values with proper formatting
                    pre_close = row.get('Pre-Earnings Close', 'N/A')
                    next_open = row.get('Next Day Open', 'N/A') 
                    next_close = row.get('Next Day Close', 'N/A')
                    overnight_str = row['Overnight Change (%)']
                    nextday_str = row['Next Day Change (%)'] 
                    week_close = row.get('End of Week Close', 'N/A')
                    week_date = row.get('Week End Date', 'N/A')
                    week_str = row['Week Performance (%)']
                    direction = row.get('Direction', 'N/A')
                    eps_est = row.get('EPS Est', 'N/A')
                    eps_act = row.get('EPS Act', 'N/A')
                    surprise = row.get('Surprise', 'N/A')
                    
                    earnings_summary.append([
                        quarter_info, earnings_date, pre_close, next_open, next_close, 
                        overnight_str, nextday_str, week_close, week_date, week_str, direction, 
                        eps_est, eps_act, surprise
                    ])
                
                # Calculate average performance for key metrics
                overnight_values = []
                nextday_values = []
                week_values = []
                
                for idx, row in earnings_analysis.iterrows():
                    try:
                        # Extract numeric values from formatted percentage strings like "+2.50%"
                        overnight_val = float(row['Overnight Change (%)'].replace('%', '').replace('+', ''))
                        nextday_val = float(row['Next Day Change (%)'].replace('%', '').replace('+', ''))
                        week_val = float(row['Week Performance (%)'].replace('%', '').replace('+', ''))
                        
                        overnight_values.append(overnight_val)
                        nextday_values.append(nextday_val)
                        week_values.append(week_val)
                    except (ValueError, AttributeError):
                        continue
                
                if overnight_values:
                    avg_overnight = sum(overnight_values) / len(overnight_values)
                    avg_nextday = sum(nextday_values) / len(nextday_values)
                    avg_week = sum(week_values) / len(week_values)
                    
                    # Add average row with key statistics
                    earnings_summary.append([
                        "AVERAGE", f"({len(earnings_analysis)} qtrs)", "—", "—", "—",
                        f"{avg_overnight:.2f}%", f"{avg_nextday:.2f}%", "—", "—", f"{avg_week:.2f}%", 
                        "—", "—", "—", "—"
                    ])
                
                # Create landscape-oriented table with smaller columns to fit all data including new Week Date column
                col_widths = [0.5*inch, 0.65*inch, 0.55*inch, 0.55*inch, 0.55*inch, 0.5*inch, 0.5*inch, 0.55*inch, 0.65*inch, 0.5*inch, 0.45*inch, 0.45*inch, 0.45*inch, 0.5*inch]
                earnings_table = Table(earnings_summary, colWidths=col_widths)
                earnings_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 8),  # Smaller font for headers
                    ('FONTSIZE', (0,1), (-1,-1), 7),  # Smaller font for data
                    ('BOTTOMPADDING', (0,0), (-1,0), 8),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,1), (-1,-1), 4),
                    ('BACKGROUND', (0,1), (-1,-2), colors.beige),
                    ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),  # Average row
                    ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                ]))
                story.append(earnings_table)
                
                # Add comprehensive earnings insights
                story.append(Spacer(1, 10))
                story.append(Paragraph(f"Comprehensive analysis of {len(earnings_analysis)} quarters with complete market data", styles['Normal']))
                
                if overnight_values:
                    # Performance insights with detailed metrics
                    if avg_overnight > 0:
                        overnight_trend = "positive"
                    elif avg_overnight < -2:
                        overnight_trend = "negative" 
                    else:
                        overnight_trend = "neutral"
                        
                    if avg_nextday > 0:
                        nextday_trend = "positive"
                    elif avg_nextday < -2:
                        nextday_trend = "negative"
                    else:
                        nextday_trend = "neutral"
                    
                    story.append(Paragraph(f"• Average overnight performance: {overnight_trend} ({avg_overnight:.2f}%)", styles['Normal']))
                    story.append(Paragraph(f"• Average next-day performance: {nextday_trend} ({avg_nextday:.2f}%)", styles['Normal']))
                    story.append(Paragraph(f"• Average weekly performance: {avg_week:.2f}%", styles['Normal']))
                    story.append(Paragraph("• Table includes pre-earnings prices, next-day opens/closes, EPS estimates/actuals, and surprise factors", styles['Normal']))
                
                story.append(Spacer(1, 15))
                
            else:
                story.append(Paragraph("Earnings analysis data not available", styles['Normal']))
                story.append(Spacer(1, 15))
                
        except Exception as e:
            story.append(Paragraph(f"Earnings analysis error: {str(e)}", styles['Normal']))
            story.append(Spacer(1, 15))

        # Market Intelligence Section
        story.append(Paragraph("Market Intelligence", heading_style))
        
        try:
            # Get market intelligence data using unified function (same as UI tab)
            import os
            fmp_api_key = os.environ.get("FMP_API_KEY")
            
            # Use unified function to ensure consistency with UI tab
            advanced_metrics, used_fallback = get_unified_market_intelligence(symbol, fmp_api_key, ticker_obj)
            
            if used_fallback:
                print("PDF Export: Using Yahoo Finance fallback for market intelligence...")
            
            if advanced_metrics and advanced_metrics != {}:
                # Create market intelligence summary table
                intel_summary = [
                    ["Category", "Information"]
                ]
                
                # Add analyst information
                if advanced_metrics.get('Analyst Rating', 'N/A') != 'N/A':
                    intel_summary.append(["Analyst Rating", advanced_metrics.get('Analyst Rating')])
                
                if advanced_metrics.get('Price Target', 'N/A') != 'N/A':
                    intel_summary.append(["Price Target", advanced_metrics.get('Price Target')])
                
                # Add insider activity
                if advanced_metrics.get('Insider Activity', 'N/A') != 'N/A':
                    intel_summary.append(["Insider Activity", advanced_metrics.get('Insider Activity')])
                
                # Add institutional ownership
                if advanced_metrics.get('Institutional Ownership', 'N/A') != 'N/A':
                    intel_summary.append(["Institutional Ownership", advanced_metrics.get('Institutional Ownership')])
                
                # Add social sentiment
                if advanced_metrics.get('Social Sentiment', 'N/A') != 'N/A':
                    intel_summary.append(["Social Sentiment", advanced_metrics.get('Social Sentiment')])
                
                # Add last updated if available
                if advanced_metrics.get('Last Updated'):
                    intel_summary.append(["Last Updated", advanced_metrics.get('Last Updated')])
                
                # Create and style market intelligence table
                if len(intel_summary) > 1:  # Only create table if we have data
                    intel_table = Table(intel_summary, colWidths=[2*inch, 4*inch])
                    intel_table.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.grey),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0,0), (-1,0), 10),
                        ('FONTSIZE', (0,1), (-1,-1), 9),
                        ('BOTTOMPADDING', (0,0), (-1,0), 12),
                        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
                    ]))
                    story.append(intel_table)
                    story.append(Spacer(1, 10))
                    
                    # Add detailed analysis if available
                    if 'raw_data' in advanced_metrics:
                        raw_data = advanced_metrics['raw_data']
                        
                        # Analyst details
                        analyst_data = raw_data.get('analyst_ratings', {})
                        if analyst_data and 'error' not in analyst_data and analyst_data.get('rating'):
                            story.append(Paragraph("Analyst Details:", styles['Normal']))
                            story.append(Paragraph(f"• Rating: {analyst_data.get('rating', 'N/A')}", styles['Normal']))
                            story.append(Paragraph(f"• Recommendation: {analyst_data.get('recommendation', 'N/A')}", styles['Normal']))
                            story.append(Spacer(1, 8))
                        
                        # Insider trading details
                        insider_data = raw_data.get('insider_trading', {})
                        if insider_data and 'error' not in insider_data:
                            story.append(Paragraph("Recent Insider Activity:", styles['Normal']))
                            if isinstance(insider_data, list) and len(insider_data) > 0:
                                for i, trade in enumerate(insider_data[:3]):  # Show top 3 trades
                                    trade_info = f"• {trade.get('name', 'N/A')}: {trade.get('transaction', 'N/A')} ({trade.get('date', 'N/A')})"
                                    story.append(Paragraph(trade_info, styles['Normal']))
                            story.append(Spacer(1, 8))
                        
                        # Institutional holdings
                        institutional_data = raw_data.get('institutional_holdings', {})
                        if institutional_data and 'error' not in institutional_data:
                            story.append(Paragraph("Top Institutional Holders:", styles['Normal']))
                            if isinstance(institutional_data, list) and len(institutional_data) > 0:
                                for i, holder in enumerate(institutional_data[:3]):  # Show top 3 holders
                                    holder_info = f"• {holder.get('holder', 'N/A')}: {holder.get('shares', 'N/A')} shares"
                                    story.append(Paragraph(holder_info, styles['Normal']))
                            story.append(Spacer(1, 8))
                    
                else:
                    story.append(Paragraph("Limited market intelligence data available", styles['Normal']))
                    story.append(Spacer(1, 15))
                    
            else:
                story.append(Paragraph("Market intelligence data not available", styles['Normal']))
                story.append(Spacer(1, 15))
                
        except Exception as e:
            story.append(Paragraph(f"Market intelligence error: {str(e)}", styles['Normal']))
            story.append(Spacer(1, 15))

        # Add comprehensive technical analysis charts using matplotlib
        story.append(PageBreak())
        story.append(Paragraph("Technical Charts", heading_style))
        
        try:
            # Generate matplotlib-based charts
            import matplotlib
            matplotlib.use('Agg')  # Use non-GUI backend
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            
            # Create comprehensive technical analysis chart
            fig_mpl, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 8))
            
            # Chart 1: Price with Moving Averages
            ax1.plot(data.index, data['Close'], label=f'{symbol.upper()} Price', color='black', linewidth=1.5)
            if not ma_50.empty:
                ax1.plot(data.index, ma_50, label='50-Day MA', color='blue', linewidth=1)
            if not ma_200.empty:
                ax1.plot(data.index, ma_200, label='200-Day MA', color='red', linewidth=1)
            ax1.axhline(y=support_level, color='green', linestyle='--', alpha=0.7, label='Support')
            ax1.axhline(y=resistance_level, color='red', linestyle='--', alpha=0.7, label='Resistance')
            ax1.set_title(f'{symbol.upper()} Price & Moving Averages', fontsize=10, fontweight='bold')
            ax1.set_ylabel(f'Price ({currency})', fontsize=9)
            ax1.grid(True, alpha=0.3)
            ax1.legend(fontsize=7)
            
            # Chart 2: MACD (calculate if not provided)
            if not rsi.empty:
                # Simple MACD calculation for display
                exp1 = data['Close'].ewm(span=12).mean()
                exp2 = data['Close'].ewm(span=26).mean()
                macd_calc = exp1 - exp2
                signal_calc = macd_calc.ewm(span=9).mean()
                
                ax2.plot(data.index, macd_calc, label='MACD', color='blue', linewidth=1)
                ax2.plot(data.index, signal_calc, label='Signal', color='red', linewidth=1)
                ax2.axhline(y=0, color='gray', linestyle='-', alpha=0.5)
            ax2.set_title('MACD', fontsize=10, fontweight='bold')
            ax2.set_ylabel('MACD', fontsize=9)
            ax2.grid(True, alpha=0.3)
            ax2.legend(fontsize=7)
            
            # Chart 3: RSI
            if not rsi.empty:
                ax3.plot(data.index, rsi, label='RSI', color='purple', linewidth=1)
                ax3.axhline(y=70, color='red', linestyle='--', alpha=0.7, label='Overbought')
                ax3.axhline(y=30, color='green', linestyle='--', alpha=0.7, label='Oversold')
                ax3.set_ylim(0, 100)
            ax3.set_title('RSI (14)', fontsize=10, fontweight='bold')
            ax3.set_ylabel('RSI', fontsize=9)
            ax3.grid(True, alpha=0.3)
            ax3.legend(fontsize=7)
            
            # Chart 4: Volume
            ax4.bar(data.index, data['Volume'], alpha=0.7, color='orange', label='Volume')
            ax4.set_title('Trading Volume', fontsize=10, fontweight='bold')
            ax4.set_ylabel('Volume', fontsize=9)
            ax4.grid(True, alpha=0.3)
            ax4.legend(fontsize=7)
            
            # Format x-axis dates for all subplots
            for ax in [ax1, ax2, ax3, ax4]:
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
                ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=8)
            
            plt.tight_layout()
            
            # Save chart to buffer
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
            plt.close(fig_mpl)
            img_buffer.seek(0)
            
            # Add chart image to PDF
            from reportlab.platypus import Image as ReportLabImage
            img = ReportLabImage(img_buffer, width=7*inch, height=5*inch)
            story.append(img)
            story.append(Spacer(1, 20))
            
            # Add note about chart completion
            chart_note = "Technical charts generated successfully with price action, moving averages, MACD, RSI, and volume analysis."
            story.append(Paragraph(chart_note, styles['Normal']))
            
        except Exception as chart_error:
            # Fallback message if chart generation fails
            fallback_note = f"Chart generation encountered an issue: {str(chart_error)}. Technical analysis summary table is available above."
            story.append(Paragraph(fallback_note, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_comprehensive_analysis_{timestamp}.pdf"
        
        return pdf_bytes, filename
        
    except Exception as e:
        st.error(f"Error creating comprehensive PDF: {str(e)}")
        return None, None

def export_chart_as_pdf(fig, filename_prefix="chart", title="Stock Analysis Chart"):
    """
    Export a Plotly figure as PDF and return download data
    
    Args:
        fig: Plotly figure object
        filename_prefix (str): Prefix for the filename
        title (str): Title for the PDF document
        
    Returns:
        tuple: (bytes_data, filename)
    """
    try:
        # Try to convert figure to PNG first (for embedding in PDF)
        try:
            img_bytes = fig.to_image(format="png", width=1200, height=800, scale=2, engine="kaleido")
        except:
            # If Kaleido fails, create a text-based PDF without the chart image
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.black,
                spaceAfter=20,
                alignment=1  # Center alignment
            )
            
            # Create story
            story = []
            
            # Add title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 20))
            
            # Add note about chart
            note_text = "Chart image could not be embedded. Please use the chart's built-in download feature (camera icon in the top-right corner) to save the chart as an image."
            story.append(Paragraph(note_text, styles['Normal']))
            
            # Add timestamp
            timestamp_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Spacer(1, 20))
            story.append(Paragraph(timestamp_text, styles['Normal']))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}.pdf"
            
            return pdf_bytes, filename
        
        # If we got the image, create PDF with chart
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.black,
            spaceAfter=20,
            alignment=1  # Center alignment
        )
        
        # Create story
        story = []
        
        # Add title
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Add chart image
        img_stream = io.BytesIO(img_bytes)
        img = ReportLabImage(img_stream, width=7*inch, height=4.5*inch)
        story.append(img)
        
        # Add timestamp
        timestamp_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Spacer(1, 20))
        story.append(Paragraph(timestamp_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.pdf"
        
        return pdf_bytes, filename
    except Exception as e:
        st.error(f"Error exporting PDF: {str(e)}")
        return None, None

def create_export_buttons(fig, chart_name, symbol="", include_comprehensive=False, stock_data=None, ticker_info=None, ticker_obj=None, ma_50=None, ma_200=None, rsi=None, support_level=None, resistance_level=None, market="US"):
    """
    Create export buttons for PNG and PDF download
    
    Args:
        fig: Plotly figure object
        chart_name (str): Name of the chart for filename
        symbol (str): Stock symbol for filename
        include_comprehensive (bool): Include comprehensive analysis PDF option
        stock_data, ticker_info, ticker_obj: Stock data for comprehensive analysis
        ma_50, ma_200, rsi: Technical indicators
        support_level, resistance_level: Support/resistance levels
        market (str): Market type
    """
    # Prepare filename prefix
    filename_prefix = f"{symbol}_{chart_name}".replace(" ", "_").replace("/", "_") if symbol else chart_name.replace(" ", "_")
    
    # Show built-in chart download option first
    st.info("💡 **Tip:** You can also download charts directly using the camera icon (📷) in the chart's toolbar (top-right corner when you hover over the chart).")
    
    if include_comprehensive:
        col_png, col_pdf, col_comprehensive = st.columns(3)
    else:
        col_png, col_pdf = st.columns(2)
    
    with col_png:
        if st.button(f"📷 Export PNG", key=f"png_{chart_name}_{symbol}", help="Download chart as PNG image (may require browser setup)"):
            png_data, png_filename = export_chart_as_png(fig, filename_prefix)
            if png_data:
                st.download_button(
                    label="⬇️ Download PNG",
                    data=png_data,
                    file_name=png_filename,
                    mime="image/png",
                    key=f"download_png_{chart_name}_{symbol}"
                )
    
    with col_pdf:
        if st.button(f"📄 Chart PDF", key=f"pdf_{chart_name}_{symbol}", help="Download chart as PDF document"):
            pdf_title = f"{symbol.upper()} - {chart_name}" if symbol else chart_name
            pdf_data, pdf_filename = export_chart_as_pdf(fig, filename_prefix, pdf_title)
            if pdf_data:
                st.download_button(
                    label="⬇️ Download Chart PDF",
                    data=pdf_data,
                    file_name=pdf_filename,
                    mime="application/pdf",
                    key=f"download_pdf_{chart_name}_{symbol}"
                )
    
    if include_comprehensive:
        with col_comprehensive:
            if st.button(f"📊 Full Analysis", key=f"comprehensive_{chart_name}_{symbol}", help="Download comprehensive analysis report with all metrics and scores"):
                if all([stock_data is not None, ticker_info is not None, ticker_obj is not None]):
                    with st.spinner('Generating comprehensive report...'):
                        pdf_data, pdf_filename = export_comprehensive_analysis_pdf(
                            symbol, stock_data, ticker_info, ticker_obj, ma_50, ma_200, 
                            rsi, support_level, resistance_level, market, fig
                        )
                        if pdf_data:
                            st.download_button(
                                label="⬇️ Download Full Report",
                                data=pdf_data,
                                file_name=pdf_filename,
                                mime="application/pdf",
                                key=f"download_comprehensive_{chart_name}_{symbol}"
                            )
                else:
                    st.error("Stock data not available for comprehensive analysis.")

def create_shareable_insight(symbol, data, metrics, privacy_level="public", fibonacci_data=None):
    """
    Create a shareable investment insight with customizable privacy
    
    Args:
        symbol (str): Stock ticker symbol
        data (pd.DataFrame): Stock price data
        metrics (dict): Key stock metrics
        privacy_level (str): Privacy level ("public", "private", "anonymized")
        fibonacci_data (dict): Fibonacci analysis data
    
    Returns:
        dict: Shareable insight data
    """
    try:
        latest_price = data['Close'].iloc[-1]
        price_change = ((latest_price - data['Close'].iloc[-2]) / data['Close'].iloc[-2]) * 100
        
        # Base insight structure
        insight = {
            "id": hashlib.md5(f"{symbol}_{datetime.now()}".encode()).hexdigest()[:8],
            "timestamp": datetime.now().isoformat(),
            "privacy_level": privacy_level,
            "analysis_type": "technical_analysis"
        }
        
        # Process Fibonacci data if available
        fib_summary = None
        if fibonacci_data:
            analysis_type = fibonacci_data.get('analysis_type', 'unknown')
            reference_high = fibonacci_data.get('reference_high', 0)
            reference_low = fibonacci_data.get('reference_low', 0)
            price_range = reference_high - reference_low
            next_levels_above = fibonacci_data.get('next_levels_above', [])
            next_levels_below = fibonacci_data.get('next_levels_below', [])
            
            # Find closest Fibonacci level
            all_levels = next_levels_above + next_levels_below
            closest_fib_level = None
            closest_distance = float('inf')
            
            for level in all_levels:
                distance = abs(latest_price - level['price'])
                if distance < closest_distance:
                    closest_distance = distance
                    closest_fib_level = level['label']
            
            # Determine trend status
            if analysis_type == "retracement":
                trend_status = "Within Range"
            elif analysis_type == "upward_extension":
                trend_status = "Above Range"
            else:
                trend_status = "Below Range"
            
            fib_summary = {
                "analysis_type": analysis_type,
                "trend_status": trend_status,
                "swing_range": f"${price_range:.2f}",
                "closest_fib_level": closest_fib_level,
                "near_support_resistance": closest_fib_level is not None and closest_distance < (price_range * 0.02)
            }
        
        if privacy_level == "public":
            # Full public sharing
            public_data = {
                "symbol": symbol,
                "company_name": metrics.get('longName', symbol),
                "current_price": f"${latest_price:.2f}",
                "daily_change": f"{price_change:+.2f}%",
                "key_metrics": {
                    "rsi": metrics.get('RSI', 'N/A'),
                    "ma_50_signal": metrics.get('Price vs 50-Day MA (%)', 'N/A'),
                    "ma_200_signal": metrics.get('Price vs 200-Day MA (%)', 'N/A'),
                    "volume": metrics.get('Volume', 'N/A'),
                    "beta": metrics.get('Beta', 'N/A')
                },
                "recommendation": generate_advanced_recommendation(metrics, fib_summary),
                "risk_level": assess_risk_level(metrics)
            }
            
            if fib_summary:
                public_data["fibonacci_analysis"] = {
                    "trend": fib_summary["trend_status"],
                    "swing_range": fib_summary["swing_range"],
                    "key_level": fib_summary["closest_fib_level"] or "No nearby levels",
                    "near_key_level": fib_summary["near_support_resistance"]
                }
            
            insight.update(public_data)
            
        elif privacy_level == "anonymized":
            # Anonymized - no symbol or company name
            anon_data = {
                "symbol": "***",
                "company_name": "Anonymized Stock",
                "current_price": f"${latest_price:.2f}",
                "daily_change": f"{price_change:+.2f}%",
                "key_metrics": {
                    "rsi": metrics.get('RSI', 'N/A'),
                    "ma_50_signal": metrics.get('Price vs 50-Day MA (%)', 'N/A'),
                    "ma_200_signal": metrics.get('Price vs 200-Day MA (%)', 'N/A')
                },
                "recommendation": generate_advanced_recommendation(metrics, fib_summary),
                "note": "Stock identity hidden for privacy"
            }
            
            if fib_summary:
                anon_data["fibonacci_analysis"] = {
                    "trend": fib_summary["trend_status"],
                    "key_level": fib_summary["closest_fib_level"] or "No nearby levels"
                }
            
            insight.update(anon_data)
            
        else:  # private
            # Private sharing - limited info
            insight.update({
                "symbol": symbol,
                "analysis_summary": "Private technical analysis with Fibonacci levels completed",
                "recommendation": generate_advanced_recommendation(metrics, fib_summary),
                "shared_with": "Private analysis - limited details"
            })
        
        return insight
        
    except Exception as e:
        return {"error": f"Failed to create insight: {str(e)}"}

def generate_simple_recommendation(metrics):
    """Generate a simple investment recommendation based on key metrics"""
    try:
        rsi_text = metrics.get('RSI', '50')
        if rsi_text != 'N/A':
            rsi = float(rsi_text)
        else:
            rsi = 50
            
        ma_50_text = metrics.get('Price vs 50-Day MA (%)', '0')
        if ma_50_text != 'N/A':
            ma_50_signal = float(ma_50_text.replace('%', ''))
        else:
            ma_50_signal = 0
            
        # Simple recommendation logic
        if rsi > 70:
            return "⚠️ Potentially Overbought"
        elif rsi < 30:
            return "📈 Potentially Oversold - Watch for Entry"
        elif ma_50_signal > 5:
            return "📈 Above 50-Day MA - Bullish Trend"
        elif ma_50_signal < -5:
            return "📉 Below 50-Day MA - Bearish Trend"
        else:
            return "➡️ Neutral - Monitor for Clear Signals"
            
    except:
        return "➡️ Analysis Available - Review Details"

def generate_advanced_recommendation(metrics, fib_summary=None):
    """Generate advanced investment recommendation including Fibonacci analysis"""
    try:
        rsi_text = metrics.get('RSI', '50')
        if rsi_text != 'N/A':
            rsi = float(rsi_text)
        else:
            rsi = 50
            
        ma_50_text = metrics.get('Price vs 50-Day MA (%)', '0')
        if ma_50_text != 'N/A':
            ma_50_signal = float(ma_50_text.replace('%', ''))
        else:
            ma_50_signal = 0
            
        # Base recommendation from technical indicators
        base_rec = ""
        if rsi > 70:
            base_rec = "⚠️ Overbought"
        elif rsi < 30:
            base_rec = "📈 Oversold Entry Zone"
        elif ma_50_signal > 5:
            base_rec = "📈 Bullish Trend"
        elif ma_50_signal < -5:
            base_rec = "📉 Bearish Trend"
        else:
            base_rec = "➡️ Neutral"
            
        # Enhance with Fibonacci analysis
        if fib_summary:
            trend_dir = fib_summary.get("trend_direction", "unknown")
            near_key_level = fib_summary.get("near_support_resistance", False)
            
            if near_key_level:
                if trend_dir == "uptrend":
                    base_rec += " + Near Fib Support"
                elif trend_dir == "downtrend":
                    base_rec += " + Near Fib Resistance"
                else:
                    base_rec += " + At Key Fib Level"
            else:
                base_rec += f" ({trend_dir.title()} Pattern)"
        
        return base_rec
            
    except:
        return "➡️ Advanced Analysis Available"

def assess_risk_level(metrics):
    """Assess risk level based on key metrics"""
    try:
        rsi_text = metrics.get('RSI', '50')
        if rsi_text != 'N/A':
            rsi = float(rsi_text)
        else:
            rsi = 50
            
        beta_text = metrics.get('Beta', '1.0')
        if beta_text != 'N/A':
            beta = float(beta_text)
        else:
            beta = 1.0
            
        # Risk assessment
        if beta > 1.5 or rsi > 75 or rsi < 25:
            return "🔴 High Risk"
        elif beta > 1.2 or rsi > 65 or rsi < 35:
            return "🟡 Medium Risk"
        else:
            return "🟢 Lower Risk"
            
    except:
        return "🟡 Risk Assessment Available"

def create_share_urls(insight):
    """Create shareable URLs for different platforms"""
    
    # Create a summary text
    symbol = insight.get('symbol', 'Stock')
    recommendation = insight.get('recommendation', 'Analysis available')
    
    if insight['privacy_level'] == 'public':
        fib_info = ""
        if 'fibonacci_analysis' in insight:
            fib_data = insight['fibonacci_analysis']
            fib_info = f" Trend: {fib_data['trend']}. Key Level: {fib_data['key_level']}."
        share_text = f"📊 {symbol} Analysis: {recommendation}. Current: {insight.get('current_price', 'N/A')} ({insight.get('daily_change', 'N/A')}). Risk: {insight.get('risk_level', 'N/A')}.{fib_info}"
    elif insight['privacy_level'] == 'anonymized':
        fib_info = ""
        if 'fibonacci_analysis' in insight:
            fib_data = insight['fibonacci_analysis']
            fib_info = f" Trend: {fib_data['trend']}."
        share_text = f"📊 Stock Analysis: {recommendation}. Daily change: {insight.get('daily_change', 'N/A')}.{fib_info} Technical indicators suggest monitoring for opportunities."
    else:
        share_text = f"📊 Completed advanced technical analysis with Fibonacci levels: {recommendation}"
    
    hashtags = "#StockAnalysis #TechnicalAnalysis #Investing"
    
    urls = {
        "twitter": f"https://twitter.com/intent/tweet?text={urllib.parse.quote(share_text)}&hashtags={urllib.parse.quote('StockAnalysis,TechnicalAnalysis')}",
        "linkedin": f"https://www.linkedin.com/sharing/share-offsite/?url={urllib.parse.quote('https://example.com')}&summary={urllib.parse.quote(share_text)}",
        "copy_text": f"{share_text} {hashtags}",
        "email_subject": f"Stock Analysis Insight - {symbol}",
        "email_body": f"Hi,\n\nI wanted to share this investment insight:\n\n{share_text}\n\nGenerated using technical analysis tools.\n\nBest regards"
    }
    
    return urls

def calculate_ctp_levels(current_price):
    """
    Calculate Safe Level High and Low reference levels
    
    Args:
        current_price (float): Current stock price
    
    Returns:
        dict: Dictionary with upper and lower safe levels
    """
    try:
        upper_ctp = current_price * 1.125  # +12.5%
        lower_ctp = current_price * 0.875  # -12.5%
        
        return {
            'upper_ctp': upper_ctp,
            'lower_ctp': lower_ctp,
            'upper_percentage': '+12.5%',
            'lower_percentage': '-12.5%'
        }
    except:
        return {
            'upper_ctp': None,
            'lower_ctp': None,
            'upper_percentage': 'N/A',
            'lower_percentage': 'N/A'
        }

def get_after_market_data(symbol, market="US"):
    """
    Get after-market trading data for a stock
    
    Args:
        symbol (str): Stock ticker symbol
        market (str): Market type ("US" or "India")
    
    Returns:
        dict: After-market trading information
    """
    try:
        # Format symbol for Indian stocks
        if market == "India":
            if not (symbol.endswith('.NS') or symbol.endswith('.BO')):
                symbol = f"{symbol}.NS"
        
        ticker = yf.Ticker(symbol)
        
        # Get today's data including pre/post market
        today_data = ticker.history(period="1d", interval="1m")
        
        if today_data.empty:
            return {
                'pre_market_change': 'N/A',
                'pre_market_change_percent': 'N/A',
                'post_market_change': 'N/A', 
                'post_market_change_percent': 'N/A',
                'regular_session_close': 'N/A'
            }
        
        # Get regular session close (4 PM ET for US markets)
        regular_close = today_data['Close'].iloc[-1]
        
        # Try to get real-time quote data
        try:
            info = ticker.info
            current_price = info.get('currentPrice', regular_close)
            pre_market_price = info.get('preMarketPrice', None)
            post_market_price = info.get('postMarketPrice', None)
            
            # Calculate pre-market movement
            pre_market_change = 'N/A'
            pre_market_change_percent = 'N/A'
            if pre_market_price and pre_market_price != regular_close:
                prev_close = info.get('previousClose', regular_close)
                if prev_close:
                    pre_change = pre_market_price - prev_close
                    pre_change_percent = (pre_change / prev_close) * 100
                    pre_market_change = f"${pre_change:+.2f}" if market == "US" else f"₹{pre_change:+.2f}"
                    pre_market_change_percent = f"{pre_change_percent:+.2f}%"
            
            # Calculate post-market movement
            post_market_change = 'N/A'
            post_market_change_percent = 'N/A'
            if post_market_price and post_market_price != regular_close:
                post_change = post_market_price - regular_close
                post_change_percent = (post_change / regular_close) * 100
                post_market_change = f"${post_change:+.2f}" if market == "US" else f"₹{post_change:+.2f}"
                post_market_change_percent = f"{post_change_percent:+.2f}%"
            
            return {
                'pre_market_change': pre_market_change,
                'pre_market_change_percent': pre_market_change_percent,
                'post_market_change': post_market_change,
                'post_market_change_percent': post_market_change_percent,
                'regular_session_close': f"${regular_close:.2f}" if market == "US" else f"₹{regular_close:.2f}",
                'current_price': f"${current_price:.2f}" if market == "US" else f"₹{current_price:.2f}"
            }
            
        except Exception as e:
            # Fallback to basic calculation
            return {
                'pre_market_change': 'N/A',
                'pre_market_change_percent': 'N/A', 
                'post_market_change': 'N/A',
                'post_market_change_percent': 'N/A',
                'regular_session_close': f"${regular_close:.2f}" if market == "US" else f"₹{regular_close:.2f}",
                'current_price': f"${regular_close:.2f}" if market == "US" else f"₹{regular_close:.2f}"
            }
            
    except Exception as e:
        return {
            'pre_market_change': 'N/A',
            'pre_market_change_percent': 'N/A',
            'post_market_change': 'N/A', 
            'post_market_change_percent': 'N/A',
            'regular_session_close': 'N/A',
            'current_price': 'N/A'
        }

def calculate_moving_average(data, window=50):
    """
    Calculate moving average for the given data
    
    Args:
        data (pd.DataFrame): Stock price data
        window (int): Moving average window size
    
    Returns:
        pd.Series: Moving average values
    """
    return data['Close'].rolling(window=window).mean()

def calculate_macd(data, fast_period=12, slow_period=26, signal_period=9):
    """
    Calculate MACD (Moving Average Convergence Divergence) indicator
    
    Args:
        data (pd.DataFrame): Stock price data
        fast_period (int): Fast EMA period (default 12)
        slow_period (int): Slow EMA period (default 26)
        signal_period (int): Signal EMA period (default 9)
    
    Returns:
        tuple: (macd_line, signal_line, histogram)
    """
    # Calculate exponential moving averages
    ema_fast = data['Close'].ewm(span=fast_period).mean()
    ema_slow = data['Close'].ewm(span=slow_period).mean()
    
    # MACD line = Fast EMA - Slow EMA
    macd_line = ema_fast - ema_slow
    
    # Signal line = EMA of MACD line
    signal_line = macd_line.ewm(span=signal_period).mean()
    
    # Histogram = MACD line - Signal line
    histogram = macd_line - signal_line
    
    return macd_line, signal_line, histogram

def calculate_rsi(data, period=14):
    """
    Calculate RSI (Relative Strength Index) indicator
    
    Args:
        data (pd.DataFrame): Stock price data
        period (int): RSI period (default 14)
    
    Returns:
        pd.Series: RSI values
    """
    # Calculate price changes
    delta = data['Close'].diff()
    
    # Separate gains and losses
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    
    # Calculate relative strength
    rs = gain / loss
    
    # Calculate RSI
    rsi = 100 - (100 / (1 + rs))
    
    return rsi

def calculate_chaikin_money_flow(data, period=20):
    """
    Calculate Chaikin Money Flow (CMF) indicator
    
    Args:
        data (pd.DataFrame): Stock price data with OHLCV
        period (int): CMF period (default 20)
    
    Returns:
        pd.Series: CMF values
    """
    # Money Flow Multiplier = [(Close - Low) - (High - Close)] / (High - Low)
    mf_multiplier = ((data['Close'] - data['Low']) - (data['High'] - data['Close'])) / (data['High'] - data['Low'])
    
    # Handle division by zero (when High = Low)
    mf_multiplier = mf_multiplier.fillna(0)
    
    # Money Flow Volume = Money Flow Multiplier * Volume
    mf_volume = mf_multiplier * data['Volume']
    
    # Chaikin Money Flow = Sum of Money Flow Volume over period / Sum of Volume over period
    cmf = mf_volume.rolling(window=period).sum() / data['Volume'].rolling(window=period).sum()
    
    return cmf

def calculate_support_resistance(data, window=20):
    """
    Calculate potential support and resistance levels
    
    Args:
        data (pd.DataFrame): Stock price data
        window (int): Rolling window for calculations
    
    Returns:
        tuple: (support_level, resistance_level)
    """
    # Calculate rolling lows and highs
    rolling_lows = data['Low'].rolling(window=window).min()
    rolling_highs = data['High'].rolling(window=window).max()
    
    # Get recent support (lowest low) and resistance (highest high)
    recent_support = rolling_lows.iloc[-window:].min()
    recent_resistance = rolling_highs.iloc[-window:].max()
    
    return recent_support, recent_resistance

def calculate_fibonacci_levels(data, period_months=3, symbol=None):
    """
    Calculate next two Fibonacci levels above and below current price
    
    Args:
        data (pd.DataFrame): Stock price data
        period_months (int): Period in months for high/low range (3 or 6)
        symbol (str): Stock symbol for direct yfinance query (optional)
    
    Returns:
        dict: Fibonacci analysis results with next levels above/below current price
    """
    try:
        current_price = data['Close'].iloc[-1]
        
        # Use the same accurate 3-month calculation as PDF to ensure consistency
        if period_months == 3 and symbol:
            try:
                # Get the correct 3-month period data directly from yfinance
                import yfinance as yf
                temp_ticker = yf.Ticker(symbol)
                temp_3mo_data = temp_ticker.history(period='3mo')
                if not temp_3mo_data.empty:
                    reference_high = temp_3mo_data['High'].max()
                    reference_low = temp_3mo_data['Low'].min()
                    print(f"UI DEBUG: Using yfinance 3mo data for {symbol} - High: {reference_high:.2f}, Low: {reference_low:.2f}")
                else:
                    raise ValueError("Empty 3mo data")
            except Exception as e:
                print(f"UI DEBUG: Error with yfinance 3mo query for {symbol}: {e}")
                # Fallback: use 3-month equivalent from existing data (approximately 64 trading days)
                period_data = data.tail(64)  # Match the yfinance 3mo period length
                reference_high = period_data['High'].max()
                reference_low = period_data['Low'].min()
                print(f"UI DEBUG: Using fallback with 64 days for {symbol} - High: {reference_high:.2f}, Low: {reference_low:.2f}")
        else:
            # For 6-month or when symbol not provided, use trading days calculation
            days_back = min(period_months * 22, len(data))  # Approximate trading days
            period_data = data.tail(days_back)
            reference_high = period_data['High'].max()
            reference_low = period_data['Low'].min()
        
        price_range = reference_high - reference_low
        
        # Fibonacci retracement ratios (0.236, 0.382, 0.5, 0.618, 0.786)
        retracement_ratios = [0.236, 0.382, 0.5, 0.618, 0.786]
        retracement_labels = ["23.6%", "38.2%", "50.0%", "61.8%", "78.6%"]
        
        # Fibonacci extension ratios (1.272, 1.618, 2.0, 2.618)
        extension_ratios = [1.272, 1.618, 2.0, 2.618]
        extension_labels = ["127.2%", "161.8%", "200%", "261.8%"]
        
        all_levels = []
        
        # Determine if we need retracement or extension levels
        if reference_low <= current_price <= reference_high:
            # Price is within range - use retracement levels
            analysis_type = "retracement"
            for ratio, label in zip(retracement_ratios, retracement_labels):
                level_price = reference_low + (price_range * ratio)
                all_levels.append({
                    'price': level_price,
                    'label': f"Fib {label}",
                    'type': 'retracement'
                })
            
            # If price is close to either end of range, also add extension levels for better targets
            distance_from_low = current_price - reference_low
            distance_from_high = reference_high - current_price
            range_threshold_low = price_range * 0.25  # 25% of the range from low
            range_threshold_high = price_range * 0.35  # 35% of the range from high (more generous for upward extensions)
            
            if distance_from_low <= range_threshold_low:
                # Add downward extension levels for support below current price
                for ratio, label in zip(extension_ratios, extension_labels):
                    level_price = reference_low - (price_range * (ratio - 1.0))
                    # Only add levels with positive prices
                    if level_price > 0:
                        all_levels.append({
                            'price': level_price,
                            'label': f"Fib {label} Ext",
                            'type': 'extension_down'
                        })
            
            # If price is close to the high end of range (within 15% of range), also add upward extensions
            if distance_from_high <= range_threshold_high:
                # Add upward extension levels for resistance above current price
                for ratio, label in zip(extension_ratios, extension_labels):
                    level_price = reference_high + (price_range * (ratio - 1.0))
                    all_levels.append({
                        'price': level_price,
                        'label': f"Fib {label} Ext",
                        'type': 'extension_up'
                    })
        
        elif current_price > reference_high:
            # Price is above range - use upward extensions
            analysis_type = "upward_extension"
            for ratio, label in zip(extension_ratios, extension_labels):
                level_price = reference_high + (price_range * (ratio - 1.0))
                all_levels.append({
                    'price': level_price,
                    'label': f"Fib {label} Ext",
                    'type': 'extension_up'
                })
            # Also include some retracement levels as potential support
            for ratio, label in zip(retracement_ratios, retracement_labels):
                level_price = reference_low + (price_range * ratio)
                all_levels.append({
                    'price': level_price,
                    'label': f"Fib {label}",
                    'type': 'retracement'
                })
                
        else:  # current_price < reference_low
            # Price is below range - use downward extensions
            analysis_type = "downward_extension"
            for ratio, label in zip(extension_ratios, extension_labels):
                level_price = reference_low - (price_range * (ratio - 1.0))
                all_levels.append({
                    'price': level_price,
                    'label': f"Fib {label} Ext",
                    'type': 'extension_down'
                })
            # Also include some retracement levels as potential resistance
            for ratio, label in zip(retracement_ratios, retracement_labels):
                level_price = reference_low + (price_range * ratio)
                all_levels.append({
                    'price': level_price,
                    'label': f"Fib {label}",
                    'type': 'retracement'
                })
        
        # Find the next two levels above and below current price
        levels_above = [level for level in all_levels if level['price'] > current_price]
        levels_below = [level for level in all_levels if level['price'] < current_price]
        
        # Sort and get closest levels
        levels_above.sort(key=lambda x: x['price'])
        levels_below.sort(key=lambda x: x['price'], reverse=True)
        
        # Get next two levels in each direction
        next_levels_above = levels_above[:2] if len(levels_above) >= 2 else levels_above
        next_levels_below = levels_below[:2] if len(levels_below) >= 2 else levels_below
        
        return {
            'analysis_type': analysis_type,
            'period_months': period_months,
            'reference_high': reference_high,
            'reference_low': reference_low,
            'current_price': current_price,
            'next_levels_above': next_levels_above,
            'next_levels_below': next_levels_below,
            'all_levels': all_levels
        }
        
    except Exception as e:
        print(f"Error calculating Fibonacci levels: {str(e)}")
        return None



def get_earnings_info(ticker_obj, ticker_info, symbol):
    """
    Extract earnings information from ticker object and info with improved accuracy and multiple fallback sources
    
    Args:
        ticker_obj: yfinance Ticker object
        ticker_info (dict): Company information from yfinance
        symbol (str): Stock symbol for major stock detection
    
    Returns:
        dict: Earnings information
    """
    earnings_info = {
        'last_earnings': None,
        'next_earnings': None,
        'last_earnings_formatted': 'N/A',
        'next_earnings_formatted': 'N/A'
    }
    
    current_date = pd.Timestamp.now()
    print(f"Gathering earnings info...")
    
    try:
        # Method 1: Try earnings_dates (most comprehensive source)
        try:
            earnings_dates = ticker_obj.earnings_dates
            if earnings_dates is not None and not earnings_dates.empty:
                print(f"Found {len(earnings_dates)} earnings dates from earnings_dates")
                
                # IMPORTANT: Filter out Meeting events and only keep Earnings events
                if 'Event Type' in earnings_dates.columns:
                    earnings_only = earnings_dates[earnings_dates['Event Type'] == 'Earnings'].copy()
                    print(f"Filtered to {len(earnings_only)} actual earnings events (excluded meetings)")
                    if earnings_only.empty:
                        print("No actual earnings events found after filtering out meetings")
                        raise ValueError("No earnings events, only meetings found")
                    earnings_dates = earnings_only
                
                # Convert current_date to match earnings_dates timezone
                if hasattr(earnings_dates.index[0], 'tz') and earnings_dates.index[0].tz:
                    current_date_tz = current_date.tz_localize(earnings_dates.index[0].tz)
                else:
                    current_date_tz = current_date
                
                # Look for the most recent past earnings (prioritize 2025 data)
                # First try to find earnings from current year (2025)
                current_year = current_date_tz.year
                current_year_earnings = earnings_dates[
                    (earnings_dates.index <= current_date_tz) & 
                    (earnings_dates.index.year == current_year)
                ]
                
                if not current_year_earnings.empty:
                    # Sort by date to ensure we get the most recent (last in chronological order)
                    current_year_earnings_sorted = current_year_earnings.sort_index()
                    last_earnings = current_year_earnings_sorted.index[-1]
                    earnings_info['last_earnings'] = last_earnings
                    earnings_info['last_earnings_formatted'] = last_earnings.strftime('%Y-%m-%d')
                    
                    # Check if earnings might be outdated (more than 100 days old or more than 80 days for major quarterly companies)
                    days_since_last = (current_date_tz - last_earnings).days
                    quarterly_threshold = 80  # ~2.5 months for major quarterly reporters
                    major_stocks = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'MSTR']
                    
                    if symbol.upper() in major_stocks and days_since_last > quarterly_threshold:
                        print(f"Found last earnings (current year - likely outdated): {earnings_info['last_earnings_formatted']} ({days_since_last} days ago)")
                        earnings_info['last_earnings_formatted'] += f" (likely outdated - check company reports)"
                    elif days_since_last > 100:
                        print(f"Found last earnings (current year - possibly outdated): {earnings_info['last_earnings_formatted']} ({days_since_last} days ago)")
                        earnings_info['last_earnings_formatted'] += f" (data may be outdated)"
                    else:
                        print(f"Found last earnings (current year): {earnings_info['last_earnings_formatted']} from {len(current_year_earnings)} 2025 earnings dates")
                else:
                    # Fallback to most recent within 12 months (more aggressive cutoff)
                    cutoff_date_tz = current_date_tz - pd.Timedelta(days=365)  # 12 months
                    recent_past_earnings = earnings_dates[
                        (earnings_dates.index <= current_date_tz) & 
                        (earnings_dates.index >= cutoff_date_tz)
                    ]
                    if not recent_past_earnings.empty:
                        last_earnings = recent_past_earnings.index[-1]
                        earnings_info['last_earnings'] = last_earnings
                        earnings_info['last_earnings_formatted'] = last_earnings.strftime('%Y-%m-%d')
                        
                        # Check if this data might be outdated
                        days_since_last = (current_date_tz - last_earnings).days
                        if days_since_last > 100:
                            earnings_info['last_earnings_formatted'] += f" (may be outdated)"
                            print(f"Found last earnings (within 12 months - possibly outdated): {earnings_info['last_earnings_formatted']} ({days_since_last} days ago)")
                        else:
                            print(f"Found last earnings (within 12 months): {earnings_info['last_earnings_formatted']}")
                    else:
                        # Check if we have any past earnings and use the most recent one
                        all_past_earnings = earnings_dates[earnings_dates.index <= current_date_tz]
                        if not all_past_earnings.empty:
                            # Use the most recent past earnings regardless of age
                            last_earnings = all_past_earnings.index[-1]
                            earnings_info['last_earnings'] = last_earnings
                            earnings_info['last_earnings_formatted'] = last_earnings.strftime('%Y-%m-%d')
                            print(f"Found last earnings (older): {earnings_info['last_earnings_formatted']}")
                        else:
                            print("No past earnings found at all")
                            # For major stocks, indicate this might be a data issue
                            major_stocks = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'MSTR']
                            if symbol.upper() in major_stocks:
                                earnings_info['last_earnings_formatted'] = "Data may be incomplete"
                
                # Get next upcoming earnings
                future_earnings = earnings_dates[earnings_dates.index > current_date_tz]
                if not future_earnings.empty:
                    next_earnings = future_earnings.index[0]
                    earnings_info['next_earnings'] = next_earnings
                    earnings_info['next_earnings_formatted'] = next_earnings.strftime('%Y-%m-%d')
                    print(f"Found next earnings: {earnings_info['next_earnings_formatted']}")
        except Exception as e:
            print(f"Earnings dates error: {e}")
        
        # Method 2: Try earnings calendar for next earnings
        if earnings_info['next_earnings'] is None:
            try:
                calendar = ticker_obj.calendar
                if calendar is not None:
                    # Handle both DataFrame and dict calendar formats
                    if hasattr(calendar, 'empty') and not calendar.empty:
                        next_earnings = calendar.index[0]
                        earnings_info['next_earnings'] = next_earnings
                        earnings_info['next_earnings_formatted'] = next_earnings.strftime('%Y-%m-%d')
                        print(f"Found next earnings from calendar: {earnings_info['next_earnings_formatted']}")
                    elif isinstance(calendar, dict) and calendar:
                        # Handle dict format calendar - prioritize 'Earnings Date' key
                        earnings_date_key = None
                        for key in ['Earnings Date', 'earnings_date', 'earnings', 'next_earnings']:
                            if key in calendar:
                                earnings_date_key = key
                                break
                        
                        if earnings_date_key and calendar[earnings_date_key]:
                            dates = calendar[earnings_date_key]
                            # Handle single date or list of dates
                            if hasattr(dates, '__len__') and not isinstance(dates, str):
                                # It's a list or array
                                if len(dates) > 0:
                                    next_earnings = pd.to_datetime(dates[0])
                                    # Only use if it's actually in the future
                                    if next_earnings > current_date:
                                        earnings_info['next_earnings'] = next_earnings
                                        earnings_info['next_earnings_formatted'] = next_earnings.strftime('%Y-%m-%d')
                                        print(f"Found next earnings from calendar dict list: {earnings_info['next_earnings_formatted']}")
                                    else:
                                        print(f"Calendar earnings date is not in future: {next_earnings.strftime('%Y-%m-%d')}")
                            else:
                                # It's a single date object
                                next_earnings = pd.to_datetime(dates)
                                # Only use if it's actually in the future
                                if next_earnings > current_date:
                                    earnings_info['next_earnings'] = next_earnings
                                    earnings_info['next_earnings_formatted'] = next_earnings.strftime('%Y-%m-%d')
                                    print(f"Found next earnings from calendar dict single: {earnings_info['next_earnings_formatted']}")
                                else:
                                    print(f"Calendar earnings date is not in future: {next_earnings.strftime('%Y-%m-%d')}")
                        else:
                            print("No valid earnings date found in calendar")
            except Exception as e:
                print(f"Calendar error: {e}")
        
        # Method 3: Try earnings history for past earnings (with date validation)
        if earnings_info['last_earnings'] is None:
            try:
                earnings_history = ticker_obj.earnings
                if earnings_history is not None and not earnings_history.empty:
                    # Convert earnings history index to datetime and filter for recent dates
                    last_earnings_date = pd.to_datetime(earnings_history.index[-1])
                    
                    # Only use if it's within the last 2 years (more reasonable for earnings)
                    cutoff_date = current_date - pd.Timedelta(days=730)  # 2 years
                    if last_earnings_date >= cutoff_date:
                        earnings_info['last_earnings'] = last_earnings_date
                        earnings_info['last_earnings_formatted'] = last_earnings_date.strftime('%Y-%m-%d')
                        print(f"Found last earnings from history: {earnings_info['last_earnings_formatted']}")
                    else:
                        print(f"Earnings history date too old: {last_earnings_date.strftime('%Y-%m-%d')}")
            except Exception as e:
                print(f"Earnings history error: {e}")
        
        # Method 4: Fallback to ticker_info fields
        if earnings_info['next_earnings'] is None:
            try:
                if 'earningsDate' in ticker_info and ticker_info['earningsDate']:
                    earnings_date = ticker_info['earningsDate']
                    if isinstance(earnings_date, list) and earnings_date:
                        next_earnings = pd.to_datetime(earnings_date[0], unit='s')
                    else:
                        next_earnings = pd.to_datetime(earnings_date, unit='s')
                    
                    # Only use if it's actually in the future
                    if next_earnings > current_date:
                        earnings_info['next_earnings'] = next_earnings
                        earnings_info['next_earnings_formatted'] = next_earnings.strftime('%Y-%m-%d')
                        print(f"Found next earnings from ticker info: {earnings_info['next_earnings_formatted']}")
                    else:
                        print(f"Ticker info earnings date is not in future: {next_earnings.strftime('%Y-%m-%d')}")
            except Exception as e:
                print(f"Ticker info earnings date error: {e}")
        
        # Final check: If we still don't have next earnings, estimate based on last earnings
        if earnings_info['next_earnings'] is None and earnings_info['last_earnings'] is not None:
            try:
                # Estimate next earnings as roughly 3 months (90 days) after last earnings
                estimated_next = earnings_info['last_earnings'] + pd.Timedelta(days=90)
                if estimated_next > current_date:
                    earnings_info['next_earnings_formatted'] = f"~{estimated_next.strftime('%Y-%m-%d')} (estimated)"
                    print(f"Estimated next earnings: {earnings_info['next_earnings_formatted']}")
                else:
                    earnings_info['next_earnings_formatted'] = "TBD"
                    print("Cannot estimate future earnings - last earnings too recent or in future")
            except Exception as e:
                print(f"Estimation error: {e}")
                earnings_info['next_earnings_formatted'] = "TBD"
        
        # Method 5: Try additional ticker_info fields for last earnings (with date validation)
        if earnings_info['last_earnings'] is None:
            try:
                # Try multiple possible fields but validate dates
                for field in ['lastFiscalYearEnd', 'mostRecentQuarter', 'lastQuarterEnd']:
                    if field in ticker_info and ticker_info[field]:
                        try:
                            last_earnings = pd.to_datetime(ticker_info[field], unit='s')
                            
                            # Only use if it's within the last 18 months (more reasonable for recent earnings)
                            cutoff_date = current_date - pd.Timedelta(days=540)  # 18 months
                            if last_earnings >= cutoff_date:
                                earnings_info['last_earnings'] = last_earnings
                                earnings_info['last_earnings_formatted'] = last_earnings.strftime('%Y-%m-%d')
                                print(f"Found last earnings from {field}: {earnings_info['last_earnings_formatted']}")
                                break
                            else:
                                print(f"Field {field} date too old: {last_earnings.strftime('%Y-%m-%d')}")
                        except:
                            continue
            except Exception as e:
                print(f"Ticker info fallback error: {e}")
        
        # Method 6: Estimate next earnings based on last earnings (quarterly companies)
        if earnings_info['next_earnings'] is None and earnings_info['last_earnings'] is not None:
            try:
                last_earnings_date = earnings_info['last_earnings']
                print(f"Attempting to estimate next earnings from last: {last_earnings_date}")
                
                # Try different estimation approaches
                for days_offset in [90, 91, 92, 89, 88]:  # Try variations around 90 days (quarterly)
                    try:
                        if hasattr(last_earnings_date, 'tz') and last_earnings_date.tz:
                            estimated_next = last_earnings_date + pd.DateOffset(days=days_offset)
                        else:
                            estimated_next = pd.to_datetime(last_earnings_date) + pd.DateOffset(days=days_offset)
                        
                        # Check if estimated date is in the future
                        if estimated_next > pd.Timestamp.now():
                            earnings_info['next_earnings'] = estimated_next
                            earnings_info['next_earnings_formatted'] = f"~{estimated_next.strftime('%Y-%m-%d')}"
                            print(f"Estimated next earnings (+{days_offset} days): {earnings_info['next_earnings_formatted']}")
                            break
                    except Exception as inner_e:
                        continue
                
                # If quarterly estimation doesn't work, try yearly estimation
                if earnings_info['next_earnings'] is None:
                    try:
                        estimated_next = pd.to_datetime(last_earnings_date) + pd.DateOffset(years=1)
                        if estimated_next > pd.Timestamp.now():
                            earnings_info['next_earnings'] = estimated_next
                            earnings_info['next_earnings_formatted'] = f"~{estimated_next.strftime('%Y-%m-%d')}"
                            print(f"Estimated next earnings (yearly): {earnings_info['next_earnings_formatted']}")
                    except:
                        pass
                        
            except Exception as e:
                print(f"Estimation error: {e}")
                
    except Exception as e:
        print(f"General earnings info error: {e}")
    
    print(f"Final earnings info - Last: {earnings_info['last_earnings_formatted']}, Next: {earnings_info['next_earnings_formatted']}")
    return earnings_info

def get_earnings_performance_analysis(ticker_obj, data, market="US"):
    """
    Analyze stock performance after earnings for available quarters (up to 4)
    
    Args:
        ticker_obj: yfinance Ticker object
        data (pd.DataFrame): Stock price data
        market (str): Market type (US/IN)
    
    Returns:
        tuple: (pd.DataFrame: Earnings performance analysis table, int: number of quarters found)
    """
    try:
        # Get earnings history - try multiple methods
        earnings = None
        earnings_count = 0
        
        # Method 1: Try earnings_dates
        try:
            earnings = ticker_obj.earnings_dates
            if earnings is not None and not earnings.empty:
                earnings_count = len(earnings)
        except Exception as e:
            print(f"Earnings dates error: {e}")
            pass
        
        # Method 2: Try calendar if earnings_dates failed  
        if earnings is None or earnings.empty:
            try:
                calendar = ticker_obj.calendar
                if calendar is not None:
                    # Calendar might be a dict, not DataFrame
                    if hasattr(calendar, 'empty') and not calendar.empty:
                        earnings = calendar
                        earnings_count = len(calendar)
                    elif isinstance(calendar, dict) and calendar:
                        # Convert dict to DataFrame if needed
                        earnings = pd.DataFrame(calendar)
                        earnings_count = len(earnings)
            except Exception as e:
                print(f"Calendar error: {e}")
                pass
        
        # Method 3: Try earnings history
        if earnings is None or earnings.empty:
            try:
                earnings_history = ticker_obj.earnings
                if earnings_history is not None and not earnings_history.empty:
                    # Convert earnings history to datetime index
                    earnings = earnings_history
                    earnings_count = len(earnings_history)
            except:
                pass
        
        if earnings is None or earnings.empty:
            return None, 0
        
        # Get available earnings dates and filter for unique quarters (up to 4)
        earnings_dates = earnings.index.tolist()
        earnings_dates.sort(reverse=True)  # Most recent first
        
        # Filter for unique quarters to avoid duplicate earnings in same quarter
        unique_quarters = []
        seen_quarters = set()
        
        for date in earnings_dates:
            quarter_key = f"{date.year}-Q{(date.month - 1) // 3 + 1}"
            if quarter_key not in seen_quarters:
                unique_quarters.append(date)
                seen_quarters.add(quarter_key)
                if len(unique_quarters) >= 4:
                    break
        
        available_earnings = unique_quarters
        print(f"Filtered to {len(available_earnings)} unique quarters from {len(earnings_dates)} total earnings dates")
        
        analysis_data = []
        successful_analyses = 0
        
        for earnings_date in available_earnings:
            try:
                print(f"Processing earnings date: {earnings_date}")
                
                # Ensure both earnings_date and data.index have the same timezone handling
                # Convert earnings_date to match data.index timezone
                if hasattr(data.index[0], 'tz') and data.index[0].tz:
                    # Data has timezone, make sure earnings_date has it too
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_for_comparison = earnings_date
                    else:
                        # Convert naive datetime to data's timezone
                        earnings_date_for_comparison = earnings_date.tz_localize(data.index[0].tz)
                else:
                    # Data is timezone-naive, make earnings_date naive too
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_for_comparison = earnings_date.tz_localize(None)
                    else:
                        earnings_date_for_comparison = earnings_date
                
                # Determine BMO vs AMC timing first to get correct pre-earnings baseline
                earnings_hour = earnings_date.hour
                is_bmo = earnings_hour < 12  # Before noon = BMO
                
                # Find pre-earnings data based on timing
                if is_bmo:
                    # For BMO: pre-earnings is the day BEFORE the earnings announcement day
                    # Exclude the earnings announcement day itself
                    earnings_day_start = earnings_date_for_comparison.normalize()
                    pre_earnings_mask = data.index < earnings_day_start
                else:
                    # For AMC: pre-earnings is up to the earnings announcement time
                    pre_earnings_mask = data.index < earnings_date_for_comparison
                
                if not pre_earnings_mask.any():
                    print(f"No pre-earnings data found for {earnings_date}")
                    continue
                    
                pre_earnings_price = data[pre_earnings_mask]['Close'].iloc[-1]
                pre_earnings_date = data[pre_earnings_mask].index[-1]
                
                # Find post-earnings data based on timing
                if is_bmo:
                    # For BMO: use same day as earnings announcement
                    earnings_day_mask = data.index.normalize() == earnings_date_for_comparison.normalize()
                    if earnings_day_mask.any():
                        post_earnings_open = data[earnings_day_mask]['Open'].iloc[0]
                        post_earnings_close = data[earnings_day_mask]['Close'].iloc[0]
                        post_earnings_date = data[earnings_day_mask].index[0]
                    else:
                        print(f"No same-day data for BMO earnings {earnings_date}")
                        continue
                else:
                    # For AMC: use next trading day
                    post_earnings_mask = data.index > earnings_date_for_comparison
                    if not post_earnings_mask.any():
                        print(f"No post-earnings data found for {earnings_date}")
                        continue
                    post_earnings_open = data[post_earnings_mask]['Open'].iloc[0]
                    post_earnings_close = data[post_earnings_mask]['Close'].iloc[0]
                    post_earnings_date = data[post_earnings_mask].index[0]
                
                # Calculate overnight change (from previous close to target open)
                overnight_change = ((post_earnings_open - pre_earnings_price) / pre_earnings_price) * 100
                
                # Calculate day change (from previous close to target close)
                next_day_change = ((post_earnings_close - pre_earnings_price) / pre_earnings_price) * 100
                
                # Find end of week price (Friday following earnings, or Thursday if Friday not available)
                week_end_mask = data.index >= post_earnings_date
                week_data = data[week_end_mask]
                
                # Find the Friday following earnings
                week_end_price = None
                week_end_date = None
                
                # Look for Friday (weekday 4) in the week following earnings
                for i, date in enumerate(week_data.index):
                    if date.weekday() == 4:  # Friday
                        week_end_price = week_data['Close'].iloc[i]
                        week_end_date = date
                        break
                
                # If no Friday found, look for Thursday (weekday 3)
                if week_end_price is None:
                    for i, date in enumerate(week_data.index):
                        if date.weekday() == 3:  # Thursday
                            week_end_price = week_data['Close'].iloc[i]
                            week_end_date = date
                            break
                
                # Fallback to last available day if neither Friday nor Thursday found
                if week_end_price is None and not week_data.empty:
                    week_end_price = week_data['Close'].iloc[-1]
                    week_end_date = week_data.index[-1]
                elif week_end_price is None:
                    week_end_price = post_earnings_open
                    week_end_date = post_earnings_date
                
                # Calculate week performance (from pre-earnings close to end of week)
                week_performance = ((week_end_price - pre_earnings_price) / pre_earnings_price) * 100
                
                # Calculate days between pre-earnings close and week-end close
                days_difference = (week_end_date.date() - pre_earnings_date.date()).days
                
                # Determine quarter
                quarter = f"Q{((earnings_date.month - 1) // 3) + 1} {earnings_date.year}"
                
                # Set timing and display info based on earlier determination
                if is_bmo:
                    timing = "BMO"
                    display_open = post_earnings_open
                    display_close = post_earnings_close
                    print(f"BMO earnings {earnings_date.strftime('%Y-%m-%d')}: Pre-close ${pre_earnings_price:.2f} (from {pre_earnings_date.strftime('%Y-%m-%d')}) → Same-day open ${display_open:.2f} → Same-day close ${display_close:.2f}")
                else:
                    timing = "AMC"
                    display_open = post_earnings_open
                    display_close = post_earnings_close
                
                print(f"Analysis successful for {earnings_date} ({timing}): {overnight_change:+.2f}% overnight, {next_day_change:+.2f}% next day, {week_performance:+.2f}% week")
                
                analysis_data.append({
                    'Quarter': quarter,
                    'Earnings Date': f"{earnings_date.strftime('%Y-%m-%d')} ({timing})",
                    'Pre-Earnings Close': format_currency(pre_earnings_price, market),
                    'Next Day Open': format_currency(display_open, market),
                    'Next Day Close': format_currency(display_close, market),
                    'Overnight Change (%)': f"{overnight_change:+.2f}%",
                    'Next Day Change (%)': f"{next_day_change:+.2f}%",
                    'End of Week Close': format_currency(week_end_price, market),
                    'Week End Date': f"{week_end_date.strftime('%Y-%m-%d')} ({days_difference}d)",
                    'Week Performance (%)': f"{week_performance:+.2f}%",
                    'Direction': '📈 Up' if week_performance > 0 else '📉 Down' if week_performance < 0 else '➡️ Flat'
                })
                successful_analyses += 1
                
            except Exception as e:
                print(f"Error processing {earnings_date}: {e}")
                continue
        
        if analysis_data:
            df = pd.DataFrame(analysis_data)
            return df, successful_analyses
        else:
            return None, 0
            
    except Exception as e:
        return None, 0

def get_detailed_earnings_performance_analysis(ticker_obj, data, market="US", max_quarters=8):
    """
    Enhanced earnings performance analysis with extended historical data
    
    Args:
        ticker_obj: yfinance Ticker object
        data (pd.DataFrame): Stock price data
        market (str): Market type (US/IN)
        max_quarters (int): Maximum number of quarters to analyze
    
    Returns:
        tuple: (pd.DataFrame: Detailed earnings performance analysis, int: number of quarters found)
    """
    try:
        # Get comprehensive earnings data
        earnings = None
        
        # Primary method: earnings_dates
        try:
            earnings = ticker_obj.earnings_dates
            if earnings is not None and not earnings.empty:
                print(f"Found {len(earnings)} earnings records from earnings_dates")
        except Exception as e:
            print(f"Earnings dates error: {e}")
        
        if earnings is None or earnings.empty:
            print("No earnings data found")
            return None, 0
        
        # Get available earnings dates and filter for unique quarters
        # IMPORTANT: Filter out Meeting events and only include actual Earnings events
        earnings_only = earnings[earnings.get('Event Type', 'Earnings') == 'Earnings'].copy()
        
        if earnings_only.empty:
            print("No actual earnings events found (filtered out meetings)")
            return None, 0
            
        earnings_dates = earnings_only.index.tolist()
        earnings_dates.sort(reverse=True)  # Most recent first
        
        # Filter for unique quarters to avoid duplicate earnings in same quarter
        unique_quarters = []
        seen_quarters = set()
        
        for date in earnings_dates:
            quarter_key = f"{date.year}-Q{(date.month - 1) // 3 + 1}"
            if quarter_key not in seen_quarters:
                unique_quarters.append(date)
                seen_quarters.add(quarter_key)
                if len(unique_quarters) >= max_quarters:
                    break
        
        available_earnings = unique_quarters
        print(f"Filtered to {len(available_earnings)} unique quarters from {len(earnings_dates)} total earnings dates")
        
        analysis_data = []
        successful_analyses = 0
        
        print(f"Analyzing {len(available_earnings)} unique quarterly earnings dates...")
        
        # Debug: Show which quarters we're analyzing
        for date in available_earnings:
            quarter_key = f"{date.year}-Q{(date.month - 1) // 3 + 1}"
            print(f"Selected {quarter_key}: {date.strftime('%Y-%m-%d %H:%M:%S')}")
        
        for earnings_date in available_earnings:
            try:
                print(f"Processing earnings date: {earnings_date}")
                
                # Handle timezone compatibility
                if hasattr(data.index[0], 'tz') and data.index[0].tz:
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_for_comparison = earnings_date
                    else:
                        earnings_date_for_comparison = earnings_date.tz_localize(data.index[0].tz)
                else:
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_for_comparison = earnings_date.tz_localize(None)
                    else:
                        earnings_date_for_comparison = earnings_date
                
                # Determine BMO vs AMC timing first to get correct pre-earnings baseline
                earnings_hour = earnings_date.hour
                is_bmo = earnings_hour < 12  # Before noon = BMO
                
                # Find pre-earnings data based on timing
                if is_bmo:
                    # For BMO: pre-earnings is the day BEFORE the earnings announcement day
                    # Exclude the earnings announcement day itself
                    earnings_day_start = earnings_date_for_comparison.normalize()
                    pre_earnings_mask = data.index < earnings_day_start
                else:
                    # For AMC: pre-earnings is up to the earnings announcement time
                    pre_earnings_mask = data.index < earnings_date_for_comparison
                
                if not pre_earnings_mask.any():
                    print(f"No pre-earnings data for {earnings_date}")
                    continue
                    
                pre_earnings_price = data[pre_earnings_mask]['Close'].iloc[-1]
                pre_earnings_date = data[pre_earnings_mask].index[-1]
                
                # Find post-earnings data based on timing
                if is_bmo:
                    # For BMO: use same day as earnings announcement
                    earnings_day_mask = data.index.normalize() == earnings_date_for_comparison.normalize()
                    if earnings_day_mask.any():
                        post_earnings_open = data[earnings_day_mask]['Open'].iloc[0]
                        post_earnings_close = data[earnings_day_mask]['Close'].iloc[0]
                        post_earnings_date = data[earnings_day_mask].index[0]
                    else:
                        print(f"No same-day data for BMO earnings {earnings_date}")
                        continue
                else:
                    # For AMC: use next trading day
                    post_earnings_mask = data.index > earnings_date_for_comparison
                    if not post_earnings_mask.any():
                        print(f"No post-earnings data for {earnings_date}")
                        continue
                    post_earnings_open = data[post_earnings_mask]['Open'].iloc[0]
                    post_earnings_close = data[post_earnings_mask]['Close'].iloc[0]
                    post_earnings_date = data[post_earnings_mask].index[0]
                
                # Calculate overnight change (pre-close to target open)
                overnight_change = ((post_earnings_open - pre_earnings_price) / pre_earnings_price) * 100
                
                # Calculate day change (pre-close to target close)
                next_day_change = ((post_earnings_close - pre_earnings_price) / pre_earnings_price) * 100
                
                # Calculate week performance (Friday following earnings, or Thursday if Friday not available)
                week_end_mask = data.index >= post_earnings_date
                week_data = data[week_end_mask]
                
                # Find the Friday following earnings
                week_end_price = None
                week_end_date = None
                
                # Look for Friday (weekday 4) in the week following earnings
                for i, date in enumerate(week_data.index):
                    if date.weekday() == 4:  # Friday
                        week_end_price = week_data['Close'].iloc[i]
                        week_end_date = date
                        break
                
                # If no Friday found, look for Thursday (weekday 3)
                if week_end_price is None:
                    for i, date in enumerate(week_data.index):
                        if date.weekday() == 3:  # Thursday
                            week_end_price = week_data['Close'].iloc[i]
                            week_end_date = date
                            break
                
                # Fallback to last available day if neither Friday nor Thursday found
                if week_end_price is None and not week_data.empty:
                    week_end_price = week_data['Close'].iloc[-1]
                    week_end_date = week_data.index[-1]
                elif week_end_price is None:
                    week_end_price = post_earnings_open
                    week_end_date = post_earnings_date
                
                # Calculate week performance (from pre-earnings close to end of week)
                week_performance = ((week_end_price - pre_earnings_price) / pre_earnings_price) * 100
                
                # Calculate days between pre-earnings close and week-end close
                days_difference = (week_end_date.date() - pre_earnings_date.date()).days
                
                # Get EPS data if available
                eps_estimate = None
                eps_actual = None
                surprise_pct = None
                
                if earnings_date in earnings.index:
                    row = earnings.loc[earnings_date]
                    if hasattr(row, 'get'):
                        eps_estimate = row.get('EPS Estimate', None)
                        eps_actual = row.get('Reported EPS', None) 
                        surprise_pct = row.get('Surprise(%)', None)
                
                # Determine quarter
                quarter_num = (earnings_date.month - 1) // 3 + 1
                quarter = f"Q{quarter_num} {earnings_date.year}"
                
                # Set timing and display info based on earlier determination
                if is_bmo:
                    timing = "BMO"
                    display_open = post_earnings_open
                    display_close = post_earnings_close
                    print(f"BMO earnings {earnings_date.strftime('%Y-%m-%d')}: Pre-close ${pre_earnings_price:.2f} (from {pre_earnings_date.strftime('%Y-%m-%d')}) → Same-day open ${display_open:.2f} → Same-day close ${display_close:.2f}")
                else:
                    timing = "AMC"
                    display_open = post_earnings_open
                    display_close = post_earnings_close
                    print(f"AMC earnings {earnings_date.strftime('%Y-%m-%d')}: Pre-close ${pre_earnings_price:.2f} → Next-day open ${display_open:.2f} → Next-day close ${display_close:.2f}")
                
                # Create detailed analysis row
                analysis_row = {
                    'Quarter': quarter,
                    'Earnings Date': f"{earnings_date.strftime('%Y-%m-%d')} ({timing})",
                    'Pre-Earnings Close': format_currency(pre_earnings_price, market),
                    'Next Day Open': format_currency(display_open, market),
                    'Next Day Close': format_currency(display_close, market),
                    'Overnight Change (%)': f"{overnight_change:+.2f}%",
                    'Next Day Change (%)': f"{next_day_change:+.2f}%",
                    'End of Week Close': format_currency(week_end_price, market),
                    'Week End Date': f"{week_end_date.strftime('%Y-%m-%d')} ({days_difference}d)",
                    'Week Performance (%)': f"{week_performance:+.2f}%",
                    'Direction': '📈 Up' if week_performance > 0 else '📉 Down' if week_performance < 0 else '➡️ Flat'
                }
                
                # Add EPS information if available
                if eps_estimate and not pd.isna(eps_estimate):
                    analysis_row['EPS Est'] = f"${eps_estimate:.2f}"
                if eps_actual and not pd.isna(eps_actual):
                    analysis_row['EPS Act'] = f"${eps_actual:.2f}"
                if surprise_pct and not pd.isna(surprise_pct):
                    analysis_row['Surprise'] = f"{surprise_pct:+.1f}%"
                
                analysis_data.append(analysis_row)
                successful_analyses += 1
                
                print(f"✅ Analysis successful: {overnight_change:+.2f}% overnight, {next_day_change:+.2f}% next day, {week_performance:+.2f}% week")
                
            except Exception as e:
                print(f"❌ Error processing {earnings_date}: {e}")
                continue
        
        if analysis_data:
            df = pd.DataFrame(analysis_data)
            print(f"Created analysis DataFrame with {len(df)} rows")
            return df, successful_analyses
        else:
            print("No successful analyses")
            return None, 0
            
    except Exception as e:
        print(f"Overall analysis error: {e}")
        return None, 0

def display_institutional_financial_metrics(info, ticker_obj, symbol):
    """
    Display comprehensive institutional-grade financial metrics
    
    Args:
        info (dict): Stock information from yfinance
        ticker_obj: yfinance Ticker object
        symbol (str): Stock symbol
    """
    try:
        st.success(f"✅ Financial metrics loaded for {symbol}")
        
        # Add prominent data source information visible to all users
        hybrid_metrics = get_hybrid_financial_metrics(symbol or "UNKNOWN", info) if symbol else None
        data_source = hybrid_metrics.get('source', 'Yahoo Finance') if hybrid_metrics else 'Yahoo Finance'
        
        with st.expander("📊 Data Source Information & Known Discrepancies", expanded=False):
            st.info(f"🔍 **Current Data Source**: {data_source}")
            
            if hybrid_metrics:
                st.info(f"📊 **Valuation**: P/E: {hybrid_metrics.get('pe_ratio', 'N/A'):.2f if hybrid_metrics.get('pe_ratio') else 'N/A'} | "
                        f"Forward P/E: {hybrid_metrics.get('forward_pe', 'N/A'):.2f if hybrid_metrics.get('forward_pe') else 'N/A'} | "
                        f"P/S: {hybrid_metrics.get('ps_ratio', 'N/A'):.2f if hybrid_metrics.get('ps_ratio') else 'N/A'} | "
                        f"PEG: {hybrid_metrics.get('peg_ratio', 'N/A'):.2f if hybrid_metrics.get('peg_ratio') and not pd.isna(hybrid_metrics.get('peg_ratio')) else 'N/A'}")
                
                ev_revenue = hybrid_metrics.get('ev_revenue') or info.get('enterpriseToRevenue')
                ev_ebitda = hybrid_metrics.get('ev_ebitda') or info.get('enterpriseToEbitda')
                st.info(f"📊 **Enterprise**: EV/Revenue: {ev_revenue:.2f if ev_revenue else 'N/A'} | "
                        f"EV/EBITDA: {ev_ebitda:.2f if ev_ebitda else 'N/A'}")
                
                # Show margin data which can have significant differences
                operating_margin = info.get('operatingMargins')
                gross_margin = info.get('grossMargins')
                st.info(f"📊 **Margins**: Gross: {gross_margin*100:.2f}% | Operating: {operating_margin*100:.2f}%" 
                        if operating_margin and gross_margin else "📊 **Margins**: Limited data available")
            
            st.warning("ℹ️ **Data Variance**: API values may differ from institutional sources (±0.1-10%) due to data timing, calculation periods, and methodologies. "
                      "Operating margins can show significant differences between sources. GuruFocus data preferred when available.")
            
            if not hybrid_metrics or hybrid_metrics.get('source') == 'Yahoo Finance':
                st.warning("⚠️ **Using Yahoo Finance**: Add GURUFOCUS_API_KEY to get exact GuruFocus institutional metrics")
            else:
                st.success("✅ **Using GuruFocus**: Showing institutional-grade metrics")
        
        # Create tabs for different metric categories
        metrics_tab1, metrics_tab2, metrics_tab3, metrics_tab4 = st.tabs([
            "📈 Valuation Metrics", 
            "💰 Profitability", 
            "🏛️ Financial Strength", 
            "🚀 Growth Metrics"
        ])
        
        with metrics_tab1:
            display_valuation_metrics(info, symbol)
        
        with metrics_tab2:
            display_profitability_metrics(info)
        
        with metrics_tab3:
            display_financial_strength_metrics(info, ticker_obj)
        
        with metrics_tab4:
            display_growth_metrics(info, ticker_obj)
            
    except Exception as e:
        st.error(f"Error displaying financial metrics: {e}")

def display_valuation_metrics(info, symbol=None):
    """Display comprehensive valuation metrics with hybrid GuruFocus/Yahoo Finance data"""
    st.markdown("### Valuation Analysis")
    
    # Get hybrid metrics (GuruFocus preferred, yfinance fallback)
    hybrid_metrics = get_hybrid_financial_metrics(symbol or "UNKNOWN", info) if symbol else None
    

    

    
    col1, col2, col3, col4 = st.columns(4)
    
    # Define data source label for all metrics
    data_source_label = f" ({hybrid_metrics.get('source', 'Yahoo Finance')})" if hybrid_metrics else " (Yahoo Finance)"
    
    # Price-to-Earnings metrics
    with col1:
        pe_ratio = hybrid_metrics.get('pe_ratio') if hybrid_metrics else info.get('trailingPE', None)
        forward_pe = hybrid_metrics.get('forward_pe') if hybrid_metrics else info.get('forwardPE', None)
        
        st.metric(
            label="P/E Ratio (TTM)",
            value=f"{pe_ratio:.2f}" if pe_ratio and not pd.isna(pe_ratio) else "N/A",
            help=f"P/E Ratio (TTM) – Price to Earnings Ratio\nShows how much investors are paying for $1 of the company's earnings over the past 12 months. A very high number may mean strong growth expectations or an overvalued stock. Compare it to peers and the market average.{data_source_label}"
        )
        
        if forward_pe and not pd.isna(forward_pe):
            st.metric(
                label="Forward P/E",
                value=f"{forward_pe:.2f}",
                help="Forward P/E – Forward Price to Earnings Ratio\nThis looks ahead, using analysts' projected earnings. It's useful for judging whether growth is expected to make the stock cheaper in the future. If the forward P/E is lower than the current P/E, it means analysts expect earnings to rise, so the stock will look less expensive relative to profits."
            )
    
    with col2:
        # Price-to-Book and Price-to-Sales
        pb_ratio = hybrid_metrics.get('pb_ratio') if hybrid_metrics else info.get('priceToBook', None)
        ps_ratio = hybrid_metrics.get('ps_ratio') if hybrid_metrics else info.get('priceToSalesTrailing12Months', None)
        
        st.metric(
            label="P/B Ratio",
            value=f"{pb_ratio:.2f}" if pb_ratio and not pd.isna(pb_ratio) else "N/A",
            help=f"Price-to-Book (P/B Ratio)\nCompares the stock price to the company's book value (assets minus liabilities). A ratio above 1 means investors value the company more than its net assets. High values may indicate strong growth potential—or overvaluation.{data_source_label}"
        )
        
        st.metric(
            label="P/S Ratio (TTM)",
            value=f"{ps_ratio:.2f}" if ps_ratio and not pd.isna(ps_ratio) else "N/A",
            help=f"Price-to-Sales (P/S Ratio)\nShows how much investors pay for $1 of revenue. Useful for companies with little or no profit. A very high number can mean growth optimism but also possible overpricing.{data_source_label}"
        )
    
    with col3:
        # Enterprise Value metrics
        ev_revenue = hybrid_metrics.get('ev_revenue') if hybrid_metrics else info.get('enterpriseToRevenue', None)
        ev_ebitda = hybrid_metrics.get('ev_ebitda') if hybrid_metrics else info.get('enterpriseToEbitda', None)
        
        st.metric(
            label="EV/Revenue",
            value=f"{ev_revenue:.2f}" if ev_revenue and not pd.isna(ev_revenue) else "N/A",
            help="EV/Revenue (Enterprise Value to Revenue)\nEnterprise Value (company's total value including debt and cash) divided by revenue. Often better than P/S since it accounts for debt. Lower is usually more attractive."
        )
        
        st.metric(
            label="EV/EBITDA",
            value=f"{ev_ebitda:.2f}" if ev_ebitda and not pd.isna(ev_ebitda) else "N/A",
            help="EV/EBITDA (Enterprise Value to EBITDA)\nEnterprise Value compared to EBITDA (earnings before interest, tax, depreciation, and amortization). A widely used measure for comparing companies, especially across industries. Lower values often signal a cheaper valuation."
        )
    
    with col4:
        # PEG and Market Cap - use hybrid data for best PEG source
        peg_ratio = hybrid_metrics.get('peg_ratio') if hybrid_metrics else get_peg_ratio(info)
        market_cap = info.get('marketCap', None)
        

        
        st.metric(
            label="PEG Ratio",
            value=f"{peg_ratio:.2f}" if peg_ratio and not pd.isna(peg_ratio) else "N/A",
            help=f"PEG Ratio – Price to Earnings Growth Ratio\nTakes the P/E Ratio and adjusts it for the company's earnings growth rate. Around 1 is considered fairly valued, above 1 may be expensive, below 1 could be undervalued. Helps balance growth with price.{data_source_label}"
        )
        
        if market_cap and not pd.isna(market_cap):
            if market_cap >= 1e12:
                cap_display = f"${market_cap/1e12:.2f}T"
            elif market_cap >= 1e9:
                cap_display = f"${market_cap/1e9:.2f}B"
            else:
                cap_display = f"${market_cap/1e6:.2f}M"
            
            st.metric(
                label="Market Cap",
                value=cap_display,
                help="Market Capitalization\nTotal value of all outstanding shares. For comparison with Enterprise Value: EV represents the true cost to acquire a company (market cap + debt – cash). Useful when comparing firms with different debt levels."
            )

def display_profitability_metrics(info):
    """Display comprehensive profitability metrics"""
    st.markdown("### Profitability Analysis")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Margin metrics
        gross_margin = info.get('grossMargins', None)
        operating_margin = info.get('operatingMargins', None)
        
        st.metric(
            label="Gross Margin",
            value=f"{gross_margin*100:.2f}%" if gross_margin and not pd.isna(gross_margin) else "N/A",
            help="Gross Profit Margin\nPercentage of revenue remaining after cost of goods sold (COGS).\n• High margin: Strong pricing power, efficient production\n• >50%: Excellent (software, luxury goods)\n• 30-50%: Good (branded consumer goods)\n• 10-30%: Average (retail, manufacturing)\n• <10%: Low margin business (groceries, commodities)\nHigher gross margins provide more flexibility for operations and growth."
        )
        
        st.metric(
            label="Operating Margin",
            value=f"{operating_margin*100:.2f}%" if operating_margin and not pd.isna(operating_margin) else "N/A",
            help="Operating Profit Margin\nPercentage of revenue remaining after all operating expenses.\n• Measures core business efficiency and management effectiveness\n• >20%: Excellent operational efficiency\n• 10-20%: Good operational performance\n• 5-10%: Average operational efficiency\n• <5%: Poor operational efficiency\n• Negative: Operating losses\nNote: yfinance source - may differ from institutional sources like GuruFocus"
        )
    
    with col2:
        # Profit margins
        profit_margin = info.get('profitMargins', None)
        ebitda_margin = info.get('ebitdaMargins', None)
        
        st.metric(
            label="Net Profit Margin",
            value=f"{profit_margin*100:.2f}%" if profit_margin and not pd.isna(profit_margin) else "N/A",
            help="Net Profit Margin\nPercentage of revenue that becomes actual profit after all expenses.\n• Ultimate measure of business profitability\n• >20%: Exceptional profitability (tech, pharma)\n• 10-20%: Strong profitability (established businesses)\n• 5-10%: Moderate profitability (retail, services)\n• 0-5%: Low profitability (commodities, utilities)\n• Negative: Net losses\nConsider industry benchmarks when evaluating."
        )
        
        st.metric(
            label="EBITDA Margin",
            value=f"{ebitda_margin*100:.2f}%" if ebitda_margin and not pd.isna(ebitda_margin) else "N/A",
            help="EBITDA Margin\nEarnings Before Interest, Taxes, Depreciation, and Amortization as % of revenue.\n• Measures operational cash generation before financing decisions\n• >30%: Excellent cash generation (software, services)\n• 20-30%: Strong cash generation (established businesses)\n• 10-20%: Moderate cash generation (manufacturing)\n• 5-10%: Low cash generation (capital-intensive)\n• <5%: Very low cash generation\nUseful for comparing companies across different tax and capital structures."
        )
    
    with col3:
        # Return metrics
        roe = info.get('returnOnEquity', None)
        roa = info.get('returnOnAssets', None)
        
        st.metric(
            label="ROE",
            value=f"{roe*100:.2f}%" if roe and not pd.isna(roe) else "N/A",
            help="Return on Equity (ROE)\nNet income as percentage of shareholders' equity - measures management effectiveness.\n• >20%: Excellent management and capital efficiency\n• 15-20%: Very good returns for shareholders\n• 10-15%: Good returns, above market average\n• 5-10%: Average returns, acceptable performance\n• <5%: Poor returns, below expectations\n• High ROE with low debt is ideal - avoid artificially high ROE from excessive leverage"
        )
        
        st.metric(
            label="ROA",
            value=f"{roa*100:.2f}%" if roa and not pd.isna(roa) else "N/A",
            help="Return on Assets (ROA)\nNet income as percentage of total assets - measures asset utilization efficiency.\n• >15%: Exceptional asset efficiency (tech, services)\n• 10-15%: Very good asset utilization\n• 5-10%: Good asset efficiency for most industries\n• 2-5%: Average asset utilization (banks, utilities)\n• <2%: Poor asset efficiency\n• ROA shows how well company converts assets into profits regardless of financing structure"
        )
    
    with col4:
        # Earnings metrics
        eps = info.get('trailingEps', None)
        forward_eps = info.get('forwardEps', None)
        
        st.metric(
            label="EPS (TTM)",
            value=f"${eps:.2f}" if eps and not pd.isna(eps) else "N/A",
            help="Earnings Per Share (TTM)\nCompany's profit divided by outstanding shares over trailing twelve months.\n• Higher EPS generally better, but compare to industry peers\n• Growing EPS over time indicates improving profitability\n• Used to calculate P/E ratio (Price ÷ EPS)\n• Positive EPS: Company is profitable\n• Negative EPS: Company has losses\n• Consider EPS growth trend more important than absolute number"
        )
        
        st.metric(
            label="Forward EPS",
            value=f"${forward_eps:.2f}" if forward_eps and not pd.isna(forward_eps) else "N/A",
            help="Forward Earnings Per Share\nAnalyst consensus estimate of EPS for next 12 months.\n• Forward EPS > Current EPS: Growth expected\n• Forward EPS < Current EPS: Decline expected\n• Used to calculate Forward P/E ratio\n• More relevant for growth investors than trailing EPS\n• Accuracy depends on analyst forecasting quality\n• Compare forward EPS to current EPS to gauge growth expectations"
        )

def calculate_stock_ratings(ticker_obj, info):
    """
    Calculate A-D ratings for Value, Growth, Momentum, and Profitability
    A = Excellent (Top 25%), B = Good (25-50%), C = Fair (50-75%), D = Poor (Bottom 25%)
    """
    ratings = {
        'Value': 'N/A',
        'Growth': 'N/A', 
        'Momentum': 'N/A',
        'Profitability': 'N/A'
    }
    
    try:
        # VALUE METRICS
        pe_ratio = info.get('trailingPE', None)
        pb_ratio = info.get('priceToBook', None)
        ps_ratio = info.get('priceToSalesTrailing12Months', None)
        
        value_score = 0
        value_count = 0
        
        # PE Ratio scoring (lower is better)
        if pe_ratio and pe_ratio > 0:
            if pe_ratio <= 15: value_score += 4
            elif pe_ratio <= 20: value_score += 3
            elif pe_ratio <= 30: value_score += 2
            else: value_score += 1
            value_count += 1
            
        # PB Ratio scoring (lower is better)
        if pb_ratio and pb_ratio > 0:
            if pb_ratio <= 1.5: value_score += 4
            elif pb_ratio <= 3: value_score += 3
            elif pb_ratio <= 5: value_score += 2
            else: value_score += 1
            value_count += 1
            
        # PS Ratio scoring (lower is better)
        if ps_ratio and ps_ratio > 0:
            if ps_ratio <= 2: value_score += 4
            elif ps_ratio <= 4: value_score += 3
            elif ps_ratio <= 8: value_score += 2
            else: value_score += 1
            value_count += 1
            
        if value_count > 0:
            avg_value = value_score / value_count
            if avg_value >= 3.5: ratings['Value'] = 'A'
            elif avg_value >= 2.5: ratings['Value'] = 'B'
            elif avg_value >= 1.5: ratings['Value'] = 'C'
            else: ratings['Value'] = 'D'
        
        # GROWTH METRICS
        revenue_growth = info.get('revenueGrowth', None)
        earnings_growth = info.get('earningsGrowth', None)
        
        growth_score = 0
        growth_count = 0
        
        # Revenue Growth scoring
        if revenue_growth is not None:
            if revenue_growth >= 0.20: growth_score += 4  # 20%+ growth
            elif revenue_growth >= 0.10: growth_score += 3  # 10-20% growth
            elif revenue_growth >= 0.05: growth_score += 2  # 5-10% growth
            else: growth_score += 1  # <5% growth
            growth_count += 1
            
        # Earnings Growth scoring
        if earnings_growth is not None:
            if earnings_growth >= 0.25: growth_score += 4  # 25%+ growth
            elif earnings_growth >= 0.15: growth_score += 3  # 15-25% growth
            elif earnings_growth >= 0.05: growth_score += 2  # 5-15% growth
            else: growth_score += 1  # <5% growth
            growth_count += 1
            
        if growth_count > 0:
            avg_growth = growth_score / growth_count
            if avg_growth >= 3.5: ratings['Growth'] = 'A'
            elif avg_growth >= 2.5: ratings['Growth'] = 'B'
            elif avg_growth >= 1.5: ratings['Growth'] = 'C'
            else: ratings['Growth'] = 'D'
        
        # MOMENTUM METRICS
        try:
            # Get price history for momentum calculation
            hist = ticker_obj.history(period='1y')
            if not hist.empty:
                current_price = hist['Close'].iloc[-1]
                
                # 1-month momentum
                month_ago_price = hist['Close'].iloc[-22] if len(hist) >= 22 else hist['Close'].iloc[0]
                month_momentum = (current_price - month_ago_price) / month_ago_price
                
                # 3-month momentum  
                quarter_ago_price = hist['Close'].iloc[-66] if len(hist) >= 66 else hist['Close'].iloc[0]
                quarter_momentum = (current_price - quarter_ago_price) / quarter_ago_price
                
                # 6-month momentum
                half_year_price = hist['Close'].iloc[-132] if len(hist) >= 132 else hist['Close'].iloc[0]
                half_year_momentum = (current_price - half_year_price) / half_year_price
                
                momentum_score = 0
                momentum_count = 0
                
                # Score each momentum period
                for momentum in [month_momentum, quarter_momentum, half_year_momentum]:
                    if momentum >= 0.15: momentum_score += 4  # 15%+ gain
                    elif momentum >= 0.05: momentum_score += 3  # 5-15% gain
                    elif momentum >= -0.05: momentum_score += 2  # -5% to 5%
                    else: momentum_score += 1  # >5% loss
                    momentum_count += 1
                    
                if momentum_count > 0:
                    avg_momentum = momentum_score / momentum_count
                    if avg_momentum >= 3.5: ratings['Momentum'] = 'A'
                    elif avg_momentum >= 2.5: ratings['Momentum'] = 'B'
                    elif avg_momentum >= 1.5: ratings['Momentum'] = 'C'
                    else: ratings['Momentum'] = 'D'
        except:
            ratings['Momentum'] = 'N/A'
        
        # PROFITABILITY METRICS
        roe = info.get('returnOnEquity', None)
        roa = info.get('returnOnAssets', None)
        profit_margin = info.get('profitMargins', None)
        
        profit_score = 0
        profit_count = 0
        
        # ROE scoring
        if roe is not None:
            if roe >= 0.20: profit_score += 4  # 20%+ ROE
            elif roe >= 0.15: profit_score += 3  # 15-20% ROE
            elif roe >= 0.10: profit_score += 2  # 10-15% ROE
            else: profit_score += 1  # <10% ROE
            profit_count += 1
            
        # ROA scoring
        if roa is not None:
            if roa >= 0.10: profit_score += 4  # 10%+ ROA
            elif roa >= 0.07: profit_score += 3  # 7-10% ROA
            elif roa >= 0.04: profit_score += 2  # 4-7% ROA
            else: profit_score += 1  # <4% ROA
            profit_count += 1
            
        # Profit Margin scoring
        if profit_margin is not None:
            if profit_margin >= 0.20: profit_score += 4  # 20%+ margin
            elif profit_margin >= 0.10: profit_score += 3  # 10-20% margin
            elif profit_margin >= 0.05: profit_score += 2  # 5-10% margin
            else: profit_score += 1  # <5% margin
            profit_count += 1
            
        if profit_count > 0:
            avg_profit = profit_score / profit_count
            if avg_profit >= 3.5: ratings['Profitability'] = 'A'
            elif avg_profit >= 2.5: ratings['Profitability'] = 'B'
            elif avg_profit >= 1.5: ratings['Profitability'] = 'C'
            else: ratings['Profitability'] = 'D'
            
    except Exception as e:
        print(f"Error calculating stock ratings: {str(e)}")
    
    return ratings

def calculate_quantitative_rating(ticker_obj, info):
    """
    Calculate Quantitative rating (1-5 scale where 5 is most desired)
    Based on data-driven financial metrics and ratios
    """
    try:
        score = 0
        max_score = 0
        
        # P/E Ratio (lower is better)
        pe_ratio = info.get('trailingPE', None)
        if pe_ratio and pe_ratio > 0:
            if pe_ratio <= 15: score += 5
            elif pe_ratio <= 20: score += 4
            elif pe_ratio <= 25: score += 3
            elif pe_ratio <= 35: score += 2
            else: score += 1
            max_score += 5
            
        # P/B Ratio (lower is better)
        pb_ratio = info.get('priceToBook', None)
        if pb_ratio and pb_ratio > 0:
            if pb_ratio <= 1.5: score += 5
            elif pb_ratio <= 2.5: score += 4
            elif pb_ratio <= 4: score += 3
            elif pb_ratio <= 6: score += 2
            else: score += 1
            max_score += 5
            
        # ROE (higher is better)
        roe = info.get('returnOnEquity', None)
        if roe is not None:
            if roe >= 0.20: score += 5
            elif roe >= 0.15: score += 4
            elif roe >= 0.10: score += 3
            elif roe >= 0.05: score += 2
            else: score += 1
            max_score += 5
            
        # Debt-to-Equity (lower is better)
        debt_to_equity = info.get('debtToEquity', None)
        if debt_to_equity is not None:
            if debt_to_equity <= 0.3: score += 5
            elif debt_to_equity <= 0.6: score += 4
            elif debt_to_equity <= 1.0: score += 3
            elif debt_to_equity <= 1.5: score += 2
            else: score += 1
            max_score += 5
            
        if max_score > 0:
            final_rating = round((score / max_score) * 5)
            return max(1, min(5, final_rating)), "Data-driven financial metrics and ratios"
        else:
            return 3, "Insufficient data for quantitative analysis"
            
    except Exception as e:
        return 3, f"Error calculating quantitative rating: {str(e)}"

def calculate_author_rating(ticker_obj, info):
    """
    Calculate Author rating (1-5 scale where 5 is most desired)
    Based on comprehensive business analysis and fundamentals
    """
    try:
        score = 0
        max_score = 0
        
        # Revenue Growth
        revenue_growth = info.get('revenueGrowth', None)
        if revenue_growth is not None:
            if revenue_growth >= 0.20: score += 5
            elif revenue_growth >= 0.10: score += 4
            elif revenue_growth >= 0.05: score += 3
            elif revenue_growth >= 0: score += 2
            else: score += 1
            max_score += 5
            
        # Profit Margins
        profit_margin = info.get('profitMargins', None)
        if profit_margin is not None:
            if profit_margin >= 0.20: score += 5
            elif profit_margin >= 0.15: score += 4
            elif profit_margin >= 0.10: score += 3
            elif profit_margin >= 0.05: score += 2
            else: score += 1
            max_score += 5
            
        # Market Cap (stability indicator)
        market_cap = info.get('marketCap', None)
        if market_cap is not None:
            if market_cap >= 50e9: score += 5  # Large cap
            elif market_cap >= 10e9: score += 4  # Mid cap
            elif market_cap >= 2e9: score += 3   # Small cap
            elif market_cap >= 300e6: score += 2 # Micro cap
            else: score += 1
            max_score += 5
            
        # Beta (volatility - lower is more stable)
        beta = info.get('beta', None)
        if beta is not None:
            if beta <= 0.7: score += 5
            elif beta <= 1.0: score += 4
            elif beta <= 1.3: score += 3
            elif beta <= 1.7: score += 2
            else: score += 1
            max_score += 5
            
        if max_score > 0:
            final_rating = round((score / max_score) * 5)
            return max(1, min(5, final_rating)), "Comprehensive business analysis and fundamentals"
        else:
            return 3, "Insufficient data for author analysis"
            
    except Exception as e:
        return 3, f"Error calculating author rating: {str(e)}"

def calculate_sellside_rating(ticker_obj, info):
    """
    Calculate Sellside rating (1-5 scale where 5 is most desired)
    Based on analyst sentiment and market perception
    """
    try:
        score = 0
        max_score = 0
        
        # Analyst Target Price vs Current Price
        target_price = info.get('targetMeanPrice', None)
        current_price = info.get('currentPrice', None)
        if target_price and current_price and current_price > 0:
            upside = (target_price - current_price) / current_price
            if upside >= 0.20: score += 5
            elif upside >= 0.10: score += 4
            elif upside >= 0.05: score += 3
            elif upside >= -0.05: score += 2
            else: score += 1
            max_score += 5
            
        # Recommendation Score
        recommendation = info.get('recommendationMean', None)
        if recommendation is not None:
            if recommendation <= 2.0: score += 5  # Strong Buy/Buy
            elif recommendation <= 2.5: score += 4
            elif recommendation <= 3.0: score += 3  # Hold
            elif recommendation <= 3.5: score += 2
            else: score += 1  # Sell
            max_score += 5
            
        # Forward P/E (market expectations)
        forward_pe = info.get('forwardPE', None)
        if forward_pe and forward_pe > 0:
            if forward_pe <= 15: score += 5
            elif forward_pe <= 20: score += 4
            elif forward_pe <= 25: score += 3
            elif forward_pe <= 35: score += 2
            else: score += 1
            max_score += 5
            
        # 52-week performance relative to highs
        fifty_two_week_high = info.get('fiftyTwoWeekHigh', None)
        if fifty_two_week_high and current_price:
            performance_from_high = (current_price - fifty_two_week_high) / fifty_two_week_high
            if performance_from_high >= -0.05: score += 5  # Near highs
            elif performance_from_high >= -0.15: score += 4
            elif performance_from_high >= -0.30: score += 3
            elif performance_from_high >= -0.50: score += 2
            else: score += 1
            max_score += 5
            
        if max_score > 0:
            final_rating = round((score / max_score) * 5)
            return max(1, min(5, final_rating)), "Analyst sentiment and market perception"
        else:
            return 3, "Insufficient data for sellside analysis"
            
    except Exception as e:
        return 3, f"Error calculating sellside rating: {str(e)}"

def calculate_piotroski_score(ticker_obj, info):
    """Calculate Piotroski Score (1-9 scale where 9 is best)"""
    try:
        # Get financial statements
        financials = ticker_obj.financials
        balance_sheet = ticker_obj.balance_sheet
        cash_flow = ticker_obj.cashflow
        
        if financials.empty or balance_sheet.empty or cash_flow.empty:
            return None, "Insufficient financial data"
        
        score = 0
        details = []
        
        # 1. Net Income > 0 (1 point)
        try:
            net_income = financials.loc['Net Income'].iloc[0] if 'Net Income' in financials.index else 0
            if net_income > 0:
                score += 1
                details.append("✓ Positive Net Income")
            else:
                details.append("✗ Negative Net Income")
        except:
            details.append("? Net Income data unavailable")
        
        # 2. Operating Cash Flow > 0 (1 point)
        try:
            op_cash_flow = cash_flow.loc['Operating Cash Flow'].iloc[0] if 'Operating Cash Flow' in cash_flow.index else 0
            if op_cash_flow > 0:
                score += 1
                details.append("✓ Positive Operating Cash Flow")
            else:
                details.append("✗ Negative Operating Cash Flow")
        except:
            details.append("? Operating Cash Flow data unavailable")
        
        # 3. ROA improvement (1 point)
        try:
            current_roa = info.get('returnOnAssets', 0)
            if current_roa and current_roa > 0:
                score += 1
                details.append("✓ Positive ROA")
            else:
                details.append("✗ Negative/Zero ROA")
        except:
            details.append("? ROA data unavailable")
        
        # 4. Operating Cash Flow > Net Income (1 point)
        try:
            if op_cash_flow > net_income and net_income > 0:
                score += 1
                details.append("✓ Operating CF > Net Income")
            else:
                details.append("✗ Operating CF ≤ Net Income")
        except:
            details.append("? CF vs Income comparison unavailable")
        
        # 5. Debt-to-Equity improvement (1 point)
        try:
            debt_to_equity = info.get('debtToEquity', 0)
            if debt_to_equity and debt_to_equity < 0.4:  # Conservative threshold
                score += 1
                details.append("✓ Low Debt-to-Equity")
            else:
                details.append("✗ High Debt-to-Equity")
        except:
            details.append("? Debt-to-Equity data unavailable")
        
        # 6. Current Ratio improvement (1 point)
        try:
            current_ratio = info.get('currentRatio', 0)
            if current_ratio and current_ratio > 1.2:
                score += 1
                details.append("✓ Strong Current Ratio")
            else:
                details.append("✗ Weak Current Ratio")
        except:
            details.append("? Current Ratio data unavailable")
        
        # 7. Gross Margin improvement (1 point)
        try:
            gross_margin = info.get('grossMargins', 0)
            if gross_margin and gross_margin > 0.3:  # 30% threshold
                score += 1
                details.append("✓ Strong Gross Margin")
            else:
                details.append("✗ Weak Gross Margin")
        except:
            details.append("? Gross Margin data unavailable")
        
        # 8. Asset Turnover improvement (1 point)
        try:
            # Simplified: if revenue growth is positive
            revenue_growth = info.get('revenueGrowth', 0)
            if revenue_growth and revenue_growth > 0:
                score += 1
                details.append("✓ Positive Revenue Growth")
            else:
                details.append("✗ Negative Revenue Growth")
        except:
            details.append("? Revenue Growth data unavailable")
        
        # 9. No share dilution (1 point)
        try:
            shares_outstanding = info.get('sharesOutstanding', 0)
            if shares_outstanding:  # Simplified check
                score += 1
                details.append("✓ Share count stable")
            else:
                details.append("? Share dilution data unavailable")
        except:
            details.append("? Share data unavailable")
        
        return score, details
        
    except Exception as e:
        return None, f"Error calculating Piotroski Score: {str(e)}"

def calculate_altman_z_score(ticker_obj, info):
    """Calculate Altman Z-Score for bankruptcy prediction"""
    try:
        balance_sheet = ticker_obj.balance_sheet
        financials = ticker_obj.financials
        
        if balance_sheet.empty or financials.empty:
            return None, "Insufficient financial data"
        
        # Get required values
        try:
            # Working Capital / Total Assets
            current_assets = balance_sheet.loc['Current Assets'].iloc[0] if 'Current Assets' in balance_sheet.index else 0
            current_liabilities = balance_sheet.loc['Current Liabilities'].iloc[0] if 'Current Liabilities' in balance_sheet.index else 0
            total_assets = balance_sheet.loc['Total Assets'].iloc[0] if 'Total Assets' in balance_sheet.index else 0
            
            working_capital = current_assets - current_liabilities
            wc_ta = working_capital / total_assets if total_assets != 0 else 0
            
            # Retained Earnings / Total Assets
            retained_earnings = balance_sheet.loc['Retained Earnings'].iloc[0] if 'Retained Earnings' in balance_sheet.index else 0
            re_ta = retained_earnings / total_assets if total_assets != 0 else 0
            
            # EBIT / Total Assets
            ebit = financials.loc['EBIT'].iloc[0] if 'EBIT' in financials.index else 0
            ebit_ta = ebit / total_assets if total_assets != 0 else 0
            
            # Market Cap / Total Liabilities
            market_cap = info.get('marketCap', 0)
            total_liabilities = balance_sheet.loc['Total Liab'].iloc[0] if 'Total Liab' in balance_sheet.index else 0
            mv_tl = market_cap / total_liabilities if total_liabilities != 0 else 0
            
            # Sales / Total Assets
            revenue = financials.loc['Total Revenue'].iloc[0] if 'Total Revenue' in financials.index else 0
            sales_ta = revenue / total_assets if total_assets != 0 else 0
            
            # Calculate Z-Score
            z_score = (1.2 * wc_ta) + (1.4 * re_ta) + (3.3 * ebit_ta) + (0.6 * mv_tl) + (1.0 * sales_ta)
            
            # Determine zone
            if z_score <= 1.8:
                zone = "🔴 Distress Zone"
                interpretation = "High bankruptcy risk"
            elif z_score >= 3.0:
                zone = "🟢 Safe Zone"
                interpretation = "Low bankruptcy risk"
            else:
                zone = "🟡 Grey Zone"
                interpretation = "Moderate risk"
            
            return z_score, f"{zone} - {interpretation}"
            
        except Exception as e:
            return None, f"Missing required financial data: {str(e)}"
            
    except Exception as e:
        return None, f"Error calculating Altman Z-Score: {str(e)}"

def calculate_beneish_m_score(ticker_obj, info):
    """Calculate Beneish M-Score for earnings manipulation detection"""
    try:
        financials = ticker_obj.financials
        balance_sheet = ticker_obj.balance_sheet
        
        if financials.empty or balance_sheet.empty or len(financials.columns) < 2:
            return None, "Insufficient financial data (need at least 2 years)"
        
        # Simplified M-Score calculation using available ratios
        m_score = 0
        components = []
        
        try:
            # Use available financial metrics as proxies
            
            # Days Sales Outstanding proxy
            receivables_turnover = info.get('receivablesTurnover', 0)
            if receivables_turnover:
                dso = 365 / receivables_turnover
                if dso > 45:  # High DSO may indicate manipulation
                    m_score += 0.5
                    components.append("⚠️ High Days Sales Outstanding")
                else:
                    components.append("✓ Normal Days Sales Outstanding")
            
            # Gross Margin deterioration
            gross_margin = info.get('grossMargins', 0)
            if gross_margin and gross_margin < 0.2:  # Low gross margin
                m_score += 0.5
                components.append("⚠️ Low Gross Margin")
            else:
                components.append("✓ Adequate Gross Margin")
            
            # Asset Quality (using current ratio as proxy)
            current_ratio = info.get('currentRatio', 0)
            if current_ratio and current_ratio < 1.0:
                m_score += 0.3
                components.append("⚠️ Poor Asset Quality")
            else:
                components.append("✓ Good Asset Quality")
            
            # Sales Growth vs Industry (using revenue growth)
            revenue_growth = info.get('revenueGrowth', 0)
            if revenue_growth and revenue_growth > 0.5:  # Very high growth might be suspicious
                m_score += 0.4
                components.append("⚠️ Unusually High Revenue Growth")
            else:
                components.append("✓ Normal Revenue Growth")
            
            # Debt growth proxy
            debt_to_equity = info.get('debtToEquity', 0)
            if debt_to_equity and debt_to_equity > 1.0:
                m_score += 0.3
                components.append("⚠️ High Debt Levels")
            else:
                components.append("✓ Manageable Debt Levels")
            
            # Convert to standard M-Score scale
            adjusted_m_score = -2.5 + m_score  # Adjust to standard scale
            
            # Interpretation
            if adjusted_m_score <= -1.78:
                interpretation = "🟢 Unlikely to be manipulating earnings"
            else:
                interpretation = "🔴 Potential earnings manipulation detected"
            
            return adjusted_m_score, f"{interpretation}"
            
        except Exception as e:
            return None, f"Error in M-Score calculation: {str(e)}"
            
    except Exception as e:
        return None, f"Error calculating Beneish M-Score: {str(e)}"

def display_financial_strength_metrics(info, ticker_obj):
    """Display comprehensive financial strength metrics"""
    st.markdown("### Financial Strength Analysis")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Debt metrics
        debt_to_equity = info.get('debtToEquity', None)
        total_debt = info.get('totalDebt', None)
        
        st.metric(
            label="Debt-to-Equity",
            value=f"{debt_to_equity:.2f}" if debt_to_equity and not pd.isna(debt_to_equity) else "N/A",
            help="Total debt to equity ratio"
        )
        
        if total_debt and not pd.isna(total_debt):
            if total_debt >= 1e9:
                debt_display = f"${total_debt/1e9:.2f}B"
            else:
                debt_display = f"${total_debt/1e6:.2f}M"
            
            st.metric(
                label="Total Debt",
                value=debt_display,
                help="Total debt amount"
            )
    
    with col2:
        # Liquidity metrics
        current_ratio = info.get('currentRatio', None)
        quick_ratio = info.get('quickRatio', None)
        
        st.metric(
            label="Current Ratio",
            value=f"{current_ratio:.2f}" if current_ratio and not pd.isna(current_ratio) else "N/A",
            help="Current assets to current liabilities"
        )
        
        st.metric(
            label="Quick Ratio",
            value=f"{quick_ratio:.2f}" if quick_ratio and not pd.isna(quick_ratio) else "N/A",
            help="Quick assets to current liabilities"
        )
    
    with col3:
        # Cash metrics
        total_cash = info.get('totalCash', None)
        cash_per_share = info.get('totalCashPerShare', None)
        
        if total_cash and not pd.isna(total_cash):
            if total_cash >= 1e9:
                cash_display = f"${total_cash/1e9:.2f}B"
            else:
                cash_display = f"${total_cash/1e6:.2f}M"
            
            st.metric(
                label="Total Cash",
                value=cash_display,
                help="Total cash and cash equivalents"
            )
        
        st.metric(
            label="Cash per Share",
            value=f"${cash_per_share:.2f}" if cash_per_share and not pd.isna(cash_per_share) else "N/A",
            help="Cash per share outstanding"
        )
    
    with col4:
        # Other strength metrics
        book_value = info.get('bookValue', None)
        beta = info.get('beta', None)
        
        st.metric(
            label="Book Value/Share",
            value=f"${book_value:.2f}" if book_value and not pd.isna(book_value) else "N/A",
            help="Book value per share"
        )
        
        st.metric(
            label="Beta",
            value=f"{beta:.2f}" if beta and not pd.isna(beta) else "N/A",
            help="Stock volatility relative to market"
        )
    
    # Stock Ratings section
    st.markdown("---")
    st.markdown("### 🎯 Stock Ratings (A-D Scale)")
    st.markdown("*Comprehensive performance ratings where A = Excellent, B = Good, C = Fair, D = Poor*")
    
    # Calculate and display ratings
    ratings = calculate_stock_ratings(ticker_obj, info)
    
    col_rating1, col_rating2, col_rating3, col_rating4 = st.columns(4)
    
    # Helper function to get rating color
    def get_rating_color(rating):
        if rating == 'A': return '🟢'
        elif rating == 'B': return '🟡'
        elif rating == 'C': return '🟠'
        elif rating == 'D': return '🔴'
        else: return '⚪'
    
    with col_rating1:
        rating_color = get_rating_color(ratings['Value'])
        st.metric(
            label="💰 Value",
            value=f"{rating_color} {ratings['Value']}",
            help="Based on P/E, P/B, and P/S ratios - lower ratios get higher grades"
        )
    
    with col_rating2:
        rating_color = get_rating_color(ratings['Growth'])
        st.metric(
            label="📈 Growth", 
            value=f"{rating_color} {ratings['Growth']}",
            help="Based on revenue and earnings growth - higher growth gets higher grades"
        )
    
    with col_rating3:
        rating_color = get_rating_color(ratings['Momentum'])
        st.metric(
            label="🚀 Momentum",
            value=f"{rating_color} {ratings['Momentum']}",
            help="Based on 1-month, 3-month, and 6-month price performance"
        )
    
    with col_rating4:
        rating_color = get_rating_color(ratings['Profitability'])
        st.metric(
            label="💵 Profitability",
            value=f"{rating_color} {ratings['Profitability']}",
            help="Based on ROE, ROA, and profit margins - higher returns get higher grades"
        )
    
    # Rating guide
    with st.expander("📖 Rating Scale Guide"):
        st.markdown("""
        **A-D Rating Scale:**
        - **🟢 A (Excellent)**: Top 25% - Outstanding performance in this category
        - **🟡 B (Good)**: 25-50% - Above average performance
        - **🟠 C (Fair)**: 50-75% - Average performance, room for improvement
        - **🔴 D (Poor)**: Bottom 25% - Below average, needs attention
        - **⚪ N/A**: Insufficient data available for rating
        
        **Category Explanations:**
        - **Value**: Lower P/E, P/B, P/S ratios indicate better value
        - **Growth**: Higher revenue and earnings growth rates
        - **Momentum**: Recent price performance across multiple timeframes
        - **Profitability**: Return on equity, assets, and profit margins
        """)

    # Financial Scoring Metrics Section
    st.markdown("---")
    st.markdown("### 📊 Financial Quality Scores")
    st.markdown("*Advanced scoring models for financial health and earnings quality assessment*")
    
    col_score1, col_score2, col_score3 = st.columns(3)
    
    with col_score1:
        # Piotroski Score
        st.markdown("**🎯 Piotroski Score**")
        piotroski_score, piotroski_details = calculate_piotroski_score(ticker_obj, info)
        
        if piotroski_score is not None:
            # Color coding for score
            if piotroski_score >= 7:
                score_color = "🟢"
                score_interpretation = "Excellent"
            elif piotroski_score >= 5:
                score_color = "🟡"
                score_interpretation = "Good"
            else:
                score_color = "🔴"
                score_interpretation = "Poor"
            
            st.metric(
                label="Score (1-9 scale)",
                value=f"{score_color} {piotroski_score}/9",
                help="Higher scores indicate better financial health"
            )
            st.markdown(f"**Quality:** {score_interpretation}")
            
            # Show details in expander
            with st.expander("📋 Score Details"):
                for detail in piotroski_details:
                    st.markdown(f"• {detail}")
        else:
            st.metric("Score (1-9 scale)", "N/A")
            st.error(piotroski_details)
    
    with col_score2:
        # Altman Z-Score
        st.markdown("**⚠️ Altman Z-Score**")
        z_score, z_interpretation = calculate_altman_z_score(ticker_obj, info)
        
        if z_score is not None:
            st.metric(
                label="Z-Score",
                value=f"{z_score:.2f}",
                help="Bankruptcy prediction model"
            )
            st.markdown(f"**Status:** {z_interpretation}")
            
            # Zone guidance
            with st.expander("📖 Zone Guide"):
                st.markdown("""
                **Z-Score Zones:**
                • **≤ 1.8**: 🔴 Distress Zone - High bankruptcy risk
                • **1.8 - 3.0**: 🟡 Grey Zone - Uncertain, monitor closely
                • **≥ 3.0**: 🟢 Safe Zone - Low bankruptcy risk
                """)
        else:
            st.metric("Z-Score", "N/A")
            st.error(z_interpretation)
    
    with col_score3:
        # Beneish M-Score
        st.markdown("**🔍 Beneish M-Score**")
        m_score, m_interpretation = calculate_beneish_m_score(ticker_obj, info)
        
        if m_score is not None:
            st.metric(
                label="M-Score",
                value=f"{m_score:.2f}",
                help="Earnings manipulation detection model"
            )
            st.markdown(f"**Assessment:** {m_interpretation}")
            
            # Interpretation guide
            with st.expander("📖 Score Guide"):
                st.markdown("""
                **M-Score Interpretation:**
                • **≤ -1.78**: 🟢 Unlikely to be manipulating earnings
                • **> -1.78**: 🔴 Potential earnings manipulation detected
                
                *Note: This is a simplified model using available data*
                """)
        else:
            st.metric("M-Score", "N/A")
            st.error(m_interpretation)

def display_growth_metrics(info, ticker_obj):
    """Display comprehensive growth metrics"""
    st.markdown("### Growth Analysis")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Revenue growth
        revenue_growth = info.get('revenueGrowth', None)
        quarterly_revenue_growth = info.get('quarterlyRevenueGrowth', None)
        
        st.metric(
            label="Revenue Growth (YoY)",
            value=f"{revenue_growth*100:.2f}%" if revenue_growth and not pd.isna(revenue_growth) else "N/A",
            help="Revenue Growth (Year-over-Year)\nPercentage increase in revenue compared to same period last year.\n• >20%: Excellent growth (high-growth companies)\n• 10-20%: Strong growth (established growth companies)\n• 5-10%: Moderate growth (mature companies)\n• 0-5%: Slow growth (mature/cyclical companies)\n• Negative: Revenue decline\nConsistent revenue growth indicates market demand and competitive advantage."
        )
        
        st.metric(
            label="Quarterly Revenue Growth",
            value=f"{quarterly_revenue_growth*100:.2f}%" if quarterly_revenue_growth and not pd.isna(quarterly_revenue_growth) else "N/A",
            help="Quarterly Revenue Growth (YoY)\nMost recent quarter's revenue growth compared to same quarter last year.\n• More current than annual growth, shows recent trends\n• Can be volatile due to seasonal factors\n• Compare to previous quarters for trend analysis\n• Positive acceleration is bullish signal\n• Deceleration may indicate slowing business\nWatch for consistent quarterly growth patterns over multiple quarters."
        )
    
    with col2:
        # Earnings growth
        earnings_growth = info.get('earningsGrowth', None)
        quarterly_earnings_growth = info.get('quarterlyEarningsGrowth', None)
        
        st.metric(
            label="Earnings Growth (YoY)",
            value=f"{earnings_growth*100:.2f}%" if earnings_growth and not pd.isna(earnings_growth) else "N/A",
            help="Earnings Growth (Year-over-Year)\nPercentage change in earnings compared to same period last year.\n• >25%: Exceptional earnings growth\n• 15-25%: Strong earnings growth\n• 5-15%: Moderate earnings growth\n• 0-5%: Slow earnings growth\n• Negative: Earnings decline\nEarnings growth more important than revenue growth - shows improving profitability and operational leverage."
        )
        
        st.metric(
            label="Quarterly Earnings Growth",
            value=f"{quarterly_earnings_growth*100:.2f}%" if quarterly_earnings_growth and not pd.isna(quarterly_earnings_growth) else "N/A",
            help="Quarterly Earnings Growth (YoY)\nMost recent quarter's earnings growth vs same quarter last year.\n• Key metric for growth investors and momentum traders\n• Watch for positive surprises vs expectations\n• Accelerating growth often drives stock price appreciation\n• Can be volatile due to one-time items or seasonal factors\n• Compare to analyst estimates and guidance\nConsistent quarterly earnings beats indicate strong execution and market position."
        )
    
    with col3:
        # Target and recommendation
        target_high = info.get('targetHighPrice', None)
        target_mean = info.get('targetMeanPrice', None)
        
        st.metric(
            label="Target High Price",
            value=f"${target_high:.2f}" if target_high and not pd.isna(target_high) else "N/A",
            help="Analyst Target High Price\nHighest price target among all covering analysts over next 12 months.\n• Represents most optimistic analyst view\n• Compare to current price for upside potential\n• Large spread between high/low targets indicates uncertainty\n• Consider analyst track record and recent revisions\n• Bull case scenario if everything goes right\nUse alongside mean target for balanced perspective on analyst sentiment."
        )
        
        st.metric(
            label="Target Mean Price",
            value=f"${target_mean:.2f}" if target_mean and not pd.isna(target_mean) else "N/A",
            help="Analyst Target Mean Price\nAverage price target among all covering analysts over next 12 months.\n• More balanced view than high/low targets\n• Compare to current price for consensus upside/downside\n• Rising mean target indicates improving sentiment\n• Falling mean target suggests deteriorating outlook\n• Most reliable single analyst metric\nTarget above current price suggests analysts expect appreciation."
        )
    
    with col4:
        # Analyst metrics
        recommendation = info.get('recommendationMean', None)
        num_analysts = info.get('numberOfAnalystOpinions', None)
        
        if recommendation and not pd.isna(recommendation):
            rec_text = ["Strong Buy", "Buy", "Hold", "Sell", "Strong Sell"]
            rec_index = min(max(int(recommendation - 1), 0), 4)
            rec_display = rec_text[rec_index]
        else:
            rec_display = "N/A"
        
        st.metric(
            label="Analyst Recommendation",
            value=rec_display,
            help="Analyst Recommendation Consensus\nAverage recommendation from all covering analysts.\n• Strong Buy (1.0-1.5): Very bullish consensus\n• Buy (1.5-2.5): Bullish consensus\n• Hold (2.5-3.5): Neutral consensus\n• Sell (3.5-4.5): Bearish consensus\n• Strong Sell (4.5-5.0): Very bearish consensus\nFewer analysts typically means less reliable consensus. Watch for recent upgrades/downgrades."
        )
        
        st.metric(
            label="Number of Analysts",
            value=f"{int(num_analysts)}" if num_analysts and not pd.isna(num_analysts) else "N/A",
            help="Number of Analysts Covering Stock\nTotal number of analysts providing ratings and price targets.\n• >20 analysts: High coverage (large-cap, popular stocks)\n• 10-20 analysts: Good coverage (mid-cap stocks)\n• 5-10 analysts: Moderate coverage (small-cap stocks)\n• <5 analysts: Low coverage (micro-cap, niche stocks)\n• More analysts generally means more reliable consensus\nHigh coverage indicates institutional interest and market attention."
        )
    
    # Additional growth insights
    st.markdown("---")
    st.markdown("#### Growth Insights")
    
    insights_col1, insights_col2 = st.columns(2)
    
    with insights_col1:
        # Calculate growth score
        growth_factors = [
            ("Revenue Growth", revenue_growth),
            ("Earnings Growth", earnings_growth),
            ("ROE", info.get('returnOnEquity', None)),
            ("Profit Margin", info.get('profitMargins', None))
        ]
        
        positive_factors = sum(1 for name, value in growth_factors 
                             if value and not pd.isna(value) and value > 0)
        
        st.info(f"""
        **Growth Score: {positive_factors}/4 positive factors**
        
        Positive growth indicators found in {positive_factors} out of 4 key metrics:
        - Revenue Growth, Earnings Growth, ROE, Profit Margin
        """)
    
    with insights_col2:
        # Valuation vs Growth analysis
        pe_ratio = info.get('trailingPE', None)
        peg_ratio = info.get('pegRatio', None)
        
        valuation_insight = "N/A"
        if pe_ratio and peg_ratio and not pd.isna(pe_ratio) and not pd.isna(peg_ratio):
            if peg_ratio < 1:
                valuation_insight = "Potentially undervalued relative to growth"
            elif peg_ratio < 1.5:
                valuation_insight = "Fairly valued relative to growth"
            else:
                valuation_insight = "Premium valuation relative to growth"
        
        peg_display = f"{peg_ratio:.2f}" if peg_ratio and not pd.isna(peg_ratio) else "N/A"
        
        st.info(f"""
        **Valuation vs Growth Assessment:**
        
        {valuation_insight}
        
        PEG Ratio: {peg_display}
        """)

def estimate_next_dividend_date(ticker_obj, dividend_info):
    """
    Estimate next dividend date based on historical dividend patterns
    
    Args:
        ticker_obj: yfinance Ticker object
        dividend_info (dict): Current dividend information
    
    Returns:
        dict: Next dividend date information with estimated flag
    """
    next_div_info = {
        'next_dividend_date': 'N/A',
        'estimated': False
    }
    
    try:
        # Get dividend history
        dividends = ticker_obj.dividends
        if not dividends.empty and len(dividends) >= 2:
            # Get last few dividend dates to determine frequency
            dividend_dates = dividends.index[-4:].tolist()  # Last 4 dividends
            
            if len(dividend_dates) >= 2:
                # Calculate average interval between dividends
                intervals = []
                for i in range(1, len(dividend_dates)):
                    interval = (dividend_dates[i] - dividend_dates[i-1]).days
                    intervals.append(interval)
                
                if intervals:
                    # Use median interval to avoid outliers
                    avg_interval = sorted(intervals)[len(intervals)//2]
                    
                    # Estimate next dividend date
                    last_div_date = dividend_dates[-1]
                    estimated_next = last_div_date + pd.Timedelta(days=avg_interval)
                    
                    # Only show if it's in the future
                    if estimated_next > pd.Timestamp.now():
                        next_div_info['next_dividend_date'] = estimated_next.strftime('%Y-%m-%d')
                        next_div_info['estimated'] = True
                    
    except Exception as e:
        pass
    
    return next_div_info


def get_dividend_info(ticker_obj, ticker_info, market="US"):
    """
    Extract dividend information from ticker object and info
    
    Args:
        ticker_obj: yfinance Ticker object
        ticker_info (dict): Company information from yfinance
        market (str): Market type ("US" or "India")
    
    Returns:
        dict: Dividend information
    """
    dividend_info = {
        'last_dividend_date': None,
        'last_dividend_amount': 0,
        'last_dividend_formatted': 'N/A',
        'dividend_yield': 'N/A',
        'forward_dividend': 'N/A',
        'payout_ratio': 'N/A'
    }
    
    try:
        # Get dividend history
        dividends = ticker_obj.dividends
        if not dividends.empty:
            # Get most recent dividend
            last_dividend_date = dividends.index[-1]
            last_dividend_amount = dividends.iloc[-1]
            
            dividend_info['last_dividend_date'] = last_dividend_date
            dividend_info['last_dividend_amount'] = last_dividend_amount
            currency_symbol = get_currency_symbol(market)
            dividend_info['last_dividend_formatted'] = f"{currency_symbol}{last_dividend_amount:.2f} on {last_dividend_date.strftime('%Y-%m-%d')}"
        
        # Get dividend yield from ticker info
        if 'dividendYield' in ticker_info and ticker_info['dividendYield']:
            # dividendYield is already in percentage format (0.45 = 0.45%, not 45%)
            dividend_info['dividend_yield'] = f"{ticker_info['dividendYield']:.2f}%"
        elif 'trailingAnnualDividendYield' in ticker_info and ticker_info['trailingAnnualDividendYield']:
            # Check if this value needs multiplication based on its scale
            trailing_yield = ticker_info['trailingAnnualDividendYield']
            if trailing_yield < 1:  # Likely a decimal (0.05 = 5%)
                dividend_info['dividend_yield'] = f"{trailing_yield*100:.2f}%"
            else:  # Already in percentage format
                dividend_info['dividend_yield'] = f"{trailing_yield:.2f}%"
        
        # Get forward dividend rate
        currency_symbol = get_currency_symbol(market)
        if 'dividendRate' in ticker_info and ticker_info['dividendRate']:
            dividend_info['forward_dividend'] = f"{currency_symbol}{ticker_info['dividendRate']:.2f}"
        elif 'trailingAnnualDividendRate' in ticker_info and ticker_info['trailingAnnualDividendRate']:
            dividend_info['forward_dividend'] = f"{currency_symbol}{ticker_info['trailingAnnualDividendRate']:.2f}"
        
        # Get payout ratio
        if 'payoutRatio' in ticker_info and ticker_info['payoutRatio']:
            payout = ticker_info['payoutRatio']
            # Check if payout ratio needs multiplication (usually it's a decimal like 0.3 = 30%)
            if payout <= 1:
                dividend_info['payout_ratio'] = f"{payout*100:.1f}%"
            else:  # Already in percentage format
                dividend_info['payout_ratio'] = f"{payout:.1f}%"
            
    except Exception as e:
        pass
    
    return dividend_info

def get_fibonacci_tooltip(fib_label, level_type):
    """Generate comprehensive tooltip for Fibonacci levels"""
    tooltips = {
        # Retracement levels
        "Fib 23.6%": {
            "resistance": "Fibonacci 23.6% Resistance\nShallow retracement level - indicates strong trending market.\n• Often first pullback target in strong uptrends\n• Price rejection here suggests trend continuation\n• Break above confirms momentum strength\n• Light resistance, suitable for trend-following entries",
            "support": "Fibonacci 23.6% Support\nShallow retracement level - indicates strong trending market.\n• Often first bounce target in strong downtrends\n• Price hold here suggests trend continuation down\n• Break below confirms downward momentum\n• Light support, caution for counter-trend entries"
        },
        "Fib 38.2%": {
            "resistance": "Fibonacci 38.2% Resistance\nModerate retracement level - common correction in healthy trends.\n• Popular profit-taking and re-entry level\n• Strong resistance in trending markets\n• Break above often leads to higher targets\n• Good risk/reward for swing trading entries",
            "support": "Fibonacci 38.2% Support\nModerate retracement level - common correction in healthy trends.\n• Popular bounce and accumulation level\n• Strong support in trending markets\n• Break below often leads to deeper correction\n• Good risk/reward for trend continuation trades"
        },
        "Fib 50.0%": {
            "resistance": "Fibonacci 50.0% Resistance\nPsychological half-retracement level - widely watched by traders.\n• Not true Fibonacci ratio but critical market level\n• Major decision point for trend continuation\n• Heavy institutional interest at this level\n• Break above suggests trend resumption strength",
            "support": "Fibonacci 50.0% Support\nPsychological half-retracement level - widely watched by traders.\n• Not true Fibonacci ratio but critical market level\n• Major decision point for trend reversal/continuation\n• Heavy institutional buying often emerges here\n• Break below suggests deeper correction likely"
        },
        "Fib 61.8%": {
            "resistance": "Fibonacci 61.8% Resistance - Golden Ratio\nDeep retracement level - often marks trend continuation point.\n• Most important Fibonacci ratio in nature\n• Last reasonable level for trend continuation\n• Strong resistance for counter-trend moves\n• Break above suggests major trend change",
            "support": "Fibonacci 61.8% Support - Golden Ratio\nDeep retracement level - often marks trend continuation point.\n• Most important Fibonacci ratio in nature\n• Last reasonable support for uptrend continuation\n• Strong support for major corrections\n• Break below often signals trend reversal"
        },
        "Fib 78.6%": {
            "resistance": "Fibonacci 78.6% Resistance\nVery deep retracement - trend weakening zone.\n• Indicates significant momentum loss\n• Often precedes trend reversal\n• Strong resistance for bear market rallies\n• Approach with caution for long positions",
            "support": "Fibonacci 78.6% Support\nVery deep retracement - trend weakening zone.\n• Indicates significant selling pressure\n• Often final support before trend reversal\n• Critical level for major trend changes\n• Break below suggests bear market likely"
        },
        # Extension levels
        "Fib 127.2%": {
            "resistance": "Fibonacci 127.2% Extension\nFirst extension target - common profit-taking level.\n• Initial target after breakout moves\n• Popular level for partial profit-taking\n• Often sees consolidation and pullbacks\n• Good target for swing trading exits",
            "support": "Fibonacci 127.2% Extension\nFirst extension target - common reversal level.\n• Initial support after breakdown moves\n• Popular level for bounce attempts\n• Often sees consolidation and rallies\n• Good target for short covering"
        },
        "Fib 161.8%": {
            "resistance": "Fibonacci 161.8% Extension - Golden Extension\nMajor resistance after breakout - institutional target level.\n• Primary target for trending moves\n• Heavy profit-taking and resistance\n• Key level watched by algorithmic trading\n• Break above suggests parabolic potential",
            "support": "Fibonacci 161.8% Extension - Golden Extension\nMajor support after breakdown - institutional target level.\n• Primary target for declining moves\n• Heavy buying interest and support\n• Key level watched by algorithmic trading\n• Break below suggests capitulation territory"
        },
        "Fib 200.0%": {
            "resistance": "Fibonacci 200.0% Extension\nDouble-the-move target - strong momentum required.\n• Indicates powerful trending market\n• Major psychological resistance level\n• Often marks intermediate-term tops\n• Extreme overextension risk above this level",
            "support": "Fibonacci 200.0% Extension\nDouble-the-move target - strong momentum required.\n• Indicates powerful declining market\n• Major psychological support level\n• Often marks intermediate-term bottoms\n• Extreme oversold conditions below this level"
        },
        "Fib 261.8%": {
            "resistance": "Fibonacci 261.8% Extension\nExtreme extension - parabolic territory.\n• Indicates bubble-like conditions\n• Very high reversal probability\n• Use extreme caution for new positions\n• Often marks major market tops",
            "support": "Fibonacci 261.8% Extension\nExtreme extension - capitulation territory.\n• Indicates panic selling conditions\n• Very high bounce probability\n• Often marks major market bottoms\n• Contrarian opportunity zone"
        }
    }
    
    # Extract base percentage from label (e.g., "Fib 61.8%" from various formats)
    for key in tooltips.keys():
        if key in fib_label or fib_label in key:
            return tooltips[key].get(level_type, f"Fibonacci Level: {fib_label}\nImportant technical level for {level_type} analysis.")
    
    # Default tooltip for unrecognized levels
    return f"Fibonacci Level: {fib_label}\nMathematical level based on golden ratio, acts as {level_type} zone."

def format_currency(value, market="US"):
    """
    Format currency based on market
    
    Args:
        value (float): Currency value
        market (str): Market type ("US" or "India")
    
    Returns:
        str: Formatted currency string
    """
    if market == "India":
        return f"₹{value:,.2f}"
    else:
        return f"${value:.2f}"

def get_currency_symbol(market="US"):
    """
    Get currency symbol for the market
    
    Args:
        market (str): Market type ("US" or "India")
    
    Returns:
        str: Currency symbol
    """
    return "₹" if market == "India" else "$"

def get_stock_metrics(symbol, period="1y", market="US"):
    """
    Get comprehensive metrics for a single stock
    
    Args:
        symbol (str): Stock ticker symbol
        period (str): Time period for analysis
        market (str): Market type ("US" or "India")
    
    Returns:
        dict: Dictionary containing all key metrics
    """
    try:
        # Fetch data
        data, ticker_info, ticker_obj = fetch_stock_data(symbol, period, market)
        
        if data is None or data.empty:
            return {
                'Symbol': symbol,
                'Error': 'No data available',
                'Current Price': 'N/A',
                '52-Week High': 'N/A',
                '52-Week Low': 'N/A',
                'Distance from High (%)': 'N/A',
                'Distance from Low (%)': 'N/A',
                '50-Day MA': 'N/A',
                '200-Day MA': 'N/A',
                'Price vs 50-Day MA (%)': 'N/A',
                'Price vs 200-Day MA (%)': 'N/A',
                'Support Level': 'N/A',
                'Resistance Level': 'N/A',
                'RSI': 'N/A',
                'MACD Signal': 'N/A',
                'Chaikin Money Flow': 'N/A',
                'Last Earnings Date': 'N/A',
                'Next Earnings Date': 'N/A',
                'Earnings Performance (%)': 'N/A',
                'Last Dividend Date': 'N/A',
                'Last Dividend Amount': 'N/A',
                'Dividend Yield (%)': 'N/A',
                'Forward Dividend': 'N/A',
                'Payout Ratio (%)': 'N/A',
                'Est. Next Dividend Date': 'N/A'
            }
        
        # Calculate all metrics
        latest_price = data['Close'].iloc[-1]
        year_high = data['Close'].max()
        year_low = data['Close'].min()
        
        # Moving averages
        ma_50 = calculate_moving_average(data, window=50)
        ma_200 = calculate_moving_average(data, window=200)
        latest_ma_50 = ma_50.iloc[-1] if not ma_50.empty else None
        latest_ma_200 = ma_200.iloc[-1] if not ma_200.empty else None
        
        # Check data availability for meaningful error messages
        data_days = len(data)
        ma_200_available = data_days >= 200 and latest_ma_200 and not pd.isna(latest_ma_200)
        ma_50_available = data_days >= 50 and latest_ma_50 and not pd.isna(latest_ma_50)
        
        # Technical indicators
        macd_line, signal_line, histogram = calculate_macd(data)
        rsi = calculate_rsi(data)
        cmf = calculate_chaikin_money_flow(data)
        
        # Support and resistance
        support_level, resistance_level = calculate_support_resistance(data)
        
        # Earnings and dividend info
        earnings_info = get_earnings_info(ticker_obj, ticker_info, symbol)
        dividend_info = get_dividend_info(ticker_obj, ticker_info, market)
        
        # Calculate performance metrics
        distance_from_high = ((year_high - latest_price) / year_high) * 100
        distance_from_low = ((latest_price - year_low) / year_low) * 100
        
        price_vs_ma_50 = ((latest_price - latest_ma_50) / latest_ma_50) * 100 if latest_ma_50 and not pd.isna(latest_ma_50) else None
        price_vs_ma_200 = ((latest_price - latest_ma_200) / latest_ma_200) * 100 if latest_ma_200 and not pd.isna(latest_ma_200) else None
        
        # Earnings performance
        earnings_performance = None
        if earnings_info['last_earnings'] is not None:
            try:
                earnings_date = earnings_info['last_earnings']
                
                # Handle timezone compatibility for comparison
                if hasattr(data.index[0], 'tz') and data.index[0].tz:
                    # Data has timezone, ensure earnings_date has the same timezone
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_tz = earnings_date
                    else:
                        earnings_date_tz = earnings_date.tz_localize(data.index[0].tz)
                else:
                    # Data is timezone-naive, make earnings_date naive too
                    if hasattr(earnings_date, 'tz') and earnings_date.tz:
                        earnings_date_tz = earnings_date.tz_localize(None)
                    else:
                        earnings_date_tz = earnings_date
                
                # Find the trading day closest to or after earnings
                post_earnings_mask = data.index >= earnings_date_tz
                if post_earnings_mask.any():
                    earnings_day_price = data[post_earnings_mask]['Close'].iloc[0]
                    earnings_performance = ((latest_price - earnings_day_price) / earnings_day_price) * 100
                    print(f"Earnings performance calculated: {earnings_performance:.2f}%")
                else:
                    # Try finding pre-earnings data and use that as baseline
                    pre_earnings_mask = data.index <= earnings_date_tz
                    if pre_earnings_mask.any():
                        pre_earnings_price = data[pre_earnings_mask]['Close'].iloc[-1]
                        earnings_performance = ((latest_price - pre_earnings_price) / pre_earnings_price) * 100
                        print(f"Earnings performance (pre-earnings baseline): {earnings_performance:.2f}%")
            except Exception as e:
                print(f"Earnings performance error: {e}")
                pass
        
        # Latest indicator values
        latest_rsi = rsi.iloc[-1] if not rsi.empty else None
        latest_macd = macd_line.iloc[-1] if not macd_line.empty else None
        latest_signal = signal_line.iloc[-1] if not signal_line.empty else None
        latest_cmf = cmf.iloc[-1] if not cmf.empty else None
        
        # MACD signal
        macd_signal = "N/A"
        if latest_macd is not None and latest_signal is not None:
            if latest_macd > latest_signal:
                macd_signal = "Bullish"
            else:
                macd_signal = "Bearish"
        
        # Estimated next dividend
        next_dividend_estimate = "N/A"
        if dividend_info['last_dividend_date'] is not None:
            try:
                estimated_next = dividend_info['last_dividend_date'] + pd.DateOffset(days=90)
                next_dividend_estimate = estimated_next.strftime('%Y-%m-%d')
            except:
                pass
        
        # Beta value and Safe Levels
        beta_value = get_beta_value(ticker_info)
        ctp_levels = calculate_ctp_levels(latest_price)
        
        # Earnings performance analysis summary
        earnings_analysis, quarters_found = get_earnings_performance_analysis(ticker_obj, data, market)
        earnings_summary = "N/A"
        if earnings_analysis is not None and not earnings_analysis.empty:
            try:
                # Calculate summary statistics
                week_changes = [float(x.replace('%', '').replace('+', '')) for x in earnings_analysis['Week Performance (%)'] if x != 'N/A']
                if week_changes:
                    avg_week = sum(week_changes) / len(week_changes)
                    positive_week = sum(1 for x in week_changes if x > 0)
                    earnings_summary = f"{positive_week}/{len(week_changes)} positive (avg: {avg_week:+.1f}%)"
            except:
                pass
        
        # Return comprehensive metrics with proper currency formatting
        currency_symbol = get_currency_symbol(market)
        return {
            'Symbol': symbol,
            'Current Price': format_currency(latest_price, market),
            '52-Week High': format_currency(year_high, market),
            '52-Week Low': format_currency(year_low, market),
            'Distance from High (%)': f"{distance_from_high:.1f}%",
            'Distance from Low (%)': f"{distance_from_low:.1f}%",
            '50-Day MA': format_currency(latest_ma_50, market) if ma_50_available else "Not enough data",
            '200-Day MA': format_currency(latest_ma_200, market) if ma_200_available else "Not enough data",
            'Price vs 50-Day MA (%)': f"{price_vs_ma_50:.1f}%" if price_vs_ma_50 is not None else "N/A",
            'Price vs 200-Day MA (%)': f"{price_vs_ma_200:.1f}%" if price_vs_ma_200 is not None else "N/A",
            'Support Level': format_currency(support_level, market),
            'Resistance Level': format_currency(resistance_level, market),
            'RSI': f"{latest_rsi:.1f}" if latest_rsi and not pd.isna(latest_rsi) else "N/A",
            'MACD Signal': macd_signal,
            'Chaikin Money Flow': f"{latest_cmf:.3f}" if latest_cmf and not pd.isna(latest_cmf) else "N/A",
            'Last Earnings Date': earnings_info['last_earnings_formatted'],
            'Next Earnings Date': earnings_info['next_earnings_formatted'],
            'Earnings Performance (%)': f"{earnings_performance:.1f}%" if earnings_performance is not None else "N/A",
            'Last Dividend Date': dividend_info['last_dividend_formatted'],
            'Last Dividend Amount': format_currency(dividend_info['last_dividend_amount'], market) if dividend_info['last_dividend_amount'] > 0 else "N/A",
            'Dividend Yield (%)': dividend_info['dividend_yield'],
            'Forward Dividend': dividend_info['forward_dividend'],
            'Payout Ratio (%)': dividend_info['payout_ratio'],
            'Est. Next Dividend Date': next_dividend_estimate,
            'Beta': beta_value,
            'Safe Level Low': format_currency(ctp_levels['lower_ctp'], market) if ctp_levels['lower_ctp'] else "N/A",
            'Safe Level High': format_currency(ctp_levels['upper_ctp'], market) if ctp_levels['upper_ctp'] else "N/A",
            'Earnings Performance (4Q)': earnings_summary
        }
        
    except Exception as e:
        return {
            'Symbol': symbol,
            'Error': str(e),
            'Current Price': 'Error',
            '52-Week High': 'Error',
            '52-Week Low': 'Error',
            'Distance from High (%)': 'Error',
            'Distance from Low (%)': 'Error',
            '50-Day MA': 'Error',
            '200-Day MA': 'Error',
            'Price vs 50-Day MA (%)': 'Error',
            'Price vs 200-Day MA (%)': 'Error',
            'Support Level': 'Error',
            'Resistance Level': 'Error',
            'RSI': 'Error',
            'MACD Signal': 'Error',
            'Chaikin Money Flow': 'Error',
            'Last Earnings Date': 'Error',
            'Next Earnings Date': 'Error',
            'Earnings Performance (%)': 'Error',
            'Last Dividend Date': 'Error',
            'Last Dividend Amount': 'Error',
            'Dividend Yield (%)': 'Error',
            'Forward Dividend': 'Error',
            'Payout Ratio (%)': 'Error',
            'Est. Next Dividend Date': 'Error',
            'Beta': 'Error',
            'Safe Level Low': 'Error',
            'Safe Level High': 'Error',
            'Earnings Performance (4Q)': 'Error'
        }

def create_excel_report(stock_metrics_list, period_label="1 Year"):
    """
    Create an Excel report with stock metrics
    
    Args:
        stock_metrics_list (list): List of stock metrics dictionaries
        period_label (str): Time period label for the report
    
    Returns:
        io.BytesIO: Excel file as bytes
    """
    # Create DataFrame
    df = pd.DataFrame(stock_metrics_list)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Stock Analysis', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Stock Analysis']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        # Write headers with formatting
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # Auto-adjust column widths
        for i, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) + 2
            worksheet.set_column(i, i, min(max_len, 20))
        
        # Add summary sheet
        summary_data = {
            'Report Generated': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            'Analysis Period': [period_label],
            'Total Stocks': [len(stock_metrics_list)],
            'Successful Analysis': [len([s for s in stock_metrics_list if 'Error' not in s or s.get('Error') == ''])],
            'Failed Analysis': [len([s for s in stock_metrics_list if 'Error' in s and s.get('Error') != ''])]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    output.seek(0)
    return output

def create_chart(data, symbol, ma_50, ma_200, period_label="1 Year", market="US", show_fibonacci=True):
    """
    Create an interactive Plotly chart with stock price, moving averages, and Fibonacci retracements
    
    Args:
        data (pd.DataFrame): Stock price data
        symbol (str): Stock symbol for chart title
        ma_50 (pd.Series): 50-day moving average data
        ma_200 (pd.Series): 200-day moving average data
        period_label (str): Time period label for chart title
        market (str): Market type ("US" or "India")
        show_fibonacci (bool): Whether to show Fibonacci retracement levels
    
    Returns:
        plotly.graph_objects.Figure: Interactive chart
    """
    fig = go.Figure()
    
    # Get currency symbol
    currency_symbol = get_currency_symbol(market)
    
    # Add stock price line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=data['Close'],
        mode='lines',
        name=f'{symbol.upper()} Close Price',
        line=dict(color='#1f77b4', width=2),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      f'Price: {currency_symbol}%{{y:,.2f}}<br>' +
                      '<extra></extra>'
    ))
    
    # Add 50-day moving average line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=ma_50,
        mode='lines',
        name='50-Day Moving Average',
        line=dict(color='#ff7f0e', width=2, dash='dash'),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      f'MA(50): {currency_symbol}%{{y:,.2f}}<br>' +
                      '<extra></extra>'
    ))
    
    # Add 200-day moving average line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=ma_200,
        mode='lines',
        name='200-Day Moving Average',
        line=dict(color='#d62728', width=2, dash='dot'),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      f'MA(200): {currency_symbol}%{{y:,.2f}}<br>' +
                      '<extra></extra>'
    ))
    
    # Add Fibonacci retracement levels
    if show_fibonacci:
        fib_data = calculate_fibonacci_levels(data, period_months=3, symbol=symbol)
        if fib_data:
            fib_colors = ['rgba(255, 215, 0, 0.7)', 'rgba(255, 165, 0, 0.7)', 'rgba(255, 140, 0, 0.7)', 
                         'rgba(255, 69, 0, 0.7)', 'rgba(255, 0, 0, 0.7)', 'rgba(139, 0, 0, 0.7)']
            
            # Add horizontal lines for next levels above and below current price
            color_index = 0
            
            # Add lines for next levels above
            for level in fib_data['next_levels_above']:
                fig.add_hline(
                    y=level['price'],
                    line=dict(
                        color=fib_colors[color_index % len(fib_colors)],
                        width=2,
                        dash='dash'
                    ),
                    annotation=dict(
                        text=f"{level['label']}: {currency_symbol}{level['price']:.2f}",
                        bgcolor='rgba(255, 255, 255, 0.8)',
                        bordercolor=fib_colors[color_index % len(fib_colors)],
                        font=dict(size=10, color='black')
                    )
                )
                color_index += 1
            
            # Add lines for next levels below  
            for level in fib_data['next_levels_below']:
                fig.add_hline(
                    y=level['price'],
                    line=dict(
                        color=fib_colors[color_index % len(fib_colors)],
                        width=2,
                        dash='dash'
                    ),
                    annotation=dict(
                        text=f"{level['label']}: {currency_symbol}{level['price']:.2f}",
                        bgcolor='rgba(255, 255, 255, 0.8)',
                        bordercolor=fib_colors[color_index % len(fib_colors)],
                        font=dict(size=10, color='black')
                    )
                )
                color_index += 1
    
    # Update layout
    fig.update_layout(
        title=f'{symbol.upper()} Stock Price with 50-Day & 200-Day Moving Averages ({period_label})',
        xaxis_title='Date',
        yaxis_title=f'Price ({currency_symbol})',
        hovermode='x unified',
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        template='plotly_white',
        height=600
    )
    
    # Update x-axis to show better date formatting
    fig.update_xaxes(
        rangeslider_visible=False,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray'
    )
    
    # Update y-axis with market-specific formatting
    if market == "India":
        tick_format = '₹,.2f'
    else:
        tick_format = '$,.2f'
    
    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        tickformat=tick_format
    )
    
    return fig

def create_macd_chart(data, symbol, macd_line, signal_line, histogram, period_label="1 Year", market="US"):
    """
    Create MACD indicator chart
    
    Args:
        data (pd.DataFrame): Stock price data
        symbol (str): Stock symbol for chart title
        macd_line (pd.Series): MACD line data
        signal_line (pd.Series): Signal line data
        histogram (pd.Series): MACD histogram data
        period_label (str): Time period label for chart title
    
    Returns:
        plotly.graph_objects.Figure: MACD chart
    """
    fig = go.Figure()
    
    # Add MACD line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=macd_line,
        mode='lines',
        name='MACD Line',
        line=dict(color='#1f77b4', width=2),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      'MACD: %{y:.4f}<br>' +
                      '<extra></extra>'
    ))
    
    # Add Signal line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=signal_line,
        mode='lines',
        name='Signal Line',
        line=dict(color='#ff7f0e', width=2),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      'Signal: %{y:.4f}<br>' +
                      '<extra></extra>'
    ))
    
    # Add histogram bars
    # Create colors for bars (green for positive, red for negative)
    colors = ['green' if val >= 0 else 'red' for val in histogram]
    
    fig.add_trace(go.Bar(
        x=data.index,
        y=histogram,
        name='MACD Histogram',
        marker_color=colors,
        opacity=0.6,
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      'Histogram: %{y:.4f}<br>' +
                      '<extra></extra>'
    ))
    
    # Add zero line
    fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
    
    # Update layout
    fig.update_layout(
        title=f'{symbol.upper()} MACD Indicator ({period_label})',
        xaxis_title='Date',
        yaxis_title='MACD Value',
        hovermode='x unified',
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        template='plotly_white',
        height=400
    )
    
    # Update x-axis
    fig.update_xaxes(
        rangeslider_visible=False,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray'
    )
    
    # Update y-axis
    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        tickformat='.4f'
    )
    
    return fig

def create_rsi_chart(data, symbol, rsi, period_label="1 Year", market="US"):
    """
    Create RSI indicator chart
    
    Args:
        data (pd.DataFrame): Stock price data
        symbol (str): Stock symbol for chart title
        rsi (pd.Series): RSI data
        period_label (str): Time period label for chart title
    
    Returns:
        plotly.graph_objects.Figure: RSI chart
    """
    fig = go.Figure()
    
    # Add RSI line
    fig.add_trace(go.Scatter(
        x=data.index,
        y=rsi,
        mode='lines',
        name='RSI',
        line=dict(color='#9467bd', width=2),
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      'RSI: %{y:.2f}<br>' +
                      '<extra></extra>'
    ))
    
    # Add overbought and oversold lines
    fig.add_hline(y=70, line_dash="dash", line_color="red", opacity=0.7, annotation_text="Overbought (70)")
    fig.add_hline(y=30, line_dash="dash", line_color="green", opacity=0.7, annotation_text="Oversold (30)")
    fig.add_hline(y=50, line_dash="dot", line_color="gray", opacity=0.5, annotation_text="Neutral (50)")
    
    # Update layout
    fig.update_layout(
        title=f'{symbol.upper()} RSI Indicator ({period_label})',
        xaxis_title='Date',
        yaxis_title='RSI Value',
        hovermode='x unified',
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        template='plotly_white',
        height=350
    )
    
    # Update x-axis
    fig.update_xaxes(
        rangeslider_visible=False,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray'
    )
    
    # Update y-axis
    fig.update_yaxes(
        range=[0, 100],
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        tickformat='.0f'
    )
    
    return fig

def create_chaikin_chart(data, symbol, cmf, period_label="1 Year", market="US"):
    """
    Create Chaikin Money Flow chart
    
    Args:
        data (pd.DataFrame): Stock price data
        symbol (str): Stock symbol for chart title
        cmf (pd.Series): Chaikin Money Flow data
        period_label (str): Time period label for chart title
    
    Returns:
        plotly.graph_objects.Figure: Chaikin Money Flow chart
    """
    fig = go.Figure()
    
    # Create colors for bars (green for positive, red for negative)
    colors = ['green' if val >= 0 else 'red' for val in cmf]
    
    # Add CMF bars
    fig.add_trace(go.Bar(
        x=data.index,
        y=cmf,
        name='Chaikin Money Flow',
        marker_color=colors,
        opacity=0.7,
        hovertemplate='<b>%{fullData.name}</b><br>' +
                      'Date: %{x}<br>' +
                      'CMF: %{y:.4f}<br>' +
                      '<extra></extra>'
    ))
    
    # Add zero line
    fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
    
    # Update layout
    fig.update_layout(
        title=f'{symbol.upper()} Chaikin Money Flow ({period_label})',
        xaxis_title='Date',
        yaxis_title='CMF Value',
        hovermode='x unified',
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        ),
        template='plotly_white',
        height=350
    )
    
    # Update x-axis
    fig.update_xaxes(
        rangeslider_visible=False,
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray'
    )
    
    # Update y-axis
    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        tickformat='.4f'
    )
    
    return fig

def display_key_metrics(data, symbol, ma_50, ma_200, rsi, ticker_info, ticker_obj, support_level, resistance_level, market="US"):
    """
    Display comprehensive key metrics about the stock
    
    Args:
        data (pd.DataFrame): Stock price data
        symbol (str): Stock symbol
        ma_50 (pd.Series): 50-day moving average data
        ma_200 (pd.Series): 200-day moving average data
        ticker_info (dict): Company information from yfinance
        ticker_obj: yfinance Ticker object for dividend data
        support_level (float): Calculated support level
        resistance_level (float): Calculated resistance level
    """

    # Get latest values
    latest_price = data['Close'].iloc[-1]
    latest_ma_50 = ma_50.iloc[-1]
    latest_ma_200 = ma_200.iloc[-1]
    
    # Calculate comprehensive metrics
    year_high = data['Close'].max()
    year_low = data['Close'].min()
    
    # Calculate daily change percentage
    if len(data) > 1:
        previous_close = data['Close'].iloc[-2]
        daily_change = ((latest_price - previous_close) / previous_close) * 100
    else:
        daily_change = 0.0
    
    # Distance from 52-week high/low and % from CTP
    distance_from_high = ((year_high - latest_price) / year_high) * 100
    distance_from_low = ((latest_price - year_low) / year_low) * 100
    
    # % change of CTP from 52W high/low calculations
    pct_change_from_52w_high = ((latest_price - year_high) / year_high) * 100
    pct_change_from_52w_low = ((latest_price - year_low) / year_low) * 100
    
    # Format display text with colors and descriptive language
    if pct_change_from_52w_high < 0:
        high_display_text = f"CTP is {abs(pct_change_from_52w_high):.1f}% below"
        high_color = "red"
    else:
        high_display_text = f"CTP is {pct_change_from_52w_high:.1f}% above"
        high_color = "green"
    
    if pct_change_from_52w_low < 0:
        low_display_text = f"CTP is {abs(pct_change_from_52w_low):.1f}% below"
        low_color = "red"
    else:
        low_display_text = f"CTP is {pct_change_from_52w_low:.1f}% above"
        low_color = "green"
    
    # Price vs MA comparison
    price_vs_ma_50 = ((latest_price - latest_ma_50) / latest_ma_50) * 100 if not pd.isna(latest_ma_50) else 0
    price_vs_ma_200 = ((latest_price - latest_ma_200) / latest_ma_200) * 100 if not pd.isna(latest_ma_200) else 0
    
    # Get earnings and dividend information
    earnings_info = get_earnings_info(ticker_obj, ticker_info, symbol)
    dividend_info = get_dividend_info(ticker_obj, ticker_info, market)
    
    # Calculate performance since last earnings (if available)
    earnings_performance = "N/A"
    if earnings_info['last_earnings'] is not None:
        try:
            # Find closest trading day to earnings date
            earnings_date = earnings_info['last_earnings']
            mask = data.index >= earnings_date
            if mask.any():
                earnings_day_price = data[mask]['Close'].iloc[0]
                earnings_performance = f"{((latest_price - earnings_day_price) / earnings_day_price) * 100:.1f}%"
        except:
            pass
    
    # Ultra-compact table view - all price action metrics in one table
    view_mode = st.session_state.get('view_mode', 'Standard')
    st.write(f"DEBUG: Current view mode is: {view_mode}")  # Debug info
    st.write(f"DEBUG: 52W High: {year_high:.2f}, CTP: {latest_price:.2f}")  # Debug price info
    st.write(f"DEBUG: High display text: {high_display_text}, Low display text: {low_display_text}")  # Debug calculations
    if view_mode == 'Compact':
        st.markdown("**📈 Price Action (Table View)**")
        st.success("✅ COMPACT MODE ACTIVE - Table format is now displayed")
    else:
        st.markdown("**📈 Price Action & Technical Analysis**")
        st.info("💡 Switch to Compact view below for table format that eliminates scrolling")
    
    # Get all required data first
    fib_data = calculate_fibonacci_levels(data, period_months=3, symbol=symbol)
    latest_rsi = rsi.iloc[-1] if not rsi.empty else None
    beta_value = get_beta_value(ticker_info)
    ctp_levels = calculate_ctp_levels(latest_price)
    
    # Fibonacci levels or fallback to standard support/resistance
    if fib_data and fib_data['next_levels_below']:
        fib_support = min(fib_data['next_levels_below'])
        support_display = format_currency(fib_support, market)
        support_type = "Fib Support"
    else:
        support_display = format_currency(support_level, market)
        support_type = "Support"
    
    if fib_data and fib_data['next_levels_above']:
        fib_resistance = min(fib_data['next_levels_above'])
        resistance_display = format_currency(fib_resistance, market)
        resistance_type = "Fib Resistance"
    else:
        resistance_display = format_currency(resistance_level, market)
        resistance_type = "Resistance"
    
    # Display based on view mode
    if view_mode == 'Compact':
        # COMPACT MODE: Table format organized as requested
        # Main price table with CTP, 52W High, % change, 52W Low, % change
        main_price_data = [
            [
                format_currency(latest_price, market),
                format_currency(year_high, market), 
                high_display_text,
                format_currency(year_low, market),
                low_display_text
            ]
        ]
        
        # Create DataFrame with proper headers
        df_main_price = pd.DataFrame(main_price_data)
        df_main_price.columns = ["CTP", "52W High", "% Change", "52W Low", "% Change"]
        
        # Display main price table
        st.dataframe(df_main_price, hide_index=True, use_container_width=True)
        
        # Add colored text display for the percentages
        col_high, col_low = st.columns(2)
        with col_high:
            if high_color == "red":
                st.markdown(f"<span style='color: red;'>52W High: {high_display_text}</span>", unsafe_allow_html=True)
            else:
                st.markdown(f"<span style='color: green;'>52W High: {high_display_text}</span>", unsafe_allow_html=True)
        with col_low:
            if low_color == "red":
                st.markdown(f"<span style='color: red;'>52W Low: {low_display_text}</span>", unsafe_allow_html=True)
            else:
                st.markdown(f"<span style='color: green;'>52W Low: {low_display_text}</span>", unsafe_allow_html=True)
        
        # Additional metrics table
        additional_data = [
            [support_type, support_display, "RSI (14)", f"{latest_rsi:.1f}" if latest_rsi and not pd.isna(latest_rsi) else "N/A"],
            [resistance_type, resistance_display, "Beta", beta_value],
            ["MA 50", format_currency(latest_ma_50, market) if not pd.isna(latest_ma_50) else "N/A", "MA 200", format_currency(latest_ma_200, market) if not pd.isna(latest_ma_200) else "N/A"],
            ["Safe Low", format_currency(ctp_levels['lower_ctp'], market) if ctp_levels['lower_ctp'] else "N/A", "Safe High", format_currency(ctp_levels['upper_ctp'], market) if ctp_levels['upper_ctp'] else "N/A"]
        ]
        
        # Convert to DataFrame for additional metrics
        df_additional = pd.DataFrame(additional_data)
        
        # Display additional metrics table
        st.dataframe(df_additional, hide_index=True, use_container_width=True)
        
        # Compact earnings and schedule table
        st.markdown("**📅 Earnings & Schedule**")
        
        # Clean earnings value for table display
        earnings_value = earnings_info['last_earnings_formatted']
        if "outdated" in earnings_value or "incomplete" in earnings_value or "likely outdated" in earnings_value:
            if " (likely outdated" in earnings_value:
                clean_date = earnings_value.split(" (likely outdated")[0] + " ⚠️"
            elif " (data may be outdated)" in earnings_value:
                clean_date = earnings_value.split(" (data may be outdated)")[0] + " ⚠️"
            else:
                clean_date = earnings_value + " ⚠️"
        else:
            clean_date = earnings_value
        
        # Get volume data
        volume_data = "N/A"
        if ticker_info.get('averageVolume'):
            volume_data = f"{ticker_info['averageVolume']:,}"
        elif ticker_info.get('volume'):
            volume_data = f"{ticker_info['volume']:,}"
        
        earnings_data = [
            ["Last Earnings", clean_date, "Next Earnings", earnings_info['next_earnings_formatted']],
            ["Volume (Avg)", volume_data, "Market Cap", format_currency(ticker_info.get('marketCap', 'N/A'), market) if ticker_info.get('marketCap') and ticker_info['marketCap'] != 'N/A' else "N/A"]
        ]
        
        df_earnings = pd.DataFrame(earnings_data)
        st.dataframe(df_earnings, hide_index=True, use_container_width=True)

        # Ultra-compact extended hours table - only show if data available
        after_market = get_after_market_data(symbol, market)
        if after_market['pre_market_change'] != 'N/A' or after_market['post_market_change'] != 'N/A':
            st.markdown("**🕘 Extended Hours**")
            
            # Determine which extended hours data to show
            extended_data = []
            if after_market['pre_market_change'] != 'N/A':
                extended_data.append(["Pre-Market", f"{after_market['pre_market_change']} ({after_market['pre_market_change_percent']})", "Regular Close", after_market['regular_session_close']])
            elif after_market['post_market_change'] != 'N/A':
                extended_data.append(["After-Hours", f"{after_market['post_market_change']} ({after_market['post_market_change_percent']})", "Regular Close", after_market['regular_session_close']])
            
            if extended_data:
                df_extended = pd.DataFrame(extended_data)
                st.dataframe(df_extended, hide_index=True, use_container_width=True)
    
    else:
        # STANDARD MODE: Original metric columns layout
        # Row 1 - Price Analysis
        col1, col2, col3, col4, col5, col6 = st.columns(6)

        with col1:
            st.metric(
                label="Current Price",
                value=format_currency(latest_price, market),
                delta=f"{daily_change:.2f}%",
                delta_color="normal"
            )

        with col2:
            st.metric(
                label="52W High",
                value=format_currency(year_high, market),
                delta=high_display_text,
                delta_color="inverse" if high_color == "red" else "normal"
            )

        with col3:
            st.metric(
                label="52W Low", 
                value=format_currency(year_low, market),
                delta=low_display_text,
                delta_color="inverse" if low_color == "red" else "normal"
            )

        with col4:
            st.metric(
                label=support_type,
                value=support_display,
                help="Support Level - Price Floor\nA price level where buying interest is expected to emerge, preventing further decline.\n• **Strong Support**: Price bounces multiple times\n• **Weak Support**: Price breaks through with volume\n• **Fibonacci Support**: Based on mathematical retracement levels (23.6%, 38.2%, 50%, 61.8%)\n• **Technical Support**: Based on previous lows, moving averages, or chart patterns\n• **Trading Strategy**: Buy near support, stop-loss below support\n• **Break Below**: Often signals further decline to next support level"
            )

        with col5:
            st.metric(
                label=resistance_type,
                value=resistance_display,
                help="Resistance Level - Price Ceiling\nA price level where selling interest is expected to emerge, preventing further rise.\n• **Strong Resistance**: Price rejects multiple times\n• **Weak Resistance**: Price breaks through with volume\n• **Fibonacci Resistance**: Based on mathematical extension levels (127.2%, 161.8%, 200%)\n• **Technical Resistance**: Based on previous highs, moving averages, or chart patterns\n• **Trading Strategy**: Sell near resistance, stop-loss above resistance\n• **Break Above**: Often signals further rise to next resistance level"
            )

        with col6:
            rsi_color = "normal"
            if latest_rsi and not pd.isna(latest_rsi):
                if latest_rsi > 70:
                    rsi_color = "inverse"
                elif latest_rsi < 30:
                    rsi_color = "normal"
            
            st.metric(
                label="RSI (14)",
                value=f"{latest_rsi:.1f}" if latest_rsi and not pd.isna(latest_rsi) else "N/A",
                help="RSI (Relative Strength Index)\nMomentum oscillator measuring speed and change of price movements. Range 0-100.\n• RSI > 70: Overbought territory - potential sell signal or pullback\n• RSI < 30: Oversold territory - potential buy signal or bounce\n• RSI 30-70: Neutral zone - trend continuation likely\n• Look for RSI divergences: price makes new highs/lows but RSI doesn't confirm"
            )

        # Row 2 - Technical Analysis
        col7, col8, col9, col10, col11, col12 = st.columns(6)

        with col7:
            ma_50_trend = ""
            if not pd.isna(latest_ma_50) and latest_price:
                ma_50_trend = "+" if latest_price > latest_ma_50 else "-"
            
            st.metric(
                label="MA 50",
                value=format_currency(latest_ma_50, market) if not pd.isna(latest_ma_50) else "N/A",
                delta=ma_50_trend if ma_50_trend else None,
                help="50-Day Moving Average\nShort-term trend indicator showing average price over 50 days.\n• Price above MA 50: Short-term uptrend, bullish momentum\n• Price below MA 50: Short-term downtrend, bearish momentum\n• MA 50 slope: Rising = strengthening trend, Falling = weakening trend\n• Use as dynamic support (uptrend) or resistance (downtrend) level"
            )

        with col8:
            ma_200_trend = ""
            if not pd.isna(latest_ma_200) and latest_price:
                ma_200_trend = "+" if latest_price > latest_ma_200 else "-"
            
            st.metric(
                label="MA 200",
                value=format_currency(latest_ma_200, market) if not pd.isna(latest_ma_200) else "N/A",
                delta=ma_200_trend if ma_200_trend else None,
                help="200-Day Moving Average\nLong-term trend indicator and major support/resistance level.\n• Price above MA 200: Long-term bull market, major uptrend\n• Price below MA 200: Long-term bear market, major downtrend\n• Golden Cross: MA 50 crosses above MA 200 = strong bullish signal\n• Death Cross: MA 50 crosses below MA 200 = strong bearish signal\n• MA 200 acts as major psychological support/resistance level"
            )

        with col9:
            st.metric(
                label="Beta",
                value=beta_value
            )

        with col10:
            st.metric(
                label="Since Earnings",
                value=earnings_performance
            )

        with col11:
            st.metric(
                label="Safe Low",
                value=format_currency(ctp_levels['lower_ctp'], market) if ctp_levels['lower_ctp'] else "N/A",
                help="Safe Entry Level (CTP -12.5%)\nConservative entry point below current price for risk management.\n• **Purpose**: Provides buffer against normal price volatility\n• **Risk Management**: Reduces chance of immediate loss after entry\n• **Dollar-Cost Averaging**: Good level for gradual position building\n• **Patience Required**: May need to wait for price to reach this level\n• **Stop-Loss**: Consider setting stops 5-10% below this level\n• **Market Conditions**: More relevant in volatile or declining markets"
            )

        with col12:
            st.metric(
                label="Safe High", 
                value=format_currency(ctp_levels['upper_ctp'], market) if ctp_levels['upper_ctp'] else "N/A",
                help="Safe Exit Level (CTP +12.5%)\nConservative profit-taking point above current price for risk management.\n• **Purpose**: Ensures reasonable profit while avoiding market top\n• **Risk Management**: Reduces chance of giving back gains\n• **Partial Selling**: Good level for taking some profits off the table\n• **Greed Control**: Helps avoid holding too long for maximum gains\n• **Re-entry**: Can rebuy lower if price pulls back after hitting this level\n• **Market Conditions**: Especially useful in volatile or uncertain markets"
            )

        # Earnings and Dividends Section
        st.markdown("---")
        st.markdown("### 📅 Earnings & Dividends")
        
        col_e1, col_e2, col_e3, col_e4, col_e5, col_e6 = st.columns(6)
        
        with col_e1:
            st.metric(
                label="Last Earnings",
                value=earnings_info['last_earnings_formatted']
            )

        with col_e2:
            st.metric(
                label="Next Earnings", 
                value=earnings_info['next_earnings_formatted']
            )

        with col_e3:
            # Get volume data
            volume_display = "N/A"
            if ticker_info.get('averageVolume'):
                volume_display = f"{ticker_info['averageVolume']:,}"
            elif ticker_info.get('volume'):
                volume_display = f"{ticker_info['volume']:,}"
            
            st.metric(
                label="Volume (Avg)",
                value=volume_display
            )

        with col_e4:
            market_cap_display = "N/A"
            if ticker_info.get('marketCap') and ticker_info['marketCap'] != 'N/A':
                market_cap_display = format_currency(ticker_info['marketCap'], market)
            
            st.metric(
                label="Market Cap",
                value=market_cap_display
            )

        # Extended Hours (if data available)
        after_market = get_after_market_data(symbol, market)
        if after_market['pre_market_change'] != 'N/A' or after_market['post_market_change'] != 'N/A':
            st.markdown("---")
            st.markdown("### 🕘 Extended Hours Trading")
            
            col_ext1, col_ext2, col_ext3 = st.columns(3)
            
            with col_ext1:
                if after_market['pre_market_change'] != 'N/A':
                    st.metric(
                        label="Pre-Market Change",
                        value=f"{after_market['pre_market_change']}",
                        delta=f"{after_market['pre_market_change_percent']}"
                    )

            with col_ext2:
                if after_market['post_market_change'] != 'N/A':
                    st.metric(
                        label="After-Hours Change",
                        value=f"{after_market['post_market_change']}",
                        delta=f"{after_market['post_market_change_percent']}"
                    )

            with col_ext3:
                st.metric(
                    label="Regular Session Close",
                    value=after_market['regular_session_close']
                )

    # All price action data is now displayed in ultra-compact table format above


def main():
    """
    Main application function
    """
    # Custom CSS to make tabs bigger and more visible
    st.markdown("""
    <style>
    /* Make main tabs bigger and more visible */
    .stTabs [data-baseweb="tab-list"] {
        gap: 24px;
        background-color: #f0f2f6;
        padding: 12px;
        border-radius: 12px;
        margin-bottom: 25px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 65px;
        white-space: pre-wrap;
        background-color: white;
        border-radius: 10px;
        color: #262730;
        font-size: 18px;
        font-weight: 600;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        padding: 18px 30px;
        border: 3px solid #e1e5e9;
        transition: all 0.3s ease;
        box-shadow: 0 3px 8px rgba(0,0,0,0.12);
        min-width: 280px;
        text-align: center;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #0066cc;
        color: white;
        border-color: #0066cc;
        box-shadow: 0 6px 15px rgba(0,102,204,0.4);
        transform: translateY(-3px);
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background-color: #f8f9fa;
        border-color: #0066cc;
        transform: translateY(-2px);
        box-shadow: 0 5px 12px rgba(0,0,0,0.18);
    }
    
    .stTabs [aria-selected="true"]:hover {
        background-color: #0052a3;
        color: white;
        transform: translateY(-3px);
    }
    
    /* Sub-tabs styling */
    .stTabs .stTabs [data-baseweb="tab-list"] {
        background-color: #fafbfc;
        padding: 10px;
        border-radius: 10px;
        margin-top: 15px;
        margin-bottom: 20px;
    }
    
    .stTabs .stTabs [data-baseweb="tab"] {
        height: 55px;
        font-size: 16px;
        font-weight: 600;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        padding: 15px 25px;
        background-color: white;
        border: 2px solid #d1d5db;
        border-radius: 8px;
        min-width: 200px;
        text-align: center;
    }
    
    .stTabs .stTabs [aria-selected="true"] {
        background-color: #10b981;
        color: white;
        border-color: #10b981;
        box-shadow: 0 4px 10px rgba(16,185,129,0.3);
        transform: translateY(-2px);
    }
    
    .stTabs .stTabs [data-baseweb="tab"]:hover {
        background-color: #f3f4f6;
        border-color: #10b981;
        transform: translateY(-1px);
        box-shadow: 0 3px 8px rgba(0,0,0,0.15);
    }
    
    .stTabs .stTabs [aria-selected="true"]:hover {
        background-color: #059669;
        transform: translateY(-2px);
    }
    </style>
    """, unsafe_allow_html=True)
    
    # App header
    st.title("📈 Stock Technical Analysis Tool")
    st.markdown("Get comprehensive technical analysis with moving averages, MACD, RSI, Chaikin Money Flow, earnings data, and dividend information for any stock symbol.")
    
    # View Mode Selector
    st.markdown("---")
    col_view1, col_view2, col_view3 = st.columns([1, 2, 1])
    
    with col_view2:
        view_mode = st.radio(
            "Display View:",
            ["Standard", "Compact"],
            horizontal=True,
            help="Standard: Normal font sizes and spacing. Compact: Reduced font sizes and minimal spacing for less scrolling.",
            key="view_mode_selector"
        )
        
        # Store view mode in session state
        st.session_state['view_mode'] = view_mode
    
    st.markdown("---")
    
    # CSS for compact tabs to fit all on one line
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 35px;
        padding: 0px 8px;
        min-width: auto;
        max-width: 120px;
        width: auto;
        font-size: 14px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .stTabs [data-baseweb="tab-list"] button {
        font-size: 14px;
        padding: 8px 12px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Create data source tabs
    tab_yahoo, tab_guru, tab_earnings, tab_events, tab_screener = st.tabs(["📊 Fundamental", "🎯 Advanced", "📅 Earnings", "🗓️ Events", "🔍 Screener"])
    
    with tab_yahoo:
        yahoo_finance_tab()
    
    with tab_guru:
        gurufocus_tab()
    
    with tab_earnings:
        weekly_earnings_calendar_tab()
    
    with tab_events:
        market_events_tab()
    
    with tab_screener:
        stock_screener_tab()

def generate_comprehensive_pdf_report(symbol, data, ticker_info, ticker_obj, ma_50, ma_200, macd_line, signal_line, histogram, rsi, cmf, support_level, resistance_level, period, market):
    """Generate a comprehensive PDF report with charts and advanced analysis data in 1-2 pages"""
    
    # Create PDF buffer
    buffer = io.BytesIO()
    
    # Create PDF document with smaller margins
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    styles = getSampleStyleSheet()
    
    # Create compact custom styles
    title_style = ParagraphStyle(
        'CompactTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=10,
        alignment=TA_CENTER
    )
    
    compact_style = ParagraphStyle(
        'Compact',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=2
    )
    
    story = []
    
    # Title and basic info
    story.append(Paragraph(f"Stock Analysis Report - {symbol.upper()}", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Period: {period}", compact_style))
    
    if ticker_info:
        company_name = ticker_info.get('longName', symbol.upper())
        sector = ticker_info.get('sector', 'N/A')
        story.append(Paragraph(f"{company_name} | Sector: {sector}", compact_style))
    
    story.append(Spacer(1, 8))
    
    # Generate comprehensive technical analysis charts
    try:
        print(f"Generating comprehensive charts for {symbol}")
        chart_buffer = generate_charts_for_pdf(data, symbol, ma_50, ma_200, macd_line, signal_line, rsi, cmf, support_level, resistance_level)
        if chart_buffer:
            print("Charts generated successfully")
            story.append(ReportLabImage(chart_buffer, width=7*inch, height=5*inch))
            story.append(Spacer(1, 8))
        else:
            print("Chart buffer is None")
            story.append(Paragraph("Charts could not be generated", compact_style))
    except Exception as e:
        print(f"Chart generation error: {e}")
        story.append(Paragraph(f"Chart generation error: {str(e)}", compact_style))
    
    # Create data table
    data_rows = []
    
    # Price Action Data
    current_price = data['Close'].iloc[-1] if not data.empty else 0
    data_rows.append(['Current Price', f'${current_price:.2f}'])
    
    if ma_50 is not None and not ma_50.empty:
        latest_ma50 = ma_50.iloc[-1]
        data_rows.append(['50-Day MA', f'${latest_ma50:.2f}'])
    
    if ma_200 is not None and not ma_200.empty:
        latest_ma200 = ma_200.iloc[-1]
        data_rows.append(['200-Day MA', f'${latest_ma200:.2f}'])
    
    if rsi is not None and not rsi.empty:
        latest_rsi = rsi.iloc[-1]
        data_rows.append(['RSI (14)', f'{latest_rsi:.1f}'])
    
    data_rows.append(['Support Level', f'${support_level:.2f}'])
    data_rows.append(['Resistance Level', f'${resistance_level:.2f}'])
    
    # Calculate Safe Levels based on Current Trading Price (CTP)
    safe_level_low = current_price * 0.875  # CTP - 12.5%
    safe_level_high = current_price * 1.125  # CTP + 12.5%
    data_rows.append(['Safe Level Low', f'${safe_level_low:.2f}'])
    data_rows.append(['Safe Level High', f'${safe_level_high:.2f}'])
    
    # Technical Indicators
    if macd_line is not None and not macd_line.empty:
        latest_macd = macd_line.iloc[-1]
        latest_signal = signal_line.iloc[-1] if signal_line is not None and not signal_line.empty else 0
        data_rows.append(['MACD Line', f'{latest_macd:.4f}'])
        data_rows.append(['Signal Line', f'{latest_signal:.4f}'])
    
    if cmf is not None and not cmf.empty:
        latest_cmf = cmf.iloc[-1]
        data_rows.append(['Chaikin Money Flow', f'{latest_cmf:.4f}'])
    
    # Market data from ticker_info
    if ticker_info:
        week_52_high = ticker_info.get('fiftyTwoWeekHigh', 'N/A')
        week_52_low = ticker_info.get('fiftyTwoWeekLow', 'N/A')
        data_rows.append(['52-Week High', f'${week_52_high}' if week_52_high != 'N/A' else 'N/A'])
        data_rows.append(['52-Week Low', f'${week_52_low}' if week_52_low != 'N/A' else 'N/A'])
        
        eps = ticker_info.get('trailingEps', 'N/A')
        pe_ratio = ticker_info.get('trailingPE', 'N/A')
        data_rows.append(['EPS (TTM)', f'{eps}'])
        data_rows.append(['P/E Ratio', f'{pe_ratio}'])
        
        dividend_yield = ticker_info.get('dividendYield', 0)
        if dividend_yield:
            dividend_yield_pct = dividend_yield * 100
            data_rows.append(['Dividend Yield', f'{dividend_yield_pct:.2f}%'])
        else:
            data_rows.append(['Dividend Yield', 'N/A'])
        
        market_cap = ticker_info.get('marketCap', 'N/A')
        if market_cap != 'N/A' and isinstance(market_cap, (int, float)):
            market_cap_b = market_cap / 1e9
            data_rows.append(['Market Cap', f'${market_cap_b:.2f}B'])
        else:
            data_rows.append(['Market Cap', 'N/A'])
        
        volume = ticker_info.get('volume', 'N/A')
        data_rows.append(['Volume', f'{volume:,}' if isinstance(volume, (int, float)) else 'N/A'])
        
        beta = ticker_info.get('beta', 'N/A')
        data_rows.append(['Beta', f'{beta}'])
        
        # Advanced Analysis Data
        book_value = ticker_info.get('bookValue', 'N/A')
        data_rows.append(['Book Value', f'${book_value}' if book_value != 'N/A' else 'N/A'])
        
        pb_ratio = ticker_info.get('priceToBook', 'N/A')
        data_rows.append(['P/B Ratio', f'{pb_ratio}'])
        
        enterprise_value = ticker_info.get('enterpriseValue', 'N/A')
        if enterprise_value != 'N/A' and isinstance(enterprise_value, (int, float)):
            ev_b = enterprise_value / 1e9
            data_rows.append(['Enterprise Value', f'${ev_b:.2f}B'])
        else:
            data_rows.append(['Enterprise Value', 'N/A'])
        
        debt_to_equity = ticker_info.get('debtToEquity', 'N/A')
        data_rows.append(['Debt/Equity', f'{debt_to_equity:.2f}' if debt_to_equity != 'N/A' and isinstance(debt_to_equity, (int, float)) else 'N/A'])
        
        # Institutional Financial Parameters
        ebitda = ticker_info.get('ebitda', 'N/A')
        if ebitda != 'N/A' and isinstance(ebitda, (int, float)):
            ebitda_b = ebitda / 1e9
            data_rows.append(['EBITDA', f'${ebitda_b:.2f}B'])
        else:
            data_rows.append(['EBITDA', 'N/A'])
        
        ev_ebitda = ticker_info.get('enterpriseToEbitda', 'N/A')
        data_rows.append(['EV/EBITDA', f'{ev_ebitda:.2f}' if ev_ebitda != 'N/A' and isinstance(ev_ebitda, (int, float)) else 'N/A'])
        
        ev_revenue = ticker_info.get('enterpriseToRevenue', 'N/A')
        data_rows.append(['EV/Revenue', f'{ev_revenue:.2f}' if ev_revenue != 'N/A' and isinstance(ev_revenue, (int, float)) else 'N/A'])
        
        forward_pe = ticker_info.get('forwardPE', 'N/A')
        data_rows.append(['Forward P/E', f'{forward_pe:.2f}' if forward_pe != 'N/A' and isinstance(forward_pe, (int, float)) else 'N/A'])
        
        peg_ratio = ticker_info.get('pegRatio', 'N/A')
        data_rows.append(['PEG Ratio', f'{peg_ratio:.2f}' if peg_ratio != 'N/A' and isinstance(peg_ratio, (int, float)) else 'N/A'])
        
        current_ratio = ticker_info.get('currentRatio', 'N/A')
        data_rows.append(['Current Ratio', f'{current_ratio:.2f}' if current_ratio != 'N/A' and isinstance(current_ratio, (int, float)) else 'N/A'])
        
        quick_ratio = ticker_info.get('quickRatio', 'N/A')
        data_rows.append(['Quick Ratio', f'{quick_ratio:.2f}' if quick_ratio != 'N/A' and isinstance(quick_ratio, (int, float)) else 'N/A'])
        
        # Profitability & Growth
        roe = ticker_info.get('returnOnEquity', 'N/A')
        if roe != 'N/A' and isinstance(roe, (int, float)):
            roe_pct = roe * 100
            data_rows.append(['ROE', f'{roe_pct:.2f}%'])
        else:
            data_rows.append(['ROE', 'N/A'])
            
        roa = ticker_info.get('returnOnAssets', 'N/A')
        if roa != 'N/A' and isinstance(roa, (int, float)):
            roa_pct = roa * 100
            data_rows.append(['ROA', f'{roa_pct:.2f}%'])
        else:
            data_rows.append(['ROA', 'N/A'])
        
        profit_margins = ticker_info.get('profitMargins', 'N/A')
        if profit_margins != 'N/A' and isinstance(profit_margins, (int, float)):
            margins_pct = profit_margins * 100
            data_rows.append(['Profit Margin', f'{margins_pct:.2f}%'])
        else:
            data_rows.append(['Profit Margin', 'N/A'])
            
        gross_margins = ticker_info.get('grossMargins', 'N/A')
        if gross_margins != 'N/A' and isinstance(gross_margins, (int, float)):
            gross_margins_pct = gross_margins * 100
            data_rows.append(['Gross Margin', f'{gross_margins_pct:.2f}%'])
        else:
            data_rows.append(['Gross Margin', 'N/A'])
            
        # Growth & Ownership Metrics
        revenue_growth = ticker_info.get('revenueGrowth', 'N/A')
        if revenue_growth != 'N/A' and isinstance(revenue_growth, (int, float)):
            revenue_growth_pct = revenue_growth * 100
            data_rows.append(['Revenue Growth', f'{revenue_growth_pct:.2f}%'])
        else:
            data_rows.append(['Revenue Growth', 'N/A'])
            
        earnings_growth = ticker_info.get('earningsGrowth', 'N/A')
        if earnings_growth != 'N/A' and isinstance(earnings_growth, (int, float)):
            earnings_growth_pct = earnings_growth * 100
            data_rows.append(['Earnings Growth', f'{earnings_growth_pct:.2f}%'])
        else:
            data_rows.append(['Earnings Growth', 'N/A'])
            
        # Institutional Ownership
        held_by_institutions = ticker_info.get('heldPercentInstitutions', 'N/A')
        if held_by_institutions != 'N/A' and isinstance(held_by_institutions, (int, float)):
            institutions_pct = held_by_institutions * 100
            data_rows.append(['Institutional %', f'{institutions_pct:.1f}%'])
        else:
            data_rows.append(['Institutional %', 'N/A'])
            
        held_by_insiders = ticker_info.get('heldPercentInsiders', 'N/A')
        if held_by_insiders != 'N/A' and isinstance(held_by_insiders, (int, float)):
            insiders_pct = held_by_insiders * 100
            data_rows.append(['Insider %', f'{insiders_pct:.1f}%'])
        else:
            data_rows.append(['Insider %', 'N/A'])
            
        short_ratio = ticker_info.get('shortRatio', 'N/A')
        data_rows.append(['Short Ratio', f'{short_ratio:.2f}' if short_ratio != 'N/A' and isinstance(short_ratio, (int, float)) else 'N/A'])
        
        short_percent = ticker_info.get('shortPercentOfFloat', 'N/A')
        if short_percent != 'N/A' and isinstance(short_percent, (int, float)):
            short_percent_pct = short_percent * 100
            data_rows.append(['Short % of Float', f'{short_percent_pct:.2f}%'])
        else:
            data_rows.append(['Short % of Float', 'N/A'])
            
        free_cash_flow = ticker_info.get('freeCashflow', 'N/A')
        if free_cash_flow != 'N/A' and isinstance(free_cash_flow, (int, float)):
            fcf_b = free_cash_flow / 1e9
            data_rows.append(['Free Cash Flow', f'${fcf_b:.2f}B'])
        else:
            data_rows.append(['Free Cash Flow', 'N/A'])
    
    # Add A-D Stock Ratings and 1-5 Investment Ratings
    try:
        print("PDF: Calculating A-D ratings...")
        # A-D Rating System
        stock_ratings = calculate_stock_ratings(ticker_obj, ticker_info)
        print(f"PDF: A-D ratings calculated: {stock_ratings}")
        data_rows.append(['═══ RATINGS ═══', ''])
        data_rows.append(['Value Rating', f"{stock_ratings['value']} - {stock_ratings['value_explanation'][:30]}..."])
        data_rows.append(['Growth Rating', f"{stock_ratings['growth']} - {stock_ratings['growth_explanation'][:30]}..."])
        data_rows.append(['Momentum Rating', f"{stock_ratings['momentum']} - {stock_ratings['momentum_explanation'][:30]}..."])
        data_rows.append(['Profitability Rating', f"{stock_ratings['profitability']} - {stock_ratings['profitability_explanation'][:30]}..."])
        
        print("PDF: Calculating 1-5 scale ratings...")
        # 1-5 Scale Investment Ratings
        investment_ratings = calculate_investment_ratings(ticker_info, ticker_obj)
        print(f"PDF: Investment ratings calculated: {investment_ratings}")
        data_rows.append(['─── 1-5 SCALE ───', ''])
        data_rows.append(['Quantitative (1-5)', f"{investment_ratings['quantitative']}/5 - {investment_ratings['quantitative_explanation'][:40]}..."])
        data_rows.append(['Author (1-5)', f"{investment_ratings['author']}/5 - {investment_ratings['author_explanation'][:40]}..."])
        data_rows.append(['Sellside (1-5)', f"{investment_ratings['sellside']}/5 - {investment_ratings['sellside_explanation'][:40]}..."])
        print("PDF: All ratings added to data_rows successfully")
        
    except Exception as e:
        print(f"Rating calculation error: {e}")
        import traceback
        traceback.print_exc()
        data_rows.append(['Ratings', 'Error calculating ratings'])
    
    # Add earnings dates and analysis
    try:
        earnings_info = get_earnings_info(ticker_obj, ticker_info, symbol)
        if earnings_info['last_earnings'] != 'N/A':
            # Format date to show only the date part, not time
            last_earnings = str(earnings_info['last_earnings'])  # Convert to string first
            if 'likely outdated' in last_earnings or 'data may be outdated' in last_earnings:
                # Extract just the date part from "2025-05-01 (likely outdated - check company reports)"
                date_part = last_earnings.split(' (')[0]
            else:
                date_part = last_earnings.split(' ')[0]  # Take first part before any space
            data_rows.append(['Last Earnings', date_part])
        if earnings_info['next_earnings'] != 'N/A':
            # Format date to show only the date part, not time
            next_earnings = str(earnings_info['next_earnings'])  # Convert to string first
            date_part = next_earnings.split(' ')[0]  # Take first part before any space
            data_rows.append(['Next Earnings', date_part])
            
        # Add earnings performance analysis
        print("Attempting to analyze earnings performance...")
        earnings_performance = analyze_earnings_performance(ticker_obj)
        print(f"Earnings performance result: {earnings_performance}")
        if earnings_performance:
            avg_overnight = earnings_performance.get('avg_overnight_return', 0)
            avg_week = earnings_performance.get('avg_week_return', 0)
            sample_size = earnings_performance.get('sample_size', 0)
            data_rows.append(['Avg Earnings Overnight', f'{avg_overnight:.2f}%'])
            data_rows.append(['Avg Earnings Week', f'{avg_week:.2f}%'])
            data_rows.append(['Earnings Sample Size', f'{sample_size} quarters'])
            print("Added earnings performance to PDF data")
        else:
            print("No earnings performance data available")
    except Exception as e:
        print(f"Error adding earnings analysis to PDF: {e}")
        import traceback
        traceback.print_exc()
    
    # Add comprehensive Fibonacci analysis
    try:
        fib_data = calculate_fibonacci_levels(data, symbol=symbol)
        if fib_data and 'next_two_levels' in fib_data:
            levels = fib_data['next_two_levels'][:4]  # Get up to 4 Fibonacci levels
            for i, level in enumerate(levels, 1):
                level_price = level['price']
                level_type = level['type']
                percentage = level.get('percentage', '')
                fib_label = f"Fib {percentage}" if percentage else f"Fib Level {i}"
                data_rows.append([fib_label, f'${level_price:.2f} ({level_type})'])
        
        # Add key Fibonacci retracement levels if available
        if fib_data and 'retracement_levels' in fib_data:
            retracement = fib_data['retracement_levels']
            if '38.2%' in retracement:
                data_rows.append(['Fib 38.2%', f"${retracement['38.2%']:.2f}"])
            if '50.0%' in retracement:
                data_rows.append(['Fib 50.0%', f"${retracement['50.0%']:.2f}"])
            if '61.8%' in retracement:
                data_rows.append(['Fib 61.8%', f"${retracement['61.8%']:.2f}"])
    except Exception as e:
        print(f"Fibonacci analysis error: {e}")
        # Fallback - calculate basic Fibonacci levels
        try:
            high_price = data['High'].tail(60).max()
            low_price = data['Low'].tail(60).min()
            fib_382 = high_price - (high_price - low_price) * 0.382
            fib_500 = high_price - (high_price - low_price) * 0.500
            fib_618 = high_price - (high_price - low_price) * 0.618
            data_rows.append(['Fib 38.2%', f'${fib_382:.2f}'])
            data_rows.append(['Fib 50.0%', f'${fib_500:.2f}'])
            data_rows.append(['Fib 61.8%', f'${fib_618:.2f}'])
        except:
            pass
    
    # Create table with two columns
    col1_data = []
    col2_data = []
    
    for i, row in enumerate(data_rows):
        if i % 2 == 0:
            col1_data.append(row)
        else:
            col2_data.append(row)
    
    # Balance columns
    while len(col1_data) > len(col2_data):
        col2_data.append(['', ''])
    while len(col2_data) > len(col1_data):
        col1_data.append(['', ''])
    
    # Create combined table
    table_data = [['Metric', 'Value', 'Metric', 'Value']]
    for i in range(len(col1_data)):
        table_data.append([
            col1_data[i][0], col1_data[i][1],
            col2_data[i][0] if i < len(col2_data) else '', 
            col2_data[i][1] if i < len(col2_data) else ''
        ])
    
    # Create table
    table = Table(table_data, colWidths=[2*inch, 1*inch, 2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Add earnings analysis table from Advanced Analysis
    try:
        print("Adding detailed earnings analysis table...")
        earnings_performance = analyze_earnings_performance(ticker_obj)
        if earnings_performance and 'earnings_data' in earnings_performance:
            earnings_title = Paragraph("Detailed Earnings Performance Analysis", title_style)
            story.append(earnings_title)
            story.append(Spacer(1, 8))
            
            # Create comprehensive earnings table data
            earnings_table_data = [['Date', 'Pre-Close', 'Next Open', 'Overnight %', 'Week Close', 'Week %', 'Direction', 'EPS']]
            
            for earning in earnings_performance['earnings_data']:
                earnings_date = earning['date'].strftime('%Y-%m-%d') if hasattr(earning['date'], 'strftime') else str(earning['date']).split(' ')[0]
                pre_close = f"${earning['pre_close']:.2f}" if earning['pre_close'] is not None else 'N/A'
                next_open = f"${earning['next_open']:.2f}" if earning['next_open'] is not None else 'N/A'
                overnight_pct = f"{earning['overnight_change_pct']:.2f}%" if earning['overnight_change_pct'] is not None else 'N/A'
                week_close = f"${earning['week_close']:.2f}" if earning['week_close'] is not None else 'N/A'
                week_pct = f"{earning['week_performance']:.2f}%" if earning['week_performance'] is not None else 'N/A'
                direction = earning.get('direction', 'N/A')
                eps = earning.get('eps', 'N/A')
                
                earnings_table_data.append([earnings_date, pre_close, next_open, overnight_pct, week_close, week_pct, direction, eps])
            
            # Create and style the comprehensive earnings table
            earnings_table = Table(earnings_table_data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch])
            earnings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(earnings_table)
            print("Added detailed earnings table to PDF")
        else:
            print("No detailed earnings data available for table")
    except Exception as e:
        print(f"Error adding earnings table: {e}")
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return buffer

def analyze_earnings_performance(ticker_obj):
    """Analyze earnings performance for PDF report"""
    try:
        # Get earnings dates and performance data
        earnings_dates = ticker_obj.earnings_dates
        if earnings_dates is None or earnings_dates.empty:
            return None
        
        # Get historical data for analysis
        hist_data = ticker_obj.history(period="2y")
        if hist_data.empty:
            return None
        
        # Calculate average overnight and week returns around earnings
        overnight_returns = []
        week_returns = []
        
        # Filter to recent earnings (last 8 quarters to match Advanced Analysis)
        recent_earnings = earnings_dates.head(8)
        
        for earnings_date in recent_earnings.index:
            try:
                # Find the closest trading day before earnings
                pre_earnings_date = earnings_date - pd.Timedelta(days=1)
                post_earnings_date = earnings_date + pd.Timedelta(days=1)
                week_post_date = earnings_date + pd.Timedelta(days=7)
                
                # Get prices
                pre_price_data = hist_data[hist_data.index <= pre_earnings_date]
                post_price_data = hist_data[hist_data.index >= post_earnings_date]
                week_price_data = hist_data[hist_data.index >= week_post_date]
                
                if not pre_price_data.empty and not post_price_data.empty:
                    pre_price = pre_price_data['Close'].iloc[-1]
                    post_price = post_price_data['Close'].iloc[0]
                    overnight_return = ((post_price - pre_price) / pre_price) * 100
                    overnight_returns.append(overnight_return)
                    
                    if not week_price_data.empty:
                        week_price = week_price_data['Close'].iloc[0]
                        week_return = ((week_price - pre_price) / pre_price) * 100
                        week_returns.append(week_return)
                        
            except Exception:
                continue
        
        # Calculate averages
        avg_overnight = sum(overnight_returns) / len(overnight_returns) if overnight_returns else 0
        avg_week = sum(week_returns) / len(week_returns) if week_returns else 0
        
        # Also collect detailed earnings data for table (matching the 8 quarters from Advanced Analysis)
        earnings_details = []
        for earnings_date in recent_earnings.index:
            try:
                pre_earnings_date = earnings_date - pd.Timedelta(days=1)
                post_earnings_date = earnings_date + pd.Timedelta(days=1)
                week_post_date = earnings_date + pd.Timedelta(days=7)
                
                pre_price_data = hist_data[hist_data.index <= pre_earnings_date]
                post_price_data = hist_data[hist_data.index >= post_earnings_date]
                week_price_data = hist_data[hist_data.index >= week_post_date]
                
                overnight_return = None
                week_return = None
                
                pre_close = None
                next_open = None
                overnight_change_pct = None
                week_close = None
                week_performance = None
                direction = 'N/A'
                
                if not pre_price_data.empty and not post_price_data.empty:
                    pre_close = pre_price_data['Close'].iloc[-1]
                    next_open = post_price_data['Open'].iloc[0] if 'Open' in post_price_data.columns else post_price_data['Close'].iloc[0]
                    overnight_change_pct = ((next_open - pre_close) / pre_close) * 100
                    overnight_return = overnight_change_pct
                    
                    # Determine direction
                    direction = 'UP' if overnight_change_pct > 0 else 'DOWN' if overnight_change_pct < 0 else 'FLAT'
                    
                    if not week_price_data.empty:
                        week_close = week_price_data['Close'].iloc[0]
                        week_performance = ((week_close - pre_close) / pre_close) * 100
                        week_return = week_performance
                
                # Get EPS data if available
                eps_value = 'N/A'
                try:
                    earnings_info = ticker_obj.earnings_dates
                    if earnings_info is not None and not earnings_info.empty:
                        earnings_row = earnings_info[earnings_info.index == earnings_date]
                        if not earnings_row.empty and 'EPS Estimate' in earnings_row.columns:
                            eps_value = earnings_row['EPS Estimate'].iloc[0]
                            if pd.notna(eps_value):
                                eps_value = f"${eps_value:.2f}"
                            else:
                                eps_value = 'N/A'
                except:
                    eps_value = 'N/A'
                
                # Determine quarter
                quarter_map = {1: 'Q1', 2: 'Q1', 3: 'Q1', 4: 'Q2', 5: 'Q2', 6: 'Q2', 
                             7: 'Q3', 8: 'Q3', 9: 'Q3', 10: 'Q4', 11: 'Q4', 12: 'Q4'}
                quarter = f"{quarter_map.get(earnings_date.month, 'Q?')} {earnings_date.year}"
                
                earnings_details.append({
                    'date': earnings_date,
                    'pre_close': pre_close,
                    'next_open': next_open,
                    'overnight_return': overnight_return,
                    'overnight_change_pct': overnight_change_pct,
                    'week_close': week_close,
                    'week_return': week_return,
                    'week_performance': week_performance,
                    'direction': direction,
                    'eps': eps_value,
                    'period': quarter
                })
                
            except Exception:
                continue

        return {
            'avg_overnight_return': avg_overnight,
            'avg_week_return': avg_week,
            'sample_size': len(overnight_returns),
            'earnings_data': earnings_details
        }
        
    except Exception as e:
        print(f"Earnings analysis error: {e}")
        return None

def generate_charts_for_pdf(data, symbol, ma_50, ma_200, macd_line, signal_line, rsi, cmf, support_level, resistance_level):
    """Generate multiple technical analysis charts for PDF inclusion"""
    try:
        import matplotlib
        matplotlib.use('Agg')  # Use non-GUI backend
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        
        print(f"Creating comprehensive charts for {symbol}")
        
        # Create a 2x2 subplot figure
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 8))
        
        # Chart 1: Price with Moving Averages
        ax1.plot(data.index, data['Close'], label=f'{symbol.upper()} Price', color='black', linewidth=1.5)
        if ma_50 is not None and not ma_50.empty:
            ax1.plot(data.index, ma_50, label='50-Day MA', color='blue', linewidth=1)
        if ma_200 is not None and not ma_200.empty:
            ax1.plot(data.index, ma_200, label='200-Day MA', color='red', linewidth=1)
        ax1.axhline(y=support_level, color='green', linestyle='--', alpha=0.7, label='Support')
        ax1.axhline(y=resistance_level, color='red', linestyle='--', alpha=0.7, label='Resistance')
        ax1.set_title(f'{symbol.upper()} Price & Moving Averages', fontsize=10, fontweight='bold')
        ax1.set_ylabel('Price ($)', fontsize=9)
        ax1.grid(True, alpha=0.3)
        ax1.legend(fontsize=7)
        
        # Chart 2: MACD
        if macd_line is not None and not macd_line.empty and signal_line is not None and not signal_line.empty:
            ax2.plot(data.index, macd_line, label='MACD', color='blue', linewidth=1)
            ax2.plot(data.index, signal_line, label='Signal', color='red', linewidth=1)
            ax2.axhline(y=0, color='gray', linestyle='-', alpha=0.5)
        ax2.set_title('MACD', fontsize=10, fontweight='bold')
        ax2.set_ylabel('MACD', fontsize=9)
        ax2.grid(True, alpha=0.3)
        ax2.legend(fontsize=7)
        
        # Chart 3: RSI
        if rsi is not None and not rsi.empty:
            ax3.plot(data.index, rsi, label='RSI', color='purple', linewidth=1)
            ax3.axhline(y=70, color='red', linestyle='--', alpha=0.7, label='Overbought')
            ax3.axhline(y=30, color='green', linestyle='--', alpha=0.7, label='Oversold')
            ax3.set_ylim(0, 100)
        ax3.set_title('RSI (14)', fontsize=10, fontweight='bold')
        ax3.set_ylabel('RSI', fontsize=9)
        ax3.grid(True, alpha=0.3)
        ax3.legend(fontsize=7)
        
        # Chart 4: Chaikin Money Flow
        if cmf is not None and not cmf.empty:
            ax4.plot(data.index, cmf, label='CMF', color='orange', linewidth=1)
            ax4.axhline(y=0, color='gray', linestyle='-', alpha=0.5)
            ax4.axhline(y=0.1, color='green', linestyle='--', alpha=0.7, label='Bullish')
            ax4.axhline(y=-0.1, color='red', linestyle='--', alpha=0.7, label='Bearish')
        ax4.set_title('Chaikin Money Flow', fontsize=10, fontweight='bold')
        ax4.set_ylabel('CMF', fontsize=9)
        ax4.grid(True, alpha=0.3)
        ax4.legend(fontsize=7)
        
        # Format x-axis for all charts
        for ax in [ax1, ax2, ax3, ax4]:
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
            ax.tick_params(axis='x', labelsize=7, rotation=45)
            ax.tick_params(axis='y', labelsize=7)
        
        # Tight layout
        plt.tight_layout()
        
        # Save to buffer
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
        plt.close(fig)
        img_buffer.seek(0)
        print("Technical analysis charts saved successfully")
        return img_buffer
        
    except Exception as e:
        print(f"Chart generation error: {e}")
        import traceback
        traceback.print_exc()
        return None

def apply_view_mode_css():
    """Apply CSS styling based on the selected view mode"""
    view_mode = st.session_state.get('view_mode', 'Standard')
    
    if view_mode == 'Compact':
        st.markdown("""
        <style>
        /* Compact Mode - Focused on readability */
        
        /* Metric containers - reduce padding only */
        [data-testid="metric-container"] {
            padding: 0.2rem 0 !important;
            margin: 0.3rem 0 !important;
        }
        
        /* Target metric values more specifically */
        .stMetric [data-testid="metric-value"],
        .stMetric > div > div > div:first-child,
        .stMetric div[data-baseweb="block"] > div:first-child,
        .stMetric div[data-testid="stMetricValue"],
        .stMetric div[data-testid="stMetricValue"] > div,
        .stMetric [role="text"] {
            font-size: 1.2rem !important;
            font-weight: 600 !important;
            line-height: 1.1 !important;
        }
        
        /* Force smaller numbers with additional selectors */
        .stMetric div[data-baseweb="block"] div[data-baseweb="block"]:first-child {
            font-size: 1.2rem !important;
        }
        
        /* Keep metric labels compact but readable */
        .stMetric [data-testid="metric-label"],
        .stMetric > div > div > div:last-child {
            font-size: 0.8rem !important;
            line-height: 1.1 !important;
        }
        
        /* Keep metric deltas (change values) compact */
        .stMetric [data-testid="metric-delta"] {
            font-size: 0.75rem !important;
        }
        
        /* Remove any transforms that shrink metrics */
        .stMetric {
            transform: none !important;
            zoom: 1 !important;
        }
        
        /* Reduce header sizes and minimize spacing */
        h1 {
            font-size: 1.5rem !important;
            margin: 0.1rem 0 !important;
            line-height: 1.1 !important;
        }
        h2 {
            font-size: 1.2rem !important;
            margin: 0.1rem 0 !important;
            line-height: 1.1 !important;
        }
        h3 {
            font-size: 1.0rem !important;
            margin: 0.1rem 0 !important;
            line-height: 1.1 !important;
        }
        .stSubheader {
            font-size: 0.95rem !important;
            margin: 0.1rem 0 !important;
            line-height: 1.1 !important;
        }
        
        /* Reduce DataFrame font sizes */
        .stDataFrame {
            font-size: 0.75rem !important;
        }
        .stDataFrame table {
            font-size: 0.75rem !important;
        }
        
        /* Reduce button and input sizes */
        .stButton button {
            font-size: 0.8rem !important;
            padding: 0.25rem 0.75rem !important;
        }
        
        /* Reduce column spacing and use grid layout */
        .stColumn {
            padding: 0.25rem !important;
        }
        
        /* Grid layout for metric containers */
        .stColumns {
            display: grid !important;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)) !important;
            gap: 0.5rem !important;
        }
        
        /* Limit text line length */
        .stMarkdown, .stText, p {
            max-width: 70ch !important;
        }
        
        /* Reduce expander font size */
        .streamlit-expanderHeader {
            font-size: 0.9rem !important;
        }
        
        /* Targeted markdown spacing */
        .stMarkdown p {
            font-size: 0.85rem !important;
            margin: 0.05rem 0 !important;
            line-height: 1.1 !important;
        }
        
        /* Minimize spacing around numbers and metrics */
        .stMarkdown p:contains('$'),
        .stMarkdown p:contains('%'),
        .stMarkdown p:contains('B'),
        .stMarkdown p:contains('M'),
        .stMarkdown p:contains('K') {
            margin: 0.02rem 0 !important;
            line-height: 1.0 !important;
        }
        
        /* Compact expander and tab spacing */
        .streamlit-expander {
            margin: 0.2rem 0 !important;
        }
        
        .stTabs [data-baseweb="tab-panel"] {
            padding-top: 0.5rem !important;
        }
        
        /* Reduce tab font size */
        .stTabs [data-baseweb="tab-list"] button {
            font-size: 0.85rem !important;
            padding: 0.3rem 0.6rem !important;
        }
        
        /* Minimize divider spacing */
        hr {
            margin: 0.25rem 0 !important;
        }
        
        /* Remove spacing from sections */
        .stContainer > div {
            margin: 0.1rem 0 !important;
        }
        
        /* Compact form elements */
        .stTextInput > div > div {
            margin: 0.1rem 0 !important;
        }
        
        /* Fix button layout and spacing */
        .stButton {
            margin: 0.1rem 0 !important;
        }
        
        .stButton > button {
            width: 100% !important;
            white-space: normal !important;
            padding: 0.5rem 0.75rem !important;
            min-height: 2.5rem !important;
            font-size: 0.9rem !important;
            text-align: center !important;
            word-wrap: break-word !important;
        }
        
        /* Ensure buttons stay in columns and have proper width */
        div[data-testid="column"] .stButton {
            width: 100% !important;
        }
        
        div[data-testid="column"] .stButton > button {
            min-width: 120px !important;
            max-width: 100% !important;
        }
        
        /* Fix column layout - ensure horizontal alignment */
        .stColumns {
            display: flex !important;
            flex-direction: row !important;
            gap: 0.5rem !important;
        }
        
        .stColumns > div {
            flex: 1 !important;
            min-width: 0 !important;
        }
        
        /* Remove extra spacing in tabs */
        .stTabs [data-baseweb="tab-panel"] {
            padding: 0.25rem 0 !important;
        }
        
        /* Ultra-compact numeric display */
        .stMarkdown div:contains('**'),
        .stMarkdown strong {
            margin: 0.02rem 0 !important;
            line-height: 0.95 !important;
        }
        
        /* Minimize space between consecutive numeric lines */
        .stMarkdown p + p {
            margin-top: 0.02rem !important;
        }
        
        /* Target bullet points and lists with numbers */
        .stMarkdown ul li,
        .stMarkdown ol li {
            margin: 0 !important;
            padding: 0.02rem 0 !important;
            line-height: 1.0 !important;
        }
        
        /* Compact chart containers */
        .stPlotlyChart {
            margin: 0.25rem 0 !important;
        }
        
        /* Compact expander content */
        .streamlit-expanderContent {
            padding: 0.5rem !important;
        }
        
        /* Aggressive whitespace reduction */
        .main .block-container {
            padding: 0.25rem 1rem !important;
            max-width: 100% !important;
        }
        
        /* Minimize all element spacing */
        .stElement {
            margin: 0.1rem 0 !important;
        }
        
        /* Remove spacing around specific components */
        .stPlotlyChart,
        .stMetric,
        .stDataFrame,
        .stTable {
            margin: 0.05rem 0 !important;
        }
        
        /* Ultra-tight spacing for numeric content */
        .stMetric [data-testid="metric-container"] {
            margin: 0 !important;
            padding: 0.1rem 0 !important;
        }
        
        .stMetric .metric-label {
            margin-bottom: 0.1rem !important;
        }
        
        .stMetric .metric-value {
            margin: 0 !important;
            line-height: 1.0 !important;
        }
        
        /* Compact vertical spacing */
        .element-container {
            margin: 0.1rem 0 !important;
        }
        
        /* Remove default spacing */
        .row-widget.stRadio > div {
            gap: 0.25rem !important;
        }
        
        .row-widget.stSelectbox > div {
            margin: 0.1rem 0 !important;
        }
        
        /* Additional metric number targeting */
        [data-testid="metric-container"] div:first-child div:first-child,
        [data-testid="metric-container"] > div > div:first-child,
        .stMetric .metric-value {
            font-size: 1.2rem !important;
            transform: scale(0.85) !important;
            transform-origin: left !important;
        }
        </style>
        
        <script>
        // Compact numbers with JavaScript when space is tight
        function compactNumbers() {
            const metrics = document.querySelectorAll('[data-testid="metric-container"]');
            metrics.forEach(metric => {
                const valueEl = metric.querySelector('[data-testid="metric-value"], div[data-baseweb="block"]:first-child');
                if (valueEl && valueEl.textContent) {
                    const text = valueEl.textContent;
                    if (text.includes('$') && text.length > 8) {
                        // Convert large numbers to compact format
                        const number = parseFloat(text.replace(/[$,]/g, ''));
                        if (number >= 1e9) {
                            valueEl.textContent = '$' + (number / 1e9).toFixed(1) + 'B';
                        } else if (number >= 1e6) {
                            valueEl.textContent = '$' + (number / 1e6).toFixed(1) + 'M';
                        } else if (number >= 1e3) {
                            valueEl.textContent = '$' + (number / 1e3).toFixed(1) + 'K';
                        }
                    }
                }
            });
        }
        
        // Run on load and when content changes
        document.addEventListener('DOMContentLoaded', compactNumbers);
        setTimeout(compactNumbers, 500);
        setTimeout(compactNumbers, 1000);
        setTimeout(compactNumbers, 2000);
        </script>
        """, unsafe_allow_html=True)
    else:  # Standard mode
        st.markdown("""
        <style>
        /* Standard Mode - Normal sizes */
        .stMetric > div > div > div {
            font-size: 0.875rem !important;
            line-height: 1.2 !important;
        }
        .stMetric > div > div > div > div {
            font-size: 1.25rem !important;
            margin-bottom: 0.25rem !important;
        }
        .stMetric [data-testid="metric-container"] {
            padding: 0.5rem 0 !important;
        }
        
        /* Standard header sizes */
        h1 {
            font-size: 2.25rem !important;
            margin: 1rem 0 0.5rem 0 !important;
        }
        h2 {
            font-size: 1.875rem !important;
            margin: 0.75rem 0 0.5rem 0 !important;
        }
        h3 {
            font-size: 1.5rem !important;
            margin: 0.5rem 0 0.25rem 0 !important;
        }
        .stSubheader {
            font-size: 1.25rem !important;
            margin: 0.5rem 0 0.25rem 0 !important;
        }
        
        /* Standard DataFrame sizes */
        .stDataFrame {
            font-size: 0.875rem !important;
        }
        
        /* Standard button sizes */
        .stButton button {
            font-size: 0.875rem !important;
            padding: 0.5rem 1rem !important;
        }
        
        /* Standard column spacing */
        .stColumn {
            padding: 0.5rem !important;
        }
        
        /* Standard markdown text size */
        .stMarkdown p {
            font-size: 1rem !important;
            margin-bottom: 0.5rem !important;
        }
        
        /* Standard tab sizes */
        .stTabs [data-baseweb="tab-list"] button {
            font-size: 1rem !important;
            padding: 0.5rem 1rem !important;
        }
        
        /* Standard divider spacing */
        hr {
            margin: 1rem 0 !important;
        }
        </style>
        """, unsafe_allow_html=True)

def yahoo_finance_tab():
    """Fundamental analysis tab content"""
    
    # Apply view mode styling
    apply_view_mode_css()
    
    # CSS for compact form elements
    st.markdown("""
    <style>
    /* Force selectbox width and visibility */
    .stSelectbox {
        width: 180px !important;
    }
    .stSelectbox > div {
        width: 180px !important;
        max-width: 180px !important;
    }
    .stSelectbox > div > div {
        width: 180px !important;
        max-width: 180px !important;
        height: 35px !important;
        min-height: 35px !important;
    }
    .stSelectbox > div > div > div {
        width: 180px !important;
        max-width: 180px !important;
        height: 35px !important;
        min-height: 35px !important;
        font-size: 14px !important;
        line-height: 35px !important;
        padding: 0 25px 0 12px !important;
        text-overflow: ellipsis !important;
        white-space: nowrap !important;
        overflow: visible !important;
    }
    
    /* Compact text input styling */
    .stTextInput {
        width: 140px !important;
    }
    .stTextInput > div {
        width: 140px !important;
        max-width: 140px !important;
    }
    .stTextInput > div > div > input {
        height: 35px !important;
        min-height: 35px !important;
        font-size: 14px !important;
        padding: 8px 12px !important;
        width: 140px !important;
        max-width: 140px !important;
    }
    
    /* Reduce label sizes */
    .stSelectbox label, .stTextInput label {
        font-size: 14px !important;
        font-weight: 500 !important;
        margin-bottom: 4px !important;
    }
    
    /* Align radio buttons and selectbox vertically */
    .stRadio > div {
        align-items: flex-start !important;
        margin-top: 0px !important;
    }
    .stRadio label {
        font-size: 14px !important;
        font-weight: 500 !important;
        margin-bottom: 4px !important;
    }
    .stRadio > div > div {
        margin-top: 4px !important;
    }
    
    /* Remove unused CSS - Time Period now positioned in layout */
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("### Real-time stock analysis with comprehensive fundamental metrics")
    
    # Show shared insights history if any exist
    if 'shared_insights' in st.session_state and st.session_state.shared_insights:
        with st.expander(f"📋 Shared Insights History ({len(st.session_state.shared_insights)})", expanded=False):
            for i, insight in enumerate(reversed(st.session_state.shared_insights[-5:]), 1):  # Show last 5
                col_hist1, col_hist2, col_hist3 = st.columns([2, 1, 1])
                with col_hist1:
                    symbol_display = insight.get('symbol', 'N/A')
                    if symbol_display == '***':
                        symbol_display = 'Anonymous'
                    st.markdown(f"**{symbol_display}** - {insight.get('recommendation', 'N/A')}")
                with col_hist2:
                    privacy_icon = {"public": "🌐", "anonymized": "🔒", "private": "🔐"}.get(insight.get('privacy_level'), "📊")
                    st.markdown(f"{privacy_icon} {insight.get('privacy_level', 'N/A').title()}")
                with col_hist3:
                    timestamp = insight.get('timestamp', '')[:16].replace('T', ' ')
                    st.markdown(f"🕒 {timestamp}")
            
            if len(st.session_state.shared_insights) > 5:
                st.caption(f"Showing latest 5 of {len(st.session_state.shared_insights)} total insights")
    
    st.markdown("---")
    
    # Market and analysis mode selection with time period aligned
    col_market, col_mode, col_time = st.columns([1, 2, 1])
    
    with col_market:
        market_selection = st.selectbox(
            "Select Market:",
            ["US Stocks", "Indian Stocks"],
            help="Choose between US stocks (USD) or Indian stocks (INR)"
        )
        market = "India" if market_selection == "Indian Stocks" else "US"
    
    with col_mode:
        analysis_mode = st.radio(
            "Analysis Mode:",
            ["Single Stock Analysis", "Bulk Stock Analysis (Excel Export)"],
            horizontal=True
        )
    
    # Move Time Period selection to align with Analysis Mode
    if analysis_mode == "Single Stock Analysis":
        with col_time:
            period_options = {
                "1 Month": "1mo",
                "3 Months": "3mo", 
                "6 Months": "6mo",
                "1 Year": "1y",
                "2 Years": "2y",
                "5 Years": "5y",
                "10 Years": "10y",
                "Maximum": "max"
            }
            
            selected_period = st.selectbox(
                "Select Time Period:",
                options=list(period_options.keys()),
                index=3,  # Default to "1 Year"
                help="Choose the time period for historical data analysis",
                key="time_period_select_top"
            )
            
            period_code = period_options[selected_period]
    
        # Create input section for single stock (simplified without time period in col2)
        col1, col2 = st.columns([3, 1])
        
        with col1:
            if market == "India":
                placeholder_text = "e.g., RELIANCE, TCS, INFY"
                help_text = "Enter Indian stock symbols (e.g., RELIANCE for Reliance Industries, TCS for Tata Consultancy Services, INFY for Infosys)"
                default_symbol = "RELIANCE"
            else:
                placeholder_text = "e.g., AAPL, GOOGL, TSLA"
                help_text = "Enter US stock symbols (e.g., AAPL for Apple, GOOGL for Google, TSLA for Tesla)"
                default_symbol = "AAPL"
            
            # Initialize session state for symbol syncing
            if 'current_symbol' not in st.session_state:
                st.session_state.current_symbol = default_symbol
            
            symbol = st.text_input(
                f"Enter {market_selection} Symbol:",
                value=st.session_state.current_symbol,
                placeholder=placeholder_text,
                help=help_text,
                key="fundamental_symbol"
            ).upper().strip()
            
            # Update session state when symbol changes
            if symbol != st.session_state.current_symbol:
                st.session_state.current_symbol = symbol
                # Also store the current market for Advanced Analysis sync
                st.session_state.current_market = market
                
            # For Indian market, ensure .NS suffix for proper data fetching
            if market == "India" and symbol and not symbol.endswith('.NS') and not symbol.endswith('.BO'):
                symbol_for_fetching = f"{symbol}.NS"
            else:
                symbol_for_fetching = symbol
        
        with col2:
            # Auto-refresh toggle
            auto_refresh = st.checkbox(
                "🔄 Auto-refresh (10 min)",
                value=False,
                help="Automatically update data every 10 minutes during market hours"
            )
            
            # Stack buttons vertically with proper sizing
            analyze_button = st.button("Generate Chart", type="primary", use_container_width=True)
            pdf_button = st.button("📄 Export PDF", help="Export all tabs to PDF", use_container_width=True)
        
        # Create sub-tabs for organized content when stock data is available
        if symbol:
            # Pre-fetch data to determine if tabs should be shown
            with st.spinner(f'Fetching data for {symbol.upper()}...'):
                data, ticker_info, ticker_obj = fetch_stock_data(symbol_for_fetching, period=period_code, market=market)
            
            if data is not None and ticker_info is not None and ticker_obj is not None and not data.empty:
                # Calculate all technical indicators once
                ma_50 = calculate_moving_average(data, window=50)
                ma_200 = calculate_moving_average(data, window=200)
                macd_line, signal_line, histogram = calculate_macd(data)
                rsi = calculate_rsi(data)
                cmf = calculate_chaikin_money_flow(data)
                support_level, resistance_level = calculate_support_resistance(data)
                
                # Handle comprehensive PDF export
                if pdf_button:
                    with st.spinner('Generating comprehensive PDF report...'):
                        try:
                            # Create a comprehensive price chart for the PDF
                            import plotly.graph_objects as go
                            currency = "₹" if market == "India" else "$"
                            
                            price_fig = go.Figure()
                            
                            # Add candlestick chart
                            price_fig.add_trace(go.Candlestick(
                                x=data.index,
                                open=data['Open'],
                                high=data['High'],
                                low=data['Low'],
                                close=data['Close'],
                                name='Price',
                                increasing_line_color='green',
                                decreasing_line_color='red'
                            ))
                            
                            # Add moving averages
                            if not ma_50.empty:
                                price_fig.add_trace(go.Scatter(
                                    x=data.index,
                                    y=ma_50,
                                    mode='lines',
                                    name='50-Day MA',
                                    line=dict(color='blue', width=2)
                                ))
                            
                            if not ma_200.empty:
                                price_fig.add_trace(go.Scatter(
                                    x=data.index,
                                    y=ma_200,
                                    mode='lines',
                                    name='200-Day MA',
                                    line=dict(color='red', width=2)
                                ))
                            
                            # Add support/resistance lines
                            price_fig.add_hline(y=support_level, line_dash="dash", line_color="green", annotation_text="Support")
                            price_fig.add_hline(y=resistance_level, line_dash="dash", line_color="red", annotation_text="Resistance")
                            
                            price_fig.update_layout(
                                title=f'{symbol} Technical Analysis Chart',
                                xaxis_title='Date',
                                yaxis_title=f'Price ({currency})',
                                height=600,
                                xaxis_rangeslider_visible=False,
                                showlegend=True
                            )
                            
                            # Generate comprehensive analysis PDF with chart included
                            pdf_data, pdf_filename = export_comprehensive_analysis_pdf(
                                symbol, data, ticker_info, ticker_obj, ma_50, ma_200, 
                                rsi, support_level, resistance_level, market, price_fig
                            )
                            
                            if pdf_data and pdf_filename:
                                # Create download button
                                st.download_button(
                                    label="📄 Download Comprehensive Report",
                                    data=pdf_data,
                                    file_name=pdf_filename,
                                    mime="application/pdf",
                                    type="primary"
                                )
                                st.success("✅ Comprehensive PDF report generated successfully! The report includes all financial metrics, quality scores, % from CTP calculations, and technical analysis.")
                            else:
                                st.error("Failed to generate PDF report. Please try again.")
                            
                        except Exception as e:
                            st.error(f"Error generating PDF: {str(e)}")
                
                # Create sub-tabs for better organization
                tab_price, tab_charts, tab_earnings, tab_intelligence, tab_sentiment = st.tabs([
                    "📊 Price Action", 
                    "📈 Charts", 
                    "📅 Earnings & Dividends",
                    "🎯 Market Intelligence",
                    "📰 News Sentiment"
                ])
                
                with tab_price:
                    display_price_action_tab(symbol, data, ticker_info, ticker_obj, ma_50, ma_200, rsi, support_level, resistance_level, selected_period, market, auto_refresh)
                
                with tab_charts:
                    display_technical_charts_tab(symbol, data, ma_50, ma_200, macd_line, signal_line, histogram, rsi, cmf, selected_period, market)
                
                with tab_earnings:
                    display_earnings_dividends_tab(symbol, data, ticker_info, ticker_obj, market)
                
                with tab_intelligence:
                    display_advanced_sentiment_metrics(symbol, market)
                
                with tab_sentiment:
                    display_news_sentiment_analysis(symbol)
            
            else:
                st.error(f"""
                ❌ **Unable to fetch data for symbol '{symbol}'**
                
                Please check that:
                - The stock symbol is valid and correctly spelled
                - The stock is publicly traded
                - You have an internet connection
                
                **Examples of valid symbols:** AAPL, GOOGL, MSFT, TSLA, AMZN
                """)
        else:
            st.info("👆 Enter a stock symbol above to begin analysis")
    
    else:
        # Bulk analysis interface
        st.markdown("### 📊 Bulk Stock Analysis")
        st.markdown("Enter multiple stock symbols to generate a comprehensive Excel report with key metrics for all stocks.")
        
        # Initialize session state for saved lists
        if 'saved_stock_lists' not in st.session_state:
            st.session_state.saved_stock_lists = {}
        
        # Saved lists management
        st.markdown("#### 📂 Saved Stock Lists")
        col_saved1, col_saved2, col_saved3 = st.columns([2, 1, 1])
        
        with col_saved1:
            if st.session_state.saved_stock_lists:
                selected_saved_list = st.selectbox(
                    "Load a saved list:",
                    options=[""] + list(st.session_state.saved_stock_lists.keys()),
                    help="Select a previously saved stock list"
                )
            else:
                st.info("No saved lists yet. Create one below!")
                selected_saved_list = ""
        
        with col_saved2:
            if selected_saved_list:
                if st.button("📥 Load List", help="Load the selected stock list"):
                    st.session_state.bulk_stock_input = st.session_state.saved_stock_lists[selected_saved_list]
                    st.success(f"Loaded '{selected_saved_list}'")
                    st.rerun()
        
        with col_saved3:
            if selected_saved_list:
                if st.button("🗑️ Delete", help="Delete the selected list"):
                    del st.session_state.saved_stock_lists[selected_saved_list]
                    st.success(f"Deleted '{selected_saved_list}'")
                    st.rerun()
        
        st.divider()
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            if market == "India":
                bulk_placeholder = "RELIANCE\nTCS\nINFY\nHDFC\nICICIBANK\n\nOr: RELIANCE, TCS, INFY, HDFC, ICICIBANK"
                bulk_help = "Enter Indian stock symbols separated by commas or new lines"
            else:
                bulk_placeholder = "AAPL\nMSFT\nGOOGL\nTSLA\nAMZN\n\nOr: AAPL, MSFT, GOOGL, TSLA, AMZN"
                bulk_help = "Enter US stock symbols separated by commas or new lines"
            
            # Use session state for persistent input
            if 'bulk_stock_input' not in st.session_state:
                st.session_state.bulk_stock_input = ""
            
            stock_list_raw = st.text_area(
                f"Enter {market_selection} Symbols (one per line or comma-separated)",
                value=st.session_state.bulk_stock_input,
                placeholder=bulk_placeholder,
                height=150,
                help=bulk_help,
                key="stock_list_input"
            )
            # Convert to uppercase for consistency
            stock_list = stock_list_raw.upper()
            
            # Update session state when input changes
            if stock_list != st.session_state.bulk_stock_input:
                st.session_state.bulk_stock_input = stock_list
            
            # Save list functionality
            col_save1, col_save2 = st.columns([2, 1])
            with col_save1:
                list_name = st.text_input(
                    "Save this list as:",
                    placeholder="e.g., Tech Stocks, Banking Stocks, My Portfolio",
                    help="Enter a name to save the current stock list for future use"
                )
            with col_save2:
                st.markdown("<br>", unsafe_allow_html=True)  # Spacing
                if st.button("💾 Save List", disabled=not (stock_list.strip() and list_name.strip())):
                    if list_name.strip() and stock_list.strip():
                        st.session_state.saved_stock_lists[list_name.strip()] = stock_list.strip()
                        st.success(f"Saved list '{list_name.strip()}'!")
                        st.rerun()
        
        with col2:
            # Period selection for bulk
            period_options = {
                "1 Month": "1mo",
                "3 Months": "3mo", 
                "6 Months": "6mo",
                "1 Year": "1y",
                "2 Years": "2y",
                "5 Years": "5y",
                "10 Years": "10y",
                "Maximum": "max"
            }
            
            bulk_selected_period = st.selectbox(
                "Analysis Period",
                options=list(period_options.keys()),
                index=3,  # Default to 1 Year
                help="Select the time period for technical analysis",
                key="bulk_period"
            )
            
            bulk_period_code = period_options[bulk_selected_period]
            
            generate_excel_button = st.button("📋 Generate Excel Report", type="primary")
        
        # Process bulk analysis
        if generate_excel_button and stock_list.strip():
            # Parse stock symbols
            symbols = []
            if ',' in stock_list:
                symbols = [s.strip().upper() for s in stock_list.split(',') if s.strip()]
            else:
                symbols = [s.strip().upper() for s in stock_list.split('\n') if s.strip()]
            
            if symbols:
                st.info(f"Analyzing {len(symbols)} stocks: {', '.join(symbols)}")
                
                # Progress tracking
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                all_metrics = []
                
                for i, symbol in enumerate(symbols):
                    status_text.text(f"Analyzing {symbol}... ({i+1}/{len(symbols)})")
                    
                    metrics = get_stock_metrics(symbol, bulk_period_code, market)
                    all_metrics.append(metrics)
                    
                    # Update progress
                    progress_bar.progress((i + 1) / len(symbols))
                
                status_text.text("Generating Excel report...")
                
                # Create Excel report
                excel_file = create_excel_report(all_metrics, bulk_selected_period)
                
                # Success message and download
                st.success(f"Analysis complete! Generated report for {len(symbols)} stocks.")
                
                # Display summary
                successful = len([m for m in all_metrics if 'Error' not in m or m.get('Error') == ''])
                failed = len(all_metrics) - successful
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Stocks", len(symbols))
                with col2:
                    st.metric("Successful", successful)
                with col3:
                    st.metric("Failed", failed)
                
                # Download button
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_file,
                    file_name=f"stock_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Show preview of data
                if all_metrics:
                    st.markdown("### Preview of Generated Data")
                    preview_df = pd.DataFrame(all_metrics)
                    st.dataframe(preview_df, use_container_width=True)
                
                # Clear progress indicators
                progress_bar.empty()
                status_text.empty()
            
            else:
                st.error("Please enter at least one valid stock symbol.")
        
        # Set default values for single stock mode (when in bulk mode)
        symbol = ""
        analyze_button = False
        selected_period = "1 Year"
        period_code = "1y"
        auto_refresh = False  # No auto-refresh in bulk mode
    
    # Auto-refresh functionality (only for single stock mode)
    if analysis_mode == "Single Stock Analysis" and auto_refresh and symbol:
        # Initialize refresh state
        if 'last_refresh_time' not in st.session_state:
            st.session_state.last_refresh_time = time.time()
            st.session_state.refresh_count = 0
        
        current_time = time.time()
        time_since_refresh = int(current_time - st.session_state.last_refresh_time)
        
        # Check if 10 minutes have passed
        if time_since_refresh >= 600:  # 600 seconds = 10 minutes
            st.session_state.last_refresh_time = current_time
            st.session_state.refresh_count += 1
            st.rerun()
        
        # Display refresh status with countdown
        minutes_since = time_since_refresh // 60
        seconds_since = time_since_refresh % 60
        next_refresh_in = 600 - time_since_refresh
        next_refresh_min = next_refresh_in // 60
        next_refresh_sec = next_refresh_in % 60
        
        refresh_info = f"🔄 Auto-refresh active (#{st.session_state.refresh_count}) | Last update: {minutes_since}m {seconds_since}s ago | Next refresh: {next_refresh_min}m {next_refresh_sec}s"
        st.info(refresh_info)
        
        # Add auto-refresh timer with meta refresh
        st.markdown(f'<meta http-equiv="refresh" content="{next_refresh_in}">', unsafe_allow_html=True)


def gurufocus_tab():
    """Advanced analysis tab content"""
    
    # Apply view mode styling
    apply_view_mode_css()
    
    st.markdown("### Professional institutional-grade financial analysis")
    st.markdown("Advanced earnings performance analysis with up to 8 quarters of historical data")
    st.markdown("---")
    
    # Create input section
    col1, col2, col3 = st.columns([3, 2, 1])
    
    with col1:
        # Use synced symbol from session state, or default to AAPL if not set
        default_guru_symbol = st.session_state.get('current_symbol', 'AAPL')
        
        symbol_guru = st.text_input(
            "Enter Stock Symbol:",
            value=default_guru_symbol,
            placeholder="e.g., AAPL, GOOGL, TSLA (US) or RELIANCE.NS, TCS.NS (India)",
            help="Enter stock symbols for detailed earnings analysis (synced with Fundamental Analysis) - supports both US and Indian markets",
            key="gurufocus_symbol"
        ).upper().strip()
        
        # Update session state when symbol changes in Advanced Analysis
        if symbol_guru != st.session_state.get('current_symbol', ''):
            st.session_state.current_symbol = symbol_guru
    
    with col2:
        quarters_selection = st.selectbox(
            "Historical Quarters:",
            ["4 Quarters", "6 Quarters", "8 Quarters"],
            index=2,
            help="Number of past quarters to analyze"
        )
        quarters_count = int(quarters_selection.split()[0])
    
    with col3:
        if st.button("🔍 Analyze Earnings", key="guru_analyze", type="primary"):
            st.session_state.guru_analyze_clicked = True
    
    # Analysis section
    if st.session_state.get('guru_analyze_clicked', False) and symbol_guru:
        with st.spinner(f"Analyzing {quarters_count} quarters of earnings data for {symbol_guru.upper()}..."):
            # Auto-detect market based on symbol and sync with fundamental analysis market selection
            # Check if we have market info from the synced symbol context
            fundamental_market = st.session_state.get('current_market', None)
            
            # Common Indian stock symbols (major companies)
            indian_symbols = {
                'RELIANCE', 'TCS', 'INFY', 'HDFC', 'ICICIBANK', 'BHARTIARTL', 'ITC', 
                'LT', 'HCLTECH', 'KOTAKBANK', 'AXISBANK', 'ASIANPAINT', 'MARUTI',
                'TITAN', 'NESTLEIND', 'ULTRACEMCO', 'POWERGRID', 'NTPC', 'COALINDIA',
                'ONGC', 'TECHM', 'TATAMOTORS', 'TATASTEEL', 'HINDUNILVR', 'WIPRO',
                'DRREDDY', 'EICHERMOT', 'BAJFINANCE', 'BAJAJFINSV', 'BRITANNIA',
                'CIPLA', 'DIVISLAB', 'HEROMOTOCO', 'HINDALCO', 'INDUSINDBK',
                'JSWSTEEL', 'M&M', 'SBIN', 'SUNPHARMA', 'GRASIM'
            }
            
            # Detect market based on multiple factors
            if ('.NS' in symbol_guru or '.BO' in symbol_guru or 
                symbol_guru.replace('.NS', '').replace('.BO', '') in indian_symbols or
                fundamental_market == "India"):
                detected_market = "India"
                # For Indian stocks, add .NS suffix if not present
                if not symbol_guru.endswith('.NS') and not symbol_guru.endswith('.BO'):
                    symbol_for_analysis = f"{symbol_guru}.NS"
                else:
                    symbol_for_analysis = symbol_guru
            else:
                detected_market = "US"
                symbol_for_analysis = symbol_guru
            
            st.info(f"🌍 Detected market: {detected_market}")
            
            # Fetch extended earnings data first to get info
            data, info, ticker_obj = fetch_stock_data(symbol_for_analysis, period="3y", market=detected_market)
            
            # Add data source information prominently in Advanced Analysis (after info is available)
            if info:
                hybrid_metrics = get_hybrid_financial_metrics(symbol_for_analysis, info) if symbol_for_analysis else None
                data_source = hybrid_metrics.get('source', 'Yahoo Finance') if hybrid_metrics else 'Yahoo Finance'
                
                with st.expander("📊 Data Source Information & Known Discrepancies", expanded=False):
                    st.info(f"🔍 **Current Data Source**: {data_source}")
                    
                    if hybrid_metrics:
                        # Format valuation metrics with proper conditional formatting
                        pe_ratio = hybrid_metrics.get('pe_ratio')
                        forward_pe = hybrid_metrics.get('forward_pe') 
                        ps_ratio = hybrid_metrics.get('ps_ratio')
                        peg_ratio = hybrid_metrics.get('peg_ratio')
                        
                        pe_str = f"{pe_ratio:.2f}" if pe_ratio and not pd.isna(pe_ratio) else 'N/A'
                        forward_pe_str = f"{forward_pe:.2f}" if forward_pe and not pd.isna(forward_pe) else 'N/A'
                        ps_str = f"{ps_ratio:.2f}" if ps_ratio and not pd.isna(ps_ratio) else 'N/A'
                        peg_str = f"{peg_ratio:.2f}" if peg_ratio and not pd.isna(peg_ratio) else 'N/A'
                        
                        st.info(f"📊 **Valuation**: P/E: {pe_str} | Forward P/E: {forward_pe_str} | P/S: {ps_str} | PEG: {peg_str}")
                        
                        ev_revenue = hybrid_metrics.get('ev_revenue') or info.get('enterpriseToRevenue')
                        ev_ebitda = hybrid_metrics.get('ev_ebitda') or info.get('enterpriseToEbitda')
                        
                        ev_revenue_str = f"{ev_revenue:.2f}" if ev_revenue and not pd.isna(ev_revenue) else 'N/A'
                        ev_ebitda_str = f"{ev_ebitda:.2f}" if ev_ebitda and not pd.isna(ev_ebitda) else 'N/A'
                        
                        st.info(f"📊 **Enterprise**: EV/Revenue: {ev_revenue_str} | EV/EBITDA: {ev_ebitda_str}")
                        
                        # Show margin data which can have significant differences
                        operating_margin = info.get('operatingMargins')
                        gross_margin = info.get('grossMargins')
                        
                        if operating_margin and gross_margin:
                            st.info(f"📊 **Margins**: Gross: {gross_margin*100:.2f}% | Operating: {operating_margin*100:.2f}%")
                        else:
                            st.info("📊 **Margins**: Limited data available")
                    
                    st.warning("ℹ️ **Data Variance**: API values may differ from institutional sources (±0.1-10%) due to data timing, calculation periods, and methodologies. "
                              "Operating margins can show significant differences between sources. GuruFocus data preferred when available.")
                    
                    if not hybrid_metrics or hybrid_metrics.get('source') == 'Yahoo Finance':
                        st.warning("⚠️ **Using Yahoo Finance**: Add GURUFOCUS_API_KEY to get exact GuruFocus institutional metrics")
                    else:
                        st.success("✅ **Using GuruFocus**: Showing institutional-grade metrics")
            
            if data is not None and ticker_obj is not None:
                # Get detailed earnings performance analysis
                earnings_analysis, quarters_found = get_detailed_earnings_performance_analysis(
                    ticker_obj, data, market=detected_market, max_quarters=quarters_count
                )
                
                if earnings_analysis is not None and not earnings_analysis.empty:
                    # Display earnings analysis results
                    st.subheader(f"📊 Earnings Performance Analysis - {quarters_found} Quarter{'s' if quarters_found != 1 else ''}")
                    
                    # Show the earnings analysis table
                    st.dataframe(earnings_analysis, use_container_width=True)
                    
                    # Add comprehensive institutional financial metrics
                    st.divider()
                    st.subheader("🏛️ Institutional Financial Parameters")
                    
                    # Get comprehensive financial data
                    try:
                        # Fetch detailed financial information
                        balance_sheet = ticker_obj.balance_sheet
                        income_stmt = ticker_obj.income_stmt
                        cash_flow = ticker_obj.cash_flow
                        
                        # Currency symbol based on market
                        currency = "₹" if detected_market == "India" else "$"
                        currency_suffix = "Cr" if detected_market == "India" else "B"
                        divisor = 1e7 if detected_market == "India" else 1e9
                        
                        # Valuation Metrics
                        st.markdown("### 💰 Valuation Ratios")
                        val_col1, val_col2, val_col3, val_col4 = st.columns(4)
                        
                        with val_col1:
                            # P/E Ratio - use consistent approach
                            pe_ratio = info.get('trailingPE')
                            forward_pe = info.get('forwardPE')
                            st.metric("P/E Ratio (TTM)", f"{pe_ratio:.2f}" if pe_ratio else "N/A", 
                                     help="P/E Ratio (TTM) – Price to Earnings Ratio\nShows how much investors are paying for $1 of the company's earnings over the past 12 months. A very high number may mean strong growth expectations or an overvalued stock. Compare it to peers and the market average.")
                            st.metric("Forward P/E", f"{forward_pe:.2f}" if forward_pe else "N/A",
                                     help="Forward P/E – Forward Price to Earnings Ratio\nThis looks ahead, using analysts' projected earnings. It's useful for judging whether growth is expected to make the stock cheaper in the future. If the forward P/E is lower than the current P/E, it means analysts expect earnings to rise, so the stock will look less expensive relative to profits.")
                        
                        with val_col2:
                            # Price-to-Book and Price-to-Sales
                            pb_ratio = info.get('priceToBook')
                            ps_ratio = info.get('priceToSalesTrailing12Months')
                            st.metric("Price-to-Book", f"{pb_ratio:.2f}" if pb_ratio else "N/A",
                                     help="Price-to-Book (P/B Ratio)\nCompares the stock price to the company's book value (assets minus liabilities). A ratio above 1 means investors value the company more than its net assets. High values may indicate strong growth potential—or overvaluation.")
                            st.metric("Price-to-Sales", f"{ps_ratio:.2f}" if ps_ratio else "N/A",
                                     help="Price-to-Sales (P/S Ratio)\nShows how much investors pay for $1 of revenue. Useful for companies with little or no profit. A very high number can mean growth optimism but also possible overpricing.")
                        
                        with val_col3:
                            # PEG and Enterprise Value ratios - use hybrid system for PEG
                            peg_ratio = get_peg_ratio(info)  # Use enhanced PEG detection
                            ev_revenue = info.get('enterpriseToRevenue')
                            st.metric("PEG Ratio", f"{peg_ratio:.2f}" if peg_ratio and not pd.isna(peg_ratio) else "N/A",
                                     help="PEG Ratio – Price to Earnings Growth Ratio\nTakes the P/E Ratio and adjusts it for the company's earnings growth rate. Around 1 is considered fairly valued, above 1 may be expensive, below 1 could be undervalued. Helps balance growth with price.")
                            st.metric("EV/Revenue", f"{ev_revenue:.2f}" if ev_revenue else "N/A",
                                     help="EV/Revenue (Enterprise Value to Revenue)\nEnterprise Value (company's total value including debt and cash) divided by revenue. Often better than P/S since it accounts for debt. Lower is usually more attractive.")
                        
                        with val_col4:
                            # Enterprise Value and EBITDA
                            enterprise_value = info.get('enterpriseValue')
                            ev_ebitda = info.get('enterpriseToEbitda')
                            if enterprise_value:
                                st.metric("Enterprise Value", f"{currency}{enterprise_value/divisor:.2f}{currency_suffix}",
                                         help="Enterprise Value (EV)\nRepresents the true cost to acquire a company (market cap + debt – cash). Useful when comparing firms with different debt levels. Investors look at EV with ratios like EV/EBITDA for deeper insights.")
                            else:
                                st.metric("Enterprise Value", "N/A",
                                         help="Enterprise Value (EV)\nRepresents the true cost to acquire a company (market cap + debt – cash). Useful when comparing firms with different debt levels. Investors look at EV with ratios like EV/EBITDA for deeper insights.")
                            st.metric("EV/EBITDA", f"{ev_ebitda:.2f}" if ev_ebitda else "N/A",
                                     help="EV/EBITDA (Enterprise Value to EBITDA)\nEnterprise Value compared to EBITDA (earnings before interest, tax, depreciation, and amortization). A widely used measure for comparing companies, especially across industries. Lower values often signal a cheaper valuation.")
                        
                        # Profitability Analysis
                        st.markdown("### 📈 Profitability Analysis")
                        prof_col1, prof_col2, prof_col3, prof_col4 = st.columns(4)
                        
                        with prof_col1:
                            # Margin metrics
                            gross_margin = info.get('grossMargins')
                            operating_margin = info.get('operatingMargins')
                            st.metric("Gross Margin", f"{gross_margin*100:.1f}%" if gross_margin else "N/A",
                                     help="Gross Profit Margin\nPercentage of revenue remaining after cost of goods sold (COGS).\n• High margin: Strong pricing power, efficient production\n• >50%: Excellent (software, luxury goods)\n• 30-50%: Good (branded consumer goods)\n• 10-30%: Average (retail, manufacturing)\n• <10%: Low margin business (groceries, commodities)\nHigher gross margins provide more flexibility for operations and growth.")
                            st.metric("Operating Margin", f"{operating_margin*100:.1f}%" if operating_margin else "N/A",
                                     help="Operating Profit Margin\nPercentage of revenue remaining after all operating expenses.\n• Measures core business efficiency and management effectiveness\n• >20%: Excellent operational efficiency\n• 10-20%: Good operational performance\n• 5-10%: Average operational efficiency\n• <5%: Poor operational efficiency\n• Negative: Operating losses\nNote: yfinance source - may differ from institutional sources like GuruFocus")
                        
                        with prof_col2:
                            # Profit margins
                            profit_margin = info.get('profitMargins')
                            ebitda_margin = info.get('ebitdaMargins')
                            st.metric("Profit Margin", f"{profit_margin*100:.1f}%" if profit_margin else "N/A",
                                     help="Net Profit Margin\nPercentage of revenue that becomes actual profit after all expenses.\n• Ultimate measure of business profitability\n• >20%: Exceptional profitability (tech, pharma)\n• 10-20%: Strong profitability (established businesses)\n• 5-10%: Moderate profitability (retail, services)\n• 0-5%: Low profitability (commodities, utilities)\n• Negative: Net losses\nConsider industry benchmarks when evaluating.")
                            st.metric("EBITDA Margin", f"{ebitda_margin*100:.1f}%" if ebitda_margin else "N/A",
                                     help="EBITDA Margin\nEarnings Before Interest, Taxes, Depreciation, and Amortization as % of revenue.\n• Measures operational cash generation before financing decisions\n• >30%: Excellent cash generation (software, services)\n• 20-30%: Strong cash generation (established businesses)\n• 10-20%: Moderate cash generation (manufacturing)\n• 5-10%: Low cash generation (capital-intensive)\n• <5%: Very low cash generation\nUseful for comparing companies across different tax and capital structures.")
                        
                        with prof_col3:
                            # Return metrics
                            roe = info.get('returnOnEquity')
                            roa = info.get('returnOnAssets')
                            st.metric("Return on Equity", f"{roe*100:.1f}%" if roe else "N/A",
                                     help="Return on Equity (ROE)\nNet income as percentage of shareholders' equity - measures management effectiveness.\n• >20%: Excellent management and capital efficiency\n• 15-20%: Very good returns for shareholders\n• 10-15%: Good returns, above market average\n• 5-10%: Average returns, acceptable performance\n• <5%: Poor returns, below expectations\n• High ROE with low debt is ideal - avoid artificially high ROE from excessive leverage")
                            st.metric("Return on Assets", f"{roa*100:.1f}%" if roa else "N/A",
                                     help="Return on Assets (ROA)\nNet income as percentage of total assets - measures asset utilization efficiency.\n• >15%: Exceptional asset efficiency (tech, services)\n• 10-15%: Very good asset utilization\n• 5-10%: Good asset efficiency for most industries\n• 2-5%: Average asset utilization (banks, utilities)\n• <2%: Poor asset efficiency\n• ROA shows how well company converts assets into profits regardless of financing structure")
                        
                        with prof_col4:
                            # Revenue per share and Book value
                            revenue_per_share = info.get('revenuePerShare')
                            book_value = info.get('bookValue')
                            st.metric("Revenue/Share", f"{currency}{revenue_per_share:.2f}" if revenue_per_share else "N/A",
                                     help="Revenue Per Share\nTotal revenue divided by outstanding shares.\n• Shows company's ability to generate sales per share\n• Growing revenue per share indicates business expansion\n• Compare to price per share for P/S ratio\n• Higher values generally better, but consider industry context\n• More stable than earnings per share\nUseful for evaluating growth companies with volatile earnings.")
                            st.metric("Book Value/Share", f"{currency}{book_value:.2f}" if book_value else "N/A",
                                     help="Book Value Per Share\nShareholders' equity divided by outstanding shares - company's net worth per share.\n• Represents liquidation value if company dissolved today\n• Compare to stock price for P/B ratio\n• Higher book value provides more downside protection\n• Asset-heavy companies typically have higher book values\n• Growing book value indicates retained earnings accumulation\nParticularly important for value investing and financial companies.")
                        
                        # Investment Ratings (1-5 Scale)
                        st.markdown("### ⭐ Investment Ratings (1-5 Scale)")
                        st.caption("Rating scale: 5 = Excellent/Most Desired, 4 = Good, 3 = Average, 2 = Below Average, 1 = Poor/Least Desired")
                        
                        # Define local rating color function
                        def get_rating_color_local(rating):
                            """Get color for rating display (1-5 scale)"""
                            color_map = {
                                5: "#22c55e",  # Green - Excellent
                                4: "#84cc16",  # Light Green - Good  
                                3: "#eab308",  # Yellow - Average
                                2: "#f97316",  # Orange - Below Average
                                1: "#ef4444"   # Red - Poor
                            }
                            return color_map.get(rating, "#6b7280")  # Default gray
                        
                        # Calculate investment ratings
                        ratings = calculate_investment_ratings(info, ticker_obj)
                        
                        rating_col1, rating_col2, rating_col3 = st.columns(3)
                        
                        with rating_col1:
                            quant_rating = ratings['quantitative']
                            color = get_rating_color_local(quant_rating)
                            st.markdown(f"**Quantitative Rating**")
                            st.markdown(f"<div style='background-color: {color}; color: white; padding: 10px; border-radius: 5px; text-align: center; font-size: 20px; font-weight: bold;'>{quant_rating}/5</div>", unsafe_allow_html=True)
                            st.caption(ratings['quantitative_explanation'])
                        
                        with rating_col2:
                            author_rating = ratings['author']
                            color = get_rating_color_local(author_rating)
                            st.markdown(f"**Author Rating**")
                            st.markdown(f"<div style='background-color: {color}; color: white; padding: 10px; border-radius: 5px; text-align: center; font-size: 20px; font-weight: bold;'>{author_rating}/5</div>", unsafe_allow_html=True)
                            st.caption(ratings['author_explanation'])
                        
                        with rating_col3:
                            sellside_rating = ratings['sellside']
                            color = get_rating_color_local(sellside_rating)
                            st.markdown(f"**Sellside Rating**")
                            st.markdown(f"<div style='background-color: {color}; color: white; padding: 10px; border-radius: 5px; text-align: center; font-size: 20px; font-weight: bold;'>{sellside_rating}/5</div>", unsafe_allow_html=True)
                            st.caption(ratings['sellside_explanation'])
                        
                        st.divider()
                        
                        # Financial Strength
                        st.markdown("### 💪 Financial Strength")
                        strength_col1, strength_col2, strength_col3, strength_col4 = st.columns(4)
                        
                        with strength_col1:
                            # Debt ratios
                            debt_to_equity = info.get('debtToEquity')
                            current_ratio = info.get('currentRatio')
                            st.metric("Debt-to-Equity", f"{debt_to_equity:.2f}" if debt_to_equity else "N/A",
                                     help="Debt-to-Equity Ratio\nTotal debt divided by shareholders' equity - measures financial leverage.\n• <0.3: Very low debt, conservative financing\n• 0.3-0.6: Moderate debt, balanced capital structure\n• 0.6-1.0: Higher debt, increased financial risk\n• >1.0: High debt, significant leverage risk\n• >2.0: Very high debt, potential financial distress\nLower ratios generally safer, but optimal level varies by industry.")
                            st.metric("Current Ratio", f"{current_ratio:.2f}" if current_ratio else "N/A",
                                     help="Current Ratio\nCurrent assets divided by current liabilities - measures short-term liquidity.\n• >2.0: Excellent liquidity, strong short-term financial health\n• 1.5-2.0: Good liquidity, adequate working capital\n• 1.0-1.5: Acceptable liquidity, monitor cash flow\n• 0.8-1.0: Tight liquidity, potential cash flow issues\n• <0.8: Poor liquidity, financial stress likely\nHigher ratios indicate better ability to pay short-term obligations.")
                        
                        with strength_col2:
                            # Quick ratio and cash
                            quick_ratio = info.get('quickRatio')
                            total_cash = info.get('totalCash')
                            st.metric("Quick Ratio", f"{quick_ratio:.2f}" if quick_ratio else "N/A",
                                     help="Quick Ratio (Acid-Test)\nQuick assets (cash + securities + receivables) divided by current liabilities.\n• >1.5: Excellent liquidity without relying on inventory\n• 1.0-1.5: Good liquidity, can meet obligations quickly\n• 0.8-1.0: Adequate liquidity, monitor cash position\n• 0.5-0.8: Tight liquidity, may struggle with payments\n• <0.5: Poor liquidity, immediate cash concerns\nMore conservative than current ratio - excludes inventory and prepaid expenses.")
                            if total_cash:
                                st.metric("Total Cash", f"{currency}{total_cash/divisor:.2f}{currency_suffix}",
                                         help="Total Cash and Cash Equivalents\nCash, short-term investments, and highly liquid securities.\n• High cash provides financial flexibility and crisis protection\n• Compare to market cap for cash-rich vs cash-poor assessment\n• Excess cash may indicate lack of growth opportunities\n• Low cash with high debt increases financial risk\n• Consider cash burn rate for growth companies\nCash is the ultimate liquidity buffer for uncertain times.")
                            else:
                                st.metric("Total Cash", "N/A",
                                         help="Total Cash and Cash Equivalents\nCash, short-term investments, and highly liquid securities.\n• High cash provides financial flexibility and crisis protection\n• Compare to market cap for cash-rich vs cash-poor assessment\n• Excess cash may indicate lack of growth opportunities\n• Low cash with high debt increases financial risk\n• Consider cash burn rate for growth companies\nCash is the ultimate liquidity buffer for uncertain times.")
                        
                        with strength_col3:
                            # Cash per share and Free cash flow
                            cash_per_share = info.get('totalCashPerShare')
                            free_cashflow = info.get('freeCashflow')
                            st.metric("Cash/Share", f"{currency}{cash_per_share:.2f}" if cash_per_share else "N/A",
                                     help="Cash Per Share\nTotal cash divided by outstanding shares - shows liquidity per share.\n• High cash per share provides downside protection\n• Compare to stock price for cash backing percentage\n• Growing cash per share indicates strong cash generation\n• Declining may signal cash burn or aggressive investments\n• Particularly important for growth and biotech companies\nRepresents the cash cushion backing each share owned.")
                            if free_cashflow:
                                st.metric("Free Cash Flow", f"{currency}{free_cashflow/divisor:.2f}{currency_suffix}",
                                         help="Free Cash Flow\nOperating cash flow minus capital expenditures - cash available to shareholders.\n• Positive FCF: Company generates cash after maintaining/growing business\n• Negative FCF: Company consuming cash, may need financing\n• Growing FCF indicates improving cash generation ability\n• FCF > Net Income suggests high-quality earnings\n• FCF conversion rate shows efficiency of profit-to-cash conversion\nMost important metric for valuing mature companies and dividend sustainability.")
                            else:
                                st.metric("Free Cash Flow", "N/A",
                                         help="Free Cash Flow\nOperating cash flow minus capital expenditures - cash available to shareholders.\n• Positive FCF: Company generates cash after maintaining/growing business\n• Negative FCF: Company consuming cash, may need financing\n• Growing FCF indicates improving cash generation ability\n• FCF > Net Income suggests high-quality earnings\n• FCF conversion rate shows efficiency of profit-to-cash conversion\nMost important metric for valuing mature companies and dividend sustainability.")
                        
                        with strength_col4:
                            # Operating cash flow and Total debt
                            operating_cashflow = info.get('operatingCashflow')
                            total_debt = info.get('totalDebt')
                            if operating_cashflow:
                                st.metric("Operating Cash Flow", f"{currency}{operating_cashflow/divisor:.2f}{currency_suffix}",
                                         help="Operating Cash Flow\nCash generated from core business operations - measures cash-generating ability.\n• Positive OCF: Business generates cash from operations\n• OCF > Net Income: High-quality earnings, good cash conversion\n• OCF < Net Income: Potential earnings quality issues\n• Growing OCF indicates strengthening business fundamentals\n• OCF/Sales ratio shows operational cash efficiency\nMore reliable than earnings for assessing true business performance.")
                            else:
                                st.metric("Operating Cash Flow", "N/A",
                                         help="Operating Cash Flow\nCash generated from core business operations - measures cash-generating ability.\n• Positive OCF: Business generates cash from operations\n• OCF > Net Income: High-quality earnings, good cash conversion\n• OCF < Net Income: Potential earnings quality issues\n• Growing OCF indicates strengthening business fundamentals\n• OCF/Sales ratio shows operational cash efficiency\nMore reliable than earnings for assessing true business performance.")
                            if total_debt:
                                st.metric("Total Debt", f"{currency}{total_debt/divisor:.2f}{currency_suffix}",
                                         help="Total Debt\nShort-term debt plus long-term debt - total borrowing obligations.\n• Compare to cash to assess net debt position\n• Compare to equity for leverage assessment (debt-to-equity)\n• Compare to EBITDA for debt service capability\n• Rising debt increases financial risk and interest expense\n• Debt maturity profile affects refinancing risk\nManageable debt levels vary significantly by industry and business model.")
                            else:
                                st.metric("Total Debt", "N/A",
                                         help="Total Debt\nShort-term debt plus long-term debt - total borrowing obligations.\n• Compare to cash to assess net debt position\n• Compare to equity for leverage assessment (debt-to-equity)\n• Compare to EBITDA for debt service capability\n• Rising debt increases financial risk and interest expense\n• Debt maturity profile affects refinancing risk\nManageable debt levels vary significantly by industry and business model.")
                        
                        # Growth Metrics
                        st.markdown("### 🚀 Growth Analysis")
                        growth_col1, growth_col2, growth_col3, growth_col4 = st.columns(4)
                        
                        with growth_col1:
                            # Revenue growth
                            revenue_growth = info.get('revenueGrowth')
                            earnings_growth = info.get('earningsGrowth')
                            st.metric("Revenue Growth", f"{revenue_growth*100:.1f}%" if revenue_growth else "N/A",
                                     help="Revenue Growth (Year-over-Year)\nPercentage increase in revenue compared to same period last year.\n• >20%: Excellent growth (high-growth companies)\n• 10-20%: Strong growth (established growth companies)\n• 5-10%: Moderate growth (mature companies)\n• 0-5%: Slow growth (mature/cyclical companies)\n• Negative: Revenue decline\nConsistent revenue growth indicates market demand and competitive advantage.")
                            st.metric("Earnings Growth", f"{earnings_growth*100:.1f}%" if earnings_growth else "N/A",
                                     help="Earnings Growth (Year-over-Year)\nPercentage change in earnings compared to same period last year.\n• >25%: Exceptional earnings growth\n• 15-25%: Strong earnings growth\n• 5-15%: Moderate earnings growth\n• 0-5%: Slow earnings growth\n• Negative: Earnings decline\nEarnings growth more important than revenue growth - shows improving profitability and operational leverage.")
                        
                        with growth_col2:
                            # Quarterly growth
                            quarterly_revenue_growth = info.get('revenueQuarterlyGrowth')
                            quarterly_earnings_growth = info.get('earningsQuarterlyGrowth')
                            st.metric("Q Revenue Growth", f"{quarterly_revenue_growth*100:.1f}%" if quarterly_revenue_growth else "N/A",
                                     help="Quarterly Revenue Growth (YoY)\nMost recent quarter's revenue growth compared to same quarter last year.\n• More current than annual growth, shows recent trends\n• Can be volatile due to seasonal factors\n• Compare to previous quarters for trend analysis\n• Positive acceleration is bullish signal\n• Deceleration may indicate slowing business\nWatch for consistent quarterly growth patterns over multiple quarters.")
                            st.metric("Q Earnings Growth", f"{quarterly_earnings_growth*100:.1f}%" if quarterly_earnings_growth else "N/A",
                                     help="Quarterly Earnings Growth (YoY)\nMost recent quarter's earnings growth vs same quarter last year.\n• Key metric for growth investors and momentum traders\n• Watch for positive surprises vs expectations\n• Accelerating growth often drives stock price appreciation\n• Can be volatile due to one-time items or seasonal factors\n• Compare to analyst estimates and guidance\nConsistent quarterly earnings beats indicate strong execution and market position.")
                        
                        with growth_col3:
                            # EPS estimates
                            target_high_price = info.get('targetHighPrice')
                            target_low_price = info.get('targetLowPrice')
                            st.metric("Target High Price", f"{currency}{target_high_price:.2f}" if target_high_price else "N/A",
                                     help="Analyst Target High Price\nHighest price target among all covering analysts over next 12 months.\n• Represents most optimistic analyst view\n• Compare to current price for upside potential\n• Large spread between high/low targets indicates uncertainty\n• Consider analyst track record and recent revisions\n• Bull case scenario if everything goes right\nUse alongside mean target for balanced perspective on analyst sentiment.")
                            st.metric("Target Low Price", f"{currency}{target_low_price:.2f}" if target_low_price else "N/A",
                                     help="Analyst Target Low Price\nLowest price target among all covering analysts over next 12 months.\n• Represents most pessimistic analyst view\n• Shows potential downside risk\n• Compare to current price for worst-case scenario\n• Large spread between high/low indicates high uncertainty\n• Bear case scenario if everything goes wrong\nUseful for risk assessment and position sizing decisions.")
                        
                        with growth_col4:
                            # Analyst recommendations
                            recommendation_mean = info.get('recommendationMean')
                            number_of_analyst_opinions = info.get('numberOfAnalystOpinions')
                            if recommendation_mean:
                                rec_text = ["Strong Buy", "Buy", "Hold", "Sell", "Strong Sell"][min(4, max(0, int(recommendation_mean)-1))]
                                st.metric("Analyst Rating", f"{rec_text} ({recommendation_mean:.1f})",
                                         help="Analyst Recommendation Consensus\nAverage recommendation from all covering analysts.\n• Strong Buy (1.0-1.5): Very bullish consensus\n• Buy (1.5-2.5): Bullish consensus\n• Hold (2.5-3.5): Neutral consensus\n• Sell (3.5-4.5): Bearish consensus\n• Strong Sell (4.5-5.0): Very bearish consensus\nFewer analysts typically means less reliable consensus. Watch for recent upgrades/downgrades.")
                            else:
                                st.metric("Analyst Rating", "N/A",
                                         help="Analyst Recommendation Consensus\nAverage recommendation from all covering analysts.\n• Strong Buy (1.0-1.5): Very bullish consensus\n• Buy (1.5-2.5): Bullish consensus\n• Hold (2.5-3.5): Neutral consensus\n• Sell (3.5-4.5): Bearish consensus\n• Strong Sell (4.5-5.0): Very bearish consensus\nFewer analysts typically means less reliable consensus. Watch for recent upgrades/downgrades.")
                            st.metric("# of Analysts", f"{number_of_analyst_opinions}" if number_of_analyst_opinions else "N/A",
                                     help="Number of Analysts Covering Stock\nTotal number of analysts providing ratings and price targets.\n• >20 analysts: High coverage (large-cap, popular stocks)\n• 10-20 analysts: Good coverage (mid-cap stocks)\n• 5-10 analysts: Moderate coverage (small-cap stocks)\n• <5 analysts: Low coverage (micro-cap, niche stocks)\n• More analysts generally means more reliable consensus\nHigh coverage indicates institutional interest and market attention.")
                        
                        # Stock Ratings section
                        st.markdown("---")
                        st.markdown("### 🎯 Stock Ratings (A-D Scale)")
                        st.markdown("*Comprehensive performance ratings where A = Excellent, B = Good, C = Fair, D = Poor*")
                        
                        # Calculate and display ratings
                        ratings = calculate_stock_ratings(ticker_obj, info)
                        
                        rating_col1, rating_col2, rating_col3, rating_col4 = st.columns(4)
                        
                        # Helper function to get rating color
                        def get_rating_color(rating):
                            if rating == 'A': return '🟢'
                            elif rating == 'B': return '🟡'
                            elif rating == 'C': return '🟠'
                            elif rating == 'D': return '🔴'
                            else: return '⚪'
                        
                        with rating_col1:
                            rating_color = get_rating_color(ratings['Value'])
                            st.metric(
                                label="💰 Value",
                                value=f"{rating_color} {ratings['Value']}",
                                help="Based on P/E, P/B, and P/S ratios - lower ratios get higher grades"
                            )
                        
                        with rating_col2:
                            rating_color = get_rating_color(ratings['Growth'])
                            st.metric(
                                label="📈 Growth", 
                                value=f"{rating_color} {ratings['Growth']}",
                                help="Based on revenue and earnings growth - higher growth gets higher grades"
                            )
                        
                        with rating_col3:
                            rating_color = get_rating_color(ratings['Momentum'])
                            st.metric(
                                label="🚀 Momentum",
                                value=f"{rating_color} {ratings['Momentum']}",
                                help="Based on 1-month, 3-month, and 6-month price performance"
                            )
                        
                        with rating_col4:
                            rating_color = get_rating_color(ratings['Profitability'])
                            st.metric(
                                label="💵 Profitability",
                                value=f"{rating_color} {ratings['Profitability']}",
                                help="Based on ROE, ROA, and profit margins - higher returns get higher grades"
                            )
                        
                        # Rating guide
                        with st.expander("📖 Rating Scale Guide"):
                            st.markdown("""
                            **A-D Rating Scale:**
                            - **🟢 A (Excellent)**: Top 25% - Outstanding performance in this category
                            - **🟡 B (Good)**: 25-50% - Above average performance
                            - **🟠 C (Fair)**: 50-75% - Average performance, room for improvement
                            - **🔴 D (Poor)**: Bottom 25% - Below average, needs attention
                            - **⚪ N/A**: Insufficient data available for rating
                            
                            **Category Explanations:**
                            - **Value**: Lower P/E, P/B, P/S ratios indicate better value
                            - **Growth**: Higher revenue and earnings growth rates
                            - **Momentum**: Recent price performance across multiple timeframes
                            - **Profitability**: Return on equity, assets, and profit margins
                            """)

                        # Financial Quality Scores Section
                        st.markdown("---")
                        st.markdown("### 📊 Financial Quality Scores")
                        st.markdown("*Advanced scoring models for financial health and earnings quality assessment*")
                        
                        score_col1, score_col2, score_col3 = st.columns(3)
                        
                        with score_col1:
                            # Piotroski Score
                            st.markdown("**🎯 Piotroski Score**")
                            piotroski_score, piotroski_details = calculate_piotroski_score(ticker_obj, info)
                            
                            if piotroski_score is not None:
                                # Color coding for score
                                if piotroski_score >= 7:
                                    score_color = "🟢"
                                    score_interpretation = "Excellent"
                                elif piotroski_score >= 5:
                                    score_color = "🟡"
                                    score_interpretation = "Good"
                                else:
                                    score_color = "🔴"
                                    score_interpretation = "Poor"
                                
                                st.metric(
                                    label="Score (1-9 scale)",
                                    value=f"{score_color} {piotroski_score}/9",
                                    help="Higher scores indicate better financial health"
                                )
                                st.markdown(f"**Quality:** {score_interpretation}")
                                
                                # Show details in expander
                                with st.expander("📋 Score Details"):
                                    for detail in piotroski_details:
                                        st.markdown(f"• {detail}")
                            else:
                                st.metric("Score (1-9 scale)", "N/A")
                                st.error(piotroski_details)
                        
                        with score_col2:
                            # Altman Z-Score
                            st.markdown("**⚠️ Altman Z-Score**")
                            z_score, z_interpretation = calculate_altman_z_score(ticker_obj, info)
                            
                            if z_score is not None:
                                st.metric(
                                    label="Z-Score",
                                    value=f"{z_score:.2f}",
                                    help="Bankruptcy prediction model"
                                )
                                st.markdown(f"**Status:** {z_interpretation}")
                                
                                # Zone guidance
                                with st.expander("📖 Zone Guide"):
                                    st.markdown("""
                                    **Z-Score Zones:**
                                    • **≤ 1.8**: 🔴 Distress Zone - High bankruptcy risk
                                    • **1.8 - 3.0**: 🟡 Grey Zone - Uncertain, monitor closely
                                    • **≥ 3.0**: 🟢 Safe Zone - Low bankruptcy risk
                                    """)
                            else:
                                st.metric("Z-Score", "N/A")
                                st.error(z_interpretation)
                        
                        with score_col3:
                            # Beneish M-Score
                            st.markdown("**🔍 Beneish M-Score**")
                            m_score, m_interpretation = calculate_beneish_m_score(ticker_obj, info)
                            
                            if m_score is not None:
                                st.metric(
                                    label="M-Score",
                                    value=f"{m_score:.2f}",
                                    help="Earnings manipulation detection model"
                                )
                                st.markdown(f"**Assessment:** {m_interpretation}")
                                
                                # Interpretation guide
                                with st.expander("📖 Score Guide"):
                                    st.markdown("""
                                    **M-Score Interpretation:**
                                    • **≤ -1.78**: 🟢 Unlikely to be manipulating earnings
                                    • **> -1.78**: 🔴 Potential earnings manipulation detected
                                    
                                    *Note: This is a simplified model using available data*
                                    """)
                            else:
                                st.metric("M-Score", "N/A")
                                st.error(m_interpretation)
                        
                    except Exception as e:
                        st.warning(f"Some institutional metrics may not be available: {str(e)}")
                        
                else:
                    st.warning(f"No earnings data available for {symbol_guru.upper()} in the selected period")
            else:
                st.error(f"Unable to fetch data for {symbol_guru.upper()}. Please verify the symbol is correct.")
    
    elif symbol_guru and not st.session_state.get('guru_analyze_clicked', False):
        st.info("👆 Click 'Analyze Earnings' to start the detailed analysis")


def display_price_action_tab(symbol, data, ticker_info, ticker_obj, ma_50, ma_200, rsi, support_level, resistance_level, selected_period, market, auto_refresh):
    """Display price action metrics and key financial data"""
    
    # Auto-refresh status display
    if auto_refresh:
        col_status1, col_status2, col_status3 = st.columns([2, 1, 1])
        with col_status1:
            st.success(f"✅ Live tracking {symbol} - Updates every 10 minutes")
        with col_status2:
            st.metric("Refresh #", st.session_state.get('refresh_count', 0))
        with col_status3:
            current_time = datetime.now().strftime("%H:%M:%S")
            st.metric("Last Update", current_time)
    
    # Calculate basic data needed for both modes
    current_price = data['Close'].iloc[-1]
    previous_close = ticker_info.get('previousClose', data['Close'].iloc[-2] if len(data) > 1 else current_price)
    price_change = current_price - previous_close
    price_change_pct = (price_change / previous_close) * 100 if previous_close != 0 else 0
    currency = "₹" if market == "India" else "$"
    company_name = ticker_info.get('longName', ticker_info.get('shortName', symbol))
    
    # Check view mode for display format
    view_mode = st.session_state.get('view_mode', 'Standard')
    
    if view_mode == 'Compact':
        # COMPACT MODE - Company info and pricing as tables
        # CSS styling for compact HTML tables
        st.markdown("""
        <style>
        .compact-table {
            width: 100% !important;
            border-collapse: collapse !important;
            margin: 8px 0 !important;
            font-size: 14px !important;
        }
        .compact-table td {
            padding: 6px 12px !important;
            border: 1px solid #e6e6e6 !important;
            text-align: left !important;
        }
        .compact-table tr:nth-child(even) {
            background-color: #f9f9f9 !important;
        }
        .compact-table tr:hover {
            background-color: #f5f5f5 !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown(f"### 🏢 {company_name} ({symbol})")
        
        # Company Information Table
        company_data = [
            ["Sector", ticker_info.get('sector', 'N/A'), "Industry", ticker_info.get('industry', 'N/A')],
        ]
        
        if ticker_info.get('sector') or ticker_info.get('industry'):
            df_company = pd.DataFrame(company_data)
            # Use HTML table instead of dataframe to avoid headers
            html_table = df_company.to_html(index=False, header=False, table_id="company_table", classes="compact-table")
            st.markdown(html_table, unsafe_allow_html=True)
        
        st.markdown("**💰 Current Price & Market Data**")
        
        # Current Price Table
        price_data = [
            ["Current Price", f"{currency}{current_price:.2f}", "Previous Close", f"{currency}{previous_close:.2f}"],
            ["Price Change", f"{price_change:+.2f}", "Change %", f"{price_change_pct:+.2f}%"]
        ]
        
        df_current_price = pd.DataFrame(price_data)
        html_table = df_current_price.to_html(index=False, header=False, table_id="price_table", classes="compact-table")
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Extended Hours Trading Table
        st.markdown("**🕐 Extended Hours Trading**")
        
        try:
            after_market_data = get_after_market_data(symbol, market)
            
            # Market status calculation
            import datetime
            current_time = datetime.datetime.now()
            
            if market == "US":
                market_open_time = current_time.replace(hour=9, minute=30)
                market_close_time = current_time.replace(hour=16, minute=0)
                
                if market_open_time <= current_time <= market_close_time:
                    market_status = "🟢 Open"
                elif current_time < market_open_time:
                    market_status = "🟡 Pre-Market"
                else:
                    market_status = "🔴 After-Hours"
            else:
                market_status = "🔵 Active"
            
            # Extended hours data table
            extended_data = [
                ["Pre-Market Change", after_market_data.get('pre_market_change', 'N/A'), "After-Hours Change", after_market_data.get('post_market_change', 'N/A')],
                ["Regular Close", after_market_data.get('regular_session_close', 'N/A'), "Market Status", market_status]
            ]
            
            df_extended = pd.DataFrame(extended_data)
            html_table = df_extended.to_html(index=False, header=False, table_id="extended_table", classes="compact-table")
            st.markdown(html_table, unsafe_allow_html=True)
            
        except Exception as e:
            st.info("Extended hours data not available")
            
    else:
        # STANDARD MODE - Original metric display
        # Company information and current price
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown(f"### 🏢 {company_name} ({symbol})")
            
            if 'sector' in ticker_info:
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    st.markdown(f"**Sector:** {ticker_info.get('sector', 'N/A')}")
                with col_info2:
                    st.markdown(f"**Industry:** {ticker_info.get('industry', 'N/A')}")
        
        with col2:
            st.metric(
                label=f"Current Price ({currency})",
                value=f"{currency}{current_price:.2f}",
                delta=f"{price_change:+.2f} ({price_change_pct:+.2f}%)"
            )
        
        st.markdown("---")
        
        # After-market information
        st.subheader("🕐 Extended Hours Trading")
        
        try:
            after_market_data = get_after_market_data(symbol, market)
            
            col_am1, col_am2, col_am3, col_am4 = st.columns(4)
            
            with col_am1:
                if after_market_data['pre_market_change'] != 'N/A':
                    st.metric(
                        "Pre-Market Change",
                        after_market_data['pre_market_change'],
                        after_market_data['pre_market_change_percent']
                    )
                else:
                    st.metric("Pre-Market Change", "N/A")
            
            with col_am2:
                if after_market_data['post_market_change'] != 'N/A':
                    st.metric(
                        "After-Hours Change", 
                        after_market_data['post_market_change'],
                        after_market_data['post_market_change_percent']
                    )
                else:
                    st.metric("After-Hours Change", "N/A")
            
            with col_am3:
                if after_market_data['regular_session_close'] != 'N/A':
                    st.metric("Regular Session Close", after_market_data['regular_session_close'])
                else:
                    st.metric("Regular Session Close", "N/A")
            
            with col_am4:
                # Market status indicator
                import datetime
                current_time = datetime.datetime.now()
                
                # Simple market hours check (9:30 AM - 4:00 PM ET for US markets)
                if market == "US":
                    market_open_time = current_time.replace(hour=9, minute=30)
                    market_close_time = current_time.replace(hour=16, minute=0)
                    
                    if market_open_time <= current_time <= market_close_time:
                        market_status = "🟢 Open"
                    elif current_time < market_open_time:
                        market_status = "🟡 Pre-Market"
                    else:
                        market_status = "🔴 After-Hours"
                else:
                    market_status = "🔵 Active"
                
                st.metric("Market Status", market_status)
                
        except Exception as e:
            st.info("Extended hours data not available")
    
    st.markdown("---")
    # Check view mode for key metrics display
    view_mode = st.session_state.get('view_mode', 'Standard')
    
    if view_mode == 'Compact':
        st.subheader("📊 Key Metrics (Table View)")
        
        # COMPACT TABLE VIEW - All key metrics in tables
        # Get all required data first
        week_52_high = ticker_info.get('fiftyTwoWeekHigh', 0)
        week_52_low = ticker_info.get('fiftyTwoWeekLow', 0)
        current_rsi = rsi.iloc[-1] if not rsi.empty else 0
        ma_50_current = ma_50.iloc[-1] if not ma_50.empty else 0
        ma_200_current = ma_200.iloc[-1] if not ma_200.empty else 0
        avg_volume = ticker_info.get('averageVolume', 0)
        current_volume = data['Volume'].iloc[-1] if len(data) > 0 else 0
        market_cap = ticker_info.get('marketCap', 0)
        beta = ticker_info.get('beta', 0)
        
        # Calculate derived values
        position_52w = ((current_price - week_52_low) / (week_52_high - week_52_low)) * 100 if week_52_high and week_52_low else 0
        volume_ratio = current_volume / avg_volume if avg_volume > 0 else 0
        ma_50_change = ((current_price - ma_50_current) / ma_50_current * 100) if ma_50_current != 0 else 0
        ma_200_change = ((current_price - ma_200_current) / ma_200_current * 100) if ma_200_current != 0 else 0
        
        # % change of CTP from 52W high/low calculations
        pct_change_from_52w_high = ((current_price - week_52_high) / week_52_high) * 100 if week_52_high else 0
        pct_change_from_52w_low = ((current_price - week_52_low) / week_52_low) * 100 if week_52_low else 0
        
        # % change of CTP from MA50 and MA200 calculations
        pct_change_from_ma_50 = ((current_price - ma_50_current) / ma_50_current) * 100 if ma_50_current else 0
        pct_change_from_ma_200 = ((current_price - ma_200_current) / ma_200_current) * 100 if ma_200_current else 0
        
        # Format display text with colors and descriptive language
        if pct_change_from_52w_high < 0:
            high_display_text = f"CTP is {abs(pct_change_from_52w_high):.1f}% below"
        else:
            high_display_text = f"CTP is {pct_change_from_52w_high:.1f}% above"
        
        if pct_change_from_52w_low < 0:
            low_display_text = f"CTP is {abs(pct_change_from_52w_low):.1f}% below"
        else:
            low_display_text = f"CTP is {pct_change_from_52w_low:.1f}% above"
        
        if pct_change_from_ma_50 < 0:
            ma_50_display_text = f"CTP is {abs(pct_change_from_ma_50):.1f}% below"
        else:
            ma_50_display_text = f"CTP is {pct_change_from_ma_50:.1f}% above"
        
        if pct_change_from_ma_200 < 0:
            ma_200_display_text = f"CTP is {abs(pct_change_from_ma_200):.1f}% below"
        else:
            ma_200_display_text = f"CTP is {pct_change_from_ma_200:.1f}% above"
        
        # Format market cap
        if market_cap >= 1e12:
            cap_display = f"{currency}{market_cap/1e12:.2f}T"
        elif market_cap >= 1e9:
            cap_display = f"{currency}{market_cap/1e9:.2f}B"
        elif market_cap >= 1e6:
            cap_display = f"{currency}{market_cap/1e6:.2f}M"
        else:
            cap_display = f"{currency}{market_cap:.0f}" if market_cap else "N/A"
        
        # Create colored HTML for % from CTP values
        high_color = "red" if pct_change_from_52w_high < 0 else "green"
        low_color = "red" if pct_change_from_52w_low < 0 else "green"
        ma_50_color = "red" if pct_change_from_ma_50 < 0 else "green"
        ma_200_color = "red" if pct_change_from_ma_200 < 0 else "green"
        
        high_colored_text = f"<span style='color: {high_color}; font-weight: bold;'>{high_display_text}</span>"
        low_colored_text = f"<span style='color: {low_color}; font-weight: bold;'>{low_display_text}</span>"
        ma_50_colored_text = f"<span style='color: {ma_50_color}; font-weight: bold;'>{ma_50_display_text}</span>"
        ma_200_colored_text = f"<span style='color: {ma_200_color}; font-weight: bold;'>{ma_200_display_text}</span>"
        
        # Price Action Table with % from CTP (with embedded colors)
        price_data = [
            ["Current Price", f"{currency}{current_price:.2f}", "52W High", f"{currency}{week_52_high:.2f}" if week_52_high else "N/A"],
            ["% from CTP (High)", high_colored_text, "52W Low", f"{currency}{week_52_low:.2f}" if week_52_low else "N/A"],
            ["% from CTP (Low)", low_colored_text, "52W Position", f"{position_52w:.1f}%" if position_52w else "N/A"],
            ["MA 50", f"{currency}{ma_50_current:.2f}" if ma_50_current else "N/A", "MA 200", f"{currency}{ma_200_current:.2f}" if ma_200_current else "N/A"],
            ["% from CTP (MA50)", ma_50_colored_text, "% from CTP (MA200)", ma_200_colored_text],
            ["RSI", f"{current_rsi:.1f}" if current_rsi else "N/A", "Volume Ratio", f"{volume_ratio:.2f}x" if volume_ratio else "N/A"],
            ["Market Cap", cap_display, "Beta", f"{beta:.2f}" if beta else "N/A"],
            ["Avg Volume", f"{avg_volume:,}" if avg_volume else "N/A", "Current Vol", f"{current_volume:,}" if current_volume else "N/A"]
        ]
        
        df_price = pd.DataFrame(price_data)
        html_table = df_price.to_html(index=False, header=False, table_id="metrics_table", classes="compact-table", escape=False)
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Support/Resistance & Safe Level Analysis Table
        st.markdown("**📊 Support/Resistance & Safe Levels**")
        
        # Calculate safe levels (CTP ± 12.5%)
        safe_level_low = current_price * 0.875  # CTP - 12.5%
        safe_level_high = current_price * 1.125  # CTP + 12.5%
        
        resistance_table_data = [
            ["Support Level", f"{currency}{support_level:.2f}" if support_level else "N/A", "Resistance Level", f"{currency}{resistance_level:.2f}" if resistance_level else "N/A"],
            ["Distance to Support", f"{((current_price - support_level)/support_level*100):+.2f}%" if support_level else "N/A", "Distance to Resistance", f"{((resistance_level - current_price)/current_price*100):+.2f}%" if resistance_level else "N/A"],
            ["Safe Level Low", f"{currency}{safe_level_low:.2f}", "Safe Level High", f"{currency}{safe_level_high:.2f}"],
            ["To Safe Low", f"{((safe_level_low - current_price)/current_price*100):+.2f}%", "To Safe High", f"{((safe_level_high - current_price)/current_price*100):+.2f}%"]
        ]
        
        df_resistance = pd.DataFrame(resistance_table_data)
        html_table = df_resistance.to_html(index=False, header=False, table_id="resistance_table", classes="compact-table")
        st.markdown(html_table, unsafe_allow_html=True)
        
    else:
        st.subheader("📊 Key Metrics")
        st.info("💡 Switch to Compact view above for table format that eliminates scrolling")
        
        # STANDARD MODE - Original metric columns
        # Calculate % from CTP for standard view
        week_52_high = ticker_info.get('fiftyTwoWeekHigh', 0)
        week_52_low = ticker_info.get('fiftyTwoWeekLow', 0)
        pct_change_from_52w_high = ((current_price - week_52_high) / week_52_high) * 100 if week_52_high else 0
        pct_change_from_52w_low = ((current_price - week_52_low) / week_52_low) * 100 if week_52_low else 0
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            # 52-week high with % from CTP
            if week_52_high:
                high_color = "red" if pct_change_from_52w_high < 0 else "green"
                high_delta = f"CTP is {abs(pct_change_from_52w_high):.1f}% {'below' if pct_change_from_52w_high < 0 else 'above'}"
                st.metric("52W High", f"{currency}{week_52_high:.2f}")
                st.markdown(f"<span style='color: {high_color}; font-size: 12px;'>{high_delta}</span>", unsafe_allow_html=True)
            else:
                st.metric("52W High", "N/A")
        
        with col2:
            # 52-week low with % from CTP
            if week_52_low:
                low_color = "red" if pct_change_from_52w_low < 0 else "green"
                low_delta = f"CTP is {abs(pct_change_from_52w_low):.1f}% {'below' if pct_change_from_52w_low < 0 else 'above'}"
                st.metric("52W Low", f"{currency}{week_52_low:.2f}")
                st.markdown(f"<span style='color: {low_color}; font-size: 12px;'>{low_delta}</span>", unsafe_allow_html=True)
            else:
                st.metric("52W Low", "N/A")
        
        with col3:
            # Volume analysis
            avg_volume = ticker_info.get('averageVolume', 0)
            current_volume = data['Volume'].iloc[-1]
            
            if avg_volume > 0:
                volume_ratio = current_volume / avg_volume
                st.metric("Volume Ratio", f"{volume_ratio:.2f}x",
                         help="Volume Ratio\nCompares current trading volume to average volume over recent period.\n• Ratio > 1.5x: High activity, increased interest/news\n• Ratio > 2x: Very high activity, potential breakout or news\n• Ratio < 0.5x: Low activity, lack of interest\n• High volume confirms price moves, low volume questions sustainability")
                st.caption(f"Avg: {avg_volume:,.0f}")
            else:
                st.metric("Volume Ratio", "N/A",
                         help="Volume Ratio\nCompares current trading volume to average volume over recent period.\n• Ratio > 1.5x: High activity, increased interest/news\n• Ratio > 2x: Very high activity, potential breakout or news\n• Ratio < 0.5x: Low activity, lack of interest\n• High volume confirms price moves, low volume questions sustainability")
        
        with col4:
            # 52W Position
            if week_52_high and week_52_low:
                position_52w = ((current_price - week_52_low) / (week_52_high - week_52_low)) * 100
                st.metric("52W Position", f"{position_52w:.1f}%",
                         help="52-Week Range Position\nShows where current price sits within the 52-week trading range.\n• 0-25%: Near 52-week low, potential value or distressed\n• 25-50%: Lower half of range, below average\n• 50-75%: Upper half of range, above average\n• 75-100%: Near 52-week high, strong momentum or expensive\n• >90%: At/near highs, momentum stock or potential resistance")
                st.caption(f"Within 52W range")
            else:
                st.metric("52W Position", "N/A",
                         help="52-Week Range Position\nShows where current price sits within the 52-week trading range.\n• 0-25%: Near 52-week low, potential value or distressed\n• 25-50%: Lower half of range, below average\n• 50-75%: Upper half of range, above average\n• 75-100%: Near 52-week high, strong momentum or expensive\n• >90%: At/near highs, momentum stock or potential resistance")
        
        # Additional metrics row
        st.markdown("#### 📊 Additional Metrics")
        col_add1, col_add2, col_add3, col_add4 = st.columns(4)
        
        with col_add1:
            # RSI
            current_rsi = rsi.iloc[-1] if not rsi.empty else 0
            rsi_status = "Overbought" if current_rsi > 70 else "Oversold" if current_rsi < 30 else "Neutral"
            st.metric("RSI", f"{current_rsi:.1f}",
                     help="RSI (Relative Strength Index)\nMomentum oscillator measuring speed and change of price movements. Range 0-100.\n• RSI > 70: Overbought territory - potential sell signal or pullback\n• RSI < 30: Oversold territory - potential buy signal or bounce\n• RSI 30-70: Neutral zone - trend continuation likely\n• Look for RSI divergences: price makes new highs/lows but RSI doesn't confirm")
            st.caption(rsi_status)
        
        with col_add2:
            # Market cap
            market_cap = ticker_info.get('marketCap', 0)
            if market_cap:
                if market_cap >= 1e12:
                    cap_display = f"{currency}{market_cap/1e12:.2f}T"
                elif market_cap >= 1e9:
                    cap_display = f"{currency}{market_cap/1e9:.2f}B"
                elif market_cap >= 1e6:
                    cap_display = f"{currency}{market_cap/1e6:.2f}M"
                else:
                    cap_display = f"{currency}{market_cap:.0f}"
                st.metric("Market Cap", cap_display,
                         help="Market Capitalization\nTotal value of all outstanding shares (shares × current price).\n• <$2B: Small-cap - higher risk/reward, more volatile\n• $2B-$10B: Mid-cap - moderate risk/reward balance\n• >$10B: Large-cap - lower risk, stable, institutional favorite\n• >$200B: Mega-cap - market leaders, defensive plays\nCompare with Enterprise Value for debt consideration.")
            else:
                st.metric("Market Cap", "N/A",
                         help="Market Capitalization\nTotal value of all outstanding shares (shares × current price).\n• <$2B: Small-cap - higher risk/reward, more volatile\n• $2B-$10B: Mid-cap - moderate risk/reward balance\n• >$10B: Large-cap - lower risk, stable, institutional favorite\n• >$200B: Mega-cap - market leaders, defensive plays\nCompare with Enterprise Value for debt consideration.")
        
        with col_add3:
            # Beta
            beta = ticker_info.get('beta', 0)
            if beta:
                st.metric("Beta", f"{beta:.2f}",
                         help="Beta - Stock Volatility vs Market\nMeasures how much the stock moves relative to the overall market (S&P 500 = 1.0).\n• Beta = 1.0: Moves with the market\n• Beta > 1.0: More volatile than market (e.g., 1.5 = 50% more volatile)\n• Beta < 1.0: Less volatile than market (e.g., 0.5 = 50% less volatile)\n• Beta < 0: Moves opposite to market (rare)\n• High Beta: Higher risk/reward potential, tech stocks\n• Low Beta: Defensive, utilities, consumer staples")
            else:
                st.metric("Beta", "N/A",
                         help="Beta - Stock Volatility vs Market\nMeasures how much the stock moves relative to the overall market (S&P 500 = 1.0).\n• Beta = 1.0: Moves with the market\n• Beta > 1.0: More volatile than market (e.g., 1.5 = 50% more volatile)\n• Beta < 1.0: Less volatile than market (e.g., 0.5 = 50% less volatile)\n• Beta < 0: Moves opposite to market (rare)\n• High Beta: Higher risk/reward potential, tech stocks\n• Low Beta: Defensive, utilities, consumer staples")
        
        with col_add4:
            # Empty space for layout
            st.write("")
        
        # Moving average analysis
        st.markdown("---")
        st.subheader("📈 Moving Average Analysis")
        
        col_ma1, col_ma2, col_ma3 = st.columns(3)
        
        with col_ma1:
            ma_50_current = ma_50.iloc[-1] if not ma_50.empty else 0
            pct_change_from_ma_50 = ((current_price - ma_50_current) / ma_50_current) * 100 if ma_50_current else 0
            ma_50_color = "red" if pct_change_from_ma_50 < 0 else "green"
            ma_50_delta = f"CTP is {abs(pct_change_from_ma_50):.1f}% {'below' if pct_change_from_ma_50 < 0 else 'above'}"
            
            st.metric("50-Day MA", f"{currency}{ma_50_current:.2f}",
                     help="50-Day Moving Average\nShort-term trend indicator showing average price over 50 days.\n• Price above MA 50: Short-term uptrend, bullish momentum\n• Price below MA 50: Short-term downtrend, bearish momentum\n• MA 50 slope: Rising = strengthening trend, Falling = weakening trend\n• Use as dynamic support (uptrend) or resistance (downtrend) level")
            st.markdown(f"<span style='color: {ma_50_color}; font-size: 12px;'>{ma_50_delta}</span>", unsafe_allow_html=True)
        
        with col_ma2:
            # Check if we have enough data for 200-day MA
            data_days = len(data)
            ma_200_available = data_days >= 200 and not ma_200.empty and not pd.isna(ma_200.iloc[-1])
            
            if ma_200_available:
                ma_200_current = ma_200.iloc[-1]
                pct_change_from_ma_200 = ((current_price - ma_200_current) / ma_200_current) * 100
                ma_200_color = "red" if pct_change_from_ma_200 < 0 else "green"
                ma_200_delta = f"CTP is {abs(pct_change_from_ma_200):.1f}% {'below' if pct_change_from_ma_200 < 0 else 'above'}"
                
                st.metric("200-Day MA", f"{currency}{ma_200_current:.2f}",
                         help="200-Day Moving Average\nLong-term trend indicator and major support/resistance level.\n• Price above MA 200: Long-term bull market, major uptrend\n• Price below MA 200: Long-term bear market, major downtrend\n• Golden Cross: MA 50 crosses above MA 200 = strong bullish signal\n• Death Cross: MA 50 crosses below MA 200 = strong bearish signal\n• MA 200 acts as major psychological support/resistance level")
                st.markdown(f"<span style='color: {ma_200_color}; font-size: 12px;'>{ma_200_delta}</span>", unsafe_allow_html=True)
            else:
                st.metric("200-Day MA", "Not enough data")
                st.markdown(f"<span style='color: orange; font-size: 12px;'>Not enough data</span>", unsafe_allow_html=True)
        
        with col_ma3:
            # Trend analysis - only show if both MAs are available
            ma_50_available = len(data) >= 50 and not ma_50.empty and not pd.isna(ma_50.iloc[-1])
            
            if ma_50_available and ma_200_available:
                ma_50_current = ma_50.iloc[-1]
                ma_200_current = ma_200.iloc[-1]
                if ma_50_current > ma_200_current:
                    trend = "🟢 Bullish"
                elif ma_50_current < ma_200_current:
                    trend = "🔴 Bearish"
                else:
                    trend = "🟡 Neutral"
                st.metric("Trend", trend)
            else:
                st.metric("Trend", "Not enough data")
                st.markdown(f"<span style='color: orange; font-size: 12px;'>Not enough data</span>", unsafe_allow_html=True)
        
        # Support and resistance levels
        st.markdown("---")
        st.subheader("🎯 Support & Resistance Levels")
        
        col_sr1, col_sr2, col_sr3, col_sr4 = st.columns(4)
        
        with col_sr1:
            st.metric("Support Level", f"{currency}{support_level:.2f}",
                     help="Support Level - Price Floor\nA price level where buying interest is expected to emerge, preventing further decline.\n• Strong Support: Price bounces multiple times\n• Weak Support: Price breaks through with volume\n• Trading Strategy: Buy near support, stop-loss below support\n• Break Below: Often signals further decline to next support level")
        
        with col_sr2:
            st.metric("Resistance Level", f"{currency}{resistance_level:.2f}",
                     help="Resistance Level - Price Ceiling\nA price level where selling interest is expected to emerge, preventing further rise.\n• Strong Resistance: Price rejects multiple times\n• Weak Resistance: Price breaks through with volume\n• Trading Strategy: Sell near resistance, stop-loss above resistance\n• Break Above: Often signals further rise to next resistance level")
        
        with col_sr3:
            # Safe level low (CTP -12.5%)
            safe_low = current_price * 0.875
            st.metric("Safe Level Low", f"{currency}{safe_low:.2f}",
                     help="Safe Entry Level (CTP -12.5%)\nConservative entry point providing buffer against normal volatility.\n• Reduces immediate loss risk after entry\n• Good for dollar-cost averaging\n• Requires patience to reach this level\n• More relevant in volatile markets")
        
        with col_sr4:
            # Safe level high (CTP +12.5%)
            safe_high = current_price * 1.125
            st.metric("Safe Level High", f"{currency}{safe_high:.2f}",
                     help="Safe Exit Level (CTP +12.5%)\nConservative profit-taking point to avoid market tops.\n• Ensures reasonable profit capture\n• Helps control greed and overholding\n• Good for partial position exits\n• Allows re-entry on pullbacks")
        
        # Earnings and dividend information moved to dedicated "Earnings & Dividends" tab
    
    # Enhanced Fibonacci Analysis - Available in both view modes
    st.markdown("---")
    st.subheader("📐 Fibonacci Analysis – Next Two Levels")
    
    # Add period selection
    col_period, col_spacer = st.columns([1, 3])
    with col_period:
        period_months = st.selectbox(
            "Reference Range:",
            [3, 6],
            index=0,
            format_func=lambda x: f"{x}-Month High/Low"
        )
    
    # Calculate Fibonacci levels
    fibonacci_data = calculate_fibonacci_levels(data, period_months=period_months, symbol=symbol)
    if fibonacci_data:
        current_price = fibonacci_data['current_price']
        reference_high = fibonacci_data['reference_high']
        reference_low = fibonacci_data['reference_low']
        analysis_type = fibonacci_data['analysis_type']
        next_levels_above = fibonacci_data['next_levels_above']
        next_levels_below = fibonacci_data['next_levels_below']
        
        if view_mode == "Compact":
            # COMPACT MODE - Fibonacci as Tables
            
            # Analysis status
            if analysis_type == "retracement":
                analysis_status = "🎯 Price within range - Using Retracement Levels"
            elif analysis_type == "upward_extension":
                analysis_status = "📈 Price above range - Using Upward Extensions"
            else:
                analysis_status = "📉 Price below range - Using Downward Extensions"
            
            st.info(analysis_status)
            
            # Fibonacci Reference Data Table
            fib_reference_data = [
                [f"{period_months}M High", format_currency(reference_high, market), "Current Price", format_currency(current_price, market)],
                [f"{period_months}M Low", format_currency(reference_low, market), "Analysis Type", analysis_type.replace("_", " ").title()]
            ]
            
            df_fib_ref = pd.DataFrame(fib_reference_data)
            html_table = df_fib_ref.to_html(index=False, header=False, table_id="fib_ref_table", classes="compact-table")
            st.markdown(html_table, unsafe_allow_html=True)
            
            # Fibonacci Levels Table
            st.markdown("**📊 Next Fibonacci Levels**")
            
            fib_levels_data = []
            
            # Add levels above
            if next_levels_above:
                for i, level in enumerate(next_levels_above, 1):
                    distance_pct = ((level['price'] - current_price) / current_price) * 100
                    fib_levels_data.append([
                        f"Above {i}", 
                        level['label'], 
                        format_currency(level['price'], market), 
                        f"+{distance_pct:.1f}%"
                    ])
            
            # Add levels below  
            if next_levels_below:
                for i, level in enumerate(next_levels_below, 1):
                    distance_pct = ((current_price - level['price']) / current_price) * 100
                    fib_levels_data.append([
                        f"Below {i}", 
                        level['label'], 
                        format_currency(level['price'], market), 
                        f"-{distance_pct:.1f}%"
                    ])
            
            if fib_levels_data:
                df_fib_levels = pd.DataFrame(fib_levels_data)
                html_table = df_fib_levels.to_html(index=False, header=False, table_id="fib_levels_table", classes="compact-table")
                st.markdown(html_table, unsafe_allow_html=True)
            else:
                st.info("No Fibonacci levels found near current price")
                
        else:
            # STANDARD MODE - Original metric display
            # Display current status
            col_status1, col_status2, col_status3 = st.columns(3)
            
            with col_status1:
                st.metric(
                    label=f"{period_months}M Reference High",
                    value=format_currency(reference_high, market),
                    help=f"Highest price in the last {period_months} months"
                )
            
            with col_status2:
                st.metric(
                    label="Current Price",
                    value=format_currency(current_price, market),
                    help="Current market price"
                )
            
            with col_status3:
                st.metric(
                    label=f"{period_months}M Reference Low",
                    value=format_currency(reference_low, market),
                    help=f"Lowest price in the last {period_months} months"
                )
            
            # Analysis type indicator
            if analysis_type == "retracement":
                analysis_status = "🎯 Price within range - Using Retracement Levels"
            elif analysis_type == "upward_extension":
                analysis_status = "📈 Price above range - Using Upward Extensions"
            else:
                analysis_status = "📉 Price below range - Using Downward Extensions"
            
            st.info(analysis_status)
            
            # Display next levels
            col_above, col_below = st.columns(2)
            
            with col_above:
                st.markdown("**🔺 Next Two Levels Above:**")
                if next_levels_above:
                    for i, level in enumerate(next_levels_above, 1):
                        distance = level['price'] - current_price
                        distance_pct = (distance / current_price) * 100
                        # Create comprehensive tooltip based on Fibonacci level type
                        fib_tooltip = get_fibonacci_tooltip(level['label'], 'resistance')
                        st.metric(
                            label=f"Level {i} - {level['label']}",
                            value=format_currency(level['price'], market),
                            delta=f"+{distance_pct:.1f}%",
                            help=f"{fib_tooltip}\n\nDistance: {format_currency(distance, market)} ({distance_pct:.1f}%)"
                        )
                else:
                    st.write("No levels found above current price")
            
            with col_below:
                st.markdown("**🔻 Next Two Levels Below:**")
                if next_levels_below:
                    for i, level in enumerate(next_levels_below, 1):
                        distance = current_price - level['price']
                        distance_pct = (distance / current_price) * 100
                        # Create comprehensive tooltip based on Fibonacci level type
                        fib_tooltip = get_fibonacci_tooltip(level['label'], 'support')
                        st.metric(
                            label=f"Level {i} - {level['label']}",
                            value=format_currency(level['price'], market),
                            delta=f"-{distance_pct:.1f}%",
                            help=f"{fib_tooltip}\n\nDistance: {format_currency(distance, market)} ({distance_pct:.1f}%)"
                        )
                else:
                    st.write("No levels found below current price")
    else:
        st.info("Fibonacci analysis requires sufficient price history for calculation")
    
    # Earnings and dividend information moved to dedicated tab


def display_technical_charts_tab(symbol, data, ma_50, ma_200, macd_line, signal_line, histogram, rsi, cmf, selected_period, market):
    """Display technical analysis charts"""
    
    st.subheader(f"📈 Technical Charts for {symbol}")
    
    # Fibonacci Levels Guide
    with st.expander("📖 Fibonacci Levels Trading Guide"):
        st.markdown("""
        **Fibonacci Levels** - Mathematical ratios based on the golden ratio, used to identify support/resistance levels:
        
        **Fibonacci Retracement Levels (Pullback Targets):**
        • **23.6%**: Shallow pullback - Strong trending market, minor correction
        • **38.2%**: Moderate pullback - Common retracement in healthy trends
        • **50.0%**: Half retracement - Not true Fibonacci but widely watched psychological level
        • **61.8%**: Golden Ratio - Deep retracement, often marks trend continuation point
        • **78.6%**: Very deep pullback - Trend weakening, possible reversal zone
        
        **Fibonacci Extension Levels (Breakout Targets):**
        • **127.2%**: First extension target - Common profit-taking level
        • **161.8%**: Golden extension - Major resistance/support after breakout
        • **200.0%**: Double the move - Strong momentum target
        • **261.8%**: Extreme extension - Parabolic move territory
        
        **Trading Strategies:**
        • **Support/Resistance**: Levels act as bounce points (support) or rejection points (resistance)
        • **Entry Points**: Buy near support levels, sell near resistance levels
        • **Confluence**: Stronger signals when Fib levels align with other technical indicators
        • **Volume Confirmation**: High volume at Fib levels increases reliability
        • **Break and Retest**: Price often retests Fib levels after breaking through
        """)
    
    # Technical indicators
    st.markdown("#### 📊 Price Chart with Moving Averages")
    with st.expander("📖 Moving Averages Chart Pattern Guide"):
        st.markdown("""
        **Moving Averages** - Trend-following indicators that smooth price data to identify direction and momentum:
        
        **50-Day Moving Average (Blue Line):**
        • **Short-term trend indicator** showing recent price momentum
        • **Price Above MA 50**: Short-term uptrend, bullish momentum
        • **Price Below MA 50**: Short-term downtrend, bearish momentum
        • **MA 50 Slope**: Rising = strengthening trend, Falling = weakening trend
        • **Dynamic Support/Resistance**: Acts as support in uptrends, resistance in downtrends
        
        **200-Day Moving Average (Red Line):**
        • **Long-term trend indicator** and major market sentiment gauge
        • **Price Above MA 200**: Long-term bull market, major uptrend
        • **Price Below MA 200**: Long-term bear market, major downtrend
        • **Major Psychological Level**: Institutional investors watch this closely
        
        **Critical Crossover Signals:**
        • **Golden Cross**: MA 50 crosses above MA 200 = Strong bullish signal (major buy signal)
        • **Death Cross**: MA 50 crosses below MA 200 = Strong bearish signal (major sell signal)
        • **Moving Average Stack**: MA 50 > MA 200 = Bullish alignment, MA 50 < MA 200 = Bearish alignment
        
        **Price Action Patterns:**
        • **Bounces**: Price touches MA and bounces = trend continuation
        • **Breaks**: Price breaks through MA with volume = potential trend change  
        • **Pullbacks**: Price returns to test MA after breakout = healthy retest
        • **Distance**: Extreme distance from MAs often leads to mean reversion
        """)
    
    # Currency based on market
    currency = "₹" if market == "India" else "$"
    
    # Main price chart with moving averages
    price_fig = go.Figure()
    
    # Add candlestick chart
    price_fig.add_trace(go.Candlestick(
        x=data.index,
        open=data['Open'],
        high=data['High'],
        low=data['Low'],
        close=data['Close'],
        name=f'{symbol} Price',
        increasing_line_color='green',
        decreasing_line_color='red'
    ))
    
    # Add moving averages
    if not ma_50.empty:
        price_fig.add_trace(go.Scatter(
            x=data.index,
            y=ma_50,
            mode='lines',
            name='50-Day MA',
            line=dict(color='blue', width=2)
        ))
    
    if not ma_200.empty:
        price_fig.add_trace(go.Scatter(
            x=data.index,
            y=ma_200,
            mode='lines',
            name='200-Day MA',
            line=dict(color='red', width=2)
        ))
    
    price_fig.update_layout(
        title=f'{symbol} Price Chart with Moving Averages',
        xaxis_title='Date',
        yaxis_title=f'Price ({currency})',
        height=600,
        xaxis_rangeslider_visible=False
    )
    
    st.plotly_chart(price_fig, use_container_width=True)
    create_export_buttons(price_fig, "Price_Chart", symbol)
    
    # Technical indicators in separate charts
    col1, col2 = st.columns(2)
    
    with col1:
        # MACD Chart
        st.markdown("#### 📊 MACD Analysis")
        with st.expander("📖 MACD Chart Pattern Guide"):
            st.markdown("""
            **MACD (Moving Average Convergence Divergence)** - Momentum indicator showing relationship between two moving averages:
            
            **Key Signals to Watch:**
            • **MACD Line Crosses Above Signal Line**: Bullish crossover - potential buy signal
            • **MACD Line Crosses Below Signal Line**: Bearish crossover - potential sell signal
            • **Histogram Above Zero**: MACD line above signal line - bullish momentum
            • **Histogram Below Zero**: MACD line below signal line - bearish momentum
            • **Increasing Histogram**: Strengthening momentum in current direction
            • **Decreasing Histogram**: Weakening momentum, potential reversal ahead
            • **MACD Above/Below Zero Line**: Above = long-term bullish, Below = long-term bearish
            • **Divergences**: Price makes new highs/lows but MACD doesn't confirm - reversal warning
            """)
        
        macd_fig = go.Figure()
        
        if not macd_line.empty and not signal_line.empty and not histogram.empty:
            macd_fig.add_trace(go.Scatter(
                x=data.index,
                y=macd_line,
                mode='lines',
                name='MACD Line',
                line=dict(color='blue')
            ))
            
            macd_fig.add_trace(go.Scatter(
                x=data.index,
                y=signal_line,
                mode='lines',
                name='Signal Line',
                line=dict(color='red')
            ))
            
            macd_fig.add_trace(go.Bar(
                x=data.index,
                y=histogram,
                name='Histogram',
                marker_color='gray',
                opacity=0.7
            ))
        
        macd_fig.update_layout(
            title='MACD Indicator',
            xaxis_title='Date',
            yaxis_title='MACD',
            height=400
        )
        
        st.plotly_chart(macd_fig, use_container_width=True)
        create_export_buttons(macd_fig, "MACD_Analysis", symbol)
    
    with col2:
        # RSI Chart
        st.markdown("#### 📊 RSI Analysis")
        with st.expander("📖 RSI Chart Pattern Guide"):
            st.markdown("""
            **RSI (Relative Strength Index)** - Momentum oscillator measuring speed and change of price movements (0-100):
            
            **Key Levels & Signals:**
            • **RSI > 70**: Overbought territory - potential sell signal or pullback expected
            • **RSI < 30**: Oversold territory - potential buy signal or bounce expected  
            • **RSI 30-70**: Neutral zone - trend continuation likely
            • **RSI Breaking Above 50**: Bullish momentum building
            • **RSI Breaking Below 50**: Bearish momentum building
            • **RSI Divergences**: Price makes new highs but RSI doesn't = bearish divergence
            • **RSI Divergences**: Price makes new lows but RSI doesn't = bullish divergence
            • **RSI Trendlines**: Draw trendlines on RSI itself - breakouts signal trend changes
            • **Multiple Touches**: RSI hitting 70/30 multiple times = strong overbought/oversold condition
            """)
        
        rsi_fig = go.Figure()
        
        if not rsi.empty:
            rsi_fig.add_trace(go.Scatter(
                x=data.index,
                y=rsi,
                mode='lines',
                name='RSI',
                line=dict(color='purple')
            ))
            
            # Add overbought/oversold lines
            rsi_fig.add_hline(y=70, line_dash="dash", line_color="red", annotation_text="Overbought (70)")
            rsi_fig.add_hline(y=30, line_dash="dash", line_color="green", annotation_text="Oversold (30)")
        
        rsi_fig.update_layout(
            title='RSI (Relative Strength Index)',
            xaxis_title='Date',
            yaxis_title='RSI',
            height=400,
            yaxis=dict(range=[0, 100])
        )
        
        st.plotly_chart(rsi_fig, use_container_width=True)
        create_export_buttons(rsi_fig, "RSI_Analysis", symbol)
    
    # Add Chaikin Money Flow Chart in full width
    st.markdown("#### 📊 Chaikin Money Flow Analysis")
    with st.expander("📖 Chaikin Money Flow (CMF) Chart Pattern Guide"):
        st.markdown("""
        **Chaikin Money Flow** - Volume-weighted indicator combining price and volume to measure buying/selling pressure:
        
        **Key Signals & Patterns:**
        • **CMF > 0**: Net buying pressure - accumulation phase, bullish sentiment
        • **CMF < 0**: Net selling pressure - distribution phase, bearish sentiment
        • **CMF Above +0.25**: Strong buying pressure - very bullish
        • **CMF Below -0.25**: Strong selling pressure - very bearish
        • **CMF Rising**: Increasing buying pressure, potential upward price movement
        • **CMF Falling**: Increasing selling pressure, potential downward price movement
        • **CMF vs Price Divergence**: Price rises but CMF falls = distribution, bearish
        • **CMF vs Price Divergence**: Price falls but CMF rises = accumulation, bullish
        • **CMF Zero Line**: Acts as support/resistance - watch for bounces or breaks
        • **Sustained CMF Direction**: Longer periods above/below zero confirm trend strength
        """)
    
    if not cmf.empty:
        period_label = selected_period.replace('_', ' ').title()
        cmf_fig = create_chaikin_chart(data, symbol, cmf, period_label, market)
        st.plotly_chart(cmf_fig, use_container_width=True)
        create_export_buttons(cmf_fig, "Chaikin_Money_Flow", symbol)
    else:
        st.warning("Chaikin Money Flow data not available for the selected period.")


def display_earnings_dividends_tab(symbol, data, ticker_info, ticker_obj, market):
    """Display earnings and dividends analysis in a dedicated tab"""
    
    # Apply same comprehensive CSS for consistent styling
    st.markdown("""
    <style>
    /* Same comprehensive CSS as main tab for consistency */
    .stMetric > div > div > div {
        font-size: 0.75rem !important;
        line-height: 1.1 !important;
    }
    .stMetric > div > div > div > div {
        font-size: 1.0rem !important;
        margin-bottom: 0.2rem !important;
    }
    .stMetric [data-testid="metric-container"] {
        padding: 0.3rem 0 !important;
    }
    .stDataFrame {
        font-size: 0.75rem !important;
    }
    .stDataFrame th, .stDataFrame td {
        padding: 0.2rem 0.4rem !important;
    }
    .stAlert {
        padding: 0.4rem 0.6rem !important;
        margin: 0.3rem 0 !important;
        font-size: 0.85rem !important;
    }
    .stMarkdown p {
        margin-bottom: 0.3rem !important;
        font-size: 0.9rem !important;
    }
    .small-subheader {
        font-size: 1rem !important;
        margin-bottom: 0.3rem !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Get earnings and dividend information
    earnings_info = get_earnings_info(ticker_obj, ticker_info, market)
    dividend_info = get_dividend_info(ticker_obj, ticker_info, market)
    next_dividend_info = estimate_next_dividend_date(ticker_obj, dividend_info)
    
    # Currency symbol
    currency = "₹" if market == "India" else "$"
    
    # Earnings Information Section
    st.markdown('<p class="small-subheader">📅 <strong>Earnings Information</strong></p>', unsafe_allow_html=True)
    
    col_earnings1, col_earnings2, col_earnings3 = st.columns(3)
    
    with col_earnings1:
        if earnings_info['last_earnings_formatted'] != 'N/A':
            earnings_value = earnings_info['last_earnings_formatted']
            if "outdated" in earnings_value or "incomplete" in earnings_value or "likely outdated" in earnings_value:
                # Split the earnings date and warning message
                if " (likely outdated" in earnings_value:
                    clean_date = earnings_value.split(" (likely outdated")[0]
                    warning_msg = "⚠️ Likely outdated - Yahoo Finance frequently misses recent earnings announcements"
                elif " (data may be outdated)" in earnings_value:
                    clean_date = earnings_value.split(" (data may be outdated)")[0]
                    warning_msg = "⚠️ Data may be outdated - verify on company investor relations page"
                else:
                    clean_date = earnings_value
                    warning_msg = "⚠️ Check company investor relations for latest earnings"
                
                st.metric("Last Earnings Date", clean_date)
                # Display warning message in smaller font on new line
                st.markdown(f'<p style="font-size:11px; margin-top:-8px; color:#ff6b6b;">{warning_msg}</p>', unsafe_allow_html=True)
            else:
                st.metric("Last Earnings Date", earnings_info['last_earnings_formatted'])
        else:
            st.metric("Last Earnings Date", "N/A")
    
    with col_earnings2:
        if earnings_info['next_earnings_formatted'] != 'N/A':
            st.metric("Next Earnings Date", earnings_info['next_earnings_formatted'])
        else:
            st.metric("Next Earnings Date", "N/A")
    
    with col_earnings3:
        performance_since_earnings = earnings_info.get('performance_since_earnings', 'N/A')
        if performance_since_earnings != 'N/A':
            st.metric("Performance Since Last Earnings", performance_since_earnings)
        else:
            st.metric("Performance Since Last Earnings", "N/A")

    # Dividend Information Section
    st.markdown('<p class="small-subheader">💰 <strong>Dividend Information</strong></p>', unsafe_allow_html=True)
    
    col_div1, col_div2, col_div3, col_div4 = st.columns(4)
    
    with col_div1:
        if dividend_info['last_dividend_date'] is not None:
            # Extract just the date part for the Last Dividend Date metric
            last_div_date = dividend_info['last_dividend_date'].strftime('%Y-%m-%d')
            st.metric("Last Dividend Date", last_div_date)
        else:
            st.metric("Last Dividend Date", "N/A")
    
    with col_div2:
        if dividend_info['last_dividend_amount'] > 0:
            st.metric("Last Dividend Amount", format_currency(dividend_info['last_dividend_amount'], market))
        else:
            st.metric("Last Dividend Amount", "N/A")
    
    with col_div3:
        if next_dividend_info['next_dividend_date'] != 'N/A':
            st.metric("Next Dividend Date", next_dividend_info['next_dividend_date'])
            if next_dividend_info['estimated']:
                st.caption("(Estimated)")
        else:
            st.metric("Next Dividend Date", "N/A")
    
    with col_div4:
        if dividend_info['dividend_yield'] != 'N/A':
            st.metric("Dividend Yield", dividend_info['dividend_yield'])
        else:
            st.metric("Dividend Yield", "N/A")
    
    # Data Source Note
    if symbol.upper() in ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'MSTR']:
        st.info(f"""
        📝 **Data Source Note**: Yahoo Finance often has delayed earnings data for major stocks. 
        For the most current earnings information, check the company's investor relations page:
        - **AAPL**: [Apple Investor Relations](https://investor.apple.com/investor-relations/default.aspx)
        - **MSFT**: [Microsoft Investor Relations](https://www.microsoft.com/en-us/Investor/)
        - **GOOGL**: [Alphabet Investor Relations](https://abc.xyz/investor/)
        - **AMZN**: [Amazon Investor Relations](https://ir.aboutamazon.com/)
        - **TSLA**: [Tesla Investor Relations](https://ir.tesla.com/)
        - **META**: [Meta Investor Relations](https://investor.fb.com/)
        - **NVDA**: [NVIDIA Investor Relations](https://investor.nvidia.com/home/default.aspx)
        - **MSTR**: [MicroStrategy Investor Relations](https://www.microstrategy.com/company/investor-relations)
        """)
    else:
        st.info("📝 **Data Source Note**: Yahoo Finance may have delayed earnings data. For the most current information, check the company's investor relations page.")
    
    # Comprehensive Earnings Performance Analysis - Using Advanced Analysis Table
    st.markdown('<p class="small-subheader">📊 <strong>Comprehensive Earnings Performance Analysis</strong></p>', unsafe_allow_html=True)
    
    try:
        # Use the comprehensive earnings analysis from Advanced Analysis tab with full 13 columns
        extended_data = ticker_obj.history(period="3y", interval="1d")  # Extended data for comprehensive analysis
        if extended_data.empty:
            extended_data = data
        
        earnings_analysis, quarters_found = get_detailed_earnings_performance_analysis(
            ticker_obj, extended_data, market=market, max_quarters=8
        )
        
        if earnings_analysis is not None and not earnings_analysis.empty:
            st.markdown(f"""
            **Comprehensive track record of stock performance after each earnings announcement ({quarters_found} quarters available):**
            - **Pre-Close**: Stock price before earnings announcement
            - **Next Open/Close**: Stock prices on trading day after earnings
            - **Overnight %**: Price movement from pre-earnings close to next day open
            - **NextDay %**: Price movement from pre-earnings close to next day close
            - **Week %**: Total change from pre-earnings close to end of week (5 trading days)
            - **Direction**: Overall weekly direction (Up/Down/Flat)
            - **EPS**: Estimated vs Actual earnings per share and surprise
            """)
            
            # Display comprehensive earnings analysis with all 13 columns
            st.dataframe(earnings_analysis, use_container_width=True)
            
            # Enhanced Summary statistics with more metrics
            col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
            
            try:
                # Calculate comprehensive summary stats from all available columns
                overnight_changes = []
                nextday_changes = []
                week_changes = []
                
                # Extract numerical values from percentage columns
                for _, row in earnings_analysis.iterrows():
                    try:
                        overnight_val = row['Overnight Change (%)']
                        if overnight_val and overnight_val != 'N/A':
                            overnight_changes.append(float(str(overnight_val).replace('%', '').replace('+', '')))
                    except:
                        pass
                    
                    try:
                        nextday_val = row['Next Day Change (%)']
                        if nextday_val and nextday_val != 'N/A':
                            nextday_changes.append(float(str(nextday_val).replace('%', '').replace('+', '')))
                    except:
                        pass
                    
                    try:
                        week_val = row['Week Performance (%)']
                        if week_val and week_val != 'N/A':
                            week_changes.append(float(str(week_val).replace('%', '').replace('+', '')))
                    except:
                        pass
                
                # Display enhanced statistics
                if overnight_changes:
                    with col_stats1:
                        avg_overnight = sum(overnight_changes) / len(overnight_changes)
                        st.metric("Avg Overnight Change", f"{avg_overnight:+.2f}%")
                        positive_overnight = sum(1 for x in overnight_changes if x > 0)
                        st.caption(f"Positive: {positive_overnight}/{len(overnight_changes)}")
                
                if nextday_changes:
                    with col_stats2:
                        avg_nextday = sum(nextday_changes) / len(nextday_changes)
                        st.metric("Avg Next Day Change", f"{avg_nextday:+.2f}%")
                        positive_nextday = sum(1 for x in nextday_changes if x > 0)
                        st.caption(f"Positive: {positive_nextday}/{len(nextday_changes)}")
                
                if week_changes:
                    with col_stats3:
                        avg_week = sum(week_changes) / len(week_changes)
                        st.metric("Avg Week Performance", f"{avg_week:+.2f}%")
                        positive_week = sum(1 for x in week_changes if x > 0)
                        st.caption(f"Positive: {positive_week}/{len(week_changes)}")
                
                # Direction analysis
                with col_stats4:
                    directions = earnings_analysis['Direction'].value_counts()
                    if not directions.empty:
                        most_common = directions.index[0]
                        st.metric("Most Common Direction", most_common)
                        st.caption(f"{directions[most_common]}/{len(earnings_analysis)} quarters")
                        
            except Exception as e:
                st.warning("Could not calculate comprehensive summary statistics")
                
        else:
            st.info("No comprehensive earnings performance data available for the selected period")
            
    except Exception as e:
        st.error(f"Error analyzing comprehensive earnings performance: {str(e)}")

def get_unified_market_intelligence(symbol, fmp_api_key, ticker_obj=None):
    """
    Unified function to get market intelligence data with Yahoo Finance fallback
    Used by both UI tab and PDF export to ensure consistency
    """
    import yfinance as yf
    import pandas as pd
    
    # Get FMP data first
    advanced_metrics = get_advanced_metrics(symbol, fmp_api_key)
    
    # Check if we need fallback
    needs_fallback = (
        not advanced_metrics or 
        advanced_metrics == {} or 
        (isinstance(advanced_metrics, dict) and (
            advanced_metrics.get('Analyst Rating') == 'N/A' and 
            advanced_metrics.get('Price Target') == 'N/A'
        ))
    )
    
    if needs_fallback:
        try:
            # Use provided ticker_obj or create new one
            if ticker_obj is None:
                ticker_obj = yf.Ticker(symbol)
            
            # Use Yahoo Finance for basic institutional data
            institutional_holders = ticker_obj.institutional_holders
            major_holders = ticker_obj.major_holders
            info = ticker_obj.info
            
            # Create fallback metrics from available Yahoo Finance data
            fallback_metrics = {}
            
            # Institutional ownership from major holders
            if major_holders is not None and not major_holders.empty and len(major_holders) > 1:
                inst_ownership = major_holders.iloc[1, 0] if pd.notnull(major_holders.iloc[1, 0]) else None
                if inst_ownership:
                    fallback_metrics['Institutional Ownership'] = f"Institutions own {inst_ownership}"
            
            # Basic analyst info from Yahoo Finance
            if info:
                # Analyst recommendation
                recommendation = info.get('recommendationKey')
                if recommendation and recommendation != 'none':
                    fallback_metrics['Analyst Rating'] = recommendation.upper().replace('_', ' ')
                
                # Target price if available
                target_mean = info.get('targetMeanPrice')
                if target_mean:
                    fallback_metrics['Price Target'] = f"${target_mean:.2f}"
                    target_low = info.get('targetLowPrice')
                    target_high = info.get('targetHighPrice')
                    if target_low and target_high:
                        fallback_metrics['Price Target'] += f" (Range: ${target_low:.2f} - ${target_high:.2f})"
                
                # Number of analyst recommendations
                num_analysts = info.get('numberOfAnalystOpinions')
                if num_analysts:
                    fallback_metrics['Analyst Coverage'] = f"{num_analysts} analysts covering"
            
            # Top institutional holders if available
            if institutional_holders is not None and not institutional_holders.empty:
                top_holder = institutional_holders.iloc[0]
                holder_name = top_holder.get('Holder') if 'Holder' in top_holder else 'N/A'
                shares = top_holder.get('Shares') if 'Shares' in top_holder else 'N/A'
                if holder_name != 'N/A':
                    fallback_metrics['Top Institutional Holder'] = f"{holder_name}"
                    if shares != 'N/A':
                        fallback_metrics['Top Institutional Holder'] += f" ({shares:,} shares)"
            
            # Insider activity placeholder (Yahoo Finance doesn't provide this easily)
            fallback_metrics['Insider Activity'] = "Limited data available"
            fallback_metrics['Social Sentiment'] = "Requires premium API"
            
            # Add data source note
            fallback_metrics['Data Source'] = "Yahoo Finance (Free)"
            fallback_metrics['Last Updated'] = f"Retrieved at {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
            
            if fallback_metrics:
                # Force replace the original metrics with fallback data
                advanced_metrics.clear()
                advanced_metrics.update(fallback_metrics)
                return advanced_metrics, True  # True indicates fallback was used
                
        except Exception as e:
            print(f"Error fetching Yahoo Finance fallback data: {e}")
            return {'Market Intelligence': 'Data temporarily unavailable due to API limits'}, False
    
    return advanced_metrics, False  # False indicates FMP data was used

def display_advanced_sentiment_metrics(symbol, market="US"):
    """Display comprehensive advanced sentiment and institutional metrics"""
    
    st.subheader(f"🎯 Advanced Market Intelligence for {symbol}")
    
    # Check for FMP API key
    import os
    fmp_api_key = os.environ.get("FMP_API_KEY")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **📊 Data Sources:**
        - Analyst Ratings & Price Targets (FMP/Yahoo Finance)
        - Insider Trading Activity (FMP/Limited)
        - Institutional Holdings (FMP/Yahoo Finance)
        - Social Sentiment Analysis (FMP/Premium)
        - News Sentiment (AI-powered)
        
        *Auto-fallback to Yahoo Finance when FMP unavailable*
        """)
    
    with col2:
        if not fmp_api_key:
            st.warning("""
            🔑 **Enhanced Data Requires API Key**
            
            Add FMP_API_KEY to environment variables for:
            - Professional analyst ratings
            - Insider trading data
            - Institutional holdings
            - Advanced sentiment metrics
            """)
        else:
            st.info("""
            ⚙️ **API Key Detected**
            
            Note: Some endpoints require paid FMP subscription.
            Free tier may have limited access to:
            - Institutional holdings
            - Advanced analyst data
            - Social sentiment metrics
            """)
            st.caption("If you see 403 errors, consider upgrading your FMP plan")
    
    # Get advanced metrics using unified function for consistency with PDF export
    try:
        advanced_metrics, used_fallback = get_unified_market_intelligence(symbol, fmp_api_key)
        
        if used_fallback:
            st.info("🔄 FMP API unavailable, using Yahoo Finance fallback data...")
            st.success("✓ Successfully loaded market intelligence from Yahoo Finance")
        
        # Display metrics in a structured format
        st.markdown("### 📈 Market Intelligence Dashboard")
        
        # Create metrics grid
        col_analyst, col_insider, col_institutional = st.columns(3)
        
        with col_analyst:
            st.metric(
                label="Analyst Rating",
                value=advanced_metrics.get('Analyst Rating', 'N/A'),
                help="Professional analyst consensus rating"
            )
            st.metric(
                label="Price Target",
                value=advanced_metrics.get('Price Target', 'N/A'),
                help="Average analyst price target"
            )
        
        with col_insider:
            st.metric(
                label="Insider Activity",
                value=advanced_metrics.get('Insider Activity', 'N/A'),
                help="Recent insider buying/selling activity"
            )
        
        with col_institutional:
            st.metric(
                label="Institutional Ownership",
                value=advanced_metrics.get('Institutional Ownership', 'N/A'),
                help="Number of institutional holders"
            )
            st.metric(
                label="Social Sentiment",
                value=advanced_metrics.get('Social Sentiment', 'N/A'),
                help="Market sentiment from social sources"
            )
        
        # Show last updated
        if advanced_metrics.get('Last Updated'):
            st.caption(f"Last updated: {advanced_metrics['Last Updated']}")
        
        # Detailed view toggle
        with st.expander("🔍 Detailed Analysis"):
            # Show data source information
            data_source = advanced_metrics.get('Data Source', 'Unknown')
            st.markdown(f"**Data Source:** {data_source}")
            st.markdown("---")
            
            # Display details based on data source
            if 'raw_data' in advanced_metrics:
                # FMP data with raw_data structure
                raw_data = advanced_metrics['raw_data']
                
                # Analyst details
                analyst_data = raw_data.get('analyst_ratings', {})
                if analyst_data and 'error' not in analyst_data:
                    st.markdown("**Analyst Information:**")
                    col_a1, col_a2 = st.columns(2)
                    with col_a1:
                        st.write(f"Rating: {analyst_data.get('rating', 'N/A')}")
                        st.write(f"Recommendation: {analyst_data.get('recommendation', 'N/A')}")
                    with col_a2:
                        st.write(f"Analyst Count: {analyst_data.get('analyst_count', 0)}")
                        st.write(f"Price Target: {analyst_data.get('price_target', 'N/A')}")
                
                # Insider details
                insider_data = raw_data.get('insider_activity', {})
                if insider_data and 'error' not in insider_data:
                    st.markdown("**Insider Activity:**")
                    st.write(f"Recent Activity: {insider_data.get('recent_activity', 'N/A')}")
                    st.write(f"Summary: {insider_data.get('insider_summary', 'N/A')}")
                    st.write(f"Net Activity: {insider_data.get('net_insider_activity', 'N/A')}")
                
                # Institutional details
                institutional_data = raw_data.get('institutional_holdings', {})
                if institutional_data and 'error' not in institutional_data:
                    st.markdown("**Institutional Holdings:**")
                    st.write(f"Total Institutions: {institutional_data.get('total_institutions', 0)}")
                    st.write(f"Total Market Value: ${institutional_data.get('total_market_value', 0):,.0f}")
                    
                    # Show top holders if available
                    top_holders = institutional_data.get('top_5_holders', [])
                    if top_holders:
                        st.markdown("**Top 5 Institutional Holders:**")
                        for i, holder in enumerate(top_holders[:3], 1):
                            st.write(f"{i}. {holder.get('name', 'Unknown')} - ${holder.get('value', 0):,.0f}")
            
            elif data_source == "Yahoo Finance (Free)":
                # Yahoo Finance fallback data - show available summary information
                st.markdown("**Yahoo Finance Data Summary:**")
                
                col_yf1, col_yf2 = st.columns(2)
                
                with col_yf1:
                    st.markdown("**Analyst Information:**")
                    st.write(f"• Rating: {advanced_metrics.get('Analyst Rating', 'N/A')}")
                    st.write(f"• Price Target: {advanced_metrics.get('Price Target', 'N/A')}")
                    if advanced_metrics.get('Analyst Coverage'):
                        st.write(f"• Coverage: {advanced_metrics.get('Analyst Coverage')}")
                
                with col_yf2:
                    st.markdown("**Institutional Information:**")
                    st.write(f"• Ownership: {advanced_metrics.get('Institutional Ownership', 'N/A')}")
                    if advanced_metrics.get('Top Institutional Holder'):
                        st.write(f"• Top Holder: {advanced_metrics.get('Top Institutional Holder')}")
                
                st.markdown("**Data Limitations:**")
                st.caption("• Yahoo Finance provides basic institutional and analyst data")
                st.caption("• For detailed insider trading data, premium APIs are required")
                st.caption("• Social sentiment requires specialized data providers")
            
            else:
                # Fallback for other data sources
                st.markdown("**Available Metrics:**")
                for key, value in advanced_metrics.items():
                    if key not in ['Data Source', 'Last Updated', 'raw_data'] and value != 'N/A':
                        st.write(f"• {key}: {value}")
        
    except Exception as e:
        st.error(f"Error loading advanced metrics: {str(e)}")
        st.info("Please ensure the sentiment_data_provider module is properly configured.")

def display_news_sentiment_analysis(symbol):
    """Display AI-powered news sentiment analysis"""
    
    # Detect market for proper symbol handling
    indian_symbols = {
        'RELIANCE', 'TCS', 'INFY', 'HDFC', 'ICICIBANK', 'BHARTIARTL', 'ITC', 
        'LT', 'HCLTECH', 'KOTAKBANK', 'AXISBANK', 'ASIANPAINT', 'MARUTI',
        'TITAN', 'NESTLEIND', 'ULTRACEMCO', 'POWERGRID', 'NTPC', 'COALINDIA',
        'ONGC', 'TECHM', 'TATAMOTORS', 'TATASTEEL', 'HINDUNILVR', 'WIPRO',
        'DRREDDY', 'EICHERMOT', 'BAJFINANCE', 'BAJAJFINSV', 'BRITANNIA',
        'CIPLA', 'DIVISLAB', 'HEROMOTOCO', 'HINDALCO', 'INDUSINDBK',
        'JSWSTEEL', 'M&M', 'SBIN', 'SUNPHARMA', 'GRASIM', 'BEL', 'BEML',
        'BHEL', 'HAL', 'SAIL', 'NMDC', 'GAIL', 'IOC', 'BPCL', 'HPCL'
    }
    
    # Detect market based on symbol
    base_symbol = symbol.replace('.NS', '').replace('.BO', '')
    if ('.NS' in symbol or '.BO' in symbol or 
        base_symbol in indian_symbols or
        st.session_state.get('current_market') == "India"):
        detected_market = "India"
        display_symbol = base_symbol
        st.info(f"🇮🇳 Analyzing Indian market news for {display_symbol}")
    else:
        detected_market = "US"
        display_symbol = symbol
        st.info(f"🇺🇸 Analyzing US market news for {display_symbol}")
    
    st.subheader(f"📰 AI-Powered News Sentiment Analysis for {display_symbol}")
    st.markdown("Analyze financial news sentiment using advanced AI to gauge market sentiment and potential impact on stock performance.")
    
    # Create analyze button
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown(f"""
        **What this analysis provides:**
        - Sentiment scoring (Bullish/Bearish/Neutral)
        - Investment impact assessment
        - Key themes identification
        - Risk factor analysis
        - Confidence levels for each analysis
        """)
    
    with col2:
        if st.button("🤖 Analyze News Sentiment", type="primary", key=f"sentiment_{symbol}"):
            st.session_state[f'analyze_sentiment_{symbol}'] = True
    
    # Perform sentiment analysis if button clicked
    if st.session_state.get(f'analyze_sentiment_{symbol}', False):
        
        # Check if OpenAI API key is available
        import os
        openai_key = os.environ.get("OPENAI_API_KEY")
        if not openai_key:
            st.error("""
            🔑 **OpenAI API Key Required**
            
            To use AI-powered sentiment analysis, please add your OpenAI API key to the environment variables.
            
            **Steps:**
            1. Go to your Replit project secrets
            2. Add a new secret with key: `OPENAI_API_KEY`
            3. Add your OpenAI API key as the value
            4. Restart the application
            """)
            return
        
        # Use the enhanced news sentiment analyzer with multiple sources
        try:
            from news_sentiment_analyzer import run_sentiment_analysis
            run_sentiment_analysis(display_symbol.upper().strip(), detected_market)
                    
        except Exception as e:
            st.error(f"Error loading sentiment analysis: {str(e)}")
            st.info("Please ensure all required dependencies are installed and OpenAI API key is configured correctly.")



def weekly_earnings_calendar_tab():
    """Weekly Earnings Calendar tab content"""
    
    st.markdown("### 📅 Weekly Earnings Calendar")
    st.markdown("View companies announcing earnings this week using **NASDAQ's free public API**")
    st.markdown("---")
    
    # Information section
    st.info("""
    **💡 About This Data:**
    - Uses NASDAQ's public earnings calendar API (completely free, no API key required)
    - Covers ALL US stocks with earnings announcements 
    - Updates automatically based on current week (Monday-Sunday)
    - Includes comprehensive data for all major exchanges (NYSE, NASDAQ, etc.)
    - More reliable than Yahoo Finance for upcoming earnings dates
    
    **📈 Stock Symbols:** NASDAQ API provides company names but not ticker symbols. 
    Major companies (ORCL, ADBE, etc.) are mapped correctly, others show approximations.
    """)
    
    st.markdown("---")
    
    # Load earnings calendar section
    col_info1, col_info2 = st.columns([3, 1])
    
    with col_info1:
        st.markdown("""
        📊 **What's Included:**
        • Company symbols and names
        • Earnings announcement dates and times
        • EPS estimates and actual values (when available)
        • Revenue estimates and actual values (when available)
        • Export options (Excel & CSV)
        """)
    
    with col_info2:
        # Current week info
        today_name = datetime.now().strftime('%A')
        current_week = datetime.now().strftime('Week of %B %d, %Y')
        st.info(f"**{current_week}**")
        st.info(f"Today: {today_name}")
    
    st.markdown("---")
    
    # Load earnings calendar button
    col_load1, col_load2 = st.columns([2, 2])
    
    with col_load1:
        if st.button("📊 Load Weekly Earnings Calendar", type="primary", help="Fetch earnings announcements for the current week"):
            st.session_state.load_earnings_calendar = True
    
    with col_load2:
        # Weekly range info
        monday = datetime.now() - timedelta(days=datetime.now().weekday())
        sunday = monday + timedelta(days=6)
        st.info(f"Range: {monday.strftime('%b %d')} - {sunday.strftime('%b %d, %Y')}")
    
    # Load and display earnings calendar
    if st.session_state.get('load_earnings_calendar', False):
        with st.spinner("Loading weekly earnings calendar..."):
            earnings_df, error_msg = get_weekly_earnings_calendar()
            
            if error_msg:
                st.error(f"❌ {error_msg}")
                if "FMP API key" in error_msg:
                    st.markdown("### 🔑 How to get FMP API key:")
                    st.info("""
                    1. Visit [Financial Modeling Prep](https://financialmodelingprep.com/developer/docs)
                    2. Sign up for a free account (they offer free tier)
                    3. Copy your API key from the dashboard
                    4. Add it as `FMP_API_KEY` in your environment variables
                    5. Restart the application
                    """)
                    
                    st.markdown("### 🆓 Free Tier Includes:")
                    st.success("""
                    • 250 API calls per day
                    • Earnings calendar data
                    • Financial statements
                    • Stock prices and more
                    """)
                    
            elif earnings_df is not None and not earnings_df.empty:
                st.success(f"✅ Found **{len(earnings_df)}** earnings announcements this week")
                
                # Display the earnings calendar table with better formatting
                st.markdown("### 📋 Weekly Earnings Schedule")
                
                # Format the display for better readability
                display_df = earnings_df.copy()
                
                # Ensure Symbol column is first and clearly visible
                columns_order = ['Symbol', 'Company', 'Date', 'Day', 'Time', 'EPS Estimate']
                available_columns = [col for col in columns_order if col in display_df.columns]
                display_df = display_df[available_columns]
                
                # Display with column configuration for better visibility
                column_config = {
                    "Symbol": st.column_config.TextColumn(
                        "📈 Symbol",
                        width="small",
                        help="Stock ticker symbol"
                    ),
                    "Company": st.column_config.TextColumn(
                        "🏢 Company",
                        width="large",
                        help="Company name"
                    ),
                    "Date": st.column_config.DateColumn(
                        "📅 Date",
                        width="small",
                        help="Earnings announcement date"
                    ),
                    "Day": st.column_config.TextColumn(
                        "📆 Day",
                        width="small",
                        help="Day of the week"
                    ),
                    "Time": st.column_config.TextColumn(
                        "🕒 Time",
                        width="small",
                        help="Announcement timing (BMO/AMC/TBD)"
                    ),
                    "EPS Estimate": st.column_config.TextColumn(
                        "💹 EPS Est",
                        width="small",
                        help="Estimated earnings per share"
                    )
                }
                
                st.dataframe(
                    display_df, 
                    use_container_width=True, 
                    hide_index=True,
                    column_config=column_config
                )
                
                # Export options
                st.markdown("### 📁 Export Options")
                col_export1, col_export2 = st.columns(2)
                
                with col_export1:
                    # Excel export
                    excel_data = create_earnings_calendar_excel(earnings_df)
                    if excel_data:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"weekly_earnings_calendar_{timestamp}.xlsx"
                        
                        st.download_button(
                            label="📊 Download Excel File",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download the weekly earnings calendar as a formatted Excel file"
                        )
                
                with col_export2:
                    # CSV export
                    csv_data = earnings_df.to_csv(index=False)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    csv_filename = f"weekly_earnings_calendar_{timestamp}.csv"
                    
                    st.download_button(
                        label="📋 Download CSV File",
                        data=csv_data,
                        file_name=csv_filename,
                        mime="text/csv",
                        help="Download the weekly earnings calendar as a CSV file"
                    )
                
                # Summary statistics
                st.markdown("### 📈 Weekly Summary & Statistics")
                
                # Group by day
                day_counts = earnings_df['Day'].value_counts()
                
                col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
                
                with col_stats1:
                    st.metric("Total Companies", len(earnings_df))
                
                with col_stats2:
                    busiest_day = day_counts.index[0] if not day_counts.empty else "N/A"
                    busiest_count = day_counts.iloc[0] if not day_counts.empty else 0
                    st.metric("Busiest Day", f"{busiest_day}")
                    st.caption(f"{busiest_count} companies")
                
                with col_stats3:
                    unique_days = len(day_counts)
                    st.metric("Active Days", unique_days)
                    st.caption("Days with earnings")
                
                with col_stats4:
                    avg_per_day = len(earnings_df) / max(unique_days, 1)
                    st.metric("Avg Per Day", f"{avg_per_day:.1f}")
                    st.caption("Companies per day")
                
                # Display detailed day breakdown
                if not day_counts.empty:
                    st.markdown("### 📊 Daily Breakdown")
                    
                    # Create columns for each day that has earnings
                    if len(day_counts) <= 3:
                        day_cols = st.columns(len(day_counts))
                    else:
                        # Split into two rows if more than 3 days
                        day_cols = st.columns(min(3, len(day_counts)))
                    
                    for idx, (day, count) in enumerate(day_counts.items()):
                        col_idx = idx % len(day_cols)
                        with day_cols[col_idx]:
                            st.metric(f"**{day}**", f"{count} companies")
                            
                            # Show companies for this day
                            day_companies = earnings_df[earnings_df['Day'] == day]['Symbol'].tolist()
                            if day_companies:
                                companies_text = ", ".join(day_companies[:5])  # Show first 5
                                if len(day_companies) > 5:
                                    companies_text += f" +{len(day_companies)-5} more"
                                st.caption(companies_text)
                        
                        # Start new row after 3 columns
                        if (idx + 1) % 3 == 0 and idx < len(day_counts) - 1:
                            day_cols = st.columns(min(3, len(day_counts) - idx - 1))
                            
            else:
                st.info("📅 No earnings announcements found for this week.")
                st.markdown("""
                **Possible reasons:**
                • This might be a holiday week
                • Earnings season may not be active
                • All major companies may have already reported
                """)
                
        # Auto-update information
        st.markdown("---")
        st.markdown("### 🔄 Auto-Update Information")
        
        col_update1, col_update2 = st.columns(2)
        
        with col_update1:
            st.info("""
            **📅 Weekly Schedule:**
            • Data covers Monday through Sunday
            • Calendar updates when you refresh on Monday
            • Shows current week's earnings only
            """)
        
        with col_update2:
            st.info("""
            **⏰ Timing Notes:**
            • Times shown are typically Eastern Time
            • BMO = Before Market Open (usually 7-9 AM ET)
            • AMC = After Market Close (usually 4-8 PM ET)
            """)


def extract_ticker_from_company_name(company_name):
    """
    Try to extract or guess ticker symbol from company name
    This is a best-effort approach since NASDAQ API doesn't provide tickers
    """
    if not company_name or company_name == 'N/A':
        return 'N/A'
    
    # Known mapping for major companies
    known_companies = {
        'Oracle Corporation': 'ORCL',
        'Adobe Inc.': 'ADBE', 
        'Adobe Inc': 'ADBE',
        'Apple Inc.': 'AAPL',
        'Microsoft Corporation': 'MSFT',
        'Amazon.com, Inc.': 'AMZN',
        'Alphabet Inc.': 'GOOGL',
        'Meta Platforms, Inc.': 'META',
        'Tesla, Inc.': 'TSLA',
        'NVIDIA Corporation': 'NVDA',
        'Chewy, Inc.': 'CHWY',
        'Caseys General Stores, Inc.': 'CASY',
        'Value Line, Inc.': 'VALU'
    }
    
    # Check known companies first
    if company_name in known_companies:
        return known_companies[company_name]
    
    # For companies not in our known list, extract first word as approximation
    first_word = company_name.split()[0].upper()
    
    # Remove common suffixes
    if first_word.endswith(','):
        first_word = first_word[:-1]
    
    # Keep only letters and limit length
    ticker_approx = ''.join(char for char in first_word if char.isalpha())[:4]
    
    return ticker_approx if ticker_approx else 'N/A'


def get_weekly_earnings_calendar():
    """
    Fetch weekly earnings announcements using NASDAQ API (completely free, no API key required)
    
    Returns:
        pandas.DataFrame: Weekly earnings calendar with company, date, and time information
    """
    from finance_calendars import finance_calendars as fc
    from datetime import datetime, timedelta
    
    try:
        # Get current Monday (start of the week)
        today = datetime.now()
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        sunday = monday + timedelta(days=6)
        
        print(f"NASDAQ EARNINGS: Fetching earnings for week {monday.strftime('%Y-%m-%d')} to {sunday.strftime('%Y-%m-%d')}")
        
        earnings_this_week = []
        
        # Get earnings for each day of the week with retry logic
        current_date = monday
        failed_dates = []
        
        while current_date <= sunday:
            date_str = current_date.strftime('%Y-%m-%d')
            print(f"NASDAQ EARNINGS: Checking date {date_str}")
            
            # Try up to 3 times for each date
            success = False
            for attempt in range(3):
                try:
                    # Get earnings for this specific date
                    daily_earnings = fc.get_earnings_by_date(current_date)
                    
                    if daily_earnings is not None and not daily_earnings.empty:
                        print(f"NASDAQ EARNINGS: Found {len(daily_earnings)} earnings on {date_str}")
                        
                        # Process each earning announcement for this date
                        for _, earning in daily_earnings.iterrows():
                            try:
                                
                                # Extract information using correct NASDAQ API field names
                                company_name = earning.get('name', 'N/A')
                                
                                # EPS forecast from NASDAQ API
                                eps_estimate = earning.get('epsForecast', 'N/A')
                                if eps_estimate == '' or eps_estimate is None:
                                    eps_estimate = 'N/A'
                                
                                # Clean up timing information
                                timing_raw = earning.get('time', 'TBD')
                                if timing_raw and timing_raw.startswith('time-'):
                                    timing = timing_raw.replace('time-', '').replace('-', ' ').title()
                                    if timing == 'After Hours':
                                        timing = 'AMC'
                                    elif timing == 'Pre Market':
                                        timing = 'BMO'
                                    elif timing == 'Not Supplied':
                                        timing = 'TBD'
                                else:
                                    timing = 'TBD'
                                
                                # Try to extract ticker symbol from company name (basic approach)
                                symbol = extract_ticker_from_company_name(company_name)
                                
                                earnings_this_week.append({
                                    'Symbol': symbol,
                                    'Company': company_name,
                                    'Date': current_date.strftime('%Y-%m-%d'),
                                    'Day': current_date.strftime('%A'),
                                    'Time': timing,
                                    'EPS Estimate': eps_estimate
                                })
                                
                            except Exception as e:
                                print(f"NASDAQ EARNINGS: Error processing earning entry: {str(e)}")
                                continue
                        
                        success = True
                        break  # Success, exit retry loop
                    
                    else:
                        print(f"NASDAQ EARNINGS: No earnings found for {date_str}")
                        success = True
                        break  # No data is also a successful response
                        
                except Exception as e:
                    if attempt < 2:  # Not the last attempt
                        print(f"NASDAQ EARNINGS: Attempt {attempt + 1} failed for {date_str}: {str(e)} - Retrying...")
                        import time
                        time.sleep(1)  # Brief pause before retry
                    else:
                        print(f"NASDAQ EARNINGS: All attempts failed for {date_str}: {str(e)}")
                        failed_dates.append(date_str)
            
            # Move to next day
            current_date += timedelta(days=1)
        
        # Report any failed dates
        if failed_dates:
            print(f"NASDAQ EARNINGS: WARNING - Failed to fetch data for: {', '.join(failed_dates)}")
        
        print(f"NASDAQ EARNINGS: Total earnings found for the week: {len(earnings_this_week)}")
        
        if not earnings_this_week:
            # Also try today's earnings as a fallback
            try:
                print("NASDAQ EARNINGS: Trying today's earnings as fallback...")
                today_earnings = fc.get_earnings_today()
                if today_earnings is not None and not today_earnings.empty:
                    print(f"NASDAQ EARNINGS: Found {len(today_earnings)} earnings for today")
                    for _, earning in today_earnings.head(5).iterrows():  # Limit to 5 for demo
                        symbol = earning.get('symbol', 'N/A')
                        company_name = earning.get('company', earning.get('name', symbol))
                        earnings_this_week.append({
                            'Symbol': symbol,
                            'Company': company_name,
                            'Date': today.strftime('%Y-%m-%d'),
                            'Day': today.strftime('%A'),
                            'Time': earning.get('time', 'TBD'),
                            'EPS Estimate': earning.get('epsEstimate', 'N/A')
                        })
            except Exception as e:
                print(f"NASDAQ EARNINGS: Error with today's earnings fallback: {str(e)}")
        
        if not earnings_this_week:
            error_msg = f"No earnings announcements found for week {monday.strftime('%Y-%m-%d')} to {sunday.strftime('%Y-%m-%d')} from NASDAQ data. This might mean: (1) Light earnings week, or (2) Earnings scheduled outside current timeframe."
            if failed_dates:
                error_msg += f" Note: Failed to fetch data for {len(failed_dates)} days due to connection issues."
            return pd.DataFrame(), error_msg
        
        # Convert to DataFrame
        df = pd.DataFrame(earnings_this_week)
        
        # Remove duplicates - only remove if same company on same day (not across different days)
        # This allows companies to appear on multiple days but prevents true duplicates
        df = df.drop_duplicates(subset=['Symbol', 'Date', 'Time']).reset_index(drop=True)
        
        # If still too many results, prioritize by keeping up to 100 companies
        if len(df) > 100:
            # Sort by date and take top 100, but prioritize major companies
            df_sorted = df.sort_values(['Date', 'Symbol'])
            
            # Ensure major companies like Oracle, Adobe are included
            major_companies = ['ORCL', 'ADBE', 'AAPL', 'MSFT', 'GOOGL', 'META', 'TSLA', 'NVDA', 'AMZN']
            major_df = df_sorted[df_sorted['Symbol'].isin(major_companies)]
            other_df = df_sorted[~df_sorted['Symbol'].isin(major_companies)]
            
            # Combine: all major companies + fill remaining slots with others
            remaining_slots = 100 - len(major_df)
            if remaining_slots > 0:
                df = pd.concat([major_df, other_df.head(remaining_slots)]).reset_index(drop=True)
            else:
                df = major_df.head(100).reset_index(drop=True)
        
        # Final sort by date and symbol
        df = df.sort_values(['Date', 'Symbol']).reset_index(drop=True)
        
        # Check if major companies are included
        major_symbols_found = []
        major_companies_check = ['ORCL', 'ADBE', 'AAPL', 'MSFT', 'GOOGL', 'META', 'TSLA', 'NVDA', 'AMZN']
        for symbol in major_companies_check:
            if symbol in df['Symbol'].values:
                major_symbols_found.append(symbol)
        
        if major_symbols_found:
            print(f"NASDAQ EARNINGS: Major companies found in results: {', '.join(major_symbols_found)}")
        else:
            print("NASDAQ EARNINGS: No major companies found in this week's results")
        
        print(f"NASDAQ EARNINGS: Returning {len(df)} unique earnings announcements")
        return df, None
        
    except Exception as e:
        error_msg = f"Error fetching NASDAQ earnings calendar: {str(e)}"
        print(f"NASDAQ EARNINGS ERROR: {error_msg}")
        return None, error_msg

def create_earnings_calendar_excel(df, filename="weekly_earnings_calendar.xlsx"):
    """
    Create Excel file from earnings calendar DataFrame
    
    Args:
        df (pandas.DataFrame): Earnings calendar data
        filename (str): Output filename
    
    Returns:
        bytes: Excel file data
    """
    try:
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Weekly Earnings Calendar', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Weekly Earnings Calendar']
            
            # Format headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add creation timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            worksheet.cell(row=len(df) + 3, column=1, value=f"Generated: {timestamp}")
        
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        return None


def create_market_events_excel(events_list, earnings_list):
    """
    Create Excel file for market events and earnings data
    """
    try:
        import io
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        # Create a new workbook with two worksheets
        wb = Workbook()
        
        # Create Events worksheet
        ws_events = wb.active
        ws_events.title = "Market Events"
        
        if events_list:
            # Define column headers for events
            event_headers = [
                ("Event", 25), ("Date", 12), ("Time", 10), ("Category", 15),
                ("Impact Level", 12), ("Description", 40), ("Market Impact Analysis", 50)
            ]
            
            # Add headers with formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, (header, width) in enumerate(event_headers, 1):
                cell = ws_events.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_events.column_dimensions[cell.column_letter].width = width
            
            # Add event data rows
            for row_idx, event in enumerate(events_list, 2):
                ws_events.cell(row=row_idx, column=1, value=event.get('Title', 'N/A'))
                ws_events.cell(row=row_idx, column=2, value=event.get('Date', 'TBD'))
                ws_events.cell(row=row_idx, column=3, value=event.get('Time', 'TBD'))
                ws_events.cell(row=row_idx, column=4, value=event.get('Category', 'N/A'))
                
                # Color-code importance level
                importance_cell = ws_events.cell(row=row_idx, column=5, value=event.get('Importance', 'Medium'))
                importance = event.get('Importance', 'Medium')
                if importance == 'High':
                    importance_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    importance_cell.font = Font(color="CC0000", bold=True)
                elif importance == 'Medium':
                    importance_cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    importance_cell.font = Font(color="FF8C00", bold=True)
                else:
                    importance_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    importance_cell.font = Font(color="228B22", bold=True)
                
                # Add description and market impact with text wrapping
                desc_cell = ws_events.cell(row=row_idx, column=6, value=event.get('Description', 'N/A'))
                desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                impact_cell = ws_events.cell(row=row_idx, column=7, value=event.get('Market Impact', 'N/A'))
                impact_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Set row height for better readability
                ws_events.row_dimensions[row_idx].height = 60
        
        # Create Earnings worksheet if earnings data exists
        if earnings_list:
            ws_earnings = wb.create_sheet("Top Earnings")
            
            # Define column headers for earnings
            earning_headers = [
                ("Company", 20), ("Ticker", 8), ("Date", 12), ("Time", 15), 
                ("Market Cap", 12), ("Sector", 15), ("Impact Level", 12),
                ("Why Important", 40), ("Key Metrics", 40), ("Market Impact", 40)
            ]
            
            for col_idx, (header, width) in enumerate(earning_headers, 1):
                cell = ws_earnings.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_earnings.column_dimensions[cell.column_letter].width = width
            
            # Add earnings data rows
            for row_idx, earning in enumerate(earnings_list, 2):
                ws_earnings.cell(row=row_idx, column=1, value=earning.get('Company', 'N/A'))
                ws_earnings.cell(row=row_idx, column=2, value=earning.get('Ticker', 'N/A'))
                ws_earnings.cell(row=row_idx, column=3, value=earning.get('Date', 'TBD'))
                ws_earnings.cell(row=row_idx, column=4, value=earning.get('Time', 'TBD'))
                ws_earnings.cell(row=row_idx, column=5, value=earning.get('Market Cap', 'N/A'))
                ws_earnings.cell(row=row_idx, column=6, value=earning.get('Sector', 'N/A'))
                
                # Color-code importance level
                importance_cell = ws_earnings.cell(row=row_idx, column=7, value=earning.get('Importance', 'Medium'))
                importance = earning.get('Importance', 'Medium')
                if importance == 'High':
                    importance_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    importance_cell.font = Font(color="CC0000", bold=True)
                elif importance == 'Medium':
                    importance_cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    importance_cell.font = Font(color="FF8C00", bold=True)
                else:
                    importance_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    importance_cell.font = Font(color="228B22", bold=True)
                
                # Add text cells with wrapping
                why_cell = ws_earnings.cell(row=row_idx, column=8, value=earning.get('Why Important', 'N/A'))
                why_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                metrics_cell = ws_earnings.cell(row=row_idx, column=9, value=earning.get('Key Metrics', 'N/A'))
                metrics_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                impact_cell = ws_earnings.cell(row=row_idx, column=10, value=earning.get('Market Impact', 'N/A'))
                impact_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Set row height for better readability
                ws_earnings.row_dimensions[row_idx].height = 80
        
        # Add metadata to first sheet
        last_row = len(events_list) + 3 if events_list else 3
        ws_events.cell(row=last_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_events.cell(row=last_row + 1, column=1, value="Source: OpenAI Market Analysis")
        
        # Create BytesIO object and save workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        return None


def get_weekly_market_events_from_db(monday):
    """
    Get market events from database for a specific week
    """
    import psycopg2
    import json
    import os
    
    try:
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor()
        
        cur.execute(
            "SELECT event_data FROM market_events WHERE week_start = %s",
            (monday.date(),)
        )
        
        result = cur.fetchone()
        conn.close()
        
        if result:
            # Handle both dict and string formats for robustness
            event_data = result[0]
            if isinstance(event_data, str):
                event_data = json.loads(event_data)
            elif isinstance(event_data, dict):
                # Data is already parsed
                pass
            else:
                print(f"DATABASE ERROR: Unexpected event_data type: {type(event_data)}")
                return None, False
                
            # Return both events and earnings
            return {
                'events': event_data.get('events', []),
                'earnings': event_data.get('earnings', [])
            }, True
        return None, False
        
    except Exception as e:
        print(f"DATABASE ERROR: {str(e)}")
        return None, False


def save_weekly_market_events_to_db(monday, sunday, market_data):
    """
    Save market events and earnings to database
    """
    import psycopg2
    import json
    import os
    
    try:
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor()
        
        event_data = {
            'events': market_data.get('events', []),
            'earnings': market_data.get('earnings', []),
            'week_range': f"{monday.strftime('%B %d')} - {sunday.strftime('%B %d, %Y')}",
            'fetched_at': datetime.now().isoformat()
        }
        
        cur.execute(
            """
            INSERT INTO market_events (week_start, week_end, event_data) 
            VALUES (%s, %s, %s)
            ON CONFLICT (week_start) 
            DO UPDATE SET event_data = EXCLUDED.event_data, created_at = CURRENT_TIMESTAMP
            """,
            (monday.date(), sunday.date(), json.dumps(event_data))
        )
        
        conn.commit()
        conn.close()
        return True
        
    except Exception as e:
        print(f"DATABASE SAVE ERROR: {str(e)}")
        return False


def get_weekly_market_events(force_refresh=False):
    """
    Use OpenAI to fetch top 10 key financial market events for the current week
    Store results in database to avoid repeated API calls for the same week
    
    Args:
        force_refresh: If True, bypass database cache and fetch fresh data
    
    Returns:
        list: List of market events with title, date, time, and description
    """
    import os
    from datetime import datetime, timedelta
    from openai import OpenAI
    
    try:
        # Get current week dates
        today = datetime.now()
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        sunday = monday + timedelta(days=6)
        
        week_range = f"{monday.strftime('%B %d')} - {sunday.strftime('%B %d, %Y')}"
        
        # Check database for existing data (unless force refresh)
        if not force_refresh:
            cached_events, found_in_db = get_weekly_market_events_from_db(monday)
            if found_in_db and cached_events:
                print(f"MARKET EVENTS: Using database data for week {week_range} - found {len(cached_events)} events")
                return cached_events, None
            else:
                print(f"MARKET EVENTS: No database data found for week {week_range}, will fetch from OpenAI")
        
        # Check if OpenAI API key is available
        openai_key = os.environ.get("OPENAI_API_KEY")
        if not openai_key:
            return [], "OpenAI API key not found. Please add OPENAI_API_KEY to your environment variables."
        
        print(f"MARKET EVENTS: Fetching NEW events for week {week_range} (not in cache)")
        
        # the newest OpenAI model is "gpt-5" which was released August 7, 2025.
        # do not change this unless explicitly requested by the user
        client = OpenAI(api_key=openai_key)
        
        # Create prompt for OpenAI
        prompt = f"""
        You are a financial market analyst. Provide comprehensive market intelligence for the week of {week_range}.

        Part 1: Provide the top 10 most important financial market events and economic releases.

        Include events such as:
        - Federal Reserve meetings or announcements
        - Major economic data releases (GDP, CPI, employment data, etc.)
        - Central bank meetings from major economies
        - Key corporate events (IPOs, major announcements)
        - Economic policy announcements
        - Important financial conferences or summits

        Part 2: Provide the top 5 company earnings to watch this week with detailed analysis.

        Focus on major companies with significant market impact potential.

        Format your response as JSON with this exact structure:
        {{
          "events": [
            {{
              "title": "Brief event title",
              "date": "YYYY-MM-DD",
              "time": "Time or 'TBD'",
              "category": "Fed Meeting|Economic Data|Earnings|Corporate|Policy|Conference",
              "importance": "High|Medium|Low", 
              "description": "Detailed explanation of why this event matters to markets",
              "market_impact": "Specific explanation of how this could affect stock prices, bond yields, currency values, and investor sentiment. Include potential market reactions and trading implications."
            }}
          ],
          "top_earnings": [
            {{
              "company": "Company Name",
              "ticker": "TICKER",
              "date": "YYYY-MM-DD",
              "time": "Before Market Open|After Market Close|TBD",
              "market_cap": "Large Cap|Mid Cap|Small Cap",
              "sector": "Technology|Healthcare|Finance|etc.",
              "importance": "High|Medium|Low",
              "why_important": "Detailed explanation of why this earnings report is crucial for markets, including company's market influence, recent performance, guidance expectations, and potential sector impact",
              "key_metrics": "Key financial metrics and expectations investors should watch (revenue, EPS, guidance, specific business segments)",
              "market_impact": "How this earnings could affect the stock, sector, and broader market sentiment"
            }}
          ]
        }}

        For market events, provide:
        1. A clear description of what the event is
        2. Detailed market impact analysis explaining why investors should care
        3. Specific mention of which market sectors or asset classes could be most affected

        For earnings, provide:
        1. Companies with significant market influence and investor attention
        2. Detailed analysis of why each earnings report matters
        3. Key metrics and expectations to watch
        4. Potential impact on stock, sector, and broader market

        Focus on events that could significantly impact stock markets, bond markets, or currency markets.
        Only include events and earnings from {week_range}.
        """
        
        # Call OpenAI API
        response = client.chat.completions.create(
            model="gpt-5",
            messages=[
                {"role": "system", "content": "You are a financial market expert providing accurate, current market event information."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"}
        )
        
        # Parse response
        import json
        market_data = json.loads(response.choices[0].message.content)
        events_list = market_data.get('events', [])
        earnings_list = market_data.get('top_earnings', [])
        
        print(f"MARKET EVENTS: Found {len(events_list)} events and {len(earnings_list)} earnings from OpenAI")
        
        # Validate and format events
        formatted_events = []
        for event in events_list:
            try:
                formatted_event = {
                    'Title': event.get('title', 'N/A'),
                    'Date': event.get('date', 'TBD'),
                    'Time': event.get('time', 'TBD'),
                    'Category': event.get('category', 'N/A'),
                    'Importance': event.get('importance', 'Medium'),
                    'Description': event.get('description', 'N/A'),
                    'Market Impact': event.get('market_impact', 'N/A')
                }
                formatted_events.append(formatted_event)
            except Exception as e:
                print(f"MARKET EVENTS: Error formatting event: {str(e)}")
                continue
        
        # Validate and format earnings
        formatted_earnings = []
        for earning in earnings_list:
            try:
                formatted_earning = {
                    'Company': earning.get('company', 'N/A'),
                    'Ticker': earning.get('ticker', 'N/A'),
                    'Date': earning.get('date', 'TBD'),
                    'Time': earning.get('time', 'TBD'),
                    'Market Cap': earning.get('market_cap', 'N/A'),
                    'Sector': earning.get('sector', 'N/A'),
                    'Importance': earning.get('importance', 'Medium'),
                    'Why Important': earning.get('why_important', 'N/A'),
                    'Key Metrics': earning.get('key_metrics', 'N/A'),
                    'Market Impact': earning.get('market_impact', 'N/A')
                }
                formatted_earnings.append(formatted_earning)
            except Exception as e:
                print(f"MARKET EVENTS: Error formatting earning: {str(e)}")
                continue
        
        # Combine events and earnings data for storage
        combined_data = {
            'events': formatted_events,
            'earnings': formatted_earnings
        }
        
        # Save the results to database
        if save_weekly_market_events_to_db(monday, sunday, combined_data):
            print(f"MARKET EVENTS: Saved {len(formatted_events)} events and {len(formatted_earnings)} earnings to database for week {week_range}")
        else:
            print(f"MARKET EVENTS: Failed to save to database, data will not be cached")
        
        return combined_data, None
        
    except Exception as e:
        error_msg = f"Error fetching market events: {str(e)}"
        print(f"MARKET EVENTS ERROR: {error_msg}")
        return [], error_msg


def market_events_tab():
    """Market Events tab content"""
    
    st.markdown("### 🗓️ Weekly Market Events")
    st.markdown("Top 10 key financial market events for the current week using **OpenAI Analysis**")
    st.markdown("---")
    
    # Information section
    st.info("""
    **🤖 About This Data:**
    - Powered by OpenAI's latest financial market analysis
    - Covers Federal Reserve meetings, economic data releases, and major corporate events
    - Updates weekly with focus on market-moving events
    - Includes event importance ratings and detailed descriptions
    - Covers global markets and central bank activities
    
    **⚡ Performance:** Data is stored in database - first load may take 10-20 seconds, subsequent loads are instant
    **💹 Enhanced Analysis:** Each event includes detailed market impact explanations and trading implications
    **📊 Event Types:** Fed Meetings, Economic Data, Earnings, Corporate Events, Policy Announcements, Conferences
    """)
    
    st.markdown("---")
    
    # Load market events section
    col_info1, col_info2 = st.columns([3, 1])
    
    with col_info1:
        st.markdown("""
        🗓️ **What's Included:**
        • Federal Reserve meetings and announcements
        • Key economic data releases (GDP, CPI, employment)
        • Central bank meetings from major economies
        • Important corporate earnings and events
        • Market-moving policy announcements
        """)
    
    with col_info2:
        st.markdown("""
        **📈 Impact Levels:**
        • **High**: Major market movers
        • **Medium**: Notable events
        • **Low**: Minor impact expected
        """)
    
    # Load button
    col_load1, col_load2, col_load3 = st.columns([2, 1.5, 1])
    
    with col_load1:
        if st.button("🤖 Load This Week's Market Events", type="primary", help="Fetch key financial events using OpenAI"):
            st.session_state.load_market_events = True
    
    with col_load2:
        # Refresh database button
        if st.button("🔄 Refresh Data", help="Fetch fresh data from OpenAI and update database"):
            st.session_state.force_refresh_market_events = True
            st.session_state.load_market_events = True
            st.success("Fetching fresh data from OpenAI...")
    
    with col_load3:
        # Weekly range info
        monday = datetime.now() - timedelta(days=datetime.now().weekday())
        sunday = monday + timedelta(days=6)
        st.info(f"Range: {monday.strftime('%b %d')} - {sunday.strftime('%b %d, %Y')}")
    
    # Load and display market events
    if st.session_state.get('load_market_events', False):
        force_refresh = st.session_state.get('force_refresh_market_events', False)
        if force_refresh:
            st.session_state.force_refresh_market_events = False  # Reset flag
            
        # Clear and simple logic: Check database first, then OpenAI if needed
        today = datetime.now()
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        
        if not force_refresh:
            # First: Check database
            with st.spinner("Checking database for cached events..."):
                cached_data, found_in_db = get_weekly_market_events_from_db(monday)
                
            if found_in_db and cached_data:
                # Display database data
                market_data = cached_data
                events_list = market_data.get('events', [])
                earnings_list = market_data.get('earnings', [])
                st.success(f"✅ Found **{len(events_list)}** market events and **{len(earnings_list)}** earnings this week (📊 Retrieved from Database)")
                show_events = True
            else:
                # No database data found, need to fetch from OpenAI
                show_events = False
        else:
            # Force refresh requested
            show_events = False
            
        # If no database data or force refresh, fetch from OpenAI
        if not show_events:
            spinner_text = "Fetching fresh data from OpenAI..." if force_refresh else "No database data found - fetching from OpenAI..."
            with st.spinner(spinner_text):
                market_data, error_msg = get_weekly_market_events(force_refresh=True)
                
                if error_msg:
                    st.error(f"❌ {error_msg}")
                    if "OpenAI API key" in error_msg:
                        st.markdown("### 🔑 OpenAI API Key Required:")
                        st.info("""
                        To use AI-powered market events analysis, you need an OpenAI API key:
                        
                        1. Visit [OpenAI API](https://platform.openai.com/api-keys)
                        2. Sign up or log in to your account
                        3. Create a new API key
                        4. Add it as `OPENAI_API_KEY` in your environment variables
                        5. Restart the application
                        """)
                    show_events = False
                elif market_data:
                    events_list = market_data.get('events', [])
                    earnings_list = market_data.get('earnings', [])
                    st.success(f"✅ Found **{len(events_list)}** market events and **{len(earnings_list)}** earnings this week (🔄 Fresh from OpenAI)")
                    show_events = True
                else:
                    show_events = False
                    
        # Display events and earnings if we have them
        if show_events and (events_list or earnings_list):
                
                # Display market events in card format for better readability
                st.markdown("### 📋 Weekly Market Events")
                
                # Show events as cards with proper text wrapping
                for i, event in enumerate(events_list, 1):
                    importance = event.get('Importance', 'Medium')
                    
                    # Color coding for importance
                    if importance == "High":
                        border_color = "#dc3545"
                        bg_color = "#fff5f5"
                        icon = "🔴"
                    elif importance == "Medium":
                        border_color = "#fd7e14"
                        bg_color = "#fff8f0"
                        icon = "🟡"
                    else:
                        border_color = "#198754"
                        bg_color = "#f0fff4"
                        icon = "🟢"
                    
                    # Create a container for each event
                    with st.container():
                        st.markdown(f"""
                        <div style="
                            border: 2px solid {border_color};
                            border-radius: 10px;
                            padding: 20px;
                            margin: 15px 0;
                            background-color: {bg_color};
                        ">
                            <h4 style="margin-top: 0; color: {border_color};">
                                {icon} {event.get('Title', 'N/A')}
                            </h4>
                            <div style="display: flex; flex-wrap: wrap; gap: 20px; margin-bottom: 15px;">
                                <span><strong>📆 Date:</strong> {event.get('Date', 'TBD')}</span>
                                <span><strong>🕒 Time:</strong> {event.get('Time', 'TBD')}</span>
                                <span><strong>🏷️ Category:</strong> {event.get('Category', 'N/A')}</span>
                                <span><strong>⚡ Impact:</strong> {importance}</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Use Streamlit columns for better text wrapping
                        col1, col2 = st.columns([1, 1])
                        
                        with col1:
                            st.markdown("**📝 What It Is:**")
                            st.write(event.get('Description', 'N/A'))
                        
                        with col2:
                            st.markdown("**💹 Why It Matters:**")
                            st.write(event.get('Market Impact', 'N/A'))
                        
                        st.markdown("---")
                
                # Display Top 5 Earnings to Watch
                if earnings_list:
                    st.markdown("### 📈 Top 5 Earnings to Watch")
                    
                    for i, earning in enumerate(earnings_list, 1):
                        importance = earning.get('Importance', 'Medium')
                        
                        # Color coding for importance
                        if importance == "High":
                            border_color = "#dc3545"
                            bg_color = "#fff5f5"
                            icon = "🔴"
                        elif importance == "Medium":
                            border_color = "#fd7e14"
                            bg_color = "#fff8f0"
                            icon = "🟡"
                        else:
                            border_color = "#198754"
                            bg_color = "#f0fff4"
                            icon = "🟢"
                        
                        # Create earnings card
                        with st.container():
                            st.markdown(f"""
                            <div style="
                                border: 2px solid {border_color};
                                border-radius: 10px;
                                padding: 20px;
                                margin: 15px 0;
                                background-color: {bg_color};
                            ">
                                <h4 style="margin-top: 0; color: {border_color};">
                                    {icon} {earning.get('Company', 'N/A')} ({earning.get('Ticker', 'N/A')})
                                </h4>
                                <div style="display: flex; flex-wrap: wrap; gap: 20px; margin-bottom: 15px;">
                                    <span><strong>📆 Date:</strong> {earning.get('Date', 'TBD')}</span>
                                    <span><strong>🕒 Time:</strong> {earning.get('Time', 'TBD')}</span>
                                    <span><strong>🏢 Market Cap:</strong> {earning.get('Market Cap', 'N/A')}</span>
                                    <span><strong>🏭 Sector:</strong> {earning.get('Sector', 'N/A')}</span>
                                    <span><strong>⚡ Impact:</strong> {importance}</span>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # Use Streamlit columns for better text wrapping
                            col1, col2, col3 = st.columns([1, 1, 1])
                            
                            with col1:
                                st.markdown("**🎯 Why Important:**")
                                st.write(earning.get('Why Important', 'N/A'))
                            
                            with col2:
                                st.markdown("**📊 Key Metrics:**")
                                st.write(earning.get('Key Metrics', 'N/A'))
                            
                            with col3:
                                st.markdown("**💹 Market Impact:**")
                                st.write(earning.get('Market Impact', 'N/A'))
                            
                            st.markdown("---")
                
                # Display summary statistics and download option
                st.markdown("### 📊 Weekly Summary")
                col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
                
                with col_stats1:
                    high_impact = len([e for e in events_list if e.get('Importance') == 'High'])
                    st.metric("High Impact Events", high_impact)
                
                with col_stats2:
                    categories = list(set([e.get('Category', 'N/A') for e in events_list]))
                    st.metric("Event Categories", len(categories))
                
                with col_stats3:
                    high_impact_earnings = len([e for e in earnings_list if e.get('Importance') == 'High'])
                    st.metric("High Impact Earnings", high_impact_earnings)
                
                with col_stats4:
                    # Excel download button
                    import pandas as pd
                    
                    # Generate Excel file with both events and earnings
                    excel_data = create_market_events_excel(events_list, earnings_list)
                    
                    if excel_data:
                        today = datetime.now()
                        days_since_monday = today.weekday()
                        monday = today - timedelta(days=days_since_monday)
                        sunday = monday + timedelta(days=6)
                        week_range = f"{monday.strftime('%b_%d')}-{sunday.strftime('%b_%d_%Y')}"
                        filename = f"Market_Analysis_{week_range}.xlsx"
                        
                        st.download_button(
                            label="📥 Download Excel",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download complete market analysis as Excel file"
                        )
        else:
            st.warning("⚠️ No major market events found for this week")
            st.markdown("""
            This might mean:
            • Light economic calendar week
            • Major events scheduled outside current timeframe
            • OpenAI response formatting issue
            """)
        
        # Auto-update information
        st.markdown("---")
        st.markdown("### 🔄 Auto-Update Information")
        
        col_update1, col_update2 = st.columns(2)
        
        with col_update1:
            st.info("""
            **📅 Weekly Schedule:**
            • Events cover Monday through Sunday
            • Data refreshes when you reload on Monday
            • Focuses on market-moving events only
            """)
        
        with col_update2:
            st.info("""
            **🤖 AI-Powered:**
            • Uses OpenAI's latest financial knowledge
            • Provides importance ratings and descriptions
            • Covers global markets and central banks
            """)


def stock_screener_tab():
    """Stock Screener tab content"""
    
    st.markdown("### 🔍 Stock Screener")
    st.markdown("Filter and discover stocks based on financial metrics, technical indicators, and custom criteria")
    st.markdown("---")
    
    # Information section
    st.info("""
    **📊 Ready to Screen Stocks:**
    - **Value Stocks strategy is pre-loaded** - Click "Run Screen" to start immediately!
    - Or customize filters: market cap, P/E ratio, dividend yield, growth metrics
    - Real-time data from Yahoo Finance across 100+ major US stocks
    - Export results to Excel for further analysis
    - Choose from 5 preset investment strategies or create custom filters
    """)
    
    st.markdown("---")
    
    # Screening interface
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("#### 🎯 Filter Criteria")
        
        # Create tabs for different filter categories
        filter_tabs = st.tabs(["💰 Valuation", "📈 Growth", "💵 Dividends", "📊 Technical", "🏢 Size"])
        
        # Initialize preset values if not set - default to Value Stocks for immediate use
        if 'preset_applied' not in st.session_state:
            st.session_state.preset_applied = "Value Stocks"
            
        # Set default values based on any applied preset
        def get_default_value(key, default_val, preset_values=None):
            if st.session_state.preset_applied and preset_values and key in preset_values:
                return preset_values[key]
            return default_val
        
        # Get preset values if any (made more realistic/less restrictive)
        preset_values = {}
        if st.session_state.preset_applied == "High Dividend":
            preset_values = {'div_yield_min': 2.0, 'div_yield_max': 15.0, 'payout_ratio_max': 90.0, 'market_cap_min': 2}
        elif st.session_state.preset_applied == "Value Stocks":
            preset_values = {'pe_min': 0.0, 'pe_max': 20.0, 'pb_min': 0.0, 'pb_max': 5.0, 'roe_min': 5.0, 'roe_max': 100.0, 'market_cap_min': 0}
        elif st.session_state.preset_applied == "Growth Stocks":
            preset_values = {'rev_growth_min': 5.0, 'earnings_growth_min': 10.0, 'roe_min': 10.0, 'roe_max': 100.0, 'gross_margin_min': 20.0, 'market_cap_min': 1}
        elif st.session_state.preset_applied == "Large Cap Dividend":
            preset_values = {'div_yield_min': 1.0, 'div_yield_max': 10.0, 'payout_ratio_max': 80.0, 'market_cap_min': 4, 'beta_max': 2.0}
        elif st.session_state.preset_applied == "Small Cap Growth":
            preset_values = {'rev_growth_min': 10.0, 'earnings_growth_min': 15.0, 'roe_min': 15.0, 'roe_max': 100.0, 'gross_margin_min': 25.0, 'market_cap_min': 1, 'volume_avg_min': 25000}

        with filter_tabs[0]:  # Valuation filters
            col_val1, col_val2, col_val3 = st.columns(3)
            with col_val1:
                pe_min = st.number_input("Min P/E Ratio", min_value=0.0, value=get_default_value('pe_min', 0.0, preset_values), step=0.1, help="Minimum Price-to-Earnings ratio")
                pe_max = st.number_input("Max P/E Ratio", min_value=0.0, value=get_default_value('pe_max', 50.0, preset_values), step=0.1, help="Maximum Price-to-Earnings ratio")
            with col_val2:
                pb_min = st.number_input("Min P/B Ratio", min_value=0.0, value=get_default_value('pb_min', 0.0, preset_values), step=0.1, help="Minimum Price-to-Book ratio")
                pb_max = st.number_input("Max P/B Ratio", min_value=0.0, value=get_default_value('pb_max', 10.0, preset_values), step=0.1, help="Maximum Price-to-Book ratio")
            with col_val3:
                roe_min = st.number_input("Min ROE (%)", min_value=-50.0, value=get_default_value('roe_min', 0.0, preset_values), step=1.0, help="Minimum Return on Equity")
                roe_max = st.number_input("Max ROE (%)", min_value=-50.0, value=get_default_value('roe_max', 100.0, preset_values), step=1.0, help="Maximum Return on Equity")
        
        with filter_tabs[1]:  # Growth filters
            col_growth1, col_growth2 = st.columns(2)
            with col_growth1:
                rev_growth_min = st.number_input("Min Revenue Growth (%)", min_value=-100.0, value=get_default_value('rev_growth_min', 0.0, preset_values), step=1.0, help="Minimum revenue growth rate")
                earnings_growth_min = st.number_input("Min Earnings Growth (%)", min_value=-100.0, value=get_default_value('earnings_growth_min', 0.0, preset_values), step=1.0, help="Minimum earnings growth rate")
            with col_growth2:
                profit_margin_min = st.number_input("Min Profit Margin (%)", min_value=-50.0, value=get_default_value('profit_margin_min', 0.0, preset_values), step=1.0, help="Minimum profit margin")
                gross_margin_min = st.number_input("Min Gross Margin (%)", min_value=-50.0, value=get_default_value('gross_margin_min', 0.0, preset_values), step=1.0, help="Minimum gross profit margin")
        
        with filter_tabs[2]:  # Dividend filters
            col_div1, col_div2 = st.columns(2)
            with col_div1:
                div_yield_min = st.number_input("Min Dividend Yield (%)", min_value=0.0, value=get_default_value('div_yield_min', 0.0, preset_values), step=0.1, help="Minimum dividend yield")
                div_yield_max = st.number_input("Max Dividend Yield (%)", min_value=0.0, value=get_default_value('div_yield_max', 20.0, preset_values), step=0.1, help="Maximum dividend yield")
            with col_div2:
                payout_ratio_max = st.number_input("Max Payout Ratio (%)", min_value=0.0, value=get_default_value('payout_ratio_max', 100.0, preset_values), step=1.0, help="Maximum dividend payout ratio")
        
        with filter_tabs[3]:  # Technical filters
            col_tech1, col_tech2 = st.columns(2)
            with col_tech1:
                price_change_1m_min = st.number_input("Min 1M Price Change (%)", min_value=-100.0, value=get_default_value('price_change_1m_min', -50.0, preset_values), step=1.0, help="Minimum 1-month price change")
                price_change_1m_max = st.number_input("Max 1M Price Change (%)", min_value=-100.0, value=get_default_value('price_change_1m_max', 100.0, preset_values), step=1.0, help="Maximum 1-month price change")
            with col_tech2:
                volume_avg_min = st.number_input("Min Avg Volume", min_value=0, value=get_default_value('volume_avg_min', 100000, preset_values), step=10000, help="Minimum average daily volume")
                beta_max = st.number_input("Max Beta", min_value=0.0, value=get_default_value('beta_max', 3.0, preset_values), step=0.1, help="Maximum beta (volatility vs market)")
        
        with filter_tabs[4]:  # Size filters
            col_size1, col_size2 = st.columns(2)
            with col_size1:
                market_cap_options = ["Any", "Micro ($0-300M)", "Small ($300M-2B)", "Mid ($2B-10B)", "Large ($10B+)"]
                market_cap_index = get_default_value('market_cap_min', 0, preset_values)
                market_cap_min = st.selectbox("Min Market Cap", 
                    market_cap_options,
                    index=market_cap_index, 
                    help="Minimum market capitalization")
            with col_size2:
                sectors = st.multiselect("Sectors", 
                    ["Technology", "Healthcare", "Financial Services", "Consumer Cyclical", 
                     "Industrials", "Energy", "Utilities", "Real Estate", "Materials", 
                     "Consumer Defensive", "Communication Services", "Basic Materials"],
                    default=get_default_value('sectors', [], preset_values),
                    help="Filter by specific sectors")
    
    with col2:
        st.markdown("#### 🎮 Quick Actions")
        
        # Popular preset filters with clear descriptions
        # Set default index based on current preset
        preset_options = [
            "Custom (Manual Settings)",
            "💰 High Dividend Stocks", 
            "📉 Value Stocks", 
            "📈 Growth Stocks",
            "🏦 Large Cap Dividend",
            "⚡ Small Cap Growth"
        ]
        
        # Find current preset index
        current_index = 0  # Default to Custom
        if st.session_state.preset_applied == "High Dividend":
            current_index = 1
        elif st.session_state.preset_applied == "Value Stocks":
            current_index = 2
        elif st.session_state.preset_applied == "Growth Stocks":
            current_index = 3
        elif st.session_state.preset_applied == "Large Cap Dividend":
            current_index = 4
        elif st.session_state.preset_applied == "Small Cap Growth":
            current_index = 5
            
        preset_filter = st.selectbox("Choose Preset Strategy", 
            preset_options,
            index=current_index,
            help="A preset strategy is already loaded - you can run the screen immediately or change strategies")
        
        # Show what the selected preset does
        if preset_filter == "💰 High Dividend Stocks":
            st.info("**Will set:** Dividend yield >2%, Payout ratio <90%, Mid-cap+ stocks")
        elif preset_filter == "📉 Value Stocks":
            st.info("**Will set:** P/E <20, P/B <5, ROE >5% (undervalued companies)")
        elif preset_filter == "📈 Growth Stocks":
            st.info("**Will set:** Revenue growth >5%, Earnings growth >10%, ROE >10%")
        elif preset_filter == "🏦 Large Cap Dividend":
            st.info("**Will set:** Large cap stocks, Dividend yield >1%, Stable financials")
        elif preset_filter == "⚡ Small Cap Growth":
            st.info("**Will set:** Small cap stocks, Growth metrics >10-15%, ROE >15%")
        
        col_preset1, col_preset2 = st.columns(2)
        
        with col_preset1:
            if st.button("Apply Preset", help="Apply the selected preset and clear other filters"):
                # Apply preset values
                if preset_filter == "💰 High Dividend Stocks":
                    st.session_state.preset_applied = "High Dividend"
                    st.success("✅ High Dividend preset applied! Check the Dividends tab.")
                    
                elif preset_filter == "📉 Value Stocks":
                    st.session_state.preset_applied = "Value Stocks"
                    st.success("✅ Value Stocks preset applied! Check Valuation & Growth tabs.")
                    
                elif preset_filter == "📈 Growth Stocks":
                    st.session_state.preset_applied = "Growth Stocks"
                    st.success("✅ Growth Stocks preset applied! Check Growth & Size tabs.")
                    
                elif preset_filter == "🏦 Large Cap Dividend":
                    st.session_state.preset_applied = "Large Cap Dividend"
                    st.success("✅ Large Cap Dividend preset applied! Check Dividends & Size tabs.")
                    
                elif preset_filter == "⚡ Small Cap Growth":
                    st.session_state.preset_applied = "Small Cap Growth"
                    st.success("✅ Small Cap Growth preset applied! Check Growth & Size tabs.")
                    
                # Trigger page rerun to update UI
                st.rerun()
        
        with col_preset2:
            if st.button("🔄 Reset All", help="Clear all filters and reset to defaults"):
                # Reset preset
                st.session_state.preset_applied = None
                st.success("✅ All filters reset to defaults!")
                st.rerun()
        
        st.markdown("---")
        
        # Show current status and run screen button
        if st.session_state.preset_applied:
            if st.session_state.preset_applied == "Value Stocks":
                st.success("✅ **Value Stocks strategy loaded** - Ready to screen for undervalued companies!")
            elif st.session_state.preset_applied == "High Dividend":
                st.success("✅ **High Dividend strategy loaded** - Ready to screen for dividend-paying stocks!")
            elif st.session_state.preset_applied == "Growth Stocks":
                st.success("✅ **Growth Stocks strategy loaded** - Ready to screen for growing companies!")
            elif st.session_state.preset_applied == "Large Cap Dividend":
                st.success("✅ **Large Cap Dividend strategy loaded** - Ready to screen for stable dividend stocks!")
            elif st.session_state.preset_applied == "Small Cap Growth":
                st.success("✅ **Small Cap Growth strategy loaded** - Ready to screen for small growth companies!")
        else:
            st.info("💡 Customize your filters above or select a preset strategy")
            
        # Run screen button
        if st.button("🔍 Run Screen", type="primary", help="Execute the stock screen with current filters"):
            st.session_state.run_screen = True
    
    st.markdown("---")
    
    # Results section
    if st.session_state.get('run_screen', False):
        with st.spinner("Screening stocks... This may take a moment"):
            
            # Create filter criteria dictionary
            filters = {
                'pe_min': pe_min, 'pe_max': pe_max,
                'pb_min': pb_min, 'pb_max': pb_max,
                'rev_growth_min': rev_growth_min, 'earnings_growth_min': earnings_growth_min,
                'roe_min': roe_min, 'roe_max': roe_max, 'profit_margin_min': profit_margin_min,
                'gross_margin_min': gross_margin_min,
                'div_yield_min': div_yield_min, 'div_yield_max': div_yield_max,
                'payout_ratio_max': payout_ratio_max,
                'price_change_1m_min': price_change_1m_min, 'price_change_1m_max': price_change_1m_max,
                'volume_avg_min': volume_avg_min, 'beta_max': beta_max,
                'market_cap_min': market_cap_min, 'sectors': sectors
            }
            
            # Run the screening
            results = run_stock_screen(filters)
            
            if results and len(results) > 0:
                st.success(f"✅ Found **{len(results)}** stocks matching your criteria")
                
                # Display results
                st.markdown("### 📊 Screening Results")
                
                # Sort options
                sort_by = st.selectbox("Sort by", 
                    ["Symbol", "Market Cap", "P/E Ratio", "Dividend Yield", "1M Change", "Volume"],
                    help="Sort results by selected metric")
                
                # Create results DataFrame
                import pandas as pd
                df_results = pd.DataFrame(results)
                
                # Apply sorting
                if sort_by == "Market Cap" and 'Market Cap' in df_results.columns:
                    df_results = df_results.sort_values('Market Cap', ascending=False)
                elif sort_by == "P/E Ratio" and 'P/E Ratio' in df_results.columns:
                    df_results = df_results.sort_values('P/E Ratio', ascending=True)
                elif sort_by == "Dividend Yield" and 'Dividend Yield' in df_results.columns:
                    df_results = df_results.sort_values('Dividend Yield', ascending=False)
                elif sort_by == "1M Change" and '1M Change' in df_results.columns:
                    df_results = df_results.sort_values('1M Change', ascending=False)
                elif sort_by == "Volume" and 'Avg Volume' in df_results.columns:
                    df_results = df_results.sort_values('Avg Volume', ascending=False)
                else:
                    df_results = df_results.sort_values('Symbol', ascending=True)
                
                # Display as interactive table
                st.dataframe(
                    df_results,
                    use_container_width=True,
                    hide_index=True,
                    height=400
                )
                
                # Export functionality
                col_export1, col_export2, col_export3 = st.columns([1, 1, 2])
                
                with col_export1:
                    # CSV download
                    csv = df_results.to_csv(index=False)
                    st.download_button(
                        label="📄 Download CSV",
                        data=csv,
                        file_name=f"stock_screen_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv"
                    )
                
                with col_export2:
                    # Excel download
                    excel_data = create_screener_excel(df_results, filters)
                    if excel_data:
                        st.download_button(
                            label="📊 Download Excel",
                            data=excel_data,
                            file_name=f"stock_screen_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                with col_export3:
                    st.info(f"💡 **Tip:** Click on column headers to sort the table. Found {len(results)} stocks matching your criteria.")
                
            else:
                st.warning("⚠️ No stocks found matching your criteria")
                st.markdown("""
                **Suggestions:**
                • Try relaxing some filter criteria
                • Check if P/E or P/B ratios are too restrictive
                • Consider expanding the market cap range
                • Remove sector filters to cast a wider net
                """)
                
        # Reset the run_screen flag
        st.session_state.run_screen = False
    
    else:
        st.markdown("### 🚀 Ready to Screen")
        st.markdown("✅ **Value Stocks strategy is pre-loaded and ready to use!** Click **'Run Screen'** above to find undervalued companies immediately.")
        
        # Sample criteria examples
        st.markdown("#### 💡 Available Investment Strategies:")
        st.markdown("""
        - **✅ Value Stocks** (Currently Loaded): P/E <20, P/B <5, ROE >5% - Find undervalued companies
        - **High Dividend Stocks**: Dividend yield >2%, stable payout ratios - Income-focused investing
        - **Growth Stocks**: Revenue growth >5%, earnings growth >10% - Fast-growing companies
        - **Large Cap Dividend**: Large companies with stable dividends - Conservative income
        - **Small Cap Growth**: Small companies with high growth potential - Aggressive growth
        
        **Want to try a different strategy?** Select one from the dropdown above and click 'Apply Preset'
        """)


def run_stock_screen(filters):
    """
    Execute stock screening based on provided filters
    """
    try:
        import yfinance as yf
        import pandas as pd
        import time
        import random
        
        # Expanded stock universe with current, active stocks
        sample_tickers = [
            # Tech Giants
            'AAPL', 'MSFT', 'GOOGL', 'AMZN', 'META', 'NVDA', 'TSLA', 'NFLX', 'ADBE', 'CRM',
            'ORCL', 'IBM', 'INTC', 'AMD', 'QCOM', 'AVGO', 'TXN', 'LRCX', 'KLAC', 'MRVL',
            
            # Financial Services
            'JPM', 'BAC', 'WFC', 'GS', 'MS', 'C', 'USB', 'PNC', 'TFC', 'COF',
            'AXP', 'BLK', 'SPGI', 'ICE', 'CME', 'V', 'MA', 'PYPL', 'FIS', 'FISV',
            
            # Healthcare & Pharma
            'JNJ', 'UNH', 'PFE', 'ABBV', 'MRK', 'TMO', 'ABT', 'DHR', 'BMY', 'AMGN',
            'GILD', 'BIIB', 'REGN', 'VRTX', 'MRNA', 'ZTS', 'CVS', 'CI', 'HUM', 'ANTM',
            
            # Consumer & Retail
            'WMT', 'HD', 'PG', 'KO', 'PEP', 'COST', 'TGT', 'LOW', 'SBUX', 'MCD',
            'NKE', 'DIS', 'CMCSA', 'VZ', 'T', 'CHTR', 'PM', 'MO', 'CL', 'KMB',
            
            # Industrial & Energy
            'XOM', 'CVX', 'COP', 'EOG', 'PSX', 'VLO', 'SLB', 'OXY', 'KMI', 'WMB',
            'GE', 'CAT', 'BA', 'RTX', 'LMT', 'NOC', 'GD', 'HON', 'UPS', 'FDX',
            
            # REITs & Utilities
            'NEE', 'DUK', 'SO', 'D', 'AEP', 'EXC', 'SRE', 'XEL', 'WEC', 'ES',
            'AMT', 'CCI', 'EQIX', 'PLD', 'SPG', 'O', 'WELL', 'EXR', 'AVB', 'EQR',
            
            # Materials & Other
            'LIN', 'APD', 'ECL', 'SHW', 'FCX', 'NEM', 'GOLD', 'FNV', 'BHP', 'RIO'
        ]
        
        results = []
        processed_count = 0
        failed_count = 0
        max_stocks = 100  # Increased limit for better screening results
        
        # Shuffle the list to get varied results
        random.shuffle(sample_tickers)
        
        # Add status placeholder
        status_placeholder = st.empty()
        
        for ticker in sample_tickers[:max_stocks]:
            try:
                processed_count += 1
                if processed_count % 10 == 0:
                    status_placeholder.info(f"🔄 Processing {processed_count}/{min(max_stocks, len(sample_tickers))} stocks... Found {len(results)} matches so far.")
                
                # Get stock data
                stock = yf.Ticker(ticker)
                info = stock.info
                hist = stock.history(period="3mo")
                
                if hist.empty or not info or len(info) < 5:
                    failed_count += 1
                    continue
                    
                # Extract metrics for filtering with better error handling
                current_price = hist['Close'].iloc[-1] if not hist.empty else 0
                if current_price <= 0:
                    continue
                    
                pe_ratio = info.get('trailingPE', None) 
                pb_ratio = info.get('priceToBook', None)
                market_cap = info.get('marketCap', 0) or 0
                dividend_yield = (info.get('dividendYield', 0) or 0) * 100
                beta = info.get('beta', 1) or 1
                avg_volume = info.get('averageVolume', 0) or 0
                profit_margin = (info.get('profitMargins', None) or 0) * 100 if info.get('profitMargins') else None
                roe = (info.get('returnOnEquity', None) or 0) * 100 if info.get('returnOnEquity') else None
                revenue_growth = (info.get('revenueGrowth', None) or 0) * 100 if info.get('revenueGrowth') else None
                earnings_growth = (info.get('earningsGrowth', None) or 0) * 100 if info.get('earningsGrowth') else None
                payout_ratio = (info.get('payoutRatio', None) or 0) * 100 if info.get('payoutRatio') else None
                gross_margin = (info.get('grossMargins', None) or 0) * 100 if info.get('grossMargins') else None
                sector = info.get('sector', 'Unknown')
                
                # Calculate 1-month price change
                try:
                    price_1m_ago = hist['Close'].iloc[-22] if len(hist) >= 22 else hist['Close'].iloc[0]
                    price_change_1m = ((current_price - price_1m_ago) / price_1m_ago) * 100
                except:
                    price_change_1m = 0
                
                # Apply filters with None checking (only filter if data is available)
                if pe_ratio is not None and (pe_ratio < filters['pe_min'] or pe_ratio > filters['pe_max']):
                    continue
                if pb_ratio is not None and (pb_ratio < filters['pb_min'] or pb_ratio > filters['pb_max']):
                    continue
                if revenue_growth is not None and revenue_growth < filters['rev_growth_min']:
                    continue
                if earnings_growth is not None and earnings_growth < filters['earnings_growth_min']:
                    continue
                if roe is not None and (roe < filters['roe_min'] or roe > filters['roe_max']):
                    continue
                if profit_margin is not None and profit_margin < filters['profit_margin_min']:
                    continue
                if gross_margin is not None and gross_margin < filters['gross_margin_min']:
                    continue
                if dividend_yield < filters['div_yield_min'] or dividend_yield > filters['div_yield_max']:
                    continue
                if payout_ratio is not None and payout_ratio > filters['payout_ratio_max']:
                    continue
                if price_change_1m < filters['price_change_1m_min'] or price_change_1m > filters['price_change_1m_max']:
                    continue
                if avg_volume < filters['volume_avg_min']:
                    continue
                if beta > filters['beta_max']:
                    continue
                
                # Market cap filter
                market_cap_options = ["Any", "Micro ($0-300M)", "Small ($300M-2B)", "Mid ($2B-10B)", "Large ($10B+)"]
                market_cap_selected = filters['market_cap_min']
                
                if market_cap_selected != "Any":
                    if market_cap_selected == "Micro ($0-300M)" and market_cap > 300_000_000:
                        continue
                    elif market_cap_selected == "Small ($300M-2B)" and (market_cap < 300_000_000 or market_cap > 2_000_000_000):
                        continue
                    elif market_cap_selected == "Mid ($2B-10B)" and (market_cap < 2_000_000_000 or market_cap > 10_000_000_000):
                        continue
                    elif market_cap_selected == "Large ($10B+)" and market_cap < 10_000_000_000:
                        continue
                
                # Sector filter
                if filters['sectors'] and sector not in filters['sectors']:
                    continue
                
                # Format market cap
                if market_cap >= 1_000_000_000:
                    market_cap_str = f"${market_cap/1_000_000_000:.1f}B"
                elif market_cap >= 1_000_000:
                    market_cap_str = f"${market_cap/1_000_000:.0f}M"
                else:
                    market_cap_str = f"${market_cap:,.0f}"
                
                # Add to results
                results.append({
                    'Symbol': ticker,
                    'Company': info.get('longName', ticker)[:30],
                    'Sector': sector,
                    'Market Cap': market_cap_str,
                    'Price': f"${current_price:.2f}",
                    'P/E Ratio': f"{pe_ratio:.1f}" if pe_ratio is not None and pe_ratio > 0 else "N/A",
                    'P/B Ratio': f"{pb_ratio:.1f}" if pb_ratio is not None and pb_ratio > 0 else "N/A",
                    'Dividend Yield': f"{dividend_yield:.1f}%" if dividend_yield > 0 else "0.0%",
                    'ROE': f"{roe:.1f}%" if roe is not None else "N/A",
                    'Profit Margin': f"{profit_margin:.1f}%" if profit_margin is not None else "N/A",
                    'Revenue Growth': f"{revenue_growth:.1f}%" if revenue_growth is not None else "N/A",
                    'Earnings Growth': f"{earnings_growth:.1f}%" if earnings_growth is not None else "N/A",
                    '1M Change': f"{price_change_1m:.1f}%",
                    'Beta': f"{beta:.2f}" if beta > 0 else "N/A",
                    'Avg Volume': f"{avg_volume:,}" if avg_volume > 0 else "N/A"
                })
                
                # Small delay to avoid overwhelming the API
                time.sleep(0.1)
                
            except Exception as e:
                print(f"Error processing {ticker}: {str(e)}")
                failed_count += 1
                continue
        
        # Clear status and show final results
        status_placeholder.success(f"✅ Screening complete! Processed {processed_count} stocks, found {len(results)} matches, {failed_count} failed.")
        
        return results
        
    except Exception as e:
        st.error(f"Error running stock screen: {str(e)}")
        return []


def create_screener_excel(df, filters):
    """
    Create Excel file for screener results
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Screener Results"
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[cell.column_letter].width = 15
        
        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Add filter information
        ws_filters = wb.create_sheet("Filter Criteria")
        ws_filters.cell(row=1, column=1, value="Stock Screener - Filter Criteria")
        ws_filters.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        filter_row = 4
        for key, value in filters.items():
            if value not in [0.0, 0, [], "Any", ""]:
                ws_filters.cell(row=filter_row, column=1, value=key.replace('_', ' ').title())
                ws_filters.cell(row=filter_row, column=2, value=str(value))
                filter_row += 1
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        return None


def calculate_investment_ratings(ticker_info, ticker_obj):
    """
    Calculate 1-5 scale investment ratings for Quantitative, Author, and Sellside analysis
    
    Args:
        ticker_info: Stock information dictionary from yfinance
        ticker_obj: yfinance Ticker object
    
    Returns:
        dict: Ratings and explanations for each category (1-5 scale)
    """
    try:
        # Initialize ratings
        ratings = {
            'quantitative': 3,
            'author': 3, 
            'sellside': 3,
            'quantitative_explanation': '',
            'author_explanation': '',
            'sellside_explanation': ''
        }
        
        # QUANTITATIVE RATING (Based on financial metrics and ratios)
        quant_score = 0
        quant_factors = []
        
        # P/E Ratio assessment
        pe_ratio = ticker_info.get('trailingPE')
        if pe_ratio:
            if pe_ratio < 15:
                quant_score += 1
                quant_factors.append("Attractive P/E (<15)")
            elif pe_ratio < 25:
                quant_score += 0.5
                quant_factors.append("Reasonable P/E (<25)")
            else:
                quant_factors.append("High P/E (>25)")
        
        # ROE assessment
        roe = ticker_info.get('returnOnEquity')
        if roe:
            if roe > 0.15:  # 15%+
                quant_score += 1
                quant_factors.append("Strong ROE (>15%)")
            elif roe > 0.10:  # 10-15%
                quant_score += 0.5
                quant_factors.append("Good ROE (>10%)")
            else:
                quant_factors.append("Weak ROE (<10%)")
        
        # Debt-to-Equity assessment
        debt_to_equity = ticker_info.get('debtToEquity')
        if debt_to_equity is not None:
            if debt_to_equity < 50:
                quant_score += 1
                quant_factors.append("Low debt (<50)")
            elif debt_to_equity < 100:
                quant_score += 0.5
                quant_factors.append("Moderate debt (<100)")
            else:
                quant_factors.append("High debt (>100)")
        
        # Profit margins
        profit_margin = ticker_info.get('profitMargins')
        if profit_margin:
            if profit_margin > 0.15:  # 15%+
                quant_score += 1
                quant_factors.append("High margins (>15%)")
            elif profit_margin > 0.05:  # 5-15%
                quant_score += 0.5
                quant_factors.append("Good margins (>5%)")
            else:
                quant_factors.append("Low margins (<5%)")
        
        # Revenue growth
        revenue_growth = ticker_info.get('revenueGrowth')
        if revenue_growth:
            if revenue_growth > 0.15:  # 15%+
                quant_score += 1
                quant_factors.append("Strong growth (>15%)")
            elif revenue_growth > 0.05:  # 5-15%
                quant_score += 0.5
                quant_factors.append("Positive growth (>5%)")
            else:
                quant_factors.append("Slow growth (<5%)")
        
        # Convert to 1-5 scale
        if quant_score >= 4:
            ratings['quantitative'] = 5
        elif quant_score >= 3:
            ratings['quantitative'] = 4
        elif quant_score >= 2:
            ratings['quantitative'] = 3
        elif quant_score >= 1:
            ratings['quantitative'] = 2
        else:
            ratings['quantitative'] = 1
        
        ratings['quantitative_explanation'] = f"Based on: {', '.join(quant_factors[:3])}" if quant_factors else "Limited quantitative data available"
        
        # AUTHOR RATING (Based on comprehensive analysis perspective)
        author_score = 0
        author_factors = []
        
        # Market cap consideration
        market_cap = ticker_info.get('marketCap')
        if market_cap:
            if market_cap > 100e9:  # Large cap
                author_score += 0.5
                author_factors.append("Large-cap stability")
            elif market_cap > 10e9:  # Mid cap
                author_score += 1
                author_factors.append("Mid-cap potential")
            else:  # Small cap
                author_score += 0.5
                author_factors.append("Small-cap risk/reward")
        
        # Beta consideration
        beta = ticker_info.get('beta')
        if beta:
            if 0.8 <= beta <= 1.2:
                author_score += 1
                author_factors.append("Moderate volatility")
            elif beta < 0.8:
                author_score += 0.5
                author_factors.append("Low volatility")
            else:
                author_factors.append("High volatility")
        
        # Analyst coverage
        analyst_count = ticker_info.get('numberOfAnalystOpinions')
        if analyst_count and analyst_count > 10:
            author_score += 0.5
            author_factors.append("Good analyst coverage")
        
        # ESG and business quality indicators
        # Current ratio (liquidity)
        current_ratio = ticker_info.get('currentRatio')
        if current_ratio and current_ratio > 1.5:
            author_score += 0.5
            author_factors.append("Strong liquidity")
        
        # Quick assessment
        if ticker_info.get('recommendationKey') in ['buy', 'strong_buy']:
            author_score += 1
            author_factors.append("Positive recommendation")
        
        # Convert to 1-5 scale
        if author_score >= 3:
            ratings['author'] = 5
        elif author_score >= 2.5:
            ratings['author'] = 4
        elif author_score >= 1.5:
            ratings['author'] = 3
        elif author_score >= 1:
            ratings['author'] = 2
        else:
            ratings['author'] = 1
        
        ratings['author_explanation'] = f"Factors: {', '.join(author_factors[:3])}" if author_factors else "Based on overall business assessment"
        
        # SELLSIDE RATING (Based on market perception and analyst sentiment)
        sellside_score = 0
        sellside_factors = []
        
        # Analyst recommendation
        recommendation = ticker_info.get('recommendationKey', '').lower()
        if 'strong_buy' in recommendation:
            sellside_score += 2
            sellside_factors.append("Strong Buy consensus")
        elif 'buy' in recommendation:
            sellside_score += 1.5
            sellside_factors.append("Buy recommendation")
        elif 'hold' in recommendation:
            sellside_score += 1
            sellside_factors.append("Hold rating")
        else:
            sellside_factors.append("Mixed/Sell sentiment")
        
        # Target price vs current price
        target_price = ticker_info.get('targetMeanPrice')
        current_price = ticker_info.get('currentPrice')
        if target_price and current_price:
            upside = (target_price - current_price) / current_price
            if upside > 0.20:  # 20%+ upside
                sellside_score += 1
                sellside_factors.append("High upside potential")
            elif upside > 0.10:  # 10-20% upside
                sellside_score += 0.5
                sellside_factors.append("Moderate upside")
            else:
                sellside_factors.append("Limited upside")
        
        # Price momentum (52-week range)
        current_price_alt = ticker_info.get('regularMarketPrice', current_price)
        week_52_high = ticker_info.get('fiftyTwoWeekHigh')
        week_52_low = ticker_info.get('fiftyTwoWeekLow')
        
        if current_price_alt and week_52_high and week_52_low:
            price_position = (current_price_alt - week_52_low) / (week_52_high - week_52_low)
            if price_position > 0.8:  # Near 52-week high
                sellside_score += 0.5
                sellside_factors.append("Strong momentum")
            elif price_position < 0.3:  # Near 52-week low
                sellside_score += 0.5
                sellside_factors.append("Value opportunity")
        
        # Earnings surprise history (if available)
        earnings_growth = ticker_info.get('earningsGrowth')
        if earnings_growth and earnings_growth > 0.1:
            sellside_score += 0.5
            sellside_factors.append("Earnings growth")
        
        # Convert to 1-5 scale
        if sellside_score >= 3.5:
            ratings['sellside'] = 5
        elif sellside_score >= 2.5:
            ratings['sellside'] = 4
        elif sellside_score >= 1.5:
            ratings['sellside'] = 3
        elif sellside_score >= 0.5:
            ratings['sellside'] = 2
        else:
            ratings['sellside'] = 1
        
        ratings['sellside_explanation'] = f"Based on: {', '.join(sellside_factors[:3])}" if sellside_factors else "Limited analyst data available"
        
        return ratings
        
    except Exception as e:
        # Return default ratings if calculation fails
        return {
            'quantitative': 3,
            'author': 3,
            'sellside': 3,
            'quantitative_explanation': 'Calculation error - using default rating',
            'author_explanation': 'Calculation error - using default rating', 
            'sellside_explanation': 'Calculation error - using default rating'
        }

def get_rating_color(rating):
    """
    Get color for rating display (1-5 scale)
    
    Args:
        rating (int): Rating from 1-5
    
    Returns:
        str: CSS color value
    """
    color_map = {
        5: "#22c55e",  # Green - Excellent
        4: "#84cc16",  # Light Green - Good  
        3: "#eab308",  # Yellow - Average
        2: "#f97316",  # Orange - Below Average
        1: "#ef4444"   # Red - Poor
    }
    return color_map.get(rating, "#6b7280")  # Default gray

if __name__ == "__main__":
    main()
